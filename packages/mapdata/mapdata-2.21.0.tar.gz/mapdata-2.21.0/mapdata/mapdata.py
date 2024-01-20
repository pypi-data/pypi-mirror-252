#!/usr/bin/python
#
# mapdata.py
#
# PURPOSE
#	Display a simple interactive map of data points, allowing points to
#	be highlighted by clicking on the map or table or by querying,
#	and allowing some simple data plots.
#
# COPYRIGHT AND LICENSE
#	Copyright (c) 2023-2024, R. Dreas Nielsen
# 	This program is free software: you can redistribute it and/or modify
# 	it under the terms of the GNU General Public License as published by
# 	the Free Software Foundation, either version 3 of the License, or
# 	(at your option) any later version.
# 	This program is distributed in the hope that it will be useful,
# 	but WITHOUT ANY WARRANTY; without even the implied warranty of
# 	MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# 	GNU General Public License for more details.
# 	The GNU General Public License is available at <http://www.gnu.org/licenses/>
#
# AUTHOR
#	Dreas Nielsen (RDN)
#
# ==================================================================

version = "2.21.0"
vdate = "2024-01-19"

copyright = "2023-2024"


import sys
import os.path
import io
import codecs
import argparse
from configparser import ConfigParser
import csv
import re
import datetime
import time
import math
import statistics
import jenkspy
import collections
import webbrowser
import threading
import queue
import sqlite3
import tempfile
import random
import uuid
import traceback
import subprocess
import multiprocessing
import copy
import tkinter as tk
import tkinter.ttk as ttk
import tkinter.font as tkfont
import tkinter.filedialog as tkfiledialog
import tkintermapview as tkmv
import textwrap
from PIL import ImageGrab
from PIL import ImageTk
import odf.opendocument
import odf.table
import odf.text
import odf.number
import odf.style
import xlrd
import openpyxl
import numpy as np
from numpy.polynomial import Polynomial
import matplotlib
matplotlib.use('TkAgg')
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
import seaborn as sns
from loess.loess_1d import loess_1d
import scipy.stats as spstats
import statsmodels.api as sm
from statsmodels.sandbox.stats.runs import runstest_1samp
from statsmodels.stats.diagnostic import normal_ad, kstest_normal


# Default name of configuration file.  Files with other names may be read.
config_file_name = "mapdata.conf"

# Configuration files read on startup
config_files = []
# Configuration file read post-startup
config_files_user = []


# Default options
multiselect = "0"
#-- Default location marker.  This may be overridden
location_marker = "triangle_open"
location_color = "black"
use_data_marker = True
use_data_color = True
#-- Selected item marker
select_symbol = "wedge"
select_color = "red"
#-- Label appearance
label_color = "black"
label_font = "Liberation Sans"
label_size = 10
label_bold = False
label_position = "below"	# above or below

# Plot configuration settings
show_regression_stats = False
wrapwidth = 20
wrap_at_underscores = False

# Name of editor, read from environment if it exists
editor = os.environ.get("EDITOR")

#-- Operational configuration
# Whether to use a temporary file for Sqlite (as opposed to memory).
temp_dbfile = False


#=====  Software patches

# Patch the tkintermapview CanvasPositionMarker 'calculate_text_y_offset()' function to
# allow labeling below the icon.  The icon anchor position is always "center" for this app.
def new_calc_text_offset(self):
	if self.icon is not None:
		if label_position == "below":
			self.text_y_offset = self.icon.height()/2 + 6 + label_size
		else:
			self.text_y_offset = -self.icon.height()/2 - 3
	else:
			self.text_y_offset = -56
tkmv.canvas_position_marker.CanvasPositionMarker.calculate_text_y_offset = new_calc_text_offset


# Patch function for ImageTk.PhotoImage.__del__ 
def new_img_del(img_self):
	try:
		name = img_self.__photo.name
		img_self.__photo.name = None
		img_self.__photo.tk.call("image", "delete", name)
	except Exception:
		pass



#=====  Global constants and variables =====

# Tile servers for map basemap layers
bm_servers = {"OpenStreetMap": "https://a.tile.openstreetmap.org/{z}/{x}/{y}.png",
			"Google streets": "https://mt0.google.com/vt/lyrs=m&hl=en&x={x}&y={y}&z={z}&s=Ga",
			"Google satellite": "https://mt0.google.com/vt/lyrs=s&hl=en&x={x}&y={y}&z={z}&s=Ga",
			"Open topo map": "https://tile.opentopomap.org/{z}/{x}/{y}.png",
			"Stamen terrain": "https://stamen-tiles.a.ssl.fastly.net/terrain/{z}/{x}/{y}.png"
			#, "Stamen toner": "https://stamen-tiles.a.ssl.fastly.net/toner/{z}/{x}/{y}.png"
			}

# API keys for tile servers that require them.  The dictionary keys should match basemap names.
api_keys = {}

# Initial basemap to use
#initial_basemap = tuple(bm_servers.keys())[0]
initial_basemap = "OpenStreetMap"

# List of initial basemap names, for use when saving config
initial_bms = list(bm_servers.keys())


# X11 bitmaps for map icons
icon_xbm = {
	'anchor': """#define anchor_width 16
#define anchor_height 16
static unsigned char anchor_bits[] = {
   0xc0, 0x03, 0x60, 0x06, 0x60, 0x06, 0xc0, 0x03, 0xfc, 0x3f, 0xfc, 0x3f,
   0x80, 0x01, 0x81, 0x81, 0x83, 0xc1, 0x87, 0xe1, 0x83, 0xc1, 0x87, 0xe1,
   0x8e, 0x71, 0xfc, 0x3f, 0xf0, 0x0f, 0xc0, 0x03};""",

	'ball': """#define ball_width 16
#define ball_height 16
static unsigned char circle_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0xf8, 0x1f, 0xfc, 0x3f, 0xfe, 0x7f, 0xfe, 0x7f,
   0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xfe, 0x7f, 0xfe, 0x7f,
   0xfc, 0x3f, 0xf8, 0x1f, 0xf0, 0x0f, 0xc0, 0x03};""",

   'binoculars': """#define binoculars_width 16
#define binoculars_height 16
static unsigned char binoculars_bits[] = {
   0x38, 0x1c, 0x38, 0x1c, 0x7c, 0x3e, 0x7c, 0x3e, 0xfc, 0x3f, 0xbc, 0x3d,
   0xbc, 0x3d, 0xfe, 0x7f, 0xfe, 0x7f, 0x7f, 0xfe, 0x7f, 0xfe, 0x3f, 0xfc,
   0x3f, 0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f, 0xfc};""",

	'bird': """#define bird_width 16
#define bird_height 16
static unsigned char bird.xbm_bits[] = {
   0x00, 0x00, 0x00, 0x1c, 0x00, 0x3f, 0x80, 0xef, 0xc0, 0x7f, 0xe0, 0x3f,
   0xf0, 0x3f, 0xf8, 0x1f, 0xff, 0x1f, 0xfc, 0x0f, 0xe0, 0x07, 0x80, 0x01,
   0x00, 0x01, 0x00, 0x01, 0x80, 0x03, 0xe0, 0x0f};""",

	'block': """#define block_width 16
#define block_height 16
static unsigned char square_bits[] = {
   0x00, 0x00, 0x00, 0x00, 0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f,
   0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f,
   0xfc, 0x3f, 0xfc, 0x3f, 0x00, 0x00, 0x00, 0x00};""",

	'bookmark': """#define bookmark_width 16
#define bookmark_height 16
static unsigned char bookmark_bits[] = {
   0xfe, 0x7f, 0xfe, 0x7f, 0xfe, 0x7f, 0xfe, 0x7f, 0xfe, 0x7f, 0xfe, 0x7f,
   0xfe, 0x7f, 0xfe, 0x7f, 0xfe, 0x7f, 0xfe, 0x7f, 0x7e, 0x7e, 0x3e, 0x7c,
   0x1e, 0x78, 0x0e, 0x70, 0x06, 0x60, 0x02, 0x40};""",

   'camera': """#define camera.xbm_width 16
#define camera.xbm_height 16
static unsigned char camera.xbm_bits[] = {
   0x00, 0x00, 0x00, 0x00, 0xf0, 0x07, 0xf6, 0x07, 0xfe, 0x7f, 0xff, 0xff,
   0x3f, 0xfe, 0x1f, 0xfc, 0xcf, 0xf9, 0xcf, 0xf9, 0xcf, 0xf9, 0x1f, 0xfc,
   0x3f, 0xfe, 0xff, 0xff, 0xfe, 0x7f, 0x00, 0x00};""",

   'cancel': """#define cancel_width 16
#define cancel_height 16
static unsigned char cancel_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0x78, 0x1e, 0x1c, 0x38, 0x0e, 0x7c, 0x06, 0x6e,
   0x07, 0xe7, 0x83, 0xc3, 0xc3, 0xc1, 0xe7, 0xe0, 0x76, 0x60, 0x3e, 0x78,
   0x1c, 0x38, 0x78, 0x3e, 0xf0, 0x0f, 0xc0, 0x03};""",

	'car': """#define car_width 16
#define car_height 16
static unsigned char car_bits[] = {
   0x00, 0x00, 0x00, 0x00, 0xf0, 0x00, 0xf8, 0x01, 0x2c, 0x03, 0x2c, 0x06,
   0x26, 0x0c, 0x26, 0x7c, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xc1, 0x83,
   0xdd, 0xbb, 0xdc, 0x3b, 0x1c, 0x38, 0x00, 0x00};""",

	'checkbox': """#define checkbox_width 16
#define checkbox_height 16
static unsigned char checkbox_bits[] = {
   0xff, 0xff, 0xff, 0xff, 0x03, 0xc0, 0x03, 0xd8, 0x03, 0xd8, 0x03, 0xdc,
   0x03, 0xcc, 0x1b, 0xce, 0x3b, 0xc6, 0x73, 0xc7, 0xe3, 0xc3, 0xc3, 0xc3,
   0x83, 0xc1, 0x03, 0xc0, 0xff, 0xff, 0xff, 0xff};""",

	'circle': """#define circle_width 16
#define circle_height 16
static unsigned char circle_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0x78, 0x1e, 0x1c, 0x38, 0x0e, 0x70, 0x06, 0x60,
   0x07, 0xe0, 0x03, 0xc0, 0x03, 0xc0, 0x07, 0xe0, 0x06, 0x60, 0x0e, 0x70,
   0x1c, 0x38, 0x78, 0x1e, 0xf0, 0x0f, 0xc0, 0x03};""",

	'circle_bar': """#define circle_bar_width 16
#define circle_bar_height 16
static unsigned char circle_bar_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0x78, 0x1e, 0x1c, 0x38, 0x0e, 0x70, 0x06, 0x60,
   0x07, 0xe0, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0x06, 0x60, 0x0e, 0x70,
   0x1c, 0x38, 0x78, 0x1e, 0xf0, 0x0f, 0xc0, 0x03};""",

	'circle_plus': """#define circle_plus_width 16
#define circle_plus_height 16
static unsigned char circle_plus_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0xf8, 0x1f, 0x9c, 0x39, 0x8e, 0x71, 0x86, 0x61,
   0x87, 0xe1, 0xff, 0xff, 0xff, 0xff, 0x87, 0xe1, 0x86, 0x61, 0x8e, 0x71,
   0x9c, 0x39, 0xf8, 0x1f, 0xf0, 0x0f, 0xc0, 0x03};""",

	'circle_x': """#define circle_x_width 16
#define circle_x_height 16
static unsigned char circle_x_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0x78, 0x1e, 0x1c, 0x38, 0x3e, 0x7c, 0x76, 0x6e,
   0xe7, 0xe7, 0xc3, 0xc3, 0xc3, 0xc3, 0xe7, 0xe7, 0x76, 0x6e, 0x3e, 0x7c,
   0x1c, 0x38, 0x78, 0x1e, 0xf0, 0x0f, 0xc0, 0x03};""",

	'clock': """#define clock_width 16
#define clock_height 16
static unsigned char clock_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0xf8, 0x1e, 0xfc, 0x3e, 0xfe, 0x7e, 0xfe, 0x7e,
   0xff, 0xfe, 0xff, 0xfe, 0x07, 0xfe, 0xff, 0xff, 0xfe, 0x7f, 0xfe, 0x7f,
   0xfc, 0x3f, 0xf8, 0x1f, 0xf0, 0x0f, 0xc0, 0x03};""",

	'deposition': """
#define deposition_width 16
#define deposition_height 16
static unsigned char deposition_bits[] = {
   0x55, 0xab, 0x00, 0x00, 0x6b, 0xdd, 0x00, 0x00, 0xf7, 0xee, 0x00, 0x00,
   0xbb, 0xbb, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff,
   0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff};""",

	'diamond': """#define diamond_width 16
#define diamond_height 16
static unsigned char diamond_bits[] = {
   0x80, 0x01, 0xc0, 0x03, 0xe0, 0x07, 0xf0, 0x0f, 0xf8, 0x1f, 0xfc, 0x3f,
   0xfe, 0x7f, 0xff, 0xff, 0xff, 0xff, 0xfe, 0x7f, 0xfc, 0x3f, 0xf8, 0x1f,
   0xf0, 0x0f, 0xe0, 0x07, 0xc0, 0x03, 0x80, 0x01};""",

	'donkey': """#define donkey.xbm_width 16
#define donkey.xbm_height 16
static unsigned char donkey.xbm_bits[] = {
   0x00, 0x00, 0x0c, 0x00, 0x08, 0x00, 0x1b, 0x00, 0x7e, 0x3c, 0xfe, 0x7f,
   0xfa, 0xff, 0xff, 0xbf, 0xff, 0x9f, 0xef, 0x9f, 0xe6, 0xbf, 0x60, 0x3c,
   0x70, 0x6c, 0x30, 0x68, 0x30, 0x48, 0x30, 0x6c};""",

   'drop': """,
#define drop.xbm_width 16
#define drop.xbm_height 16
static unsigned char drop.xbm_bits[] = {
   0x80, 0x01, 0x80, 0x01, 0xc0, 0x03, 0xc0, 0x03, 0xe0, 0x07, 0xe0, 0x07,
   0xf8, 0x1f, 0xf8, 0x1f, 0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f,
   0xfc, 0x3f, 0xf0, 0x1f, 0xf0, 0x0f, 0xc0, 0x03};""",

	'elephant': """#define elephant_width 16
#define elephant_height 16
static unsigned char elephant_bits[] = {
   0x00, 0x1c, 0x00, 0x1f, 0xfe, 0x3f, 0xfe, 0x7f, 0xff, 0xff, 0xfd, 0xef,
   0xfd, 0xff, 0xfd, 0xff, 0xfd, 0xff, 0xfd, 0xf7, 0xfc, 0xc7, 0x3c, 0x83,
   0x1c, 0xa3, 0x1c, 0xb3, 0x9e, 0xf3, 0xbe, 0x67};""",

	'eye': """#define eye_width 16
#define eye_height 16
static unsigned char eye_bits[] = {
   0x00, 0x00, 0x40, 0x02, 0x44, 0x22, 0x09, 0x90, 0xe2, 0x47, 0x38, 0x1c,
   0x8c, 0x31, 0xc6, 0x63, 0xc7, 0xe3, 0x86, 0x61, 0x0e, 0x70, 0x3c, 0x3c,
   0xf8, 0x1f, 0xe0, 0x07, 0x00, 0x00, 0x00, 0x00};""",

   'fish': """#define fish_width 16
#define fish_height 16
static unsigned char fish.xbm_bits[] = {
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x01, 0x00, 0x03, 0x01, 0x0f,
   0xe3, 0x7f, 0xf7, 0x78, 0x3e, 0xea, 0x9e, 0xfb, 0x73, 0x84, 0xc1, 0x63,
   0x01, 0x3e, 0x00, 0x18, 0x00, 0x18, 0x00, 0x08};""",

	'flag': """#define flag_width 16
#define flag_height 16
static unsigned char flag.xbm_bits[] = {
   0x00, 0x00, 0x0e, 0x00, 0x3e, 0x00, 0xfe, 0x01, 0xfe, 0x1f, 0xfe, 0xff,
   0xfe, 0xff, 0xfe, 0xff, 0xfe, 0xff, 0xfe, 0xff, 0xfe, 0xff, 0xf6, 0xff,
   0x86, 0xff, 0x06, 0xf8, 0x06, 0x00, 0x06, 0x00};""",

	'hand': """#define hand_width 16
#define hand_height 16
static unsigned char hand_bits[] = {
   0xc0, 0x00, 0xd8, 0x06, 0xd8, 0x06, 0xdb, 0x06, 0xdb, 0x06, 0xdb, 0x06,
   0xdb, 0x06, 0xdb, 0xc6, 0xff, 0xe7, 0xff, 0xf7, 0xff, 0x7f, 0xff, 0x3f,
   0xff, 0x3f, 0xff, 0x1f, 0xff, 0x0f, 0xfe, 0x0f};""",

	'heart': """#define heart_width 16
#define heart_height 16
static unsigned char heart_bits[] = {
   0x3c, 0x3c, 0x7e, 0x7e, 0x7f, 0xfe, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff,
   0xff, 0xff, 0xff, 0xff, 0xfe, 0x7f, 0xfe, 0x7f, 0xfc, 0x3f, 0xf8, 0x1f,
   0xf0, 0x0f, 0xe0, 0x07, 0xc0, 0x03, 0x80, 0x01};""",

	'hourglass': """#define hourglass_width 16
#define hourglass_height 16
static unsigned char hourglass_bits[] = {
   0xff, 0xff, 0xff, 0xff, 0x0c, 0x30, 0x0c, 0x30, 0x18, 0x18, 0xf0, 0x0f,
   0xe0, 0x07, 0xc0, 0x03, 0xc0, 0x03, 0x60, 0x06, 0x30, 0x0c, 0x98, 0x19,
   0xcc, 0x33, 0xec, 0x37, 0xff, 0xff, 0xff, 0xff};""",

	'house': """#define house_width 16
#define house_height 16
static unsigned char house_bits[] = {
   0x80, 0x01, 0xc0, 0x33, 0x60, 0x36, 0xb0, 0x3d, 0xd8, 0x3b, 0xec, 0x37,
   0xf6, 0x6f, 0xfb, 0xdf, 0xfd, 0xbf, 0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f,
   0x7c, 0x3e, 0x7c, 0x3e, 0x7c, 0x3e, 0x7c, 0x3e};""",

	'info': """#define info_width 16
#define info_height 16
static unsigned char info_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0x78, 0x1e, 0x7c, 0x3e, 0xfe, 0x7f, 0xfe, 0x7f,
   0x3f, 0xfe, 0x7f, 0xfe, 0x7f, 0xfe, 0x7f, 0xfe, 0x7e, 0x7e, 0x7e, 0x7e,
   0x3c, 0x3c, 0xf8, 0x1f, 0xf0, 0x0f, 0xc0, 0x03};""",

	'leaf': """#define leaf_width 16
#define leaf_height 16
static unsigned char leaf_bits[] = {
   0x00, 0x00, 0xc0, 0xff, 0xf0, 0xfb, 0xf8, 0xdb, 0xdc, 0xcb, 0xdc, 0xe3,
   0xdc, 0xf9, 0xdc, 0xfc, 0x5c, 0xc0, 0x1c, 0xff, 0x9c, 0x7f, 0xcc, 0x7f,
   0xec, 0x3f, 0xe6, 0x1f, 0x03, 0x00, 0x01, 0x00};""",

	'lightning': """#define lightning_width 16
#define lightning_height 16
static unsigned char Lightning_bits[] = {
   0x00, 0xc0, 0x00, 0x70, 0x00, 0x1c, 0x00, 0x07, 0x80, 0x03, 0xe0, 0x01,
   0xf0, 0x00, 0xf8, 0x03, 0xc0, 0x3f, 0x00, 0x1f, 0x80, 0x07, 0xc0, 0x01,
   0xe0, 0x00, 0x38, 0x00, 0x0e, 0x00, 0x03, 0x00};""",

	'mine': """#define mine_width 16
#define mine_height 16
static unsigned char mine_bits[] = {
   0xe0, 0xf1, 0x70, 0xf8, 0x3c, 0xfc, 0x1c, 0xf8, 0x3e, 0x7c, 0x77, 0x2e,
   0xe3, 0x07, 0xc1, 0x03, 0xc1, 0x03, 0xe0, 0x07, 0x70, 0x0e, 0x38, 0x1c,
   0x1c, 0x38, 0x0e, 0x70, 0x07, 0xe0, 0x03, 0xc0};""",

	'pennant': """#define pennant2_width 16
#define pennant2_height 16
static unsigned char pennant2_bits[] = {
   0x0e, 0x00, 0x3e, 0x00, 0xfe, 0x00, 0xfe, 0x03, 0xfe, 0x0f, 0xfe, 0x3f,
   0xfe, 0xff, 0xfe, 0x3f, 0xfe, 0x07, 0xfe, 0x00, 0x1e, 0x00, 0x06, 0x00,
   0x06, 0x00, 0x06, 0x00, 0x06, 0x00, 0x06, 0x00};""",

	'photo': """#define photo_width 16
#define photo_height 16
static unsigned char photo_bits[] = {
   0xff, 0xff, 0x01, 0x80, 0x01, 0x80, 0x01, 0x98, 0x01, 0x98, 0x01, 0x80,
   0x11, 0x80, 0x39, 0x80, 0x7d, 0x84, 0xfd, 0x8e, 0xfd, 0x9f, 0xfd, 0xbf,
   0xfd, 0xbf, 0xfd, 0xbf, 0x01, 0x80, 0xff, 0xff};""",

	'picnic': """#define picnic_width 16
#define picnic_height 16
static unsigned char picnic_bits[] = {
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0xfe, 0x7f, 0xfe, 0x7f,
   0x70, 0x0e, 0x30, 0x0c, 0x38, 0x1c, 0x38, 0x1c, 0xff, 0xff, 0xff, 0xff,
   0x06, 0x60, 0x06, 0x60, 0x03, 0xc0, 0x03, 0xc0};""",

	'plus': """#define plus_width 16
#define plus_height 16
static unsigned char plus_bits[] = {
   0x80, 0x01, 0x80, 0x01, 0x80, 0x01, 0x80, 0x01, 0x80, 0x01, 0x80, 0x01,
   0x80, 0x01, 0xff, 0xff, 0xff, 0xff, 0x80, 0x01, 0x80, 0x01, 0x80, 0x01,
   0x80, 0x01, 0x80, 0x01, 0x80, 0x01, 0x80, 0x01};""",

	'qmark_circle': """#define qmark_circle_width 16
#define qmark_circle_height 16
static unsigned char qmark_circle_bits[] = {
   0xc0, 0x03, 0x70, 0x0e, 0x18, 0x18, 0xcc, 0x33, 0xe6, 0x67, 0x72, 0x4e,
   0x33, 0xcc, 0x01, 0x87, 0x81, 0x83, 0x83, 0xc1, 0x02, 0x40, 0x86, 0x61,
   0x8c, 0x31, 0x38, 0x18, 0x70, 0x0e, 0xc0, 0x03};""",

	'raincloud': """#define raincloud_width 16
#define raincloud_height 16
static unsigned char raincloud_bits[] = {
   0x00, 0x00, 0xe0, 0x03, 0xf0, 0x07, 0xf8, 0x0f, 0xf8, 0x1f, 0xfe, 0x1f,
   0xff, 0x7f, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xfe, 0xff, 0xfc, 0x7f,
   0x98, 0x19, 0x88, 0x08, 0xcc, 0x0c, 0x44, 0x04};""",

	'rose': """#define rose_width 16
#define rose_height 16
static unsigned char rose_bits[] = {
   0x80, 0x01, 0x80, 0x01, 0xc0, 0x03, 0xc0, 0x03, 0xc0, 0x03, 0xe0, 0x07,
   0xfc, 0x3f, 0xff, 0xff, 0xff, 0xff, 0xfc, 0x3f, 0xe0, 0x07, 0xc0, 0x03,
   0xc0, 0x03, 0xc0, 0x03, 0x80, 0x01, 0x80, 0x01};""",

	'square': """#define square_width 16
#define square_height 16
static unsigned char square_bits[] = {
   0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0x07, 0xe0, 0x07, 0xe0, 0x07, 0xe0,
   0x07, 0xe0, 0x07, 0xe0, 0x07, 0xe0, 0x07, 0xe0, 0x07, 0xe0, 0x07, 0xe0,
   0x07, 0xe0, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff};""",

	'star': """#define star_width 16
#define star_height 16
static unsigned char star_bits[] = {
   0x80, 0x01, 0x80, 0x01, 0xc0, 0x03, 0xc0, 0x03, 0xc0, 0x03, 0xe0, 0x07,
   0xff, 0xff, 0xff, 0xff, 0xfc, 0x3f, 0xf0, 0x0f, 0xf8, 0x1f, 0xf8, 0x1f,
   0x7c, 0x3e, 0x3c, 0x3c, 0x0e, 0x70, 0x06, 0x60};""",

	'surprise_circle': """#define surprise_circle_width 16
#define surprise_circle_height 16
static unsigned char surprise_circle_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0x78, 0x1e, 0x9c, 0x39, 0xce, 0x73, 0xc6, 0x63,
   0xc7, 0xe3, 0xc3, 0xc3, 0x83, 0xc1, 0x87, 0xe1, 0x06, 0x60, 0x8e, 0x71,
   0x9c, 0x39, 0x78, 0x1e, 0xf0, 0x0f, 0xc0, 0x03};""",

	'target': """#define target_width 16
#define target_height 16
static unsigned char target_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0x78, 0x1e, 0x1c, 0x38, 0x0e, 0x70, 0x86, 0x61,
   0xc7, 0xe3, 0xe3, 0xc7, 0xe3, 0xc7, 0xc7, 0xe3, 0x86, 0x61, 0x0e, 0x70,
   0x1c, 0x38, 0x78, 0x1e, 0xf0, 0x0f, 0xc0, 0x03};""",

	'trash': """#define trash_width 16
#define trash_height 16
static unsigned char trash_bits[] = {
   0xf0, 0x0f, 0xff, 0xff, 0xff, 0xff, 0x06, 0x60, 0x66, 0x66, 0x66, 0x66,
   0x66, 0x66, 0x66, 0x66, 0x66, 0x66, 0x66, 0x66, 0x66, 0x66, 0x66, 0x66,
   0x66, 0x66, 0x06, 0x60, 0xfe, 0x7f, 0xfc, 0x3f};""",

	'tree': """
#define tree_width 16
#define tree_height 16
static unsigned char tree_bits[] = {
   0xf8, 0x00, 0xa8, 0x37, 0x7c, 0x7d, 0xef, 0x4a, 0x37, 0xf5, 0xdf, 0xaf,
   0xbe, 0xdb, 0xfc, 0x7f, 0xb0, 0x77, 0xc0, 0x7b, 0xc0, 0x1f, 0xc0, 0x03,
   0xc0, 0x07, 0xc0, 0x07, 0xe0, 0x0f, 0xfc, 0x0f};""",

	'triangle': """#define triangle_width 16
#define triangle_height 16
static unsigned char triangle_bits[] = {
   0x80, 0x01, 0x80, 0x01, 0xc0, 0x03, 0xc0, 0x03, 0xe0, 0x07, 0xe0, 0x07,
   0xf0, 0x0f, 0xf0, 0x0f, 0xf8, 0x1f, 0xf8, 0x1f, 0xfc, 0x3f, 0xfc, 0x3f,
   0xfe, 0x7f, 0xfe, 0x7f, 0xff, 0xff, 0xff, 0xff};""",

	'triangle_open': """#define triangle_open_width 16
#define triangle_open_height 16
static unsigned char triangle_open_bits[] = {
   0x80, 0x01, 0x80, 0x01, 0xc0, 0x03, 0xc0, 0x03, 0xe0, 0x07, 0xe0, 0x07,
   0x70, 0x0e, 0x70, 0x0e, 0x38, 0x1c, 0x38, 0x1c, 0x1c, 0x38, 0x1c, 0x38,
   0x0e, 0x70, 0xfe, 0x7f, 0xff, 0xff, 0xff, 0xff};""",

	'vapor': """
#define vapor_width 16
#define vapor_height 16
static unsigned char vapor_bits[] = {
   0x20, 0x01, 0x30, 0x03, 0xcc, 0x4c, 0xc4, 0xcc, 0x13, 0x32, 0x31, 0x23,
   0xc8, 0x8c, 0x4c, 0xc4, 0x11, 0x22, 0x33, 0x33, 0xc4, 0xc4, 0x4c, 0x4c,
   0x32, 0x33, 0x20, 0x23, 0xc8, 0x0c, 0x80, 0x04};""",

	'wave': """#define wave_width 16
#define wave_height 16
static unsigned char wave_bits[] = {
   0x00, 0x00, 0x70, 0x00, 0xf8, 0x00, 0xce, 0x00, 0x83, 0x01, 0x00, 0xc3,
   0x00, 0xe6, 0x70, 0x3e, 0xf8, 0x1c, 0xce, 0x00, 0x83, 0x01, 0x00, 0xc3,
   0x00, 0xe6, 0x00, 0x3e, 0x00, 0x1c, 0x00, 0x00};""",

	'wedge': """#define wedge_width 16
#define wedge_height 16
static unsigned char stn_marker_inv_bits[] = {
   0xff, 0xff, 0xff, 0x7f, 0xfe, 0x7f, 0xfe, 0x3f, 0xfc, 0x3f, 0xfc, 0x1f,
   0xf8, 0x1f, 0xf8, 0x0f, 0xf0, 0x0f, 0xf0, 0x07, 0xe0, 0x07, 0xe0, 0x03,
   0xc0, 0x03, 0xc0, 0x01, 0x80, 0x01, 0x80, 0x00};""",

	'wheelchair': """#define wheelchair_width 16
#define wheelchair_height 16
static unsigned char wheelchair_bits[] = {
   0x30, 0x00, 0x78, 0x00, 0x78, 0x00, 0x78, 0x00, 0x30, 0x00, 0x77, 0x1f,
   0x76, 0x1f, 0xf6, 0x06, 0xf3, 0x07, 0xf3, 0x3f, 0x03, 0x3f, 0x03, 0x36,
   0x07, 0x77, 0x06, 0x63, 0xfe, 0xe3, 0xf8, 0xe0};""",

	'x': """#define x_width 16
#define x_height 16
static unsigned char x_bits[] = {
   0x00, 0x00, 0x06, 0x60, 0x0e, 0x70, 0x1c, 0x38, 0x38, 0x1c, 0x70, 0x0e,
   0xe0, 0x07, 0xc0, 0x03, 0xc0, 0x03, 0xe0, 0x07, 0x70, 0x0e, 0x38, 0x1c,
   0x1c, 0x38, 0x0e, 0x70, 0x06, 0x60, 0x00, 0x00};""",


	'ball20': """#define ball20_width 20
#define ball20_height 20
static unsigned char ball20_bits[] = {
   0x00, 0x00, 0xf0, 0x00, 0x00, 0xf0, 0x00, 0x0f, 0xf0, 0xc0, 0x3f, 0xf0,
   0xe0, 0x7f, 0xf0, 0xf0, 0xff, 0xf0, 0xf8, 0xff, 0xf1, 0xf8, 0xff, 0xf1,
   0xfc, 0xff, 0xf3, 0xfc, 0xff, 0xf3, 0xfc, 0xff, 0xf3, 0xfc, 0xff, 0xf3,
   0xf8, 0xff, 0xf1, 0xf8, 0xff, 0xf1, 0xf0, 0xff, 0xf0, 0xe0, 0x7f, 0xf0,
   0xc0, 0x3f, 0xf0, 0x00, 0x0f, 0xf0, 0x00, 0x00, 0xf0, 0x00, 0x00, 0xf0};""",

	'block20': """#define block20_width 20
#define block20_height 20
static unsigned char block20_bits[] = {
   0x00, 0x00, 0xf0, 0x00, 0x00, 0xf0, 0xfc, 0xff, 0xf3, 0xfc, 0xff, 0xf3,
   0xfc, 0xff, 0xf3, 0xfc, 0xff, 0xf3, 0xfc, 0xff, 0xf3, 0xfc, 0xff, 0xf3,
   0xfc, 0xff, 0xf3, 0xfc, 0xff, 0xf3, 0xfc, 0xff, 0xf3, 0xfc, 0xff, 0xf3,
   0xfc, 0xff, 0xf3, 0xfc, 0xff, 0xf3, 0xfc, 0xff, 0xf3, 0xfc, 0xff, 0xf3,
   0xfc, 0xff, 0xf3, 0xfc, 0xff, 0xf3, 0x00, 0x00, 0xf0, 0x00, 0x00, 0xf0};""",

	'circle20': """#define circle20_width 20
#define circle20_height 20
static unsigned char circle20_bits[] = {
   0xc0, 0x3f, 0x00, 0xf0, 0xff, 0x00, 0xf8, 0xf0, 0x01, 0x3c, 0xc0, 0x03,
   0x1e, 0x80, 0x07, 0x0e, 0x00, 0x07, 0x07, 0x00, 0x0e, 0x07, 0x00, 0x0e,
   0x03, 0x00, 0x0c, 0x03, 0x00, 0x0c, 0x03, 0x00, 0x0c, 0x03, 0x00, 0x0c,
   0x07, 0x00, 0x0e, 0x07, 0x00, 0x0e, 0x0e, 0x00, 0x07, 0x1e, 0x80, 0x07,
   0x3c, 0xc0, 0x03, 0xf8, 0xf0, 0x01, 0xf0, 0xff, 0x00, 0xc0, 0x3f, 0x00};""",

	'square20': """#define square20_width 20
#define square20_height 20
static unsigned char square20_bits[] = {
   0xff, 0xff, 0x0f, 0xff, 0xff, 0x0f, 0xff, 0xff, 0x0f, 0x07, 0x00, 0x0e,
   0x07, 0x00, 0x0e, 0x07, 0x00, 0x0e, 0x07, 0x00, 0x0e, 0x07, 0x00, 0x0e,
   0x07, 0x00, 0x0e, 0x07, 0x00, 0x0e, 0x07, 0x00, 0x0e, 0x07, 0x00, 0x0e,
   0x07, 0x00, 0x0e, 0x07, 0x00, 0x0e, 0x07, 0x00, 0x0e, 0x07, 0x00, 0x0e,
   0x07, 0x00, 0x0e, 0xff, 0xff, 0x0f, 0xff, 0xff, 0x0f, 0xff, 0xff, 0x0f};""",


	'ball24': """#define ball24_width 24
#define ball24_height 24
static unsigned char ball24_bits[] = {
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x3c, 0x00, 0x80, 0xff, 0x01,
   0xc0, 0xff, 0x03, 0xe0, 0xff, 0x07, 0xf0, 0xff, 0x0f, 0xf8, 0xff, 0x1f,
   0xf8, 0xff, 0x1f, 0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f,
   0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f, 0xf8, 0xff, 0x1f, 0xf8, 0xff, 0x1f,
   0xf0, 0xff, 0x1f, 0xf0, 0xff, 0x0f, 0xe0, 0xff, 0x07, 0xc0, 0xff, 0x03,
   0x80, 0xff, 0x01, 0x00, 0x3c, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00};""",

	'block24': """#define block24_width 24
#define block24_height 24
static unsigned char block24_bits[] = {
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f,
   0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f,
   0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f,
   0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f,
   0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f,
   0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00};""",

	'circle24': """#define circle24_width 24
#define circle24_height 24
static unsigned char circle24_bits[] = {
   0x00, 0x7e, 0x00, 0xc0, 0xff, 0x03, 0xe0, 0xc3, 0x07, 0x70, 0x00, 0x0e,
   0x38, 0x00, 0x1c, 0x1c, 0x00, 0x38, 0x0e, 0x00, 0x70, 0x06, 0x00, 0x60,
   0x06, 0x00, 0x60, 0x03, 0x00, 0xc0, 0x03, 0x00, 0xc0, 0x03, 0x00, 0xc0,
   0x03, 0x00, 0xc0, 0x03, 0x00, 0xc0, 0x07, 0x00, 0xe0, 0x06, 0x00, 0x60,
   0x0e, 0x00, 0x60, 0x0e, 0x00, 0x70, 0x1c, 0x00, 0x38, 0x38, 0x00, 0x1c,
   0x70, 0x00, 0x0e, 0xe0, 0xc3, 0x07, 0xc0, 0xff, 0x03, 0x00, 0x7e, 0x00};""",

	'square24': """#define square24_width 24
#define square24_height 24
static unsigned char square24_bits[] = {
   0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0x07, 0x00, 0xe0,
   0x07, 0x00, 0xe0, 0x07, 0x00, 0xe0, 0x07, 0x00, 0xe0, 0x07, 0x00, 0xe0,
   0x07, 0x00, 0xe0, 0x07, 0x00, 0xe0, 0x07, 0x00, 0xe0, 0x07, 0x00, 0xe0,
   0x07, 0x00, 0xe0, 0x07, 0x00, 0xe0, 0x07, 0x00, 0xe0, 0x07, 0x00, 0xe0,
   0x07, 0x00, 0xe0, 0x07, 0x00, 0xe0, 0x07, 0x00, 0xe0, 0x07, 0x00, 0xe0,
   0x07, 0x00, 0xe0, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff};""",


	'ball28': """#define ball28_width 28
#define ball28_height 28
static unsigned char circle28_bits[] = {
   0x00, 0x00, 0x00, 0xf0, 0x00, 0x00, 0x00, 0xf0, 0x00, 0x00, 0x00, 0xf0,
   0x00, 0xfc, 0x03, 0xf0, 0x80, 0xff, 0x1f, 0xf0, 0xc0, 0xff, 0x3f, 0xf0,
   0xe0, 0xff, 0x7f, 0xf0, 0xf0, 0xff, 0xff, 0xf0, 0xf0, 0xff, 0xff, 0xf0,
   0xf0, 0xff, 0xff, 0xf0, 0xf8, 0xff, 0xff, 0xf1, 0xf8, 0xff, 0xff, 0xf1,
   0xf8, 0xff, 0xff, 0xf1, 0xf8, 0xff, 0xff, 0xf1, 0xf8, 0xff, 0xff, 0xf1,
   0xf8, 0xff, 0xff, 0xf1, 0xf8, 0xff, 0xff, 0xf1, 0xf0, 0xff, 0xff, 0xf0,
   0xf0, 0xff, 0xff, 0xf0, 0xf0, 0xff, 0xff, 0xf0, 0xe0, 0xff, 0x7f, 0xf0,
   0xe0, 0xff, 0x7f, 0xf0, 0xc0, 0xff, 0x3f, 0xf0, 0x80, 0xff, 0x1f, 0xf0,
   0x00, 0xfc, 0x03, 0xf0, 0x00, 0x00, 0x00, 0xf0, 0x00, 0x00, 0x00, 0xf0,
   0x00, 0x00, 0x00, 0xf0};""",

	'block28': """#define block28_width 28
#define block28_height 28
static unsigned char block28_bits[] = {
   0x00, 0x00, 0x00, 0xf0, 0x00, 0x00, 0x00, 0xf0, 0xfc, 0xff, 0xff, 0xf3,
   0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3,
   0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3,
   0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3,
   0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3,
   0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3,
   0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3,
   0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3,
   0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3, 0x00, 0x00, 0x00, 0xf0,
   0x00, 0x00, 0x00, 0xf0};""",

	'circle28': """#define circle28_width 28
#define circle28_height 28
static unsigned char circle24_bits[] = {
   0x00, 0xfe, 0x07, 0x00, 0x80, 0xff, 0x1f, 0x00, 0xe0, 0xff, 0x7f, 0x00,
   0xf0, 0x03, 0xfc, 0x00, 0x78, 0x00, 0xe0, 0x01, 0x3c, 0x00, 0xc0, 0x03,
   0x1c, 0x00, 0x80, 0x03, 0x0e, 0x00, 0x00, 0x07, 0x0e, 0x00, 0x00, 0x07,
   0x0f, 0x00, 0x00, 0x0f, 0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e,
   0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e,
   0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e, 0x0f, 0x00, 0x00, 0x0f,
   0x0f, 0x00, 0x00, 0x0f, 0x0e, 0x00, 0x00, 0x07, 0x1e, 0x00, 0x80, 0x07,
   0x1c, 0x00, 0x80, 0x03, 0x3c, 0x00, 0xc0, 0x03, 0x78, 0x00, 0xe0, 0x01,
   0xf0, 0x03, 0xfc, 0x00, 0xe0, 0xff, 0x7f, 0x00, 0x80, 0xff, 0x1f, 0x00,
   0x00, 0xfe, 0x07, 0x00};""",

	'square28': """#define square28_width 28
#define square28_height 28
static unsigned char square28_bits[] = {
   0xff, 0xff, 0xff, 0x0f, 0xff, 0xff, 0xff, 0x0f, 0xff, 0xff, 0xff, 0x0f,
   0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e,
   0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e,
   0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e,
   0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e,
   0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e,
   0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e,
   0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e,
   0x07, 0x00, 0x00, 0x0e, 0xff, 0xff, 0xff, 0x0f, 0xff, 0xff, 0xff, 0x0f,
   0xff, 0xff, 0xff, 0x0f};"""

	}

# X11 bitmaps for map button bar icons
expand_xbm = """#define expand_width 16
#define expand_height 16
static unsigned char expand_bits[] = {
   0x3f, 0xfc, 0x07, 0xe0, 0x0f, 0xf0, 0x1d, 0xb8, 0x39, 0x9c, 0x71, 0x8e,
   0x60, 0x06, 0x00, 0x00, 0x00, 0x00, 0x61, 0x06, 0x71, 0x8e, 0x39, 0x9c,
   0x1d, 0xb8, 0x0f, 0xf0, 0x07, 0xe0, 0x3f, 0xfc};"""

wedges_3_xbm = """#define wedges_3_width 16
#define wedges_3_height 16
static unsigned char wedges_3_bits[] = {
   0xff, 0x01, 0xfe, 0x00, 0x7c, 0x00, 0x38, 0x00, 0x10, 0x00, 0x00, 0x00,
   0x80, 0xff, 0x00, 0x7f, 0x00, 0x3e, 0x00, 0x1c, 0x00, 0x08, 0xff, 0x01,
   0xfe, 0x00, 0x7c, 0x00, 0x38, 0x00, 0x10, 0x00};"""

wedge_sm_xbm = """#define wedge_sm_width 16
#define wedge_sm_height 16
static unsigned char wedge_sm_bits[] = {
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0xf8, 0x1f, 0xf8, 0x1f, 0xf0, 0x0f,
   0xf0, 0x0f, 0xe0, 0x07, 0xe0, 0x07, 0xc0, 0x03, 0xc0, 0x03, 0x80, 0x01,
   0x80, 0x01, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00};"""

circle_xbm = """#define circle_width 16
#define circle_height 16
static unsigned char circle_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0x78, 0x1e, 0x1c, 0x38, 0x0e, 0x70, 0x06, 0x60,
   0x07, 0xe0, 0x03, 0xc0, 0x03, 0xc0, 0x07, 0xe0, 0x06, 0x60, 0x0e, 0x70,
   0x1c, 0x38, 0x78, 0x1e, 0xf0, 0x0f, 0xc0, 0x03};"""

cancel_xbm = """#define cancel_width 16
#define cancel_height 16
static unsigned char cancel_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0x78, 0x1e, 0x1c, 0x38, 0x0e, 0x7c, 0x06, 0x6e,
   0x07, 0xe7, 0x83, 0xc3, 0xc3, 0xc1, 0xe7, 0xe0, 0x76, 0x60, 0x3e, 0x78,
   0x1c, 0x38, 0x78, 0x3e, 0xf0, 0x0f, 0xc0, 0x03};"""


# Color names for map symbols.  See https://www.w3schools.com/colors/colors_names.asp.
color_names = ("aliceblue", "antiquewhite", "aqua", "aquamarine", "azure", "beige", "bisque", "black", "blanchedalmond",
		"blue", "blueviolet", "brown", "burlywood", "cadetblue", "chartreuse", "chocolate", "coral", "cornflowerblue",
		"cornsilk", "crimson", "cyan", "darkblue", "darkcyan", "darkgoldenrod", "darkgray", "darkgrey", "darkgreen",
		"darkkhaki", "darkmagenta", "darkolivegreen", "darkorange", "darkorchid", "darkred", "darksalmon", "darkseagreen",
		"darkslateblue", "darkslategray", "darkslategrey", "darkturquose", "darkviolet", "deeppink", "deepskyblue",
		"dimgray", "dimgrey", "dodgerblue", "firebrick", "floralwhite", "forestgreen", "fuschia", "gainsboro", "ghostwhite",
		"gold", "goldenrod", "gray", "grey", "green", "greenyellow", "honeydew", "hotpink", "indianred", "indigo", "ivory",
		"khaki", "lavender", "lavenderblush", "lawngreen", "lemonchiffon", "lightblue", "lightcoral", "lightcyan",
		"lightgoldenrodyellow", "lightgray", "lightgrey", "lightgreen", "lightpink", "lightsalmon", "lightseagreen",
		"lightskyblue", "lightslategray", "lightslategrey", "lightsteelblue", "lightyellow", "lime", "limegreen", "linen",
		"magenta", "maroon", "mediumaquamarine", "mediumblue", "mediumorchid", "mediumpurple", "mediumseagreen",
		"mediumslateblue", "mediumspringgreen", "mediumturquose", "mediumvioletred", "midnightblue", "mintcream", "mistyrose",
		"moccasin", "navajowhite", "navy", "oldlace", "olive", "olivedrab", "orange", "orangered", "orchid", "palegoldenrod",
		"palegreen", "paleturquose", "palevioletred", "papayawhip", "peachpuff", "peru", "pink", "plum", "powderblue",
		"purple", "rebeccapurple", "red", "rosybrown", "royalblue", "saddlebrown", "salmon", "sandybrown", "seagreen",
		"seashell", "sienna", "silver", "skyblue", "slateblue", "slategray", "slategrey", "snow", "springgreen",
		"steelblue", "tan", "teal", "thistle", "tomato", "turquoise", "violet", "wheat", "white", "whitesmoke", "yellow",
		"yellowgreen")

# A shorter list for interactive selection of the marker color
select_colors = ('aqua', 'black', 'blue', 'blueviolet', 'brown', 'chartreuse', 'cornflowerblue', 'crimson',
		'cyan', 'darkblue', 'darkgreen', 'darkmagenta', 'darkorange', 'darkred', 'darkslategray', 'deeppink',
		'forestgreen', 'fuschia', 'green', 'greenyellow', 'magenta', 'maroon', 'navy', 'orange', 'orangered',
		'purple', 'red', 'violet', 'white', 'yellow', 'yellowgreen')


# List of imported symbol names and paths
imported_symbols = []

# Keys for custom symbols are made up of the color name and the icon name, separated with a space.
custom_icons = {}


# X11 bitmap for the application window icon
win_icon_xbm = """#define window_icon_width 16
#define window_icon_height 16
static unsigned char window_icon_bits[] = {
   0xff, 0xff, 0x01, 0x80, 0x01, 0x84, 0x01, 0x8e, 0x01, 0x9f, 0x81, 0xbf,
   0x21, 0x80, 0x71, 0x80, 0xf9, 0x80, 0xfd, 0x81, 0x01, 0x84, 0x01, 0x8e,
   0x01, 0x9f, 0x81, 0xbf, 0x01, 0x80, 0xff, 0xff};"""



#=====  Functions and classes  =====

def warning(message, kwargs):
	dlg = MsgDialog("Warning", message, parent=kwargs.get('parent'), bgcolor="Gold")
	dlg.show()

def fatal_error(message, kwargs):
	dlg = MsgDialog("Fatal Error", message, parent=kwargs.get('parent'), bgcolor="Red")
	dlg.show()
	sys.exit()


class CsvFile(object):
	def __init__(self, csvfname, junk_header_lines=0, dialect=None):
		self.csvfname = csvfname
		self.junk_header_lines = junk_header_lines
		self.lineformat_set = False		# Indicates whether delimiter, quotechar, and escapechar have been set
		self.delimiter = None
		self.quotechar = None
		self.escapechar = None
		self.inf = None
		self.colnames = None
		self.rows_read = 0
		# Python 3 only
		self.reader = csv.reader(open(csvfname, mode="rt", newline=''), dialect=dialect)
	def __next__(self):
		row = next(self.reader)
		self.rows_read = self.rows_read + 1
		return row
	def next(self):
		row = next(self.reader)
		self.rows_read = self.rows_read + 1
		if self.rows_read == 1:
			self.colnames = row
		return row
	def __iter__(self):
		return self


def sort_columns(columns, sortby=0):
	# Sorts a list of one, two, or three sublists, where each sublist is a column.
	# Rows are sorted by the 'sortby' column, which is zero-based.
	# The returned value is also column-wise, but sorted by rows.
	if len(columns) == 1:
		return sorted(columns)
	nrows = len(columns[0])
	if len(columns) == 2:
		rowdata = [[columns[0][i], columns[1][i]] for i in range(nrows)]
		rowdata.sort(key = lambda c: c[sortby])
		return [[rowdata[i][0] for i in range(nrows)], [rowdata[i][1] for i in range(nrows)]]
	else:
		# len(columns) should be 3, though this is not checked
		rowdata = [[columns[0][i], columns[1][i], columns[2][i]] for i in range(nrows)]
		rowdata.sort(key = lambda c: c[sortby])
		return [[rowdata[i][0] for i in range(nrows)], [rowdata[i][1] for i in range(nrows)], [rowdata[i][2] for i in range(nrows)]]


def treeview_sort_column(tv, col, reverse):
	# Sort columns in Tkinter Treeview.  From https://stackoverflow.com/questions/1966929/tk-treeview-column-sort#1967793
    colvals = [(tv.set(k, col), k) for k in tv.get_children()]
    colvals.sort(reverse=reverse)
    # Rearrange items in sorted positions
    for index, (val, k) in enumerate(colvals):
        tv.move(k, '', index)
    # Reverse sort next time
    tv.heading(col, command=lambda: treeview_sort_column(tv, col, not reverse))

def set_tv_headers(tvtable, column_headers, colwidths, charpixels):
	pixwidths = [charpixels * col for col in colwidths]
	for i in range(len(column_headers)):
		hdr = column_headers[i]
		tvtable.column(hdr, width=pixwidths[i])
		tvtable.heading(hdr, text=hdr, command=lambda _col=hdr: treeview_sort_column(tvtable, _col, False))

def fill_tv_table(tvtable, rowset, status_label=None):
	for i, row in enumerate(rowset):
		enc_row = [c if c is not None else '' for c in row]
		tvtable.insert(parent='', index='end', iid=str(i), values=enc_row)
	if status_label is not None:
		status_label.config(text = "    %d rows" % len(rowset))

def treeview_table(parent, rowset, column_headers, select_mode="none", nrows=None):
	# Creates a TreeView table containing the specified data, with scrollbars and status bar
	# in an enclosing frame.
	# This does not grid the table frame in its parent widget.
	# Returns a tuple of 0: the frame containing the table,  and 1: the table widget itself.
	nrows = range(len(rowset))
	ncols = range(len(column_headers))
	hdrwidths = [len(column_headers[j]) for j in ncols]
	if len(rowset) > 0:
		if sys.version_info < (3,):
			datawidthtbl = [[len(rowset[i][j] if isinstance(rowset[i][j], str) else type(u"")(rowset[i][j])) for i in nrows] for j in ncols]
		else:
			datawidthtbl = [[len(rowset[i][j] if isinstance(rowset[i][j], str) else str(rowset[i][j])) for i in nrows] for j in ncols]
		datawidths = [max(cwidths) for cwidths in datawidthtbl]
	else:
		datawidths = hdrwidths
	colwidths = [max(hdrwidths[i], datawidths[i]) for i in ncols]
	# Set the font.
	ff = tkfont.nametofont("TkFixedFont")
	tblstyle = ttk.Style()
	tblstyle.configure('tblstyle', font=ff)
	charpixels = int(1.3 * ff.measure(u"0"))
	tableframe = ttk.Frame(master=parent, padding="3 3 3 3")
	statusframe = ttk.Frame(master=tableframe)
	# Create and configure the Treeview table widget
	tv_widget = ttk.Treeview(tableframe, columns=column_headers, selectmode=select_mode, show="headings")
	tv_widget.configure()["style"] = tblstyle
	if nrows is not None:
		tv_widget.configure()["height"] = nrows
	ysb = ttk.Scrollbar(tableframe, orient='vertical', command=tv_widget.yview)
	xsb = ttk.Scrollbar(tableframe, orient='horizontal', command=tv_widget.xview)
	tv_widget.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)
	# Status bar
	parent.statusbar = ttk.Label(statusframe, text="    %d rows" % len(rowset), relief=tk.RIDGE, anchor=tk.W)
	tableframe.statuslabel = parent.statusbar
	# Fill the Treeview table widget with data
	set_tv_headers(tv_widget, column_headers, colwidths, charpixels)
	fill_tv_table(tv_widget, rowset, parent.statusbar)
	# Place the table
	tv_widget.grid(column=0, row=0, sticky=tk.NSEW)
	ysb.grid(column=1, row=0, sticky=tk.NS)
	xsb.grid(column=0, row=1, sticky=tk.EW)
	statusframe.grid(column=0, row=3, sticky=tk.EW)
	tableframe.columnconfigure(0, weight=1)
	tableframe.rowconfigure(0, weight=1)
	# Place the status bar
	parent.statusbar.pack(side=tk.BOTTOM, fill=tk.X)
	#
	return tableframe, tv_widget

def export_data_table(headers, rows, sheetname="mapdata_export"):
	# Exports the selected data to a CSV or ODS file.  Returns the file name or None if canceled.
	outfile = tkfiledialog.asksaveasfilename(title="File name for saved data rows",
		filetypes=[('CSV files', '.csv'), ('ODS files', '.ods'), ('TSV files', '.tsv'), ('Plain text', '.txt'), ('LaTeX', '.tex')])
	if outfile:
		if outfile[-3:].lower() == 'csv':
			write_delimited_file(outfile, "csv", headers, rows)
		elif outfile[-3:].lower() == 'tsv':
			write_delimited_file(outfile, "tsv", headers, rows)
		elif outfile[-3:].lower() == 'txt':
			write_delimited_file(outfile, "plain", headers, rows)
		elif outfile[-3:].lower() == 'tex':
			write_delimited_file(outfile, "tex", headers, rows)
		elif outfile[-3:].lower() == 'ods':
			export_ods(outfile, headers, rows, append=True, sheetname=sheetname)
		else:
			# Force write as CSV.
			outfile = outfile + ".csv"
			write_delimited_file(outfile, "csv", headers, rows)
	return outfile

def dquote(v):
	# Returns a double-quoted value if it is not an identifier.
	if not v.isidentifier():
		return '"%s"' % v
	else:
		return v

def db_colnames(tbl_hdrs):
	# Takes a list of table headers and returns a list of database column names,
	# with double-quoting of any name that is not all alphanumeric and starts with
	# an alphabetic.
	colnames = []
	for hdr in tbl_hdrs:
		colnames.append(dquote(hdr))
	return colnames

def isint(v):
	# Missing values match and will be handled by 'conv_int()'
	if v is None or (type(v) is str and v.strip() == ''):
		return True
	if type(v) == int:
		return True
	if type(v) == float:
		return False
	try:
		int(v)
		return True
	except ValueError:
		return False
	except TypeError:
		return False

def conv_int(v):
	if v is None or (type(v) is str and v.strip() == ''):
		return None
	try:
		rv = int(v)
		return rv
	except:
		return None

def isfloat(v):
	# Missing values match and will be handled by 'conv_float()'
	if v is None or (type(v) is str and v.strip() == ''):
		return True
	try:
		float(v)
		return True
	except ValueError:
		return False
	except TypeError:
		return False

def conv_float(v):
	if v is None or (type(v) is str and v.strip() == ''):
		return None
	try:
		rv = float(v)
		return rv
	except:
		return None

def isboolean(v):
	return parse_boolean(v) is not None

def conv_datetime(v):
	if v is None or (type(v) is str and v.strip() == ''):
		return None
	try:
		rv = parse_datetime(v)
	except:
		try:
			d = parse_date(v)
			rv = datetime.datetime.combine(d, datetime.datetime.min.time())
		except:
			return None
	return rv


def dt_type(v):
	# Type of date/time: timestamp, date, or None
	if type(v) is str and v.strip() == '':
		v = None
	if parse_date(v):
		return "date"
	if parse_datetime(v):
		return "timestamp"
	if parse_datetimetz(v):
		return "timestamptz"
	return None

def data_type(v):
	# Characterizes the value v as one of a simple set of data types.
	# Returns "timestamp", "date", "timestamptz", "int", "float", "boolean", or "string"
	if v is None or (type(v) is str and v.strip() == ''):
		return None
	if isint(v):
		return "int"
	if isfloat(v):
		return "float"
	dt = dt_type(v)
	if dt is not None:
		return dt
	if isboolean(v):
		return "boolean"
	return "string"

# Lookup table for priorities among data types
data_type_pair_priorities = {
		"int" : {"int":"int", "float":"float", "date":"string", "timestamp":"string", "timestamptz":"string", "boolean":"string", "string":"string", None:"int"},
		"float" : {"int":"float", "float":"float", "date":"string", "timestamp":"string", "timestamptz":"string", "boolean":"string", "string":"string", None:"float"},
		"date" : {"int":"string", "float":"string", "date":"date", "timestamp":"timestamp", "timestamptz":"string", "boolean":"string", "string":"string", None:"date"},
		"timestamp" : {"int":"string", "float":"string", "date":"timestamp", "timestamp":"timestamp", "datetimetz":"string", "boolean":"string", "string":"string", None:"timestamp"},
		"timestamptz" : {"int":"string", "float":"string", "date":"string", "timestamp":"string", "timestamptz":"timestamptz", "boolean":"string", "string":"string", None:"timestamptz"},
		"string" : {"int":"string", "float":"string", "date":"string", "timestamp":"string", "timestamptz":"string", "boolean":"string", "string":"string", None:"string"},
		"boolean" : {"int":"string", "float":"string", "date":"string", "timestamp":"string", "timestamptz":"string", "boolean":"boolean", "string":"string", None:"boolean"},
		None : {"int":"int", "float":"float", "date":"date", "timestamp":"timestamp", "timestamptz":"timestamptz", "boolean":"boolean", "string":"string", None:None}
		}
def priority_data_type(dt1, dt2):
	return data_type_pair_priorities[dt1][dt2]


def data_type_cast_fn(data_type_str):
	if data_type_str == "string":
		return str
	elif data_type_str == "date":
		return parse_date
	elif data_type_str == "timestamp":
		return conv_datetime
	elif data_type_str == "timestamptz":
		return parse_datetimetz
	elif data_type_str == "int":
		return conv_int
	elif data_type_str == "float":
		return conv_float
	elif data_type_str == "boolean":
		return parse_boolean

def common_data_type(values):
	# Returns a data type common to all the values in the list.
	# This is "string" unless all types are the same, or consist only of ints and floats.
	# Null (None) values are ignored.  If all values are null, return "string".
	val2 = [v for v in values if v is not None and not (type(v) is str and v.strip() == '')]
	if len(val2) == 0:
		return "string"
	else:
		types = [data_type(v) for v in val2]
		typeset = set(types)
		if len(typeset) == 1:
			return types[0]
		elif len(typeset) == 2:
			uq_types = list(typeset)
			if 'int' in uq_types and 'float' in uq_types:
				return 'float'
			else:
				if 'date' in uq_types and 'timestamp' in uq_types:
					return 'timestamp'
				else:
					return "string"
		else:
			return "string"


def set_data_types_core(headers, rows):
	# Column-by-column processing is slightly faster than row-by-row processing.
	coltypes = []
	for i, colname in enumerate(headers):
		datavals = [row[i] for row in rows]
		dt = None
		for d in datavals:
			ndt = data_type(d)
			if ndt != dt:
				dt = priority_data_type(ndt, dt)
				if dt == "string":
					break
		non_null = [d for d in datavals if d is not None and not (type(d) is str and d.strip() == '')]
		nullcount = len(datavals) - len(non_null)
		uniquevals = len(set(non_null))
		coltypes.append((colname, dt, nullcount, uniquevals))
	return coltypes

def set_data_types(headers, rows, q):
	q.put(set_data_types_core(headers, rows))


# Translations to SQLite type affinity names
sqlite_type_x = {'int': 'INTEGER', 'float': 'REAL', 'string': 'TEXT', 'timestamptz': 'TEXT',
		'timestamp': 'TEXT', 'date': 'TEXT', 'boolean': 'INTEGER'}

def round_figs(x, figs=3):
	if x == 0.0 or x == -0.0 or x == float('inf') or x == float('-inf'):
		return 0.0
	else:
		return round(x, figs - int(math.floor(math.log10(abs(x)))) - 1)

def center_window(win, x_offset=0, y_offset=0):
	win.update_idletasks()
	m = re.match(r"(\d+)x(\d+)\+(-?\d+)\+(-?\d+)", win.geometry())
	if m is not None:
		wwd = int(m.group(1))
		wht = int(m.group(2))
		swd = win.winfo_screenwidth()
		sht = win.winfo_screenheight()
		xpos = (swd/2) - (wwd/2) + x_offset
		ypos = (sht/2) - (wht/2) + y_offset
		win.geometry("%dx%d+%d+%d" % (wwd, wht, xpos, ypos))

def raise_window(win):
	win.attributes('-topmost', 1)
	win.attributes('-topmost', 0)

def shift_window(win, x_offset=0, y_offset=0):
	win.update_idletasks()
	m = re.match(r"(\d+)x(\d+)\+(-?\d+)\+(-?\d+)", win.geometry())
	if m is not None:
		xpos = int(m.group(1)) + x_offset
		ypos = int(m.group(2)) + y_offset
		win.geometry("+%d+%d" % (xpos, ypos))

# Inverse normal CDF function (qnorm) by Acklam's algorithm
A = (-3.969683028665376e+01, 2.209460984245205e+02, -2.759285104469687e+02,
		1.383577518672690e+02, -3.066479806614716e+01, 2.506628277459239e+00)
B = (-5.447609879822406e+01, 1.615858368580409e+02, -1.556989798598866e+02,
		6.680131188771972e+01, -1.328068155288572e+01)
C = (-7.784894002430293e-03, -3.223964580411365e-01, -2.400758277161838e+00,
		-2.549732539343734e+00, 4.374664141464968e+00, 2.938163982698783e+00)
D = (7.784695709041462e-03, 3.224671290700398e-01, 2.445134137142996e+00,
		3.754408661907416e+00)
P_LOW = 0.02425
P_HIGH = 1.0 - P_LOW
def qnorm(p):
	if p <= 0 or p >= 1.0:
		raise ValueError("Invalid input to qnorm()")
	if p >= P_LOW and p <= P_HIGH:
		q = p - 0.5
		r = q*q
		return (((((A[0]*r+A[1])*r+A[2])*r+A[3])*r+A[4])*r+A[5])*q / \
        (((((B[0]*r+B[1])*r+B[2])*r+B[3])*r+B[4])*r+1)
	elif p < P_LOW:
		q = math.sqrt(-2 * math.log(p))
		return (((((C[0]*q+C[1])*q+C[2])*q+C[3])*q+C[4])*q+C[5]) / \
         ((((D[0]*q+D[1])*q+D[2])*q+D[3])*q+1)
	else:
		q = math.sqrt(-2 * math.log(1.0 - p))
		return -(((((C[0]*q+C[1])*q+C[2])*q+C[3])*q+C[4])*q+C[5]) / \
          ((((D[0]*q+D[1])*q+D[2])*q+D[3])*q+1)


# Functions for calculation of Fisher-Jenks breaks
def ssd(values):
    m = statistics.mean(values)
    return sum([(x-m)**2 for x in values])

def all_jenks_breaks(d, max_groups):
    d_ssd = ssd(d)
    groups = [1]
    gvf = [0]
    for i in range(2, max_groups+2):
        jnb = jenkspy.JenksNaturalBreaks(i)
        jnb.fit(d)
        grps = jnb.groups_
        g_ssd = [ssd(g) for g in grps]
        groups.append(i)
        gvf.append((d_ssd - sum(g_ssd)) / d_ssd)
    return groups, gvf

def slopes(xs, ys):
    return [(ys[i+1]-ys[i])/(xs[i+1]-xs[i]) for i in range(len(xs)-1)]

def reldiffs(v):
    return [(v[i] - v[i+1]) / ((v[i] + v[i+1])/2) for i in range(len(v)-1)]

def optimum_jenks(d, max_groups):
    # Returns the optimum number of Fisher-Jenks groups, >= 2.
    df = reldiffs(slopes(*all_jenks_breaks(d, max_groups)))
    imd = df.index(max(df))
    return list(range(2, max_groups+2))[imd]


# ==================================================================================
#		class MapUI
# The main UI element
# ==================================================================================
class MapUI(object):
	def __init__(self, src_name, message, lat_col, lon_col, crs=4326, sheet=None,
			label_col=None, symbol_col=None, color_col=None,
			map_export_file=None, export_time_sec=10):
		self.win = tk.Tk()
		self.win.withdraw()
		self.loading_dlg = LoadingDialog(self.win)
		self.loading_dlg.display("Preparing map")

		#=====  Tkinter style modifications
		# Set entry of readonly comboboxes to white instead of default grey
		self.ttkstyle = ttk.Style(self.win)
		self.ttkstyle.map('TCombobox', fieldbackground=[('disabled', 'lightgrey'), ('readonly', 'white')])

		# Size and position window.
		self.win.geometry("1200x1000")
		center_window(self.win)


		self.win.protocol("WM_DELETE_WINDOW", self.cancel)
		self.data_src_name = src_name
		self.win.title("Map of %s" % src_name)
		# Patch ImageTk.PhotoImage.__del__ 
		ImageTk.PhotoImage.__del__ = new_img_del
		# Set the font
		self.mapfont = self.makefont()
		# Set the application window icon
		#win_icon = tk.BitmapImage(data=win_icon_xbm, foreground="black", background="tan")
		#self.win.iconbitmap(win_icon)
		# The markers for all the locations in the data table
		self.loc_map_markers = []
		# The markers for the selected location(s)
		self.sel_map_markers = []
		# The number of table rows without coordinates
		self.missing_latlon = 0
		# Map bounds
		self.min_lat = None
		self.max_lat = None
		self.min_lon = None
		self.max_lon = None
		# List of PlotDialog objects, so they can be told to update themselves, or be deleted.
		self.plot_list = []
		# List of UnivarStatsDialog objects, so they can have data pushed.
		self.univar_list = []
		# List of BivarStatsDialog objects, so they can have data pushed.
		self.bivar_list = []
		# List of CorrMatrixDialog objects, so they can have data pushed.
		self.corrmat_list = []
		# List of CategCorrespDialog objects, so they can have data pushed.
		self.catcorresp_list = []
		# Database connection is set in 'add_data()'; variables are initialized here
		self.dbtmpdir = None
		self.dbname = None
		self.db = None
		# Create default markers for the map
		self.loc_marker_icon = self.set_get_loc_marker()
		# Initializes selection marker to the global settings
		self.set_sel_marker(select_symbol, select_color)
		# Create icons for the buttonbar
		expand_icon = tk.BitmapImage(data=expand_xbm, foreground="black")
		focus_icon = tk.BitmapImage(data=wedge_sm_xbm, foreground="red")
		zoom_sel_icon = tk.BitmapImage(data=wedges_3_xbm, foreground="red")
		unselect_icon = tk.BitmapImage(data=cancel_xbm, foreground="black")
		# Use stacked frames for the main application window components.  Map and table in a PanedWindow.
		msgframe = ttk.Frame(self.win, padding="3 2")
		ctrlframe = ttk.Frame(self.win, padding="3 2")
		datapanes = ttk.PanedWindow(self.win, orient=tk.VERTICAL)
		mapframe = ttk.Frame(datapanes, borderwidth=2, relief=tk.RIDGE)
		self.tblframe = ttk.Frame(datapanes, padding="3 2")
		datapanes.add(mapframe, weight=1)
		datapanes.add(self.tblframe, weight=1)
		# Allow vertical resizing of map and table frames, not of message and control frames
		self.win.columnconfigure(0, weight=1)
		self.win.rowconfigure(0, weight=0)		# msgframe
		self.win.rowconfigure(1, weight=0)		# ctrlframe
		self.win.rowconfigure(2, weight=1)		# datapanes
		# Grid all the main frames
		msgframe.grid(row=0, column=0, sticky=tk.NSEW)
		ctrlframe.grid(row=1, column=0, sticky=tk.W)
		datapanes.grid(row=2, column=0, sticky=tk.NSEW)
		# Populate the message frame
		self.msg_label = ttk.Label(msgframe, text=message)
		def wrap_msg(event):
			self.msg_label.configure(wraplength=event.width - 5)
		self.msg_label.bind("<Configure>", wrap_msg)
		self.msg_label.grid(column=0, row=0, sticky=tk.EW, padx=(3,3), pady=(3,3))
		msgframe.columnconfigure(0, weight=1)
		msgframe.rowconfigure(0, weight=1)
		# Populate the map control frame
		ctrlframe.rowconfigure(0, weight=0)
		ctrlframe.columnconfigure(0, weight=0)
		# Basemap controls and buttons
		self.basemap_label = ttk.Label(ctrlframe, text="Basemap:", anchor="w")
		self.basemap_label.grid(row=0, column=0, padx=(5, 5), pady=(2, 5), sticky=tk.W)
		global initial_basemap
		bm_name = initial_basemap
		if bm_name not in bm_servers:
			bm_name = tuple(bm_servers.keys())[0]
			initial_basemap = bm_name
		self.basemap_var = tk.StringVar(self.win, bm_name)
		self.map_option_menu = ttk.Combobox(ctrlframe, state="readonly", textvariable=self.basemap_var,
				values=self.available_tile_servers(), width=18)
		self.map_option_menu["state"] = "normal"
		self.map_option_menu.bind('<<ComboboxSelected>>', self.change_basemap)
		self.map_option_menu.grid(row=0, column=1, padx=(5, 30), pady=(2, 5), sticky=tk.W)
		# Multi-select option
		def ck_changed():
			ck = self.multiselect_var.get()
			if ck == '0':
				self.unselect_map()
				self.tbl.configure(selectmode = tk.BROWSE)
			else:
				self.tbl.configure(selectmode = tk.EXTENDED)
			self.set_status()
		# Set by global variable
		self.multiselect_var = tk.StringVar(self.win, multiselect)
		ck_multiselect = ttk.Checkbutton(ctrlframe, text="Multi-select", variable=self.multiselect_var, command=ck_changed)
		ck_multiselect.grid(row=0, column=2, sticky=tk.W, padx=(0, 20))
		# Map control buttons
		zoomsel_btn = ttk.Button(ctrlframe, text="Zoom selected", image=zoom_sel_icon, compound=tk.LEFT, command=self.zoom_selected)
		zoomsel_btn.image = zoom_sel_icon
		zoomsel_btn.grid(row=0, column=3, sticky=tk.W)
		expand_btn = ttk.Button(ctrlframe, text="Zoom full", image=expand_icon, compound=tk.LEFT, command=self.zoom_full)
		expand_btn.image = expand_icon
		expand_btn.grid(row=0, column=4, sticky=tk.W)
		focus_btn = ttk.Button(ctrlframe, text="Center", image=focus_icon, compound=tk.LEFT, command=self.focus_map)
		focus_btn.image = focus_icon
		focus_btn.grid(row=0, column=5, sticky=tk.W)
		unselect_btn = ttk.Button(ctrlframe, text="Un-select", image=unselect_icon, compound=tk.LEFT, command=self.unselect_map)
		unselect_btn.image = unselect_icon
		unselect_btn.grid(row=0, column=6, sticky=tk.W)
		# Map widget
		mapframe.rowconfigure(0, weight=1)
		mapframe.columnconfigure(0, weight=1)
		self.map_widget = tkmv.TkinterMapView(mapframe, height=600, width=600, corner_radius=0)
		if initial_basemap != "OpenMapServer":
			tileserver = self.tile_url(initial_basemap)
			self.map_widget.set_tile_server(tileserver)
		self.map_widget.grid(row=0, column=0, sticky=tk.NSEW)

		# Remove the splash screen message and show the main UI.
		self.loading_dlg.hide()
		# Re-setting the size is necessary on Windows in at least some environments.
		self.win.geometry("1200x1000")
		self.win.deiconify()

		# Get data if not provided as arguments
		if src_name is None:
			sdsd = SelDataSrcDialog(parent=self.win, mapui=self)
			src_name, label_col, lat_col, lon_col, crs, symbol_col, color_col, message, headers, rows = sdsd.select()
			if src_name is None:
				self.cancel()
		else:
			# src_name is a filename, either CSV or spreadsheet
			self.loading_dlg.display("Loading data")
			fn, ext = os.path.splitext(src_name)
			if ext.lower() == ".csv":
				try:
					headers, rows = file_data(src_name, 0)
				except:
					self.loading_dlg.hide_all()
					fatal_error("Could not read data from %s" % src_data, kwargs={'parent': self.win})
			else:
				if sheet is None:
					self.loading_dlg.hide_all()
					fatal_error("A sheet name must be specified for spreadsheets", kwargs={'parent': self.win})
				try:
					if ext.lower() == '.ods':
						headers, rows = ods_data(src_name, sheet, 0)
					else:
						headers, rows = xls_data(src_name, sheet, 0)
				except:
					self.loading_dlg.hide_all()
					fatal_error("Could not read table from %s, sheet %s" % (src_name, sheet), kwargs={'parent': self.win})
			self.loading_dlg.hide()

		self.msg_label.config(text=message)

		# Source and possibly un-projected crs
		self.src_crs = crs
		self.crs = crs

		# Populate the table frame
		self.tblframe.rowconfigure(0, weight=1)
		self.tblframe.columnconfigure(0, weight=1)
		try:
			self.tableframe, self.tbl = self.add_data(rows, headers, lat_col, lon_col, label_col, symbol_col, color_col)
		except:
			self.loading_dlg.hide_all()
			fatal_error("Cannot load data.  Check latitude, longitude, and CRS values.", kwargs={'parent': self.win})
		self.tableframe.grid(column=0, row=0, sticky=tk.NSEW)
		self.set_tbl_selectmode()
		self.set_status()
		# Add menu
		self.add_menu(table_object = self.tbl, column_headers=headers)
		self.map_widget.canvas.bind("<Control-ButtonRelease-1>", self.ctrl_click)
		self.tbl.bind('<ButtonRelease-1>', self.mark_map)
		# Other key bindings
		self.win.protocol("WM_DELETE_WINDOW", self.cancel)
		self.win.bind("<Alt-q>", self.run_query)
		# Limit resizing
		self.win.minsize(width=860, height=640)
		# Set table status message
		self.set_status()
		# Just export the map and quit?
		if map_export_file is not None:
			self.imageoutputfile = map_export_file
			self.win.after(export_time_sec * 1000, self.export_map_and_quit)
		self.loading_dlg.hide()
	def makefont(self):
		global label_font, label_size, label_bold
		fams = tkfont.families()
		if not label_font in fams:
			alt_fonts = ["Liberation Sans", "Arial", "Helvetica", "Nimbus Sans", "Liberation Sans", "Trebuchet MS", "Tahoma", "DejaVu Sans", "Bitstream Vera Sans", "Open Sans"]
			font_found = False
			for f in alt_fonts:
				if f in fams:
					label_font = f
					font_found = True
					break
			if not font_found:
				label_font = tkfont.nametofont("TkDefaultFont").actual()["family"]
		boldstr = "normal" if not label_bold else "bold"
		return tkfont.Font(family=label_font, size=label_size, weight=boldstr)
	def available_tile_servers(self):
		# Return a list of those without API keys or for which API keys are provided
		avail = []
		for k in bm_servers:
			if self.tile_url(k) is not None:
				avail.append(k)
		return avail
	def tile_url(self, source_name):
		# Return the URL with the API key replaced, unless it is not available.
		source_url = bm_servers[source_name]
		if "<api_key>" in source_url.lower():
			if source_name in api_keys:
				api_key = api_keys[source_name]
				for matched in re.findall("<api_key>", source_url, re.IGNORECASE):
					source_url = source_url.replace(matched, api_key)
				return source_url
			else:
				return None
		else:
			return source_url
	def mark_map(self, event):
		# Highlight the selected row(s) in the table and get the coordinates to map it
		if self.tbl.selection():
			new_markers = []
			for sel_row in self.tbl.selection():
				rowdata = self.tbl.item(sel_row)["values"]
				try:
					lat_val = float(rowdata[self.lat_index])
				except:
					lat_val = None
				try:
					lon_val = float(rowdata[self.lon_index])
				except:
					lon_val = None
				if lon_val is not None and lat_val is not None:
					new_marker = self.map_widget.set_marker(lat_val, lon_val, icon=self.sel_marker_icon)
					new_markers.append(new_marker)
			for m in self.sel_map_markers:
				m.delete()
			self.sel_map_markers = new_markers
		else:
			for m in self.sel_map_markers:
				m.delete()
		self.update_plot_data()
		self.set_status()
	def set_sel_marker(self, symbol, color):
		select_marker = tk.BitmapImage(data=icon_xbm[symbol], foreground=color)
		mkr_key = "%s %s" % (color, symbol)
		if mkr_key not in custom_icons:
			custom_icons[mkr_key] = tk.BitmapImage(data=icon_xbm[symbol], foreground=color)
		select_marker = custom_icons[mkr_key]
		self.sel_marker_icon = select_marker
	def redraw_sel_markers(self):
		new_markers = []
		for mkr in self.sel_map_markers:
			mposition = mkr.position
			micon = mkr.icon
			mkr.delete()
			new_marker = self.map_widget.set_marker(mposition[0], mposition[1], icon=micon)
			new_markers.append(new_marker)
		self.sel_map_markers = new_markers
	def draw_sel_markers(self):
		for mkr in self.sel_map_markers:
			mkr.draw()
	def set_get_loc_marker(self):
		mkr_key = "%s %s" % (location_color, location_marker)
		if mkr_key not in custom_icons:
			custom_icons[mkr_key] = tk.BitmapImage(data=icon_xbm[location_marker], foreground=location_color)
		return custom_icons[mkr_key]
	def redraw_loc_markers(self, tdata):
		# tdata is the treeview control containing the data table.
		self.loading_dlg.display("Redrawing markers")
		while len(self.loc_map_markers) > 0:
			self.loc_map_markers.pop().delete()
		self.draw_loc_markers(tdata)
		self.loading_dlg.hide()
	def draw_loc_markers(self, tdata):
		# tdata is the treeview control containing the data table.
		# Also set the number of rows missing coordinates and the bounding box.
		self.loading_dlg.display("Preparing map")
		self.missing_latlon = 0
		for row_id in tdata.get_children():
			rowdata = tdata.item(row_id)["values"]
			try:
				lat_val = float(rowdata[self.lat_index])
			except:
				lat_val = None
			try:
				lon_val = float(rowdata[self.lon_index])
			except:
				lon_val = None
			if lon_val is not None and lat_val is not None:
				if self.min_lat is None or lat_val < self.min_lat:
					self.min_lat = lat_val
				if self.max_lat is None or lat_val > self.max_lat:
					self.max_lat = lat_val
				if self.min_lon is None or lon_val < self.min_lon:
					self.min_lon = lon_val
				if self.max_lon is None or lon_val > self.max_lon:
					self.max_lon = lon_val
				if self.color_index is None and self.symbol_index is None:
					marker_icon = self.loc_marker_icon
				else:
					if self.color_index is None or not use_data_color:
						color = location_color
					else:
						color = rowdata[self.color_index].lower()
						if color not in color_names:
							color = location_color
					if self.symbol_index is None or not use_data_marker:
						symbol = location_marker
					else:
						symbol = rowdata[self.symbol_index].lower()
						if symbol not in icon_xbm:
							symbol = location_marker
					mkr_key = "%s %s" % (color, symbol)
					if mkr_key not in custom_icons:
						custom_icons[mkr_key] = tk.BitmapImage(data=icon_xbm[symbol], foreground=color)
					marker_icon = custom_icons[mkr_key]
				if self.label_index is not None:
					lbl = rowdata[self.label_index]
					mkr = self.map_widget.set_marker(lat_val, lon_val, icon=marker_icon,
							text=lbl, font=self.mapfont, text_color=label_color,
							command=self.map_sel_table)
					self.loc_map_markers.append(mkr)
				else:
					mkr = self.map_widget.set_marker(lat_val, lon_val, icon=marker_icon, command=self.map_sel_table)
					self.loc_map_markers.append(mkr)
			else:
				self.missing_latlon += 1
		self.update_plot_data()
		self.loading_dlg.hide()
	def ctrl_click(self, event):
		click_coords = self.map_widget.convert_canvas_coords_to_decimal_coords(event.x, event.y)
		dlg = MsgDialog("Coordinates", str(click_coords))
		dlg.show()
		#if 0 - 50 < canvas_pos_x < self.map_widget.width + 50 and 0 < canvas_pos_y < self.map_widget.height + 70:
	def add_data(self, rows, headers, lat_col, lon_col, label_col, symbol_col, color_col):
		# Re-set all data-specific variables and widgets
		self.headers = headers
		self.rows = rows
		self.lat_col = lat_col
		self.lon_col = lon_col
		self.src_lat_col = lat_col
		self.src_lon_col = lon_col
		self.lat_4326_col = None
		self.lon_4326_col = None
		self.label_col = label_col
		self.symbol_col = symbol_col
		self.color_col = color_col
		self.lat_index = headers.index(lat_col)
		self.lon_index = headers.index(lon_col)
		self.src_lat_index = headers.index(lat_col)
		self.src_lon_index = headers.index(lon_col)
		self.label_index = headers.index(label_col) if label_col is not None and label_col != '' else None
		self.symbol_index = headers.index(symbol_col) if symbol_col is not None and symbol_col != '' else None
		self.color_index = headers.index(color_col) if color_col is not None and color_col != '' else None

		if self.crs != 4326:
			try:
				from pyproj import CRS, Transformer
			except:
				self.loading_dlg.hide_all()
				fatal_error("The pyproj library is required to re-project spatial coordinates", kwargs={})
			try:
				crs_proj = CRS(self.crs)
			except:
				self.loading_dlg.hide_all()
				fatal_error("Invalid CRS (%s)" % self.crs, kwargs={})
			if self.lat_4326_col is None:
				for colname in ('lat_4326', 'latitude_4326', 'y_4326', 'unprojected_lat'):
					if colname not in headers:
						self.lat_4326_col = colname
						headers.append(colname)
						break
			if self.lon_4326_col is None:
				for colname in ('lon_4326', 'longitude_4326', 'x_4326', 'unprojected_lon'):
					if colname not in headers:
						self.lon_4326_col = colname
						headers.append(colname)
						break
			self.lat_col = self.lat_4326_col
			self.lon_col = self.lon_4326_col
			self.lat_index = headers.index(self.lat_col)
			self.lon_index = headers.index(self.lon_col)
			crs_4326 = CRS(4326)
			reproj = Transformer.from_crs(crs_proj, crs_4326, always_xy=True)
			for r in rows:
				y = r[self.src_lat_index]
				x = r[self.src_lon_index]
				if y is not None and y != 0 and x is not None and x != 0:
					try:
						newx, newy = reproj.transform(x, y)
					except:
						newx = None
						newy = None
				else:
					newx = None
					newy = None
				if len(r) < len(headers):
					r.extend([newy, newx])
				else:
					r[self.lat_index] = newy
					r[self.lon_index] = newx

		# Populate the treeview
		tframe, tdata = treeview_table(self.tblframe, rows, headers, "browse")
		self.table_row_count = len(tdata.get_children())
		# Scan the table, put points on the map, and find the map extent.
		self.min_lat = self.max_lat = self.min_lon = self.max_lon = None
		self.sel_map_markers = []
		self.missing_latlon = 0
		self.draw_loc_markers(tdata)
		# Set the map extent based on coordinates.
		self.map_widget.fit_bounding_box((self.max_lat, self.min_lon), (self.min_lat, self.max_lon))
		# Copy data from the treeview table to the database.  This includes the treeview IDs
		# Database connection
		if self.db is not None:
			self.db.close()
		if self.dbtmpdir is not None:
			self.dbtmpdir.cleanup(ignore_cleanup_errors = True)
		if temp_dbfile:
			self.dbtmpdir = tempfile.TemporaryDirectory()
			self.dbname = os.path.join(self.dbtmpdir.name, "mapdata.db")
			self.db = sqlite3.connect(self.dbname)
		else:
			self.tmpdir = None
			self.dbname = None
			self.db = sqlite3.connect(":memory:")
		cur = self.db.cursor()
		colnames = db_colnames(headers)
		colnames.append("treeviewid")
		cur.execute("create table mapdata (%s);" % ",".join(colnames))
		tbldata = []
		for row_id in tdata.get_children():
			row_vals = tdata.item(row_id)["values"]
			row_vals = [None if isinstance(x, str) and x.strip() == '' else x for x in row_vals]
			row_vals.append(row_id)
			tbldata.append(row_vals)
		params = ",".join(['?'] * len(colnames))
		cur.executemany("insert into mapdata values (%s)" % params, tbldata)
		cur.close()
		# Initial value for user-entered WHERE clause
		self.whereclause = ""

		# Determe data types for use in table statistics display and in column selection for plotting
		self.data_types = None
		if os.name == 'posix':
			self.data_types_queue = multiprocessing.Queue()
			self.data_types_process = multiprocessing.Process(target=set_data_types, args=(headers, rows, self.data_types_queue))
			self.data_types_process.start()
		else:
			self.data_types = set_data_types_core(headers, rows)

		# Return frame and data table
		return tframe, tdata
	def remove_data(self):
		while len(self.sel_map_markers) > 0:
			self.sel_map_markers.pop().delete()
		while len(self.loc_map_markers) > 0:
			self.loc_map_markers.pop().delete()
		self.map_widget.delete_all_marker()
		self.close_all_plots()
		self.tableframe.destroy()
		self.tbl.destroy()
	def set_tbl_selectmode(self):
		ck = self.multiselect_var.get()
		if ck == '0':
			self.tbl.configure(selectmode = tk.BROWSE)
		else:
			self.tbl.configure(selectmode = tk.EXTENDED)
		self.tbl.bind('<ButtonRelease-1>', self.mark_map)
	def replace_data(self, rows, headers, lat_col, lon_col, label_col, symbol_col, color_col):
		self.remove_data()
		try:
			self.tableframe, self.tbl = self.add_data(rows, headers, lat_col, lon_col, label_col, symbol_col, color_col)
		except:
			self.loading_dlg.hide_all()
			fatal_error("Cannot load data.  Check latitude, longitude, and CRS values.", kwargs={'parent': self.win})
		self.tableframe.grid(column=0, row=0, sticky=tk.NSEW)
		self.set_tbl_selectmode()
		self.set_status()
	def new_data_file(self):
		dfd = DataFileDialog()
		fn, id_col, lat_col, lon_col, crs, sym_col, col_col, msg, headers, rows = dfd.get_datafile()
		if fn is not None and fn != '':
			self.crs = crs
			self.data_src_name = os.path.abspath(fn)
			base_fn = os.path.basename(fn)
			self.win.title("Map of %s" % base_fn)
			self.replace_data(rows, headers, lat_col, lon_col, id_col, sym_col, col_col)
			if msg is not None and msg != '':
				self.msg_label['text'] = msg
	def new_spreadsheet_file(self):
		dfd = ImportSpreadsheetDialog(self.win, self)
		fn, id_col, lat_col, lon_col, crs, sym_col, col_col, msg, headers, rows = dfd.get_datafile()
		if fn is not None and fn != '':
			self.crs = crs
			self.data_src_name = os.path.abspath(fn)
			base_fn = os.path.basename(fn)
			self.win.title("Map of %s" % base_fn)
			self.replace_data(rows, headers, lat_col, lon_col, id_col, sym_col, col_col)
			if msg is not None and msg != '':
				self.msg_label['text'] = msg
	def new_db_table(self):
		dbd = DbConnectDialog(self.win, self)
		tablename, id_col, lat_col, lon_col, crs, sym_col, col_col, desc, headers, rows = dbd.get_data()
		if tablename is not None and tablename != '':
			self.crs = crs
			self.win.title("Map of %s" % tablename)
			self.replace_data(rows, headers, lat_col, lon_col, id_col, sym_col, col_col)
			if desc is not None and desc != '':
				self.msg_label['text'] = desc
	def zoom_full(self):
		self.map_widget.fit_bounding_box((self.max_lat, self.min_lon), (self.min_lat, self.max_lon))
	def zoom_selected(self):
		if len(self.sel_map_markers) > 0:
			if len(self.sel_map_markers) == 1:
				self.focus_map()
			else:
				min_lat = max_lat = min_lon = max_lon = None
				for m in self.sel_map_markers:
					lat, lon = m.position
					if min_lat is None or lat < min_lat:
						min_lat = lat
					if max_lat is None or lat > max_lat:
						max_lat = lat
					if min_lon is None or lon < min_lon:
						min_lon = lon
					if max_lon is None or lon > max_lon:
						max_lon = lon
			self.map_widget.fit_bounding_box((max_lat, min_lon), (min_lat, max_lon))
	def focus_map(self):
		# Center the map on the last marker
		if len(self.sel_map_markers) > 0:
			m = self.sel_map_markers[-1]
			self.map_widget.set_position(m.position[0], m.position[1])
	def unselect_map(self):
		for m in self.sel_map_markers:
			self.map_widget.delete(m)
		self.tbl.selection_remove(*self.tbl.selection())
		self.sel_map_markers = []
		self.update_plot_data()
		self.set_status()
	def change_basemap(self, *args):
		new_map = self.basemap_var.get()
		tileserver = self.tile_url(new_map)
		self.map_widget.set_tile_server(tileserver)
	def map_sel_table(self, marker):
		# Highlight the table row for the clicked map marker
		lat, lon = marker.position
		if self.multiselect_var.get() == '0':
			for mkr in self.sel_map_markers:
				self.map_widget.delete(mkr)
			self.sel_map_markers = []
			self.tbl.selection_remove(*self.tbl.selection())
		for row_id in self.tbl.get_children():
			rowdata = self.tbl.item(row_id)["values"]
			try:
				lat_val = float(rowdata[self.lat_index])
			except:
				lat_val = None
			try:
				lon_val = float(rowdata[self.lon_index])
			except:
				lon_val = None
			if lon_val is not None and lat_val is not None:
				if lat_val == lat and lon_val == lon:
					self.tbl.selection_add(row_id)
					self.tbl.see(row_id)
					new_marker = self.map_widget.set_marker(lat, lon, icon=self.sel_marker_icon)
					if not new_marker in self.sel_map_markers:
						self.sel_map_markers.append(new_marker)
		self.update_plot_data()
		self.set_status()
	def set_status(self):
		statusmsg = "    %d rows" % self.table_row_count
		if self.missing_latlon > 0:
			statusmsg = statusmsg + " (%d without lat/lon)" % self.missing_latlon
		if len(self.tbl.selection()) > 0:
			statusmsg = statusmsg + "  |  %s selected" % len(self.tbl.selection())
		if self.multiselect_var.get() == "1":
			statusmsg = statusmsg + "  |  Ctrl-click to select multiple rows"
		self.tblframe.statusbar.config(text = statusmsg)

	def get_all_data(self, column_list):
		# Plotting and statistics support.  Return all data for the specified columns as a list of column-oriented lists.
		res = []
		for c in column_list:
			if c in self.headers:
				i = self.headers.index(c)
				res.append([row[i] for row in self.rows])
		return res
	def get_sel_data(self, column_list):
		# Plotting and statistics support.  Return data from selected rows for the specified columns, as a list of lists.
		res = [[] for _ in column_list]
		indices = [self.headers.index(c) for c in column_list]
		for sel_row in self.tbl.selection():
			datarow = self.tbl.item(sel_row)["values"]
			for i, index in enumerate(indices):
				res[i].append(datarow[index])
		return res
	def update_plot_data(self):
		# Pushes updates to all plots and other dialogs that may use only selected data.
		# The 'push' is done by calling their own refresh method.
		for plot in self.plot_list:
			if plot.sel_only_var.get() == "1" and plot.auto_update:
				plot.q_redraw()
		for dlg in self.univar_list:
			if dlg.sel_only_var.get() == "1":
				dlg.q_recalc()
		for dlg in self.bivar_list:
			if dlg.sel_only_var.get() == "1":
				dlg.q_recalc()
		for dlg in self.corrmat_list:
			if dlg.sel_only_var.get() == "1":
				dlg.q_redraw()
		for dlg in self.catcorresp_list:
			if dlg.sel_only_var.get() == "1":
				dlg.q_recalc()
	def clone_plot(self, plot_obj):
		if self.data_types is None:
			self.loading_dlg.display("Evaluating data types")
			self.data_types = self.data_types_queue.get()
			self.data_types_process.join()
			self.data_types_process.close()
			self.loading_dlg.hide()
		clone = PlotDialog(self, self.data_types)
		self.plot_list.append(clone)
		clone.dlg.geometry(plot_obj.dlg.geometry())
		shift_window(clone.dlg, x_offset=10, y_offset=10)
		clone.type_var.set(plot_obj.type_var.get())
		clone.sel_only_var.set(plot_obj.sel_only_var.get())
		clone.autoupdate_var.set(plot_obj.autoupdate_var.get())
		clone.x_var.set(plot_obj.x_var.get())
		clone.y_var.set(plot_obj.y_var.get())
		clone.xlog_var.set(plot_obj.xlog_var.get())
		clone.ylog_var.set(plot_obj.ylog_var.get())
		clone.groupby_var.set(plot_obj.groupby_var.get())
		clone.x_sel["values"] = copy.copy(plot_obj.x_sel["values"])
		clone.y_sel["values"] = copy.copy(plot_obj.y_sel["values"])
		clone.groupby_sel["values"] = copy.copy(plot_obj.groupby_sel["values"])
		clone.x_sel["state"] = plot_obj.x_sel["state"]
		clone.y_sel["state"] = plot_obj.y_sel["state"]
		clone.groupby_sel["state"] = plot_obj.groupby_sel["state"]
		clone.xlog_ck["state"] = plot_obj.xlog_ck["state"]
		clone.ylog_ck["state"] = plot_obj.ylog_ck["state"]
		clone.data_btn["state"] = plot_obj.data_btn["state"]
		clone.plot_data_btn["state"] = plot_obj.plot_data_btn["state"]
		clone.dlg.bind("<Alt-h>", clone.do_help)
		clone.dlg.bind("<Alt-n>", clone.clone_plot)
		clone.dlg.bind("<Alt-c>", clone.do_close)
		clone.dlg.bind("<Escape>", clone.do_close)
		clone.dlg.bind("<Alt-t>", clone.set_title)
		clone.dlg.bind("<Alt-x>", clone.set_xlabel)
		clone.dlg.bind("<Alt-y>", clone.set_ylabel)
		if clone.type_var.get() == "Histogram":
			clone.dlg.bind("<Alt-b>", clone.set_bins)
		clone.alpha = plot_obj.alpha
		clone.rotated = plot_obj.rotated
		if clone.type_var.get() in ("Box plot", "Scatter plot", "Stripchart" "Kernel density (KD) plot", "Violin plot"):
			clone.dlg.bind("<Alt-a>", clone.set_alpha)
		clone.scatter_breaks = plot_obj.scatter_breaks
		clone.scatter_x_breaks = plot_obj.scatter_x_breaks
		clone.scatter_y_breaks = plot_obj.scatter_y_breaks
		clone.lineplot_breaks = plot_obj.lineplot_breaks
		clone.lineplot_x_breaks = plot_obj.lineplot_x_breaks
		clone.loess = plot_obj.loess
		clone.linreg = plot_obj.linreg
		clone.theilsen = plot_obj.theilsen
		clone.numeric_columns = plot_obj.numeric_columns
		clone.dataset = copy.copy(plot_obj.dataset)
		clone.plot_data = copy.copy(plot_obj.plot_data)
		clone.data_labels = copy.copy(plot_obj.data_labels)
		clone.plot_data_labels = copy.copy(plot_obj.plot_data_labels)
		clone.q_redraw(get_data=False)
		raise_window(clone.dlg)
		clone.dlg.focus()
	def remove_plot(self, plot_obj):
		# For use by the plot 'do_close()' method.
		try:
			self.plot_list.remove(plot_obj)
		except:
			pass
	def close_plot(self, plot_obj):
		try:
			plot_obj.do_close()
			self.remove_plot()
		except:
			pass
	def close_all_plots(self):
		while len(self.plot_list) > 0:
			self.plot_list[0].do_close()
			# The callback will remove the plot.
		self.plot_list = []
	def remove_univar(self, univar_dlg):
		try:
			self.univar_list.remove(univar_dlg)
		except:
			pass
	def remove_bivar(self, bivar_dlg):
		try:
			self.bivar_list.remove(bivar_dlg)
		except:
			pass
	def remove_corrmat(self, corrmat_dlg):
		try:
			self.corrmat_list.remove(corrmat_dlg)
		except:
			pass
	def remove_categcorresp(self, categcorresp_dlg):
		try:
			self.catcorresp_list.remove(categcorresp_dlg)
		except:
			pass

	def change_crs(self):
		crsdlg = NewCrsDialog(self.crs)
		new_crs = crsdlg.get_crs()
		if new_crs is not None:
			if new_crs != self.crs:
				try:
					from pyproj import CRS, Transformer
				except:
					self.loading_dlg.hide_all()
					fatal_error("The pyproj library is required to re-project spatial coordinates", kwargs={})
				try:
					crs_proj = CRS(new_crs)
				except:
					warning("Invalid CRS (%s)" % new_crs, kwargs={})
				else:
					if self.lat_4326_col is None:
						for colname in ('lat_4326', 'latitude_4326', 'y_4326', 'unprojected_lat'):
							if colname not in self.headers:
								self.lat_4326_col = colname
								self.headers.append(colname)
								for r in self.rows:
									r.append(None)
								break
					if self.lon_4326_col is None:
						for colname in ('lon_4326', 'longitude_4326', 'x_4326', 'unprojected_lon'):
							if colname not in self.headers:
								self.lon_4326_col = colname
								self.headers.append(colname)
								for r in self.rows:
									r.append(None)
								break
					self.lat_col = self.lat_4326_col
					self.lon_col = self.lon_4326_col
					self.lat_index = self.headers.index(self.lat_4326_col)
					self.lon_index = self.headers.index(self.lon_4326_col)
					crs_4326 = CRS(4326)
					self.crs = new_crs
					reproj = Transformer.from_crs(crs_proj, crs_4326, always_xy=True)
					for r in self.rows:
						y = r[self.src_lat_index]
						x = r[self.src_lon_index]
						if y is not None and y != 0 and x is not None and x != 0:
							try:
								newx, newy = reproj.transform(x, y)
								r[self.lat_index] = newy
								r[self.lon_index] = newx
							except:
								r[self.lat_index] = None
								r[self.lon_index] = None
						else:
							r[self.lat_index] = None
							r[self.lon_index] = None
					selected = self.tbl.selection()
					self.replace_data(self.rows, self.headers, self.src_lat_col, self.src_lon_col, self.label_col, self.symbol_col, self.color_col)
					self.tbl.selection_set(tuple(selected))
					self.mark_map({})
	def cancel(self):
		self.win.destroy()
		sys.exit()
	def export_map_and_quit(self):
		fn, ext = os.path.splitext(self.imageoutputfile)
		if ext.lower() == ".ps":
			self.export_map_to_ps(self.imageoutputfile)
		else:
			self.map_widget.update_idletasks()
			#self.win.after(200, self.save_imageoutputfile)
			self.save_imageoutputfile()
		self.win.destroy()
	def export_map_to_ps(self, outfile):
		self.map_widget.canvas.postscript(file=outfile, colormode='color')
	def save_imageoutputfile(self):
		obj = self.map_widget.canvas
		bounds = (obj.winfo_rootx(), obj.winfo_rooty(), 
				obj.winfo_rootx() + obj.winfo_width(), obj.winfo_rooty() + obj.winfo_height())
		ImageGrab.grab(bbox=bounds).save(self.imageoutputfile)
	def export_map_to_img(self, outfile):
		# Allow map to recover from blocking by the file dialog box before grabbing and exporting the canvas
		self.map_widget.update_idletasks()
		self.imageoutputfile = outfile
		self.win.after(1000, self.save_imageoutputfile)
	def run_query(self, args=None):
		dlg = QueryDialog(self.headers, self.db, self.whereclause)
		whereclause, action = dlg.get_where()
		if whereclause is not None:
			self.whereclause = whereclause
			sqlcmd = "SELECT treeviewid FROM mapdata WHERE %s" % whereclause
			cur = self.db.cursor()
			try:
				result = cur.execute(sqlcmd)
				id_list = [r[0] for r in result.fetchall()]
			except:
				cur.close()
				warning("Invalid data selection expression: %s" % whereclause, kwargs={})
			else:
				cur.close()
				# Enable multiselect
				global multiselect
				multiselect = "1"
				self.multiselect_var.set("1")
				self.tbl.configure(selectmode = tk.EXTENDED)
				if action == "Replace":
					self.unselect_map()
					self.tbl.selection_set(list(id_list))
				elif action == "Union":
					all_selections = tuple(set(self.tbl.selection()) | set(id_list))
					self.tbl.selection_set(all_selections)
				elif action == "Intersection":
					int_selections = tuple(set(self.tbl.selection()) & set(id_list))
					self.tbl.selection_set(int_selections)
				elif action == "Difference O-N":
					# Old - New
					diff_selections = tuple(set(self.tbl.selection()) - set(id_list))
					self.tbl.selection_set(diff_selections)
				else:
					# New - Old
					diff_selections = tuple(set(id_list) - set(self.tbl.selection()))
					self.tbl.selection_set(diff_selections)
				self.mark_map(None)
				self.set_status()
	def add_menu(self, table_object, column_headers):
		mnu = tk.Menu(self.win)
		self.win.config(menu=mnu)
		file_menu = tk.Menu(mnu, tearoff=0)
		tbl_menu = tk.Menu(mnu, tearoff=0)
		map_menu = tk.Menu(mnu, tearoff=0)
		sel_menu = tk.Menu(mnu, tearoff=0)
		plot_menu = tk.Menu(mnu, tearoff=0)
		stats_menu = tk.Menu(mnu, tearoff=0)
		help_menu = tk.Menu(mnu, tearoff=0)
		mnu.add_cascade(label="File", menu=file_menu, underline=0)
		mnu.add_cascade(label="Table", menu=tbl_menu, underline=0)
		mnu.add_cascade(label="Map", menu=map_menu, underline=0)
		mnu.add_cascade(label="Selections", menu=sel_menu, underline=0)
		mnu.add_cascade(label="Plot", menu=plot_menu, underline=0)
		mnu.add_cascade(label="Stats", menu=stats_menu, underline=0)
		mnu.add_cascade(label="Help", menu=help_menu, underline=0)
		def save_table():
			if table_object.selection():
				rowset = []
				for sel_row in table_object.selection():
					rowset.append(table_object.item(sel_row)["values"])
				export_data_table(column_headers, rowset, sheetname="Selected map items")
		def save_map():
			outfile = tkfiledialog.asksaveasfilename(title="File to save map",
				filetypes=[('Postscript files', '.ps'), ('JPEG files', '.jpg'), ('PNG files', '.png')])
			fn, ext = os.path.splitext(outfile)
			if len(ext) > 1 and outfile[-2:].lower() == 'ps':
				self.export_map_to_ps(outfile)
			else:
				self.export_map_to_img(outfile)
		def change_marker():
			global select_symbol, select_color
			marker_dlg = MarkerDialog(map_menu)
			symbol, color = marker_dlg.get_marker()
			if symbol is not None or color is not None:
				if symbol is None or symbol == '':
					symbol = select_symbol
				if color is None or color == '':
					color = select_color
				symb_name = "%s %s" % (color, symbol)
				if symb_name not in custom_icons:
					custom_icons[symb_name] = tk.BitmapImage(data=icon_xbm[symbol], foreground=color)
				select_symbol = symbol
				select_color = color
				self.sel_marker_icon = custom_icons[symb_name]
		def change_labeling():
			global location_marker, location_color, use_data_marker, use_data_color
			global label_font, label_size, label_bold, label_position
			lbl_dlg = LabelDialog(map_menu, self.headers, self.label_col)
			mkr, clr, datamkr, dataclr, column, ffam, fsize, fbold, pos = lbl_dlg.get_labeling()
			if mkr is not None:
				fsize = int(fsize)
				fbold = False if "0" else True
				if mkr != location_marker or clr != location_color or datamkr != use_data_marker or \
						dataclr != use_data_color or column != self.label_col or ffam != label_font or \
						fsize != label_size or fbold != label_bold or pos != label_position:
							fontchanged = ffam != label_font or fsize != label_size or fbold != label_bold
							location_marker = mkr
							location_color = clr
							use_data_marker = datamkr == '1'
							use_data_color = dataclr == '1'
							self.label_col = column if column != '' else None
							label_font = ffam
							label_size = fsize
							label_bold = fbold
							label_position = pos
							if fontchanged:
								self.mapfont = self.makefont()
							self.loc_marker_icon = self.set_get_loc_marker()
							self.label_index = self.headers.index(self.label_col) if self.label_col is not None and self.label_col != '' else None
							self.redraw_loc_markers(self.tbl)
							self.redraw_sel_markers()
		def import_symbol_file():
			sd = ImportSymbolDialog()
			name, fn = sd.run()
			if name is not None and fn is not None:
				import_symbol(name, fn)
				fqfn = os.path.abspath(fn)
				symb_spec = (name, fqfn)
				if not symb_spec in imported_symbols:
					imported_symbols.append(symb_spec)
		def read_config_file():
			fn = tkfiledialog.askopenfilename(filetypes=([('Config files', '.conf')]))
			if fn != '' and fn is not None and fn != ():
				global multiselect, select_symbol, select_color
				pre_select = multiselect
				pre_basemap = self.basemap_var.get()
				pre_symbol = select_symbol
				pre_color = select_color
				pre_loc_symbol = location_marker
				pre_loc_color = location_color
				pre_label_color = label_color
				pre_label_font = label_font
				pre_label_size = label_size
				pre_label_bold = label_bold
				pre_label_position = label_position
				read_config(fn)
				# (Re)set configuration options to global defaults
				self.map_option_menu['values'] = self.available_tile_servers()
				if multiselect != pre_select:
					self.multiselect_var.set(multiselect)
				if initial_basemap != pre_basemap:
					self.basemap_var.set(initial_basemap)
					tileserver = self.tile_url(initial_basemap)
					self.map_widget.set_tile_server(tileserver)
				if select_symbol != pre_symbol or select_color != pre_color:
					self.set_sel_marker(select_symbol, select_color)
				# Redraw markers if any setting has changed
				if location_marker != pre_loc_symbol or location_color != pre_loc_color or \
						label_color != pre_label_color or label_font != pre_label_font or \
						label_size != pre_label_size or label_bold != pre_label_bold or \
						label_position != pre_label_position:
							if label_font != pre_label_font or label_size != pre_label_size or label_bold != pre_label_bold:
								self.mapfont = self.makefont()
							self.loc_marker_icon = self.set_get_loc_marker()
							self.redraw_loc_markers(self.tbl)
							self.redraw_sel_markers()
				global config_files_user
				config_files_user.append(os.path.abspath(fn))
		def save_config():
			fn = tkfiledialog.asksaveasfilename(filetypes=([('Config files', '.conf')]))
			if fn != '':
				f = open(fn, "w")
				f.write("# Configuration file for mapdata.py\n# Created by export from mapdata.py at %s\n" % datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
				f.write("\n[basemap_tile_servers]\n")
				added_bms = [k for k in bm_servers if not k in initial_bms]
				for k in added_bms:
					f.write("%s=%s\n" % (k, bm_servers[k]))
				f.write("\n[api_keys]\n")
				for k in api_keys:
					f.write("%s=%s\n" % (k, api_keys[k]))
				f.write("\n[symbols]\n")
				for s in imported_symbols:
					f.write("%s=%s\n" % (s[0], s[1]))
				f.write("\n[defaults]\n")
				f.write("basemap=%s\n" % self.basemap_var.get())
				f.write("location_color=%s\n" % location_color)
				f.write("location_marker=%s\n" % location_marker)
				f.write("label_bold=%s\n" % ('No' if not label_bold else 'Yes'))
				f.write("label_color=%s\n" % label_color)
				f.write("label_font=%s\n" % label_font)
				f.write("label_position=%s\n" % label_position)
				f.write("label_size=%s\n" % label_size)
				f.write("multiselect=%s\n" % ('Yes' if self.multiselect_var.get() == '1' else 'No'))
				f.write("select_color=%s\n" % select_color)
				f.write("select_symbol=%s\n" % select_symbol)
				f.write("show_regression_stats=%s\n" % show_regression_stats)
				f.write("wrapwidth=%s\n" % wrapwidth)
				f.write("wrap_at_underscores=%s\n" % wrap_at_underscores)
				f.write("use_data_color=%s\n" % use_data_color)
				f.write("use_data_marker=%s\n" % use_data_marker)
				f.write("\n[misc]\n")
				if editor is not None:
					f.write("editor=%s\n" % editor)
				f.write("temp_dbfile=%s\n" % temp_dbfile)

		def set_editor():
			global editor
			dlg = GetEditorDialog(self.win, editor)
			new_editor = dlg.show()
			if new_editor is not None:
				editor = new_editor
		def show_data_types():
			if self.data_types is None:
				self.loading_dlg.display("Evaluating data types")
				self.data_types = self.data_types_queue.get()
				self.data_types_process.join()
				self.data_types_process.close()
				self.loading_dlg.hide()
			dlg = MsgDialog2("Data Types", "Data types, data completeness, and number of unique non-missing values for columns of the data table:", can_resize=True)
			tframe, tdata = treeview_table(dlg.content_frame, self.data_types, ["Column", "Type", "Missing", "Unique"], "browse")
			tframe.grid(row=0, column=0, sticky=tk.NSEW)
			dlg.show()
		def invert_selections():
			selected = self.tbl.selection()
			new_selections = []
			for iid in self.tbl.get_children():
				if not iid in selected:
					new_selections.append(iid)
			self.tbl.selection_set(tuple(new_selections))
			self.mark_map(None)
			self.set_status()
		def wait_for_data_types():
			if self.data_types is None:
				self.loading_dlg.display("Evaluating data types")
				self.data_types = self.data_types_queue.get()
				self.data_types_process.join()
				self.data_types_process.close()
				self.loading_dlg.hide()
		def new_plot():
			wait_for_data_types()
			dlg = PlotDialog(self, self.data_types)
			self.plot_list.append(dlg)
			dlg.show

		def get_plot_config():
			global show_regression_stats, wrapwidth, wrap_at_underscores
			dlg = PlotConfigDialog(show_regression_stats, wrapwidth, wrap_at_underscores)
			plotconfig = dlg.show()
			show_regression_stats = plotconfig["show_regr_stats"]
			wrapwidth = plotconfig["wrapwidth"]
			wrap_at_underscores = plotconfig["wrap_underscores"]

		def univar_stats():
			wait_for_data_types()
			dlg = UnivarStatsDialog(self, self.data_types)
			self.univar_list.append(dlg)
			dlg.show()

		def bivar_stats():
			wait_for_data_types()
			dlg = BivarStatsDialog(self, self.data_types)
			self.bivar_list.append(dlg)
			dlg.show()

		def corr_matrix():
			wait_for_data_types()
			dlg = CorrMatrixDialog(self, self.data_types)
			self.corrmat_list.append(dlg)
			dlg.show()

		def categ_corresp():
			wait_for_data_types()
			dlg = CategCorrespDialog(self, self.data_types)
			self.catcorresp_list.append(dlg)
			dlg.show()

		def online_help():
			webbrowser.open("https://mapdata.readthedocs.io/en/latest/", new=2, autoraise=True)
		def show_config_files():
			if len(config_files) == 0 and len(config_files_user) == 0:
				msg = "No configuration files have been read."
			else:
				if len(config_files) > 0:
					msg = "Configuration files read on startup:\n   %s" % "\n   ".join(config_files)
					if len(config_files_user) > 0:
						msg = msg + "\n\n"
				if len(config_files_user) > 0:
					msg = msg + "Configuration files read after startup, in sequence:\n   %s" % "\n   ".join(config_files_user)
			dlg = MsgDialog("Config files", msg)
			dlg.show()
		def show_hotkeys():
			dlg = HelpHotkeysDialog()
			dlg.show()
		def show_about():
			message="""
                     mapdata.py

           version: %s, %s
      Copyright %s, R Dreas Nielsen
               License: GNU GPL3""" % (version, vdate, copyright)
			dlg = MsgDialog("About", message)
			dlg.show()

		file_menu.add_command(label="Open CSV", command = self.new_data_file, underline=5)
		file_menu.add_command(label="Open spreadsheet", command = self.new_spreadsheet_file, underline=5)
		file_menu.add_command(label="Open database", command = self.new_db_table, underline=5)
		file_menu.add_command(label="Import symbol", command = import_symbol_file, underline=0)
		file_menu.add_command(label="Set editor", command = set_editor, underline=4)
		file_menu.add_command(label="Read config", command = read_config_file, underline=0)
		file_menu.add_command(label="Save config", command = save_config, underline=0)
		file_menu.add_command(label="Quit", command = self.cancel, underline=0)
		tbl_menu.add_command(label="Un-select all", command = self.unselect_map, underline=0)
		tbl_menu.add_command(label="Export selected", command = save_table, underline=1)
		tbl_menu.add_command(label="Data types", command = show_data_types, underline=5)
		map_menu.add_command(label="Change marker", command = change_marker, underline=7)
		map_menu.add_command(label="Change labeling", command = change_labeling, underline=7)
		map_menu.add_command(label="Zoom selected", command = self.zoom_selected, underline=5)
		map_menu.add_command(label="Zoom full", command = self.zoom_full, underline=5)
		map_menu.add_command(label="Center on selection", command = self.focus_map, underline=0)
		map_menu.add_command(label="Un-select all", command = self.unselect_map, underline=0)
		map_menu.add_command(label="Change CRS", command = self.change_crs, underline=1)
		map_menu.add_command(label="Export", command = save_map, underline=1)
		sel_menu.add_command(label="Invert", command = invert_selections, underline=0)
		sel_menu.add_command(label="Un-select all", command = self.unselect_map, underline=0)
		sel_menu.add_command(label="Set by query", command = self.run_query, underline=7)
		plot_menu.add_command(label="New", command = new_plot, underline=0)
		plot_menu.add_command(label="Close all", command = self.close_all_plots, underline=0)
		plot_menu.add_command(label="Configure", command = get_plot_config, underline=3)
		stats_menu.add_command(label="Univariate", command = univar_stats, underline=0)
		stats_menu.add_command(label="Bivariate", command = bivar_stats, underline=0)
		stats_menu.add_command(label="Corr. matrix", command = corr_matrix, underline=0)
		stats_menu.add_command(label="Categ. corresp.", command = categ_corresp, underline=1)
		help_menu.add_command(label="Online help", command = online_help, underline=7)
		help_menu.add_command(label="Config files", command = show_config_files, underline=0)
		help_menu.add_command(label="Hotkeys", command = show_hotkeys, underline=0)
		help_menu.add_command(label="About", command = show_about, underline=0)



class LoadingDialog(object):
	def __init__(self, parent):
		self.parent = parent
		self.dlg = tk.Toplevel(parent)
		self.dlg.title("MapData")
		#self.dlg.geometry("150x50")
		center_window(self.dlg)
		self.dlg.update_idletasks()
		self.dlg.withdraw()
		self.dlg.wm_overrideredirect(True)
		self.dlg.configure(bg="Gold")
		self.messages = []
		self.var_lbl = tk.StringVar(self.dlg, "")
		self.lbl_loading = tk.Label(self.dlg, bg="Gold", textvariable=self.var_lbl)
		self.lbl_loading.place(relx=0.5, rely=0.5, anchor="center")
		self.dlg.update()
		self.dots = 3
	def update_lbl(self):
		if len(self.messages) > 0:
			self.dots = self.dots % 3 + 1
			lbl = self.messages[0] + '.' * self.dots
			self.var_lbl.set(lbl)
			self.dlg.update()
			self.after_id = self.dlg.after(250, self.update_lbl)
	def display(self, message):
		self.messages.append(message)
		self.var_lbl.set(message)
		self.dlg.deiconify()
		raise_window(self.dlg)
		self.dlg.config(cursor="watch")
		self.dlg.update()
		#self.dlg.focus_force()
		#self.after_id = self.dlg.after(250, self.update_lbl)
	def hide(self):
		self.var_lbl.set("")
		#self.dlg.after_cancel(self.after_id)
		do_withdraw = True
		self.dlg.config(cursor="arrow")
		if len(self.messages) > 0:
			self.messages.pop(-1)
			if len(self.messages) > 0:
				self.var_lbl.set(self.messages[0])
				do_withdraw = False
		if do_withdraw:
			self.dlg.withdraw()
	def hide_all(self):
		self.var_lbl.set("")
		self.messages = []
		self.dlg.config(cursor="arrow")
		self.dlg.withdraw()




class LabelDialog(object):
	def __init__(self, parent, column_list, label_col):
		columns = ['']
		columns.extend(column_list)
		label_col = '' if label_col is None else label_col
		self.dlg = tk.Toplevel()
		self.dlg.title("Change Labeling")
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.NSEW, pady=(3,3))
		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.grid(row=1, column=0, sticky=tk.EW, pady=(3,3))
		btn_frame.columnconfigure(0, weight=1)
		#
		symbol_lbl = ttk.Label(prompt_frame, text="Location symbol:")
		symbol_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3))
		self.symbol_var = tk.StringVar(self.dlg, location_marker)
		symbol_vals = list(icon_xbm.keys())
		symbol_vals.sort()
		self.symbol_opts = ttk.Combobox(prompt_frame, state="readonly", textvariable=self.symbol_var,
				values=symbol_vals, width=15)
		self.symbol_opts.grid(row=0, column=1, columnspan=3, sticky=tk.W, padx=(6,3))
		color_lbl = ttk.Label(prompt_frame, text="Color:")
		color_lbl.grid(row=1, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.color_var = tk.StringVar(self.dlg, location_color)
		color_opts = ttk.Combobox(prompt_frame, state="readonly", textvariable=self.color_var,
				values=list(select_colors), width=15)
		color_opts.grid(row=1, column=1, columnspan=3, sticky=tk.W, padx=(6,3), pady=(3,3))
		#
		self.use_data_marker_var = tk.StringVar(prompt_frame, use_data_marker)
		ck_use_data_marker = ttk.Checkbutton(prompt_frame, text="Use data symbol", variable=self.use_data_marker_var)
		ck_use_data_marker.grid(row=0, column=3, columnspan=2, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.use_data_color_var = tk.StringVar(prompt_frame, use_data_color)
		ck_use_color_marker = ttk.Checkbutton(prompt_frame, text="Use data color", variable=self.use_data_color_var)
		ck_use_color_marker.grid(row=1, column=3, columnspan=2, sticky=tk.W, padx=(3,6), pady=(3,3))
		#
		self.col_var = tk.StringVar(prompt_frame, label_col)
		col_lbl = ttk.Label(prompt_frame, text = "Data column:")
		col_lbl.grid(row=2, column=0, sticky=tk.E, padx=(6,3), pady=(6,3))
		col_opts = ttk.Combobox(prompt_frame, state="readonly", textvariable=self.col_var,
				values=columns, width=40)
		col_opts.grid(row=2, column=1, columnspan=4, sticky=tk.W, padx=(3,6), pady=(3,3))
		#
		self.font_var = tk.StringVar(prompt_frame, label_font)
		font_lbl = ttk.Label(prompt_frame, text="Font:")
		font_lbl.grid(row=3, column=0, sticky=tk.E, padx=(3,3), pady=(6,3))
		fonts = list(set(list(tkfont.families())))
		fonts.sort()
		font_opts = ttk.Combobox(prompt_frame, state="readonly", textvariable=self.font_var,
				values=fonts, width=25)
		font_opts.grid(row=3, column=1, columnspan=3, sticky=tk.W, padx=(3,3), pady=(3,3))
		self.bold_var = tk.StringVar(prompt_frame, "0" if not label_bold else "1")
		ck_bold = ttk.Checkbutton(prompt_frame, text="Bold", variable=self.bold_var)
		ck_bold.grid(row=3, column=4, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.size_var = tk.IntVar(prompt_frame, label_size)
		#
		size_lbl = ttk.Label(prompt_frame, text="Size:")
		size_lbl.grid(row=4, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		size_opt = ttk.Combobox(prompt_frame, state="normal", textvariable=self.size_var,
				values=[8, 10, 12, 14, 16, 20, 24], width=3)
		size_opt.grid(row=4, column=1, sticky=tk.W, padx=(6,3), pady=(3,3))
		self.position_var = tk.StringVar(prompt_frame, label_position)
		position_lbl = ttk.Label(prompt_frame, text="Position:")
		position_lbl.grid(row=4, column=2, sticky=tk.E, padx=(3,3), pady=(3,3))
		position_sel = ttk.Combobox(prompt_frame, state="readonly", textvariable=self.position_var,
				values=["above", "below"], width=6)
		position_sel.grid(row=4, column=3, sticky=tk.W, padx=(3,6), pady=(3,3))
		#
		# Buttons
		self.canceled = False
		help_btn = ttk.Button(btn_frame, text="Help", command=self.do_help, underline=0)
		help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		ok_btn = ttk.Button(btn_frame, text="OK", command=self.do_select, underline=0)
		ok_btn.grid(row=0, column=1, sticky=tk.E, padx=(3,3))
		self.dlg.bind('<Alt-o>', self.do_select)
		cancel_btn = ttk.Button(btn_frame, text="Cancel", command=self.do_cancel)
		cancel_btn.grid(row=0, column=2, sticky=tk.E, padx=(3,6))
		self.dlg.bind("<Return>", self.do_select)
		self.dlg.bind("<Escape>", self.do_cancel)
	def do_help(self, *args):
		webbrowser.open("https://mapdata.readthedocs.io/en/latest/dialogs.html#change-labeling", new=2, autoraise=True)
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def do_select(self, *args):
		self.canceled = False
		self.dlg.destroy()
	def get_labeling(self):
		self.dlg.grab_set()
		center_window(self.dlg)
		self.dlg.resizable(False, False)
		self.symbol_opts.focus()
		self.dlg.wait_window(self.dlg)
		if not self.canceled:
			return (self.symbol_var.get(), self.color_var.get(), self.use_data_marker_var.get(),
					self.use_data_color_var.get(), self.col_var.get(), self.font_var.get(), self.size_var.get(),
					self.bold_var.get(), self.position_var.get())
		else:
			return (None,None,None,None,None,None,None,None,None)



class MarkerDialog(object):
	def __init__(self, parent):
		self.dlg = tk.Toplevel()
		self.dlg.title("Change Marker")
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.NSEW, pady=(3,3))
		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.grid(row=1, column=0, sticky=tk.EW, pady=(3,3))
		btn_frame.columnconfigure(0, weight=1)
		symbol_lbl = ttk.Label(prompt_frame, text="Marker symbol:")
		symbol_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3))
		self.symbol_var = tk.StringVar(self.dlg, select_symbol)
		symbol_vals = list(icon_xbm.keys())
		symbol_vals.sort()
		self.symbol_opts = ttk.Combobox(prompt_frame, state="readonly", textvar=self.symbol_var,
				values=symbol_vals, width=15)
		self.symbol_opts.grid(row=0, column=1, sticky=tk.W, padx=(3,6))
		color_lbl = ttk.Label(prompt_frame, text="Color:")
		color_lbl.grid(row=1, column=0, sticky=tk.E, padx=(6,3))
		self.color_var = tk.StringVar(self.dlg, select_color)
		color_vals = list(select_colors)
		color_opts = ttk.Combobox(prompt_frame, state="readonly", textvar=self.color_var,
				values=color_vals, width=15)
		color_opts.grid(row=1, column=1, sticky=tk.W, padx=(3,6))
		# Buttons
		self.canceled = False
		help_btn = ttk.Button(btn_frame, text="Help", command=self.do_help, underline=0)
		help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		ok_btn = ttk.Button(btn_frame, text="OK", command=self.do_select, underline=0)
		ok_btn.grid(row=0, column=1, sticky=tk.E, padx=(3,3))
		self.dlg.bind('<Alt-o>', self.do_select)
		cancel_btn = ttk.Button(btn_frame, text="Cancel", command=self.do_cancel)
		cancel_btn.grid(row=0, column=2, sticky=tk.E, padx=(3,6))
		self.dlg.bind("<Return>", self.do_select)
		self.dlg.bind("<Escape>", self.do_cancel)
	def do_help(self, *args):
		webbrowser.open("https://mapdata.readthedocs.io/en/latest/dialogs.html#change-marker", new=2, autoraise=True)
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def do_select(self, *args):
		self.canceled = False
		self.dlg.destroy()
	def get_marker(self):
		self.dlg.grab_set()
		center_window(self.dlg)
		self.dlg.resizable(False, False)
		self.symbol_opts.focus()
		self.dlg.wait_window(self.dlg)
		if not self.canceled:
			return (self.symbol_var.get(), self.color_var.get())
		else:
			return (None, None)


class ImportSymbolDialog(object):
	def __init__(self):
		def get_fn(*args):
			fn = tkfiledialog.askopenfilename(filetypes=([('X11 bitmaps', '.xbm')]))
			if fn != '' and fn is not None and fn != ():
				self.fn_var.set(fn)
		def check_enable(*args):
			if self.fn_var.get() != '' and self.symbol_var.get() != '':
				self.ok_btn["state"] = tk.NORMAL
			else:
				self.ok_btn["state"] = tk.DISABLED
		self.dlg = tk.Toplevel()
		self.dlg.title("Import X11 Bitmap Symbol")
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.NSEW, pady=(3,3))
		button_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		button_frame.grid(row=1, column=0, columnspan=3, sticky=tk.EW, pady=(3,3))
		button_frame.columnconfigure(0, weight=1)
		btn_frame = tk.Frame(button_frame)
		btn_frame.grid(row=0, column=0, sticky=tk.EW)
		btn_frame.columnconfigure(0, weight=1)
		# Prompts
		symbol_lbl = ttk.Label(prompt_frame, text="Symbol name:")
		symbol_lbl.grid(row=0, column=0, sticky=tk.E, padx=(3,3))
		self.symbol_var = tk.StringVar(self.dlg, "")
		self.symbol_var.trace('w', check_enable)
		self.symbol_entry = ttk.Entry(prompt_frame, textvariable=self.symbol_var, width=12)
		self.symbol_entry.grid(row=0, column=1, sticky=tk.W, padx=(3,3))
		#
		fn_label = ttk.Label(prompt_frame, text="File:")
		fn_label.grid(row=1, column=0, sticky=tk.E, padx=(3,3))
		self.fn_var = tk.StringVar(prompt_frame, '')
		self.fn_var.trace('w', check_enable)
		fn_entry = ttk.Entry(prompt_frame, textvariable=self.fn_var)
		fn_entry.configure(width=64)
		fn_entry.grid(row=1, column=1, sticky=tk.EW, padx=(3,3))
		fn_btn = ttk.Button(prompt_frame, text="Browse", command=get_fn)
		fn_btn.grid(row=1, column=2, sticky=tk.W)
		# Buttons
		self.canceled = False
		help_btn = ttk.Button(btn_frame, text="Help", command=self.do_help, underline=0)
		help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		self.ok_btn = ttk.Button(btn_frame, text="OK", command=self.do_select, underline=0)
		self.ok_btn["state"] = tk.DISABLED
		self.ok_btn.grid(row=0, column=1, sticky=tk.E, padx=(3,3))
		self.dlg.bind('<Alt-o>', self.do_select)
		cancel_btn = ttk.Button(btn_frame, text="Cancel", command=self.do_cancel)
		cancel_btn.grid(row=0, column=2, sticky=tk.E, padx=(3,6))
		self.dlg.bind("<Return>", self.do_select)
		self.dlg.bind("<Escape>", self.do_cancel)
	def do_help(self, *args):
		webbrowser.open("https://mapdata.readthedocs.io/en/latest/dialogs.html#import-symbol", new=2, autoraise=True)
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def do_select(self, *args):
		if self.ok_btn["state"] != tk.DISABLED:
			self.canceled = False
			self.dlg.destroy()
	def run(self):
		self.dlg.grab_set()
		center_window(self.dlg)
		self.dlg.resizable(False, False)
		self.symbol_entry.focus()
		self.dlg.wait_window(self.dlg)
		if not self.canceled:
			return (self.symbol_var.get(), self.fn_var.get())
		else:
			return (None, None)



class DataFileDialog(object):
	def __init__(self):
		def get_fn():
			fn = tkfiledialog.askopenfilename(filetypes=([('CSV files', '.csv')]), parent=self.dlg)
			if fn != '' and fn != () and fn is not None:
				self.fn_var.set(fn)
				csvreader = CsvFile(fn)
				self.header_list = csvreader.next()
				self.id_sel["values"] = self.header_list
				self.lat_sel["values"] = self.header_list
				self.lon_sel["values"] = self.header_list
				self.sym_sel["values"] = self.header_list
				self.col_sel["values"] = self.header_list
		def check_enable(*args):
			if self.fn_var.get() != '' and self.lat_var.get() != '' and self.lon_var.get() != '':
				ok_btn["state"] = tk.NORMAL
			else:
				ok_btn["state"] = tk.DISABLED
		def new_fn(*args):
			check_enable()
		self.header_list = []
		self.dlg = tk.Toplevel()
		self.dlg.title("Open CSV Data File for Map Display")
		# Main frames
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.NSEW, padx=(3,3), pady=(3,3))
		dir_frame = tk.Frame(prompt_frame)
		dir_frame.grid(row=0, column=0, sticky=tk.EW, padx=(3,3), pady=(3,3))
		dir_frame.rowconfigure(0, weight=1)
		dir_frame.columnconfigure(0, weight=1)
		req_frame = ttk.LabelFrame(prompt_frame, text="Required")
		req_frame.grid(row=1, column=0, sticky=tk.EW, padx=(6,3), pady=(3,3))
		req_frame.columnconfigure(0, weight=1)
		opt_frame = ttk.LabelFrame(prompt_frame, text="Optional")
		opt_frame.grid(row=2, column=0, sticky=tk.EW, padx=(6,3), pady=(9,3))
		opt_frame.columnconfigure(0, weight=1)
		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.grid(row=1, column=0, sticky=tk.EW, padx=(3,3), pady=(3,3))
		btn_frame.columnconfigure(0, weight=1)
		# Prompts
		#-- Directions
		dir_lbl = ttk.Label(dir_frame,
				text="Select a CSV file with columns containing latitude and longitude values, and optionally other information.",
				width=80, justify=tk.LEFT, wraplength=600)
		dir_lbl.grid(row=0, column=0, padx=(3,3), pady=(3,3))
		def wrap_msg(event):
			dir_lbl.configure(wraplength=event.width - 5)
		dir_lbl.bind("<Configure>", wrap_msg)
		#-- Filename
		fn_frame = tk.Frame(req_frame)
		fn_frame.grid(row=0, column=0, sticky=tk.EW, pady=(5,5))
		fn_label = ttk.Label(fn_frame, text="File:")
		fn_label.grid(row=0, column=0, sticky=tk.E, padx=(3,3))
		self.fn_var = tk.StringVar(fn_frame, '')
		self.fn_var.trace('w', new_fn)
		fn_entry = ttk.Entry(fn_frame, textvariable=self.fn_var)
		fn_entry.configure(width=64)
		fn_entry.grid(row=0, column=1, sticky=tk.EW, padx=(3,3))
		fn_btn = ttk.Button(fn_frame, text="Browse", command=get_fn)
		fn_btn.grid(row=0, column=2, sticky=tk.W, padx=(3,3))
		#-- Required columns
		column_choices = list(self.header_list)
		#
		req_col_frame = tk.Frame(req_frame)
		req_col_frame.grid(row=1, column=0, sticky=tk.EW, pady=(3,3))
		lat_label = ttk.Label(req_col_frame, text="Latitude column:")
		lat_label.grid(row=0, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.lat_var = tk.StringVar(req_col_frame, '')
		self.lat_var.trace('w', check_enable)
		self.lat_sel = ttk.Combobox(req_col_frame, state="readonly", textvariable=self.lat_var, values=self.header_list, width=12)
		self.lat_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,3), pady=(3,3))
		#
		lon_label = ttk.Label(req_col_frame, text="Longitude column:")
		lon_label.grid(row=0, column=2, sticky=tk.E, padx=(20,3), pady=(3,3))
		self.lon_var = tk.StringVar(req_frame, '')
		self.lon_var.trace('w', check_enable)
		self.lon_sel = ttk.Combobox(req_col_frame, state="readonly", textvariable=self.lon_var, values=self.header_list, width=12)
		self.lon_sel.grid(row=0, column=3, sticky=tk.W, padx=(3,3), pady=(3,3))
		#-- Optional columns
		opt_col_frame = tk.Frame(opt_frame)
		opt_col_frame.grid(row=2, column=0, sticky=tk.EW, pady=(3,3))
		id_label = ttk.Label(opt_col_frame, text="Label column:")
		id_label.grid(row=0, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.id_var = tk.StringVar(opt_col_frame, '')
		self.id_sel = ttk.Combobox(opt_col_frame, state="readonly", textvariable=self.id_var, values=self.header_list, width=12)
		self.id_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,20), pady=(3,3))
		#
		crs_label = ttk.Label(opt_col_frame, text="CRS:")
		crs_label.grid(row=0, column=2, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.crs_var = tk.IntVar(opt_col_frame, 4326)
		self.crs_var.trace('w', check_enable)
		self.crs_sel = ttk.Entry(opt_col_frame, width=8, textvariable=self.crs_var)
		self.crs_sel.grid(row=0, column=3, sticky=tk.W, padx=(3,20), pady=(3,3))
		#
		sym_label = ttk.Label(opt_col_frame, text="Symbol column:")
		sym_label.grid(row=1, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.sym_var = tk.StringVar(opt_col_frame, '')
		self.sym_sel = ttk.Combobox(opt_col_frame, state="readonly", textvariable=self.sym_var, values=self.header_list, width=12)
		self.sym_sel.grid(row=1, column=1, sticky=tk.W, padx=(3,20), pady=(3,3))
		#
		col_label = ttk.Label(opt_col_frame, text="Color column:")
		col_label.grid(row=1, column=2, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.col_var = tk.StringVar(opt_col_frame, '')
		self.col_sel = ttk.Combobox(opt_col_frame, state="readonly", textvariable=self.col_var, values=self.header_list, width=12)
		self.col_sel.grid(row=1, column=3, sticky=tk.W, padx=(3,20), pady=(3,3))
		#-- Description
		title_label = ttk.Label(opt_col_frame, text="Description:")
		title_label.grid(row=2, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.title_var = tk.StringVar(opt_col_frame, '')
		title_entry = ttk.Entry(opt_col_frame, width=60, textvariable=self.title_var)
		title_entry.grid(row=2, column=1, columnspan=3, sticky=tk.EW, padx=(3,6), pady=(3,3))
		# Buttons
		self.canceled = False
		help_btn = ttk.Button(btn_frame, text="Help", command=self.do_help, underline=0)
		help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		ok_btn = ttk.Button(btn_frame, text="OK", command=self.do_select, underline=0)
		ok_btn.grid(row=0, column=1, sticky=tk.E, padx=3)
		self.dlg.bind('<Alt-o>', self.do_select)
		cancel_btn = ttk.Button(btn_frame, text="Cancel", command=self.do_cancel)
		cancel_btn.grid(row=0, column=2, sticky=tk.E, padx=(3,6))
		ok_btn["state"] = tk.DISABLED
		self.dlg.bind("<Return>", self.do_select)
		self.dlg.bind("<Escape>", self.do_cancel)
		self.dlg.resizable(False, False)
	def do_help(self, *args):
		webbrowser.open("https://mapdata.readthedocs.io/en/latest/dialogs.html#open-csv-data-file", new=2, autoraise=True)
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def do_select(self, *args):
		if self.fn_var.get() != '' and self.lat_var.get() != '' and self.lon_var.get() != '':
			self.canceled = False
			self.dlg.destroy()
	def get_datafile(self):
		self.dlg.grab_set()
		center_window(self.dlg)
		self.dlg.resizable(False, False)
		self.dlg.wait_window(self.dlg)
		self.dlg = None
		if not self.canceled:
			headers, rows = file_data(self.fn_var.get())
			return (self.fn_var.get(), self.id_var.get(), self.lat_var.get(), self.lon_var.get(),
					self.crs_var.get(), self.sym_var.get(), self.col_var.get(), self.title_var.get(),
					headers, rows)
		else:
			return (None, None, None, None, None, None, None, None, None, None)



class ImportSpreadsheetDialog(object):
	def __init__(self, parent, mapui):
		self.parent = parent
		self.mapui = mapui
		def get_fn(*args):
			fn = tkfiledialog.askopenfilename(filetypes=([('Spreadsheets', '.ods .xlsx .xls')]), parent=self.dlg)
			if fn != '' and fn != () and fn is not None:
				self.fn_var.set(fn)
		def check_w1enable(*args):
			if self.fn_var.get() != '':
				if os.path.isfile(self.fn_var.get()):
					w1next_btn["state"] = tk.NORMAL
				else:
					w1next_btn["state"] = tk.DISABLED
			else:
				w1next_btn["state"] = tk.DISABLED
		def check_w2enable(*args):
			if self.fn_var.get() != '' and self.sheet_var.get() != '':
				w2next_btn["state"] = tk.NORMAL
			else:
				w2next_btn["state"] = tk.DISABLED
		def check_w3enable(*args):
			if self.fn_var.get() != '' and self.sheet_var.get() != '' and self.lat_var.get() != '' and self.lon_var.get() != '':
				w3ok_btn["state"] = tk.NORMAL
			else:
				w3ok_btn["state"] = tk.DISABLED
		def new_fn(*args):
			check_w1enable()
		self.sheet_list = []
		self.header_list = []
		self.dlg = tk.Toplevel(parent)
		self.dlg.title("Open Spreadsheet File for Map Display")
		# Main frames
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.EW, padx=(3,3), pady=(3,3))
		# Wizard frames 1, 2, and 3 are gridded in the same cell to make a wizard.
		self.dlg.rowconfigure(0, weight=0)
		wiz1_frame = tk.Frame(self.dlg)		# For description, filename, and sheet name
		wiz1_frame.grid(row=1, column=0, sticky=tk.NSEW, padx=(3,3), pady=(3,3))
		wiz1_frame.rowconfigure(0, weight=1)
		wiz2_frame = tk.Frame(self.dlg)		# For sheet selections
		wiz2_frame.grid(row=1, column=0, sticky=tk.NSEW, padx=(3,3), pady=(3,3))
		wiz2_frame.rowconfigure(0, weight=1)
		wiz2_frame.columnconfigure(0, weight=1)
		wiz3_frame = tk.Frame(self.dlg)		# For column selections
		wiz3_frame.grid(row=1, column=0, sticky=tk.NSEW, padx=(3,3), pady=(3,3))
		wiz3_frame.columnconfigure(0, weight=1)
		self.dlg.rowconfigure(1, weight=0)
		self.dlg.resizable(False, False)
		wiz1_frame.lift()

		# Populate directions frame
		dir_lbl = ttk.Label(prompt_frame,
				text="Select a spreadsheet file with columns containing latitude and longitude values, and optionally other information.",
				width=80, justify=tk.LEFT, wraplength=600)
		dir_lbl.grid(row=0, column=0, padx=(6,6), pady=(3,3))
		def wrap_msg(event):
			dir_lbl.configure(wraplength=event.width - 5)
		dir_lbl.bind("<Configure>", wrap_msg)

		# Populate wiz1_frame
		w1req_frame = ttk.LabelFrame(wiz1_frame, text="Required")
		w1req_frame.grid(row=0, column=0, sticky=tk.EW, padx=(6,6), pady=(3,3))
		w1req_frame.columnconfigure(0, weight=1)
		fn_frame = tk.Frame(w1req_frame)
		fn_frame.grid(row=0, column=0, sticky=tk.EW, pady=(3,3))
		fn_label = ttk.Label(fn_frame, text="File:")
		fn_label.grid(row=0, column=0, sticky=tk.E, padx=(3,3))
		self.fn_var = tk.StringVar(fn_frame, '')
		self.fn_var.trace('w', new_fn)
		fn_entry = ttk.Entry(fn_frame, textvariable=self.fn_var)
		fn_entry.configure(width=64)
		fn_entry.grid(row=0, column=1, sticky=tk.EW, padx=(3,3))
		fn_btn = ttk.Button(fn_frame, text="Browse", command=get_fn, underline=0)
		fn_btn.grid(row=0, column=2, sticky=tk.W, padx=(3,3))
		self.dlg.bind("<Alt-b>", get_fn)

		w1opt_frame = ttk.LabelFrame(wiz1_frame, text="Optional")
		w1opt_frame.grid(row=1, column=0, sticky=tk.EW, padx=(6,3), pady=(9,3))
		w1opt_frame.columnconfigure(0, weight=1)
		desc_label = ttk.Label(w1opt_frame, text="Description:")
		desc_label.grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.desc_var = tk.StringVar(w1opt_frame, '')
		desc_entry = ttk.Entry(w1opt_frame, width=60, textvariable=self.desc_var)
		desc_entry.grid(row=0, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))

		def w1_next(*args):
			if self.fn_var.get() != '':
				# Open spreadsheet, get sheet names
				self.mapui.loading_dlg.display("Importing spreadsheet")
				fn, ext = os.path.splitext(self.fn_var.get())
				ext = ext.lower()
				try:
					if ext == '.ods':
						sso = OdsFile()
					elif ext == '.xlsx':
						sso = XlsxFile()
					else:
						sso = XlsFile()
				except:
					warning("Could not open %s" % self.fn_var.get(), kwargs={'parent': self.dlg})
				else:
					sso.open(self.fn_var.get())
					self.sheet_list = sso.sheetnames()
					self.sheet_sel["values"] = self.sheet_list
					if ext in ('.ods', '.xlsx'):
						try:
							sso.close()
						except:
							pass
					else:
						try:
							sso.release_resources()
							del sso
						except:
							pass
					self.dlg.bind("<Alt-b>")
					self.dlg.bind("<Alt-n>")
					wiz2_frame.lift()
					self.dlg.bind("<Alt-b>", w2_back)
					self.dlg.bind("<Alt-n>", w2_next)
				self.mapui.loading_dlg.hide()

		w1btn_frame = tk.Frame(wiz1_frame, borderwidth=3, relief=tk.RIDGE)
		w1btn_frame.grid(row=2, column=0, sticky=tk.EW, padx=(3,3), pady=(3,3))
		w1btn_frame.columnconfigure(0, weight=1)
		self.canceled = False
		#
		w1help_btn = ttk.Button(w1btn_frame, text="Help", command=self.do_help, underline=0)
		w1help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		w1next_btn = ttk.Button(w1btn_frame, text="Next", command=w1_next, underline=0)
		w1next_btn.grid(row=0, column=1, sticky=tk.E, padx=3)
		self.dlg.bind("<Alt-n>", w1_next)
		w1cancel_btn = ttk.Button(w1btn_frame, text="Cancel", command=self.do_cancel, underline=0)
		w1cancel_btn.grid(row=0, column=2, sticky=tk.E, padx=(3,6))
		w1next_btn["state"] = tk.DISABLED
		self.dlg.bind("<Alt-c>", self.do_cancel)
		self.dlg.bind("<Escape>", self.do_cancel)


		# Populate wiz2_frame
		w2req_frame = ttk.LabelFrame(wiz2_frame, text="Required")
		w2req_frame.grid(row=0, column=0, sticky=tk.EW, padx=(6,6), pady=(3,3))
		w2req_frame.columnconfigure(0, weight=1)
		w2inner_frame = tk.Frame(w2req_frame)
		w2inner_frame.grid(row=0, column=0, sticky=tk.W)
		#
		sheet_label = ttk.Label(w2inner_frame, text="Sheet:")
		sheet_label.grid(row=0, column=0, sticky=tk.E, padx=(3,3))
		self.sheet_var = tk.StringVar(w2req_frame, '')
		self.sheet_var.trace('w', check_w2enable)
		self.sheet_sel = ttk.Combobox(w2inner_frame, state="readonly", textvariable=self.sheet_var, values=self.sheet_list, width=16)
		self.sheet_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,3))
		#
		xrows_label = ttk.Label(w2inner_frame, text="Initial rows to skip:")
		xrows_label.grid(row=1, column=0, sticky=tk.E, padx=(3,3))
		self.xrows_var = tk.IntVar(w2req_frame, 0)
		self.xrows_var.trace('w', check_w2enable)
		xrows_entry = ttk.Entry(w2inner_frame, textvariable=self.xrows_var, width=6)
		xrows_entry.grid(row=1, column=1, sticky=tk.W, padx=(3,3))

		def w2_back(*args):
			self.dlg.bind("<Alt-b>")
			self.dlg.bind("<Alt-n>")
			wiz1_frame.lift()
			self.dlg.bind("<Alt-n>", w1_next)
			self.dlg.bind("<Alt-b>", get_fn)

		def w2_next(*args):
			# Open spreadsheet, get column names
			if self.fn_var.get() != '' and self.sheet_var.get() != '':
				fn, ext = os.path.splitext(self.fn_var.get())
				try:
					if ext.lower() == '.ods':
						hdrs, data = ods_data(self.fn_var.get(), self.sheet_var.get(), junk_header_rows=self.xrows_var.get())
					else:
						hdrs, data = xls_data(self.fn_var.get(), self.sheet_var.get(), junk_header_rows=self.xrows_var.get())
				except:
					warning("Could not read table from %s, sheet %s" % (self.fn_var.get(), self.sheet_var.get()), 
							kwargs={'parent': self.dlg})
				else:
					self.headers = hdrs
					self.header_list = list(hdrs)
					self.rows = data
					# Set list box values
					self.id_sel["values"] = self.header_list
					self.lat_sel["values"] = self.header_list
					self.lon_sel["values"] = self.header_list
					self.sym_sel["values"] = self.header_list
					self.col_sel["values"] = self.header_list
					self.dlg.bind("<Alt-b>")
					self.dlg.bind("<Alt-n>")
					wiz3_frame.lift()
					self.dlg.bind("<Alt-b>", w3_back)

		w2btn_frame = tk.Frame(wiz2_frame, borderwidth=3, relief=tk.RIDGE)
		w2btn_frame.grid(row=2, column=0, sticky=tk.EW, padx=(3,3), pady=(3,3))
		w2btn_frame.columnconfigure(0, weight=1)
		#
		w2help_btn = ttk.Button(w2btn_frame, text="Help", command=self.do_help, underline=0)
		w2help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		w2prev_btn = ttk.Button(w2btn_frame, text="Back", command=w2_back, underline=0)
		w2prev_btn.grid(row=0, column=1, sticky=tk.E, padx=3)
		w2next_btn = ttk.Button(w2btn_frame, text="Next", command=w2_next, underline=0)
		w2next_btn.grid(row=0, column=2, sticky=tk.E, padx=3)
		w2cancel_btn = ttk.Button(w2btn_frame, text="Cancel", command=self.do_cancel, underline=0)
		w2cancel_btn.grid(row=0, column=3, sticky=tk.E, padx=(3,6))
		w2next_btn["state"] = tk.DISABLED
	
		# Populate wiz3_frame
		w3req_frame = ttk.LabelFrame(wiz3_frame, text="Required")
		w3req_frame.grid(row=0, column=0, sticky=tk.EW, padx=(6,6), pady=(3,3))
		w3req_frame.columnconfigure(0, weight=1)
		#
		lat_label = ttk.Label(w3req_frame, text="Latitude column:")
		lat_label.grid(row=0, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.lat_var = tk.StringVar(w3req_frame, '')
		self.lat_var.trace('w', check_w3enable)
		self.lat_sel = ttk.Combobox(w3req_frame, state="readonly", textvariable=self.lat_var, values=self.header_list, width=15)
		self.lat_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,30), pady=(3,3))
		#
		lon_label = ttk.Label(w3req_frame, text="Longitude column:")
		lon_label.grid(row=0, column=2, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.lon_var = tk.StringVar(w3req_frame, '')
		self.lon_var.trace('w', check_w3enable)
		self.lon_sel = ttk.Combobox(w3req_frame, state="readonly", textvariable=self.lon_var, values=self.header_list, width=15)
		self.lon_sel.grid(row=0, column=3, sticky=tk.W, padx=(3,6), pady=(3,3))


		w3opt_frame = ttk.LabelFrame(wiz3_frame, text="Optional")
		w3opt_frame.grid(row=1, column=0, sticky=tk.EW, padx=(6,6), pady=(9,3))
		w3opt_frame.columnconfigure(0, weight=1)
		#
		id_label = ttk.Label(w3opt_frame, text="Label column:")
		id_label.grid(row=0, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.id_var = tk.StringVar(w3opt_frame, '')
		self.id_sel = ttk.Combobox(w3opt_frame, state="readonly", textvariable=self.id_var, values=self.header_list, width=12)
		self.id_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,30), pady=(3,3))
		#
		crs_label = ttk.Label(w3opt_frame, text="CRS:")
		crs_label.grid(row=0, column=2, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.crs_var = tk.IntVar(w3opt_frame, 4326)
		self.crs_var.trace('w', check_w2enable)
		self.crs_sel = ttk.Entry(w3opt_frame, width=8, textvariable=self.crs_var)
		self.crs_sel.grid(row=0, column=3, sticky=tk.W, padx=(3,6), pady=(3,3))
		#
		sym_label = ttk.Label(w3opt_frame, text="Symbol column:")
		sym_label.grid(row=1, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.sym_var = tk.StringVar(w3opt_frame, '')
		self.sym_sel = ttk.Combobox(w3opt_frame, state="readonly", textvariable=self.sym_var, values=self.header_list, width=12)
		self.sym_sel.grid(row=1, column=1, sticky=tk.W, padx=(3,30), pady=(3,3))
		#
		col_label = ttk.Label(w3opt_frame, text="Color column:")
		col_label.grid(row=1, column=2, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.col_var = tk.StringVar(w3opt_frame, '')
		self.col_sel = ttk.Combobox(w3opt_frame, state="readonly", textvariable=self.col_var, values=self.header_list, width=12)
		self.col_sel.grid(row=1, column=3, sticky=tk.W, padx=(3,6), pady=(3,3))

		def w3_back(*args):
			self.dlg.bind("<Alt-b>")
			wiz2_frame.lift()
			self.dlg.bind("<Alt-b>", w2_back)
			self.dlg.bind("<Alt-n>", w2_next)

		w3btn_frame = tk.Frame(wiz3_frame, borderwidth=3, relief=tk.RIDGE)
		w3btn_frame.grid(row=2, column=0, sticky=tk.EW, padx=(3,3), pady=(3,3))
		w3btn_frame.columnconfigure(0, weight=1)
		#
		w3help_btn = ttk.Button(w3btn_frame, text="Help", command=self.do_help, underline=0)
		w3help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		w3prev_btn = ttk.Button(w3btn_frame, text="Back", command=w3_back)
		w3prev_btn.grid(row=0, column=1, sticky=tk.E, padx=3)
		w3ok_btn = ttk.Button(w3btn_frame, text="OK", command=self.do_select, underline=0)
		w3ok_btn.grid(row=0, column=2, sticky=tk.E, padx=3)
		self.dlg.bind('<Alt-o>', self.do_select)
		w3cancel_btn = ttk.Button(w3btn_frame, text="Cancel", command=self.do_cancel, underline=0)
		w3cancel_btn.grid(row=0, column=3, sticky=tk.E, padx=(3,6))
		w3ok_btn["state"] = tk.DISABLED
	
	def do_help(self, *args):
		webbrowser.open("https://mapdata.readthedocs.io/en/latest/dialogs.html#open-spreadsheet-data-file", new=2, autoraise=True)
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def do_select(self, *args):
		if self.fn_var.get() != '' and self.sheet_var.get() != '' and self.lat_var.get() != '' and self.lon_var.get() != '':
			self.canceled = False
			self.dlg.destroy()
	def get_datafile(self):
		self.dlg.grab_set()
		center_window(self.dlg)
		self.dlg.resizable(False, False)
		self.dlg.wait_window(self.dlg)
		self.dlg = None
		if not self.canceled:
			return (self.fn_var.get(), self.id_var.get(), self.lat_var.get(), self.lon_var.get(),
					self.crs_var.get(), self.sym_var.get(), self.col_var.get(), self.desc_var.get(),
					self.headers, self.rows)
		else:
			return (None, None, None, None, None, None, None, None, None, None)


class NewCrsDialog(object):
	def __init__(self, current_crs):
		self.dlg = tk.Toplevel()
		self.dlg.title("Change CRS")
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.NSEW, pady=(3,3))
		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.grid(row=1, column=0, sticky=tk.EW, pady=(3,3))
		btn_frame.rowconfigure(0, weight=1)
		crs_lbl = ttk.Label(prompt_frame, text="New CRS:")
		crs_lbl.grid(row=0, column=0, sticky=tk.E, padx=(3,3))
		self.crs_var = tk.IntVar(self.dlg, current_crs)
		self.crs_entry = ttk.Entry(prompt_frame, width=12, textvariable=self.crs_var)
		self.crs_entry.grid(row=0, column=1, sticky=tk.W, padx=(3,3))
		# Buttons
		self.canceled = False
		help_btn = ttk.Button(btn_frame, text="Help", command=self.do_help, underline=0)
		help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		ok_btn = ttk.Button(btn_frame, text="OK", command=self.do_select, underline=0)
		ok_btn.grid(row=0, column=1, sticky=tk.E, padx=(3,3))
		self.dlg.bind('<Alt-o>', self.do_select)
		cancel_btn = ttk.Button(btn_frame, text="Cancel", command=self.do_cancel)
		cancel_btn.grid(row=0, column=2, sticky=tk.E, padx=(3,6))
		self.dlg.bind("<Return>", self.do_select)
		self.dlg.bind("<Escape>", self.do_cancel)
	def do_help(self, *args):
		webbrowser.open("https://mapdata.readthedocs.io/en/latest/dialogs.html#change-crs", new=2, autoraise=True)
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def do_select(self, *args):
		self.canceled = False
		self.dlg.destroy()
	def get_crs(self):
		self.dlg.grab_set()
		center_window(self.dlg)
		self.dlg.resizable(False, False)
		self.crs_entry.focus()
		self.dlg.wait_window(self.dlg)
		if not self.canceled:
			return self.crs_var.get()
		else:
			return None


class QueryDialog(object):
	def __init__(self, column_headers, db_conn, init_sql=""):
		self.dlg = tk.Toplevel()
		self.dlg.title("Query Data")
		self.canceled = True
		self.dlg.columnconfigure(0, weight=1)
		self.dlg.rowconfigure(1, weight=1)
		# Frames
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.NSEW, padx=(3,3), pady=(3,3))
		prompt_frame.columnconfigure(0, weight=1)
		ctrlfrm = tk.Frame(self.dlg)
		ctrlfrm.grid(row=1, column=0, sticky=tk.NSEW)
		ctrlfrm.rowconfigure(0, weight=1)
		ctrlfrm.columnconfigure(0, weight=1)
		query_frame = tk.Frame(ctrlfrm)
		query_frame.grid(row=0, column=0, sticky=tk.NSEW, padx=(3,3), pady=(3,3))
		query_frame.rowconfigure(1, weight=1)
		query_frame.columnconfigure(0, weight=1)
		query_frame.columnconfigure(0, weight=3)
		#query_frame.columnconfigure(1, weight=3)
		col_frame = tk.Frame(ctrlfrm)
		col_frame.grid(row=0, column=1, rowspan=2, sticky=tk.NS, padx=(3,3), pady=(3,3))
		col_frame.rowconfigure(0, weight=1)
		chars_frame = tk.Frame(query_frame)
		chars_frame.grid(row=0, column=0, sticky = tk.EW, padx=(3,3))
		sql_frame = tk.Frame(query_frame)
		sql_frame.grid(row=1, column=0, sticky=tk.NSEW, padx=(3,3), pady=(3,3))
		sql_frame.rowconfigure(0, weight=1)
		sql_frame.columnconfigure(0, weight=1)
		act_frame = tk.Frame(query_frame)
		act_frame.grid(row=2, column=0, sticky=tk.EW, padx=(3,3), pady=(3,3))
		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.grid(row=2, column=0, sticky=tk.EW, pady=(3,3))
		btn_frame.columnconfigure(0, weight=1)
		# Prompt
		prompt_lbl = ttk.Label(prompt_frame, wraplength=300, justify=tk.LEFT, text="Enter an expression below to identify the data rows that you want to select.  The syntax of this expression should correspond to a SQL 'WHERE' clause.  Column names with non-alphanumeric characters should be double-quoted.  String literals should be single-quoted.  The '%' character is a wildcard.  Ctrl-Enter completes entry.")
		prompt_lbl.grid(row=0, column=0, sticky=tk.EW, padx=(3,3))
		def wrap_prompt(event):
			prompt_lbl.configure(wraplength=event.width - 5)
		prompt_lbl.bind("<Configure>", wrap_prompt)
		def clear_sql():
			self.sql_text.delete(1.0, tk.END)
			self.ok_btn["state"] = tk.DISABLED
		# SQL text-insertion buttons
		ttk.Button(chars_frame, text="=", width=1, command=lambda:self.sql_text.insert(tk.END, "=")).grid(row=0, column=0, sticky=tk.W)
		ttk.Button(chars_frame, text="<>", width=3, command=lambda:self.sql_text.insert(tk.END, "<>")).grid(row=0, column=1, sticky=tk.W)
		ttk.Button(chars_frame, text="<", width=1, command=lambda:self.sql_text.insert(tk.END, "<")).grid(row=0, column=2, sticky=tk.W)
		ttk.Button(chars_frame, text=">", width=1, command=lambda:self.sql_text.insert(tk.END, ">")).grid(row=0, column=3, sticky=tk.W)
		ttk.Button(chars_frame, text="not", width=3, command=lambda:self.sql_text.insert(tk.END, " not ")).grid(row=0, column=4, sticky=tk.W)
		ttk.Button(chars_frame, text="in", width=2, command=lambda:self.sql_text.insert(tk.END, " in ")).grid(row=0, column=5, sticky=tk.W)
		ttk.Button(chars_frame, text="(", width=1, command=lambda:self.sql_text.insert(tk.END, " (")).grid(row=0, column=6, sticky=tk.W)
		ttk.Button(chars_frame, text=",", width=1, command=lambda:self.sql_text.insert(tk.END, ",")).grid(row=0, column=7, sticky=tk.W)
		ttk.Button(chars_frame, text=")", width=1, command=lambda:self.sql_text.insert(tk.END, ")")).grid(row=0, column=8, sticky=tk.W)
		ttk.Button(chars_frame, text="'", width=1, command=lambda:self.sql_text.insert(tk.END, "'")).grid(row=0, column=9, sticky=tk.W)
		ttk.Button(chars_frame, text="and", width=3, command=lambda:self.sql_text.insert(tk.END, " and ")).grid(row=0, column=10, sticky=tk.W)
		ttk.Button(chars_frame, text="or", width=2, command=lambda:self.sql_text.insert(tk.END, " or ")).grid(row=0, column=11, sticky=tk.W)
		ttk.Button(chars_frame, text="like", width=4, command=lambda:self.sql_text.insert(tk.END, " like ")).grid(row=0, column=12, sticky=tk.W)
		ttk.Button(chars_frame, text="%", width=2, command=lambda:self.sql_text.insert(tk.END, "%")).grid(row=0, column=13, sticky=tk.W)
		ttk.Button(chars_frame, text="_", width=1, command=lambda:self.sql_text.insert(tk.END, "_")).grid(row=0, column=14, sticky=tk.W)
		ttk.Button(chars_frame, text="Clear", width=5, command=clear_sql).grid(row=0, column=15, sticky=tk.E, padx=(3,24))
		chars_frame.columnconfigure(15, weight=1)
		# SQL text entry
		self.sql = init_sql
		self.sql_text = tk.Text(sql_frame, width=60, height=10)
		if init_sql is not None and init_sql != "":
			self.sql_text.insert(tk.END, init_sql)
		self.sql_text.grid(row=0, column=0, sticky=tk.NSEW, padx=(3,0), pady=(3,3))
		self.sql_text.bind("<KeyRelease>", self.check_enable)
		sbar = tk.Scrollbar(sql_frame)
		sbar.grid(row=0, column=1, sticky=tk.NS, padx=(0,3), pady=(3,3))
		sbar.config(command=self.sql_text.yview)
		self.sql_text.config(yscrollcommand = sbar.set)
		# Column values
		col_lbl = ttk.Label(col_frame, text="Column values:")
		col_lbl.grid(row=0, column=0, sticky=tk.NW, padx=(3,3), pady=(3,3))
		col_var = tk.StringVar(col_frame, "")
		colsel = ttk.Combobox(col_frame, state="readonly", textvariable=col_var, values=column_headers, width=20)
		colsel.grid(row=1, column=0, sticky=tk.NW, padx=(3,3), pady=(3,3))
		def copycol():
			cval = col_var.get()
			if cval != '':
				self.sql_text.insert(tk.END, cval)
				self.ok_btn["state"] = tk.NORMAL
		self.colcopy = ttk.Button(col_frame, state="disabled", text="\u2B05", width=2, command=copycol)
		self.colcopy.grid(row=1, column=1, sticky=tk.W, pady=(3,3))
		tv_frame = tk.Frame(col_frame)
		tv_frame.grid(row=2, column=0, sticky=tk.NS, padx=(3,3), pady=(3,3))
		col_frame.rowconfigure(0, weight=0)
		col_frame.rowconfigure(1, weight=0)
		col_frame.rowconfigure(2, weight=1)
		col_frame.columnconfigure(2, weight=1)
		def colval_to_sql(event):
			item_iid = self.tv_tbl.identify('item', event.x, event.y)
			item_val = self.tv_tbl.item(item_iid, "values")[0]
			if not isfloat(item_val):
				item_val = "'%s'" % item_val
			self.sql_text.insert(tk.END, " "+item_val)
			self.ok_btn["state"] = tk.NORMAL
			self.sql_text.focus()
		def list_col_vals(event):
			curs = db_conn.cursor()
			colname = dquote(col_var.get())
			res = curs.execute('SELECT DISTINCT %s FROM mapdata ORDER BY %s' % (colname, colname))
			rowset = res.fetchall()
			for widget in tv_frame.winfo_children():
				widget.destroy()
			tblframe, self.tv_tbl = treeview_table(tv_frame, rowset, [col_var.get()])
			tblframe.grid(column=0, row=0, sticky=tk.NSEW)
			tv_frame.rowconfigure(0, weight=1)
			tv_frame.columnconfigure(0, weight=1)
			curs.close()
			self.tv_tbl.bind("<Double-1>", colval_to_sql)
			self.colcopy["state"] = "normal"
		colsel.bind("<<ComboboxSelected>>", list_col_vals)
		# Action selection
		self.act_var = tk.StringVar(act_frame, "Replace")
		act_lbl = ttk.Label(act_frame, text="Action:")
		act_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3))
		act_sel = ttk.Combobox(act_frame, state="readonly", textvariable=self.act_var, values=["Replace", "Union", "Intersection", "Difference O-N", "Difference N-O"], width=15)
		act_sel.grid(row=1, column=1, sticky=tk.W, padx=(3,6))
		# Buttons
		self.canceled = False
		help_btn = ttk.Button(btn_frame, text="Help", command=self.do_help, underline=0)
		help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		self.ok_btn = ttk.Button(btn_frame, text="OK", command=self.do_select, underline=0)
		self.ok_btn.grid(row=0, column=1, sticky=tk.E, padx=(3,3))
		if init_sql or "" == "":
			self.ok_btn["state"] = tk.DISABLED
		self.dlg.bind('<Alt-o>', self.do_select)
		cancel_btn = ttk.Button(btn_frame, text="Cancel", command=self.do_cancel, underline=0)
		cancel_btn.grid(row=0, column=2, sticky=tk.E, padx=(3,6))
		self.dlg.bind("<Escape>", self.do_cancel)
		self.dlg.bind("<Alt-c>", self.do_cancel)
		self.dlg.bind("<Alt-o>", self.do_select)
		self.dlg.bind("<Control-Return>", self.do_select)
	def check_enable(self, *args):
		if self.sql_text.get("1.0", "end-1c") != '':
			self.ok_btn["state"] = tk.NORMAL
		else:
			self.ok_btn["state"] = tk.DISABLED
	def do_help(self, *args):
		webbrowser.open("https://mapdata.readthedocs.io/en/latest/dialogs.html#query-data", new=2, autoraise=True)
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def do_select(self, *args):
		self.sql = self.sql_text.get("1.0", "end-1c")
		if len(self.sql) > 0:
			self.canceled = False
			self.dlg.destroy()
	def get_where(self):
		self.dlg.grab_set()
		center_window(self.dlg)
		self.sql_text.focus()
		self.dlg.wait_window(self.dlg)
		if not self.canceled:
			return (self.sql, self.act_var.get())
		else:
			return (None, None)



class PlotDialog(object):
	def __init__(self, parent, column_specs):
		self.parent = parent
		self.column_specs = column_specs
		self.dataset = None
		self.data_labels = None
		self.plot_data = None
		self.plot_data_labels = None
		self.dlg = tk.Toplevel()
		self.dlg.title("Plot")
		self.dlg.columnconfigure(0, weight=1)
		self.auto_update = True
		self.plot_title = None
		# For transparency on some plots; also initial value set by plot type in 'set_xy()'
		self.alpha = 0.45
		# For rotation of some plot types
		self.rotated = False
		# For histogram
		self.bins = 10
		# For display of groups at Jenks breaks on Q-Q plots
		self.qq_groups = False
		# For display of lines at X and Y Jenks breaks on scatter plots
		self.scatter_breaks = False
		self.scatter_x_breaks = None
		self.scatter_y_breaks = None
		self.lineplot_breaks = False
		self.lineplot_x_breaks = None
		self.loess = False
		self.linreg = False
		self.theilsen = False

		def set_autoupdate():
			if self.autoupdate_var.get() == "1":
				self.auto_update = True
				self.q_redraw()
			else:
				self.auto_update = False

		# Message
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.NSEW, pady=(3,3))
		prompt_frame.columnconfigure(0, weight=1)
		msg_lbl = ttk.Label(prompt_frame, width=70, text="Select the type of plot, data columns to use, and whether to show only selected data.")
		msg_lbl.grid(row=0, column=0, sticky=tk.W, padx=(6,6), pady=(3,3))
		def wrap_msg(event):
			msg_lbl.configure(wraplength=event.width - 5)
		msg_lbl.bind("<Configure>", wrap_msg)

		# Controls
		ctrl_frame = tk.Frame(self.dlg)
		ctrl_frame.grid(row=1, column=0, sticky=tk.N+tk.EW)

		self.type_var = tk.StringVar(ctrl_frame, "")
		type_lbl = ttk.Label(ctrl_frame, text="Plot type:")
		type_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.type_sel = ttk.Combobox(ctrl_frame, state="readonly", textvariable=self.type_var, width=20, height=15,
				values=["Box plot", "Breaks groups", "Breaks optimum", "Category counts", "Empirical CDF", "Histogram", "Kernel density (KD) plot", "Line plot", "Min-max plot", "Normal Q-Q plot", "Scatter plot", "Stripchart", "Violin plot", "Y range plot"])
		self.type_sel.grid(row=0, column=1, columnspan=2, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.type_sel["state"] = "normal"
		self.type_sel.bind("<<ComboboxSelected>>", self.set_xy)

		self.sel_only_var = tk.StringVar(ctrl_frame, "0")
		self.sel_only_ck = ttk.Checkbutton(ctrl_frame, text="Selected data only", command=self.q_redraw, variable=self.sel_only_var,
				onvalue="1", offvalue="0")
		self.sel_only_ck.grid(row=1, column=0, columnspan=2, sticky=tk.W, padx=(6,3), pady=(3,3))

		self.autoupdate_var = tk.StringVar(ctrl_frame, "1")
		self.autoupdate_ck = ttk.Checkbutton(ctrl_frame, text="Auto-update", command=set_autoupdate, variable=self.autoupdate_var,
				onvalue="1", offvalue="0")
		self.autoupdate_ck.grid(row=1, column=2, sticky=tk.W, padx=(3,3), pady=(3,3))

		self.x_var = tk.StringVar(ctrl_frame, "")
		x_lbl = ttk.Label(ctrl_frame, text="X column:")
		x_lbl.grid(row=0, column=3, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.x_sel = ttk.Combobox(ctrl_frame, state="disabled", textvariable=self.x_var, width=24)
		self.x_sel.grid(row=0, column=4, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.x_sel.bind("<<ComboboxSelected>>", self.x_changed)

		self.y_var = tk.StringVar(ctrl_frame, "")
		y_lbl = ttk.Label(ctrl_frame, text="Y column:")
		y_lbl.grid(row=1, column=3, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.y_sel = ttk.Combobox(ctrl_frame, state="disabled", textvariable=self.y_var, width=24)
		self.y_sel.grid(row=1, column=4, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.y_sel.bind("<<ComboboxSelected>>", self.y_changed)

		self.xlog_var = tk.StringVar(ctrl_frame, "0")
		self.xlog_ck = ttk.Checkbutton(ctrl_frame, text="Log X", state="disabled", command=self.q_redraw, variable=self.xlog_var,
				onvalue="1", offvalue="0")
		self.xlog_ck.grid(row=0, column=5, sticky=tk.W, padx=(6,6), pady=(3,3))

		self.ylog_var = tk.StringVar(ctrl_frame, "0")
		self.ylog_ck = ttk.Checkbutton(ctrl_frame, text="Log Y", state="disabled", command=self.q_redraw, variable=self.ylog_var,
				onvalue="1", offvalue="0")
		self.ylog_ck.grid(row=1, column=5, sticky=tk.W, padx=(6,6), pady=(3,3))

		self.groupby_var = tk.StringVar(ctrl_frame, "")
		groupby_lbl = ttk.Label(ctrl_frame, text="Group by:")
		groupby_lbl.grid(row=2, column=3, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.groupby_sel = ttk.Combobox(ctrl_frame, state="disabled", textvariable=self.groupby_var, width=24)
		self.groupby_sel.grid(row=2, column=4, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.groupby_sel.bind("<<ComboboxSelected>>", self.groupby_changed)

		# Plot
		self.content_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		self.content_frame.grid(row=2, column=0, sticky=tk.NSEW)
		self.dlg.rowconfigure(2, weight=1)
		self.dlg.columnconfigure(0, weight=1)
		self.content_frame.rowconfigure(0, weight=1)
		self.content_frame.columnconfigure(0, weight=1)
		self.plotfig = Figure(dpi=100)
		self.plotfig.set_figheight(5)
		self.plotfig_canvas = FigureCanvasTkAgg(self.plotfig, self.content_frame)
		self.plot_nav = NavigationToolbar2Tk(self.plotfig_canvas, self.content_frame)
		self.plot_axes = self.plotfig.add_subplot(111)
		self.plotfig_canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
		self.plot_nav.update()

		# Buttons
		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.columnconfigure(0, weight=1)
		btn_frame.grid(row=3, column=0, sticky=tk.EW, pady=(3,3))
		btn_frame.columnconfigure(0, weight=0)
		btn_frame.columnconfigure(1, weight=0)
		btn_frame.columnconfigure(2, weight=0)
		btn_frame.columnconfigure(3, weight=0)
		btn_frame.columnconfigure(4, weight=1)
		self.canceled = False
		self.help_btn = ttk.Button(btn_frame, text="Help", command=self.do_help, underline=0)
		self.help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		self.data_btn = ttk.Button(btn_frame, text="Source Data", state="disabled", command=self.show_data, underline=0)
		self.data_btn.grid(row=0, column=1, sticky=tk.W, padx=(3,3))
		self.plot_data_btn = ttk.Button(btn_frame, text="Plot Data", state="disabled", command=self.show_plot_data, underline=0)
		self.plot_data_btn.grid(row=0, column=2, sticky=tk.W, padx=(3,3))
		self.clone_btn = ttk.Button(btn_frame, text="Clone", command=self.clone_plot, underline=3)
		self.clone_btn.grid(row=0, column=3, sticky=tk.W, padx=(3,6))
		close_btn = ttk.Button(btn_frame, text="Close", command=self.do_close, underline=0)
		close_btn.grid(row=0, column=4, sticky=tk.E, padx=(6,6))
		self.dlg.bind("<Alt-n>", self.clone_plot)
		self.dlg.bind("<Alt-c>", self.do_close)
		self.dlg.bind("<Escape>", self.do_close)
		self.dlg.bind("<Alt-t>", self.set_title)
		self.dlg.bind("<Alt-x>", self.set_xlabel)
		self.dlg.bind("<Alt-y>", self.set_ylabel)
		center_window(self.dlg)
		raise_window(self.dlg)

	def do_help(self, *args):
		webbrowser.open("https://mapdata.readthedocs.io/en/latest/dialogs.html#plot-dialog", new=2, autoraise=True)

	def show_data(self, *args):
		# Show data that have been collected for plotting, but not summarized as needed for a particular plot type.
		if self.dataset is not None:
			dlg = MsgDialog2("Source Data", "Original data:")
			variables = len(self.dataset)
			rowwise_data = []
			for i in range(len(self.dataset[0])):
				row = []
				for j in range(variables):
					row.append(self.dataset[j][i])
				rowwise_data.append(row)
			tframe, tdata = treeview_table(dlg.content_frame, rowwise_data, self.data_labels[0:variables])
			tframe.grid(row=0, column=0, sticky=tk.NSEW)
			def save_data(*args):
				export_data_table(self.data_labels[0:variables], rowwise_data, sheetname="Source data for plotting")
			dlg.dlg.bind("<Control-s>", save_data)
			dlg.show()

	def show_plot_data(self, *args):
		# Show data as summarized for a particular plot type.
		if self.plot_data is not None:
			dlg = MsgDialog2("Data for Plotting", "Data to be plotted:")
			variables = len(self.plot_data)
			rowwise_data = []
			max_data_len = max([len(self.plot_data[i]) for i in range(variables)])
			for i in range(max_data_len):
				row = []
				for j in range(variables):
					try:
						# Boxplot data are not necessarily a full matrix
						row.append(self.plot_data[j][i])
					except:
						row.append(None)
				rowwise_data.append(row)
			tframe, tdata = treeview_table(dlg.content_frame, rowwise_data, self.plot_data_labels)
			tframe.grid(row=0, column=0, sticky=tk.NSEW)
			def save_data(*args):
				export_data_table(self.plot_data_labels, rowwise_data, sheetname="Data for plotting")
			dlg.dlg.bind("<Control-s>", save_data)
			dlg.show()
	
	def clone_plot(self, *args):
		self.parent.clone_plot(self)

	def x_changed(self, *args):
		plot_type = self.type_var.get()
		if plot_type in ("Category counts"):
			self.xlog_ck["state"] = "disabled"
		else:
			self.xlog_ck["state"] = "readonly"
		xval = self.x_var.get()
		yval = self.y_var.get()
		if xval != '' and yval == xval:
			self.y_var.set('')
		self.q_redraw(args)

	def y_changed(self, *args):
		plot_type = self.type_var.get()
		if plot_type in ("Category counts", "Histogram", "Empirical CDF", "Min-max plot"):
			self.ylog_ck["state"] = "disabled"
		else:
			self.ylog_ck["state"] = "readonly"
		xval = self.x_var.get()
		yval = self.y_var.get()
		if yval != '' and yval == xval:
			self.x_var.set('')
		self.q_redraw(args)

	def groupby_changed(self, *args):
		self.q_redraw(args)

	def set_xy(self, *args):
		# Enable X and Y value selection, and set Combobox values based on plot type and column types.
		# Also sets the groupby Combobox values if applicable.
		self.plotfig.clear()
		self.plot_title = None
		self.dlg.bind("<Alt-a>")
		self.dlg.bind("<Alt-b>")
		self.dlg.bind("<Alt-g>")
		self.dlg.bind("<Alt-l>")
		self.dlg.bind("<Alt-r>")
		self.dlg.bind("<Alt-s>")
		self.plot_axes = self.plotfig.add_subplot(111)
		self.plotfig_canvas.draw()
		self.dataset = None
		self.data_labels = None
		self.plot_data = None
		self.plot_data_labels = None
		self.data_btn["state"] = "disabled"
		self.plot_data_btn["state"] = "disabled"
		# Category columns.  Does not include date columns for most uses, but may include dates for some.
		categ_columns = [c[0] for c in self.column_specs if c[1] in ("string", "boolean")]
		categ_columns.sort()
		categ_columns2 = [c[0] for c in self.column_specs if c[1] in ("string", "boolean", "date")]
		categ_columns2.sort()
		# quant_columns includes date and timestamp columns
		quant_columns = [c[0] for c in self.column_specs if c[1] in ("int", "float", "date", "timestamp", "timestamptz")]
		quant_columns.sort()
		self.numeric_columns = [c[0] for c in self.column_specs if c[1] in ("int", "float")]
		self.numeric_columns.sort()
		date_columns = [c[0] for c in self.column_specs if c[1] in ("date", "timestamp", "timestamptz")]
		date_columns.sort()
		self.x_var.set('')
		self.y_var.set('')
		self.xlog_var.set('0')
		self.ylog_var.set('0')
		self.groupby_var.set('')
		self.xlog_ck["state"] = "disabled"
		self.ylog_ck["state"] = "disabled"
		self.x_sel["state"] = "disabled"
		self.y_sel["state"] = "disabled"
		self.groupby_sel["state"] = "disabled"
		plot_type = self.type_var.get()
		if plot_type == "Category counts":
			self.x_sel["state"] = "readonly"
			self.x_sel["values"] = categ_columns
		elif plot_type in ("Histogram", "Empirical CDF", "Normal Q-Q plot", "Breaks groups", "Breaks optimum"):
			self.x_sel["state"] = "readonly"
			self.x_sel["values"] = self.numeric_columns
			self.xlog_ck["state"] = "normal"
			if plot_type == "Histogram":
				self.dlg.bind("<Alt-b>", self.set_bins)
			if plot_type == "Normal Q-Q plot":
				self.dlg.bind("<Alt-g>", self.show_groups)
		elif plot_type in ("Box plot", "Kernel density (KD) plot", "Stripchart", "Violin plot"):
			self.x_sel["state"] = "readonly"
			self.x_sel["values"] = quant_columns
			self.xlog_ck["state"] = "normal"
			self.groupby_sel["state"] = "readonly"
			self.groupby_sel["values"] = [''] + categ_columns2
			self.alpha = 0.45
		elif plot_type == "Min-max plot":
			self.x_sel["state"] = "readonly"
			self.x_sel["values"] = quant_columns
			self.xlog_ck["state"] = "normal"
			self.y_sel["state"] = "readonly"
			self.y_sel["values"] = categ_columns2
		else:
			self.x_sel["state"] = "readonly"
			self.x_sel["values"] = quant_columns
			self.xlog_ck["state"] = "normal"
			self.y_sel["state"] = "readonly"
			self.y_sel["values"] = quant_columns
			self.ylog_ck["state"] = "normal"
			if plot_type in ("Scatter plot", "Line plot"):
				self.groupby_sel["state"] = "readonly"
				if plot_type == "Scatter plot":
					self.groupby_sel["values"] = [''] + categ_columns2 + ['* Breaks in X', '* Breaks in Y']
					self.alpha = 0.45
				else:
					self.groupby_sel["values"] = [''] + categ_columns2 + ['* Breaks in X']
					self.alpha = 1.0

	def q_redraw(self, get_data=True, *args):
		# Conditionally (re)draw the plot.
		plot_type = self.type_var.get()
		can_redraw = (plot_type in ("Category counts", "Empirical CDF", \
				"Normal Q-Q plot", "Breaks groups", "Breaks optimum", "Histogram") \
				and self.x_var.get() != '') \
				or (plot_type in ("Scatter plot", "Line plot", "Min-max plot", "Y range plot") and \
				self.x_var.get() != '' and self.y_var.get() != '') \
				or (plot_type in ("Box plot", "Kernel density (KD) plot", "Stripchart", "Violin plot") and self.x_var.get() != '')
		if can_redraw:
			self.plotfig.clear()
			self.plot_axes = self.plotfig.add_subplot(111)
			self.plotfig_canvas.draw()
			if get_data or self.dataset is None or self.plot_data is None:
				self.get_data()
			if self.dataset is not None:
				self.redraw()

	def get_data(self):
		self.data_btn["state"] = "disabled"
		self.plot_data_btn["state"] = "disabled"
		self.dataset = None
		plot_type = self.type_var.get()
		column_list = []
		if self.x_var.get() != '':
			column_list = [self.x_var.get()]
		if self.y_var.get() != '':
			column_list.append(self.y_var.get())
		if self.groupby_var.get() != '':
			column_list.append(self.groupby_var.get())
		# Get either only the selected data or all data.
		if self.sel_only_var.get() == "1":
			dataset = self.parent.get_sel_data(column_list)
		else:
			dataset = self.parent.get_all_data(column_list)
		if dataset is None or len(dataset[0]) == 0:
			self.dataset = None
			self.data_labels = None
			self.plot_data = None
			self.plot_data_labels = None
			self.data_btn["state"] = "disabled"
			self.plot_data_btn["state"] = "disabled"
		else:
			# Remove missing data
			column_indexes = range(len(dataset))
			clean_data = [[] for _ in dataset]
			for i in range(len(dataset[0])):
				ok = True
				for col in column_indexes:
					if dataset[col][i] is None or dataset[col][i] == '':
						ok = False
				if ok:
					for col in column_indexes:
						clean_data[col].append(dataset[col][i])
			dataset = None
			# Convert quantitative data types
			if plot_type != "Category counts":
				x_data_type = [cs[1] for cs in self.column_specs if cs[0] == self.x_var.get()][0]
				cast_fn = data_type_cast_fn(x_data_type)
				for i in range(len(clean_data[0])):
					clean_data[0][i] = cast_fn(clean_data[0][i])
			if self.y_sel["state"] != "disabled" and self.y_var.get() != "" and len(clean_data) > 1:
				y_data_type = [cs[1] for cs in self.column_specs if cs[0] == self.y_var.get()][0]
				cast_fn = data_type_cast_fn(y_data_type)
				for i in range(len(clean_data[1])):
					clean_data[1][i] = cast_fn(clean_data[1][i])
			# Sort the dataset by X values
			clean_data = sort_columns(clean_data)
			# Set data labels
			if self.y_var.get() != '':
				self.data_labels = [self.x_var.get(), self.y_var.get()]
				grpvar = self.groupby_var.get()
				if grpvar not in ('', '* Breaks in X', '* Breaks in Y') :
					self.data_labels.append(grpvar)
			else:
				self.data_labels = [self.x_var.get()]
			# Log-transform data if specified.
			log_data = [[] for _ in clean_data]
			log_error = False
			if self.xlog_ck["state"] != "disabled" and self.xlog_var.get() == "1":
				for i in range(len(clean_data[0])):
					try:
						log_data[0].append(math.log10(clean_data[0][i]))
					except:
						log_error = True
						self.xlog_var.set("0")
						self.xlog_ck["state"] = "disabled"
						break
				if not log_error:
					clean_data[0] = log_data[0]
					self.data_labels[0] = "Log10 of " + self.x_var.get()
			if self.ylog_ck["state"] != "disabled" and self.ylog_var.get() == "1" and len(clean_data) > 1:
				log_error = False
				for i in range(len(clean_data[1])):
					try:
						log_data[1].append(math.log10(clean_data[1][i]))
					except:
						log_error = True
						self.ylog_var.set("0")
						self.ylog_ck["state"] = "disabled"
						break
				if not log_error:
					clean_data[1] = log_data[1]
					self.data_labels[1] = "Log10 of " + self.y_var.get()
			log_data = None
			#
			self.dataset = clean_data
			self.data_btn["state"] = "normal"
			# Summarize and sort the data as needed for each type of plot.
			if plot_type == "Category counts":
				# Count of values for each X, ordered by X
				counter = collections.Counter(self.dataset[0])
				x_vals = list(counter.keys())
				x_vals.sort()
				x_counts = [counter[k] for k in x_vals]
				self.plot_data = [x_vals, x_counts]
				self.plot_data_labels = [self.data_labels[0], "Count"]
			elif plot_type in ("Box plot", "Kernel density (KD) plot", "Stripchart", "Violin plot"):
				# The groupby variable may or may not be set
				if self.groupby_var.get() != '':
					grp_vals = list(set(self.dataset[1]))
					grp_vals.sort()
					ds = list(zip(self.dataset[1], self.dataset[0]))
					plot_data = []
					for g in grp_vals:
						plot_data.append([d[1] for d in ds if d[0] == g])
					self.plot_data = plot_data
					self.plot_data_labels = grp_vals
				else:
					self.plot_data = self.dataset
					self.plot_data_labels = self.data_labels
			elif plot_type == "Breaks groups":
				x_vals = copy.copy(self.dataset[0])
				x_vals.sort()
				oj = optimum_jenks(x_vals, 8)
				jnb = jenkspy.JenksNaturalBreaks(oj)
				jnb.fit(x_vals)
				self.plot_data = [list(x) for x in jnb.groups_]
				self.plot_data_labels = [str(i+1) for i in range(len(self.plot_data))]
			elif plot_type == "Breaks optimum":
				nrows = len(self.dataset[0])
				if nrows < 4:
					warning("The data set must have at least 4 values.", {})
				else:
					x_vals = copy.copy(self.dataset[0])
					x_vals.sort()
					x, y = all_jenks_breaks(x_vals, min(len(self.dataset[0])-1, 8))
					oj = optimum_jenks(x_vals, min(nrows-1, 8))
					self.plot_data = [x, y, [x[oj-1]], [y[oj-1]]]
					self.plot_data_labels = ["Groups", "Goodness of Variance Fit", "Optimum group", "Optimum GVF"]
			elif plot_type == "Empirical CDF":
				# Y is the fraction of data points below each X value
				x_counts = np.unique(self.dataset[0], return_counts=True)
				y_vals = list(np.cumsum(x_counts[1]/np.sum(x_counts[1])))
				self.plot_data = [list(x_counts[0]), y_vals]
				self.plot_data_labels = [self.data_labels[0], "Cumulative frequency"]
			elif plot_type == "Min-max plot":
				# Min and max X for each Y
				y_vals = list(set(self.dataset[1]))
				y_vals.sort()
				plotdata = dict(zip(y_vals, [[None, None] for _ in y_vals]))
				for i in range(len(self.dataset[1])):
					x = self.dataset[0][i]
					y = self.dataset[1][i]
					x_vals = plotdata[y]
					if x_vals[0] is None or x < x_vals[0]:
						plotdata[y][0] = x
					if x_vals[1] is None or x > x_vals[1]:
						plotdata[y][1] = x
				x1 = [plotdata[y][0] for y in y_vals]
				x2 = [plotdata[y][1] for y in y_vals]
				self.plot_data = [y_vals, x1, x2]
				self.plot_data_labels = [self.data_labels[1], self.data_labels[0] + " min", self.data_labels[0] + " max"]
			elif plot_type == "Normal Q-Q plot":
				x_vals = copy.copy(self.dataset[0])
				x_vals.sort()
				x_mean = statistics.mean(x_vals)
				x_sd = statistics.stdev(x_vals)
				x_quantiles = [(x - x_mean)/x_sd for x in x_vals]
				x_len = len(x_vals)
				q = [(i + 0.5)/x_len for i in range(x_len)]
				norm_quantiles = [qnorm(p) for p in q]
				if x_len < 4:
					oj = 2
				else:
					oj = optimum_jenks(x_vals, min(x_len-1, 8))
				jnb = jenkspy.JenksNaturalBreaks(oj)
				jnb.fit(x_vals)
				self.plot_data = [x_vals, x_quantiles, norm_quantiles, jnb.labels_]
				self.plot_data_labels = [self.data_labels[0], "Sample quantiles", "Theoretical quantiles", "Group"]
			elif plot_type == "Y range plot":
				# Min and max Y for each X
				x_vals = list(set(self.dataset[0]))
				x_vals.sort()
				plotdata = dict(zip(x_vals, [[None, None] for i in x_vals]))
				for i in range(len(self.dataset[0])):
					x = self.dataset[0][i]
					y = self.dataset[1][i]
					y_vals = plotdata[x]
					if y_vals[0] is None or y < y_vals[0]:
						plotdata[x][0] = y
					if y_vals[1] is None or y > y_vals[1]:
						plotdata[x][1] = y
				y1 = [plotdata[x][0] for x in x_vals]
				y2 = [plotdata[x][1] for x in x_vals]
				self.plot_data = [x_vals, y1, y2]
				self.plot_data_labels = [self.data_labels[0], self.data_labels[1] + " min", self.data_labels[1] + " max"]
			elif plot_type == "Line plot":
				ds = sort_columns(self.dataset)
				if self.groupby_var.get() == "* Breaks in X":
					if self.x_var.get() in self.numeric_columns:
						oj = optimum_jenks(ds[0], 8)
						jnb = jenkspy.JenksNaturalBreaks(oj)
						jnb.fit(ds[0])
						self.plot_data = [ds[0], ds[1], ["Group "+str(lbl+1) for lbl in jnb.labels_]]
						self.plot_data_labels = self.data_labels + ['Breaks in X']
					else:
						# Can't compute breaks for X
						self.groupby_var.set('')
						self.plot_data = self.dataset
						self.plot_data_labels = self.data_labels
				else:
					self.plot_data = ds
					self.plot_data_labels = self.data_labels
			elif plot_type == "Scatter plot":
				if self.groupby_var.get() == "* Breaks in X":
					if self.x_var.get() in self.numeric_columns:
						ds = sort_columns(self.dataset)
						oj = optimum_jenks(ds[0], 8)
						jnb = jenkspy.JenksNaturalBreaks(oj)
						jnb.fit(ds[0])
						self.plot_data = [ds[0], ds[1], ["Group "+str(lbl+1) for lbl in jnb.labels_]]
						self.plot_data_labels = self.data_labels + ['Breaks in X']
					else:
						# Can't compute breaks for X
						self.groupby_var.set('')
						self.plot_data = self.dataset
						self.plot_data_labels = self.data_labels
				elif self.groupby_var.get() == "* Breaks in Y":
					if self.y_var.get() in self.numeric_columns:
						ds = sort_columns(self.dataset, sortby=1)
						oj = optimum_jenks(ds[1], 8)
						jnb = jenkspy.JenksNaturalBreaks(oj)
						jnb.fit(ds[1])
						self.plot_data = [ds[0], ds[1], ["Group "+str(lbl+1) for lbl in jnb.labels_]]
						self.plot_data_labels = self.data_labels + ['Breaks in Y']
					else:
						# Can't compute breaks for Y
						self.groupby_var.set('')
						self.plot_data = self.dataset
						self.plot_data_labels = self.data_labels
				else:
					self.plot_data = self.dataset
					self.plot_data_labels = self.data_labels
			elif plot_type == "Histogram":
				# No special preparation
				self.plot_data = self.dataset
				self.plot_data_labels = self.data_labels
			elif plot_type == "Y range plot":
				self.plot_data = sort_columns(self.dataset)
				self.plot_data_labels = self.data_labels
			self.plot_data_btn["state"] = "normal"

	def redraw(self):
		plot_type = self.type_var.get()
		if self.plot_data is not None and len(self.plot_data) > 0 and len(self.plot_data[0]) > 0:
			if plot_type == "Category counts":
				self.plot_axes.bar(self.plot_data[0], self.plot_data[1])
				self.plot_axes.set_xlabel(self.plot_data_labels[0])
				self.plot_axes.set_ylabel(self.plot_data_labels[1])
			elif plot_type == "Histogram":
				self.plot_axes.hist(self.plot_data[0], bins=self.bins)
				self.plot_axes.set_xlabel(self.x_var.get())
				self.plot_axes.set_ylabel("Counts")
			elif plot_type == "Scatter plot":
				self.dlg.bind("<Alt-a>", self.set_alpha)
				self.dlg.bind("<Alt-b>", self.set_scatter_breaks)
				self.dlg.bind("<Alt-l>", self.set_loess)
				self.dlg.bind("<Alt-r>", self.set_linreg)
				self.dlg.bind("<Alt-s>", self.set_theilsen)
				if self.scatter_breaks:
					if self.x_var.get() in self.numeric_columns:
						x_vals = copy.copy(self.plot_data[0])
						nx = len(x_vals)
						if nx < 4:
							warning("There must be at least 4 X values to calculate natural breaks.", {})
						else:
							x_vals.sort()
							oj = optimum_jenks(x_vals, min(nx-1, 8))
							jnb = jenkspy.JenksNaturalBreaks(oj)
							jnb.fit(x_vals)
							self.scatter_x_breaks = jnb.inner_breaks_
							for b in self.scatter_x_breaks:
								self.plot_axes.axvline(b, color='0.8', alpha=0.5)
					if self.y_var.get() in self.numeric_columns:
						y_vals = copy.copy(self.plot_data[1])
						ny = len(y_vals)
						if ny < 4:
							warning("There must be at least 4 Y values to calculate natural breaks.", {})
						else:
							y_vals.sort()
							oj = optimum_jenks(y_vals, min(ny-1, 8))
							jnb = jenkspy.JenksNaturalBreaks(oj)
							jnb.fit(y_vals)
							self.scatter_y_breaks = jnb.inner_breaks_
							for b in self.scatter_y_breaks:
								self.plot_axes.axhline(b, color='0.8', alpha=0.5)
				if self.groupby_var.get() == '':
					self.plot_axes.scatter(self.plot_data[0], self.plot_data[1], alpha=self.alpha)
				else:
					groups = list(set(self.plot_data[2]))
					groups.sort()
					datarows = len(self.plot_data[0])
					for g in groups:
						pdx = [self.plot_data[0][i] for i in range(datarows) if self.plot_data[2][i] == g]
						pdy = [self.plot_data[1][i] for i in range(datarows) if self.plot_data[2][i] == g]
						grplbl = g
						if wrap_at_underscores:
							grplbl = g.replace("_", " ")
						grplbl = "\n".join(textwrap.wrap(grplbl, width=wrapwidth))
						self.plot_axes.plot(pdx, pdy, marker='o', alpha=self.alpha, linestyle='', label=grplbl)
				bbox = self.plot_axes.get_clip_box()
				if self.loess and self.x_var.get() in self.numeric_columns:
					loess_x, loess_y, wts = loess_1d(np.array(self.plot_data[0]), np.array(self.plot_data[1]))
					self.plot_axes.plot(loess_x, loess_y, label="LOESS", color="black", linestyle="dashed", alpha=0.65)
				if self.linreg and self.x_var.get() in self.numeric_columns:
					ns, stats = Polynomial.fit(np.array(self.plot_data[0]), np.array(self.plot_data[1]), 1, full=True)
					intercept, slope = ns.convert().coef
					# Use the mean point to plot instead of the intercept because the intercept may be out of the bounding box, especially for log-transformed data
					xmean = statistics.fmean(self.plot_data[0])
					ymean = statistics.fmean(self.plot_data[1])
					self.plot_axes.axline((xmean, ymean), slope=slope, clip_box=bbox, clip_on=True, label="Linear fit", color="darkorange", linestyle="dashdot", linewidth=2, alpha=0.65)
					if show_regression_stats:
						N = len(self.plot_data[1])
						total_ss = sum([(self.plot_data[1][i] - ymean)**2 for i in range(N)])
						resid_ss = list(stats[0])[0]
						exp_ss = total_ss - resid_ss
						if total_ss == 0.0:
							r_square = "NC"
						else:
							r_square = str(round_figs(exp_ss/total_ss, 4))
						statdlg = MsgDialog(title="Linear Regression", message = "Slope: %s\nIntercept: %s\nR-square: %s\nN: %s" % \
								(round_figs(slope, 4), round_figs(intercept, 4), r_square, N))
						statdlg.show()
				if self.theilsen and self.x_var.get() in self.numeric_columns:
					ts_slope, ts_intercept, ts_high, ts_low = spstats.theilslopes(np.array(self.plot_data[1]), np.array(self.plot_data[0]))
					self.plot_axes.axline((statistics.median(self.plot_data[0]), statistics.median(self.plot_data[1])), slope=ts_slope, \
							clip_box=bbox, clip_on=True, label="Theil-Sen line", color="darkgreen", linestyle="dotted", linewidth=2, alpha=0.65)
				if self.groupby_var.get() != '' or ((self.loess or self.linreg or self.theilsen) and self.x_var.get() in self.numeric_columns):
					self.plot_axes.legend()
				self.plot_axes.set_xlabel(self.plot_data_labels[0])
				self.plot_axes.set_ylabel(self.plot_data_labels[1])
			elif plot_type == "Line plot":
				self.dlg.bind("<Alt-a>", self.set_alpha)
				self.dlg.bind("<Alt-b>", self.set_lineplot_breaks)
				self.dlg.bind("<Alt-l>", self.set_loess)
				self.dlg.bind("<Alt-r>", self.set_linreg)
				if self.lineplot_breaks:
					if self.x_var.get() in self.numeric_columns:
						x_vals = copy.copy(self.plot_data[0])
						#x_vals.sort()
						nx = len(x_vals)
						if nx < 4:
							warning("There must be at least 4 X values to calculate natural breaks.", {})
						else:
							oj = optimum_jenks(x_vals, min(nx-1, 8))
							jnb = jenkspy.JenksNaturalBreaks(oj)
							jnb.fit(x_vals)
							self.lineplot_x_breaks = jnb.inner_breaks_
							for b in self.lineplot_x_breaks:
								self.plot_axes.axvline(b, color='0.8', alpha=0.5)
				if self.groupby_var.get() == '':
					self.plot_axes.plot(self.plot_data[0], self.plot_data[1], alpha=self.alpha)
				else:
					groups = list(set(self.plot_data[2]))
					groups.sort()
					datarows = len(self.plot_data[0])
					for g in groups:
						pdx = [self.plot_data[0][i] for i in range(datarows) if self.plot_data[2][i] == g]
						pdy = [self.plot_data[1][i] for i in range(datarows) if self.plot_data[2][i] == g]
						self.plot_axes.plot(pdx, pdy, label=g, alpha=self.alpha)
				bbox = self.plot_axes.get_clip_box()
				if self.loess and self.x_var.get() in self.numeric_columns:
					loess_x, loess_y, wts = loess_1d(np.array(self.plot_data[0]), np.array(self.plot_data[1]))
					self.plot_axes.plot(loess_x, loess_y, label="LOESS", color="black", linestyle="dashed", alpha=0.65)
				if self.linreg and self.x_var.get() in self.numeric_columns:
					ns, stats = Polynomial.fit(np.array(self.plot_data[0]), np.array(self.plot_data[1]), 1, full=True)
					intercept, slope = ns.convert().coef
					# Use the mean point to plot instead of the intercept because the intercept may be out of the bounding box, especially for log-transformed data
					xmean = statistics.fmean(self.plot_data[0])
					ymean = statistics.fmean(self.plot_data[1])
					self.plot_axes.axline((xmean, ymean), slope=slope, clip_box=bbox, clip_on=True, label="Linear fit", color="darkorange", linestyle="dashdot", linewidth=2, alpha=0.65)
					if show_regression_stats:
						N = len(self.plot_data[1])
						total_ss = sum([(self.plot_data[1][i] - ymean)**2 for i in range(N)])
						resid_ss = list(stats[0])[0]
						exp_ss = total_ss - resid_ss
						if total_ss == 0.0:
							r_square = "NC"
						else:
							r_square = str(round_figs(exp_ss/total_ss, 4))
						statdlg = MsgDialog(title="Linear Regression", message = "Slope: %s\nIntercept: %s\nR-square: %s\nN: %s" % \
								(round_figs(slope, 4), round_figs(intercept, 4), r_square, N))
						statdlg.show()
				if self.theilsen and self.x_var.get() in self.numeric_columns:
					ts_slope, ts_intercept, ts_high, ts_low = spstats.theilslopes(np.array(self.plot_data[1]), np.array(self.plot_data[0]))
					self.plot_axes.axline((statistics.median(self.plot_data[0]), statistics.median(self.plot_data[1])), slope=ts_slope, \
							clip_box=bbox, clip_on=True, label="Theil-Sen line", color="darkgreen", linestyle="dotted", linewidth=2, alpha=0.65)
				if self.groupby_var.get() != '' or ((self.loess or self.linreg or self.theilsen) and self.x_var.get() in self.numeric_columns):
					self.plot_axes.legend()
				self.plot_axes.set_xlabel(self.plot_data_labels[0])
				self.plot_axes.set_ylabel(self.plot_data_labels[1])
			elif plot_type == "Breaks groups":
				for i in range(len(self.plot_data)):
					xs = [i+1] * len(self.plot_data[i])
					self.plot_axes.scatter(xs, self.plot_data[i])
				ticks = [int(l) for l in self.plot_data_labels]
				ticks.insert(0,0)
				ticks.append(ticks[-1]+1)
				self.plot_axes.set_xticks(ticks)
				lbls = copy.copy(self.plot_data_labels)
				lbls.insert(0, "")
				lbls.append("")
				self.plot_axes.set_xticklabels(lbls)
				self.plot_axes.set_xlabel("Group")
				self.plot_axes.set_ylabel(self.data_labels[0])
			elif plot_type == "Breaks optimum":
				self.plot_axes.plot(self.plot_data[0], self.plot_data[1])
				self.plot_axes.scatter(self.plot_data[2], self.plot_data[3])
				self.plot_axes.set_xlabel(self.plot_data_labels[0])
				self.plot_axes.set_ylabel(self.plot_data_labels[1])
			elif plot_type == "Empirical CDF":
				self.plot_axes.stackplot(self.plot_data[0], self.plot_data[1])
				self.plot_axes.set_xlabel(self.plot_data_labels[0])
				self.plot_axes.set_ylabel(self.plot_data_labels[1])
			elif plot_type == "Kernel density (KD) plot":
				self.dlg.bind("<Alt-a>", self.set_alpha)
				if self.groupby_var.get() == '':
					sns.kdeplot({self.x_var.get(): self.dataset[0]}, x=self.x_var.get(), fill=True, ax=self.plot_axes)
					self.plot_axes.set_xlabel(self.x_var.get())
				else:
					grplbls = copy.copy(self.dataset[1])
					if wrap_at_underscores:
						grplbls = [lbl.replace("_", " ") for lbl in grplbls]
					grplbls = [textwrap.wrap(lbl, width=wrapwidth) for lbl in grplbls]
					grplbls = ["\n".join(lbl) for lbl in grplbls]
					sns.kdeplot({self.groupby_var.get(): grplbls, self.x_var.get(): self.dataset[0]},
							x=self.x_var.get(), hue=self.groupby_var.get(), multiple="layer", fill=True, alpha=self.alpha, ax=self.plot_axes,
							warn_singular=False)
					self.plot_axes.set_xlabel(self.x_var.get())
				self.plot_axes.set_ylabel("Density")
			elif plot_type == "Min-max plot":
				self.dlg.bind("<Alt-r>", self.set_rotated)
				if not self.rotated:
					self.plot_axes.hlines(self.plot_data[0], self.plot_data[1], self.plot_data[2], linewidths=5.0)
					if self.xlog_ck["state"] != "disabled" and self.xlog_var.get() == "1":
						self.plot_axes.set_xlabel("Log10 of " + self.x_var.get())
					else:
						self.plot_axes.set_xlabel(self.x_var.get())
					self.plot_axes.set_ylabel(self.plot_data_labels[0])
				else:
					self.plot_axes.vlines(self.plot_data[0], self.plot_data[1], self.plot_data[2], linewidths=5.0)
					if self.xlog_ck["state"] != "disabled" and self.xlog_var.get() == "1":
						self.plot_axes.set_ylabel("Log10 of " + self.x_var.get())
					else:
						self.plot_axes.set_ylabel(self.x_var.get())
					self.plot_axes.set_xlabel(self.plot_data_labels[0])
			elif plot_type == "Normal Q-Q plot":
				if self.qq_groups:
					self.plot_axes.scatter(self.plot_data[2], self.plot_data[1], c=self.plot_data[3], cmap="tab10")
				else:
					self.plot_axes.scatter(self.plot_data[2], self.plot_data[1])
				pmin = min(self.plot_data[2][0], self.plot_data[1][0])
				pmax = max(self.plot_data[2][-1], self.plot_data[1][-1])
				self.plot_axes.plot([pmin, pmax], [pmin, pmax])
				self.plot_axes.set_xlabel(self.plot_data_labels[2])
				self.plot_axes.set_ylabel(self.plot_data_labels[1])
			elif plot_type == "Y range plot":
				self.plot_axes.fill_between(self.plot_data[0], self.plot_data[1], self.plot_data[2])
				self.plot_axes.set_xlabel(self.x_var.get())
				if self.xlog_ck["state"] != "disabled" and self.xlog_var.get() == "1":
					self.plot_axes.set_xlabel("Log10 of " + self.x_var.get())
				else:
					self.plot_axes.set_xlabel(self.x_var.get())
				if self.ylog_ck["state"] != "disabled" and self.ylog_var.get() == "1":
					self.plot_axes.set_ylabel("Log10 of " + self.y_var.get())
				else:
					self.plot_axes.set_ylabel(self.y_var.get())
			elif plot_type == "Box plot":
				self.dlg.bind("<Alt-r>", self.set_rotated)
				orient = not self.rotated
				grplbls = copy.copy(self.plot_data_labels)
				if wrap_at_underscores:
					grplbls = [lbl.replace("_", " ") for lbl in grplbls]
				grplbls = [textwrap.wrap(lbl, width=wrapwidth) for lbl in grplbls]
				grplbls = ["\n".join(lbl) for lbl in grplbls]
				self.plot_axes.boxplot(self.plot_data, labels=grplbls, vert=orient)
				if not self.rotated:
					self.plot_axes.set_xlabel(self.groupby_var.get())
					self.plot_axes.set_ylabel(self.data_labels[0])
				else:
					self.plot_axes.set_ylabel(self.groupby_var.get())
					self.plot_axes.set_xlabel(self.data_labels[0])
			elif plot_type == "Stripchart":
				self.dlg.bind("<Alt-a>", self.set_alpha)
				self.dlg.bind("<Alt-r>", self.set_rotated)
				orientation = "v" if not self.rotated else "h"
				if self.groupby_var.get() != '':
					grplbls = copy.copy(self.dataset[1])
					if wrap_at_underscores:
						grplbls = [lbl.replace("_", " ") for lbl in grplbls]
					grplbls = [textwrap.wrap(lbl, width=wrapwidth) for lbl in grplbls]
					grplbls = ["\n".join(lbl) for lbl in grplbls]
					if not self.rotated:
						sns.stripplot({self.groupby_var.get(): grplbls, \
								self.x_var.get(): self.dataset[0]}, \
								x=self.groupby_var.get(), y=self.x_var.get(), alpha=self.alpha, \
								ax=self.plot_axes, orient=orientation)
						self.plot_axes.set_xlabel(self.groupby_var.get())
						self.plot_axes.set_ylabel(self.data_labels[0])
					else:
						sns.stripplot({self.groupby_var.get(): grplbls, \
								self.x_var.get(): self.dataset[0]}, \
								y=self.groupby_var.get(), x=self.x_var.get(), alpha=self.alpha, \
								ax=self.plot_axes, orient=orientation)
						self.plot_axes.set_xlabel(self.groupby_var.get())
						self.plot_axes.set_ylabel(self.data_labels[0])
				else:
					if not self.rotated:
						sns.stripplot({self.x_var.get(): self.dataset[0]}, y=self.x_var.get(), \
								alpha=self.alpha, ax=self.plot_axes, orient=orientation)
						self.plot_axes.set_xlabel(self.x_var.get())
						self.plot_axes.set_ylabel(self.data_labels[0])
					else:
						sns.stripplot({self.x_var.get(): self.dataset[0]}, x=self.x_var.get(), \
								alpha=self.alpha, ax=self.plot_axes, orient=orientation)
						self.plot_axes.set_ylabel(self.x_var.get())
						self.plot_axes.set_xlabel(self.data_labels[0])
			elif plot_type == "Violin plot":
				self.dlg.bind("<Alt-r>", self.set_rotated)
				orientation = "v" if not self.rotated else "h"
				if self.groupby_var.get() != '':
					grplbls = copy.copy(self.dataset[1])
					if wrap_at_underscores:
						grplbls = [lbl.replace("_", " ") for lbl in grplbls]
					grplbls = [textwrap.wrap(lbl, width=wrapwidth) for lbl in grplbls]
					grplbls = ["\n".join(lbl) for lbl in grplbls]
					if not self.rotated:
						sns.violinplot({self.groupby_var.get(): grplbls, self.x_var.get(): self.dataset[0]},
								x=self.groupby_var.get(), y=self.x_var.get(), ax=self.plot_axes)
						self.plot_axes.set_xlabel(self.groupby_var.get())
						self.plot_axes.set_ylabel(self.data_labels[0])
					else:
						sns.violinplot({self.groupby_var.get(): grplbls, self.x_var.get(): self.dataset[0]},
								y=self.groupby_var.get(), x=self.x_var.get(), ax=self.plot_axes)
						self.plot_axes.set_ylabel(self.groupby_var.get())
						self.plot_axes.set_xlabel(self.data_labels[0])
				else:
					if not self.rotated:
						sns.violinplot({self.x_var.get(): self.dataset[0]}, y=self.x_var.get(), alpha=self.alpha, ax=self.plot_axes)
						self.plot_axes.set_xlabel(self.x_var.get())
						self.plot_axes.set_ylabel(self.data_labels[0])
					else:
						sns.violinplot({self.x_var.get(): self.dataset[0]}, x=self.x_var.get(), alpha=self.alpha, ax=self.plot_axes)
						self.plot_axes.set_ylabel(self.x_var.get())
						self.plot_axes.set_xlabel(self.data_labels[0])
			if self.plot_title is not None:
				self.plot_axes.set_title(self.plot_title)
			self.plotfig_canvas.draw()
			self.plot_nav.update()

	def set_title(self, *args):
		dlg = OneEntryDialog(self.dlg, "Plot Title", "Enter a title for the plot:")
		title = dlg.show()
		if title is not None:
			self.plot_title = title
			self.plot_axes.set_title(title)
			self.plotfig_canvas.draw()
	def set_xlabel(self, *args):
		dlg = OneEntryDialog(self.dlg, "X-Axis Label", "Enter a title for the X-axis label:")
		xlabel = dlg.show()
		if xlabel is not None:
			self.plot_axes.set_xlabel(xlabel)
			self.plotfig_canvas.draw()
	def set_ylabel(self, *args):
		dlg = OneEntryDialog(self.dlg, "Y-Axis Label", "Enter a title for the Y-axis label:")
		ylabel = dlg.show()
		if ylabel is not None:
			self.plot_axes.set_ylabel(ylabel)
			self.plotfig_canvas.draw()
	def set_bins(self, *args):
		dlg = OneIntDialog(self.dlg, "Histogram Bins", "Enter the number of histogram bins", min_value=2, max_value=50, initial=self.bins)
		num_bins = dlg.show()
		if num_bins is not None:
			self.bins = num_bins
			if self.type_var.get() == "Histogram":
				self.q_redraw()
	def set_alpha(self, *args):
		dlg = OneFloatDialog(self.dlg, "Transparency", "Enter the transparency (alpha) value", min_value=0.0, max_value=1.0, initial=self.alpha)
		new_alpha = dlg.show()
		if new_alpha is not None:
			self.alpha = new_alpha
			if self.type_var.get() in ("Box plot", "Scatter plot", "Kernel density (KD) plot", "Line plot", "Stripchart", "Violin plot"):
				self.q_redraw()
	def set_scatter_breaks(self, *args):
		# Toggle display of vertical and horizontal lines at natural breaks values on a scatter plot.
		self.scatter_breaks = not self.scatter_breaks
		if self.type_var.get() == "Scatter plot":
			self.q_redraw()
	def set_lineplot_breaks(self, *args):
		# Toggle display of vertical lines at natural breaks values on a scatter plot.
		self.lineplot_breaks = not self.lineplot_breaks
		if self.type_var.get() == "Line plot":
			self.q_redraw()
	def set_loess(self, *args):
		# Toggle display of loess line on a scatter plot.
		self.loess = not self.loess
		if self.type_var.get() in ("Line plot", "Scatter plot"):
			self.q_redraw()
	def set_linreg(self, *args):
		# Toggle display of linear regression line on a scatter plot.
		self.linreg = not self.linreg
		if self.type_var.get() in ("Line plot", "Scatter plot"):
			self.q_redraw()
	def set_theilsen(self, *args):
		self.theilsen = not self.theilsen
		if self.type_var.get() in ("Line plot", "Scatter plot"):
			self.q_redraw()
	def set_rotated(self, *args):
		self.rotated = not self.rotated
		if self.type_var.get() in ("Min-max plot", "Box plot", "Stripchart", "Violin plot"):
			self.q_redraw()

	def show_groups(self, *args):
		if self.type_var.get() == "Normal Q-Q plot":
			self.qq_groups = not self.qq_groups
			self.q_redraw()
	def do_close(self, *args):
		self.parent.remove_plot(self)
		try:
			self.dlg.destroy()
		except:
			pass
	def show(self):
		self.dlg.update_idle_tasks()
		self.dlg.minsize(width=500, height=500)
		self.dlg.wait_window(self.dlg)


def rosners_stat(x):
	# Calculates the statistic for Rosner's test.
	# Returns the standard score (Z score) for the value with the maximum absolute
	# deviation from the mean, and the index into the data list for that value.
	# From https://www.itl.nist.gov/div898/handbook/eda/section3/eda35h3.htm
	mn, sd = statistics.fmean(x), statistics.stdev(x)
	absdevs = [abs(v - mn) for v in x]
	maxdev = max(absdevs)
	maxidx = absdevs.index(maxdev)
	return maxdev/sd, maxidx

def rosners_critical(n, alpha, outlier_no):
	# From https://www.itl.nist.gov/div898/handbook/eda/section3/eda35h3.htm
	tval = spstats.t.ppf(1 - alpha / (2*(n-outlier_no+1)), n-outlier_no-1)
	return ((n-1) * tval) / (math.sqrt((n - outlier_no - 1 + tval**2) * (n - outlier_no + 1)))

def rosners_test(x, alpha, max_outliers):
	# Carries out Rosner's test for up to max_outliers.
	# Returns the number of outliers, from 0 up to max_outliers.
	d = copy.copy(x)
	n_outliers = 0
	for i in range(max_outliers):
		rs, ix = rosners_stat(d)
		cv = rosners_critical(len(d), alpha, i+1)
		if rs > cv:
			n_outliers = i+1
		del d[i]
	# No non-outlier was found
	return n_outliers

def tukey_outliers(x):
	d = np.array(x)
	q25, q75 = np.percentile(d, 25), np.percentile(d, 75)
	bound = 1.5 * (q75 - q25)
	lbound, ubound = q25 - bound, q75 + bound
	return len([v for v in x if v < lbound or v > ubound])



class UnivarStatsDialog(object):
	def __init__(self, parent, column_specs):
		self.parent = parent
		self.column_specs = column_specs
		self.dlg = tk.Toplevel()
		self.dlg.title("Univariate Statistics")
		self.dlg.rowconfigure(0, weight=0)
		self.dlg.rowconfigure(1, weight=0)
		self.dlg.rowconfigure(2, weight=1)
		self.dlg.rowconfigure(3, weight=0)
		self.dlg.columnconfigure(0, weight=1)
		# Data
		self.dataset = None
		self.data_labels = None
		self.numeric_columns = [c[0] for c in self.column_specs if c[1] in ("int", "float")]
		self.numeric_columns.sort()
		self.dnames = ["Variable", " N ", "Min.", "Max.", "Mean", "Median", "Mode", \
				"Geo. mean", "Std.Dev.", "C.V.", "5th %ile", "95th %ile", "Anderson-Darling p", \
				"Lillefors p", "Rosner's outliers", "Tukey outliers"]
		self.logdnames = ["Variable", " N ", "Min.", "Max.", "Mean", "Median", "Mode", \
				"Std.Dev.", "C.V.", "5th %ile", "95th %ile", "Anderson-Darling p", "Lillefors p", \
				"Rosner's outliers", "Tukey outliers"]
		self.statdata = []
		self.dlg.bind("<Control-s>")
		self.dlg.bind("<Control-z>")
		# Message
		prompt_frame = tk.Frame(self.dlg, borderwidth=5)
		prompt_frame.grid(row=0, column=0, sticky=tk.N+tk.EW, pady=(3,3))
		prompt_frame.columnconfigure(0, weight=1)
		msg_lbl = ttk.Label(prompt_frame, width=100, text="Select one or more variables from the left to see univariate statistics.  Use Ctrl-click or Shift-click to select multiple rows.")
		msg_lbl.grid(row=0, column=0, sticky=tk.W, padx=(6,6), pady=(3,3))
		def wrap_msg(event):
			msg_lbl.configure(wraplength=event.width - 5)
		msg_lbl.bind("<Configure>", wrap_msg)

		# Top controls
		# Only the 'Selected only' checkbox
		ctrl_frame = tk.Frame(self.dlg, borderwidth=5)
		ctrl_frame.rowconfigure(0, weight=0)
		ctrl_frame.columnconfigure(0, weight=1)
		ctrl_frame.grid(row=1, column=0, sticky=tk.N+tk.EW)
		self.sel_only_var = tk.StringVar(ctrl_frame, "0")
		self.sel_only_ck = ttk.Checkbutton(ctrl_frame, text="Selected data only", command=self.q_recalc, variable=self.sel_only_var,
				onvalue="1", offvalue="0")
		self.sel_only_ck.grid(row=0, column=0, sticky=tk.W, padx=(6,3), pady=(3,3))

		# The data_frame encompasses the two panes of the variable frame and the content frame
		data_frame = tk.Frame(self.dlg, borderwidth=5)
		data_frame.rowconfigure(0, weight=1)
		data_frame.columnconfigure(0, weight=1)
		data_frame.grid(row=2, column=0, sticky=tk.NSEW)
		frame_panes = ttk.PanedWindow(data_frame, orient=tk.HORIZONTAL)
		frame_panes.grid(row=0, column=0, sticky=tk.NSEW)

		# Variable frame for list of quantitative columns/variables
		var_frame = tk.Frame(frame_panes, borderwidth=2, relief=tk.RIDGE)
		var_frame.grid(row=0, column=0, sticky=tk.NSEW)
		var_frame.rowconfigure(0, weight=1)
		var_frame.columnconfigure(0, weight=1)
		frame_panes.add(var_frame, weight=1)
		# Add multi-select list of variables to the leftmost pane
		self.column_frame, self.column_table = treeview_table(var_frame, rowset=[[v] for v in self.numeric_columns], \
				column_headers=['Variable'], select_mode=tk.EXTENDED, nrows=min(10, len(self.numeric_columns)))
		self.column_frame.grid(row=0, column=0, sticky=tk.NSEW)
		self.column_table.bind('<ButtonRelease-1>', self.q_recalc)

		# Content frame for output.  This contains a tabbed Notebook widget,
		# with separate pages for statistics of untransformed and log-transformed data.
		self.content_frame = tk.Frame(frame_panes, borderwidth=3, relief=tk.RIDGE)
		self.content_frame.grid(row=0, column=1, sticky=tk.NSEW)
		self.content_frame.rowconfigure(0, weight=1)
		self.content_frame.columnconfigure(0, weight=1)
		frame_panes.add(self.content_frame, weight=12)
		output_pages = ttk.Notebook(self.content_frame)
		output_pages.grid(row=0, column=0, sticky=tk.NSEW)
		self.data_page = tk.Frame(output_pages)
		self.log_page = tk.Frame(output_pages)
		self.data_page.name = "untransformed"
		self.log_page.name = "logtransformed"
		self.data_page.rowconfigure(0, weight=1)
		self.data_page.columnconfigure(0, weight=1)
		self.log_page.rowconfigure(0, weight=1)
		self.log_page.columnconfigure(0, weight=1)
		self.log_page.grid(row=0, column=0, sticky=tk.NSEW)
		self.log_page.grid(row=0, column=0, sticky=tk.NSEW)
		output_pages.add(self.data_page, text="Untransformed")
		output_pages.add(self.log_page, text="Log-transformed")

		# initialize content frame with headings, no data
		self.recalc()

		# Buttons
		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.columnconfigure(0, weight=1)
		btn_frame.grid(row=3, column=0, sticky=tk.EW, pady=(3,3))
		btn_frame.columnconfigure(0, weight=0)
		btn_frame.columnconfigure(1, weight=1)
		self.canceled = False
		self.help_btn = ttk.Button(btn_frame, text="Help", command=self.do_help, underline=0)
		self.help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.data_btn = ttk.Button(btn_frame, text="Source Data", state="disabled", command=self.show_data, underline=0)
		self.data_btn.grid(row=0, column=1, sticky=tk.W, padx=(3,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		close_btn = ttk.Button(btn_frame, text="Close", command=self.do_close, underline=0)
		close_btn.grid(row=0, column=2, sticky=tk.E, padx=(6,6))
		self.dlg.bind("<Alt-c>", self.do_close)
		self.dlg.bind("<Escape>", self.do_close)
		center_window(self.dlg)
		raise_window(self.dlg)

	def do_help(self, *args):
		webbrowser.open("https://mapdata.readthedocs.io/en/latest/dialogs.html#univariate-statistics-dialog", new=2, autoraise=True)

	def q_recalc(self, get_data=True, *args):
		if self.dataset is None or get_data:
			self.clear_output()
			self.get_data()
		if self.dataset is not None and len(self.dataset[0]) > 1:
			self.data_btn["state"] = "normal"
			self.recalc()
		else:
			self.data_btn["state"] = "disabled"

	def get_data(self):
		# Get the selected data into 'dataset'
		self.dataset = None
		column_list = []
		for sel_row in self.column_table.selection():
			datarow = self.column_table.item(sel_row)["values"]
			column_list.append(datarow[0])
		if len(column_list) > 0:
			# Get either only the selected data or all data.
			if self.sel_only_var.get() == "1":
				dataset = self.parent.get_sel_data(column_list)
			else:
				dataset = self.parent.get_all_data(column_list)
			if dataset is None or len(dataset[0]) == 0:
				self.dataset = None
				self.data_labels = None
			else:
				self.dataset = dataset
				self.data_labels = column_list

	def show_data(self, *args):
		if self.dataset is not None:
			dlg = MsgDialog2("Source Data", "Selected data:")
			variables = len(self.dataset)
			rowwise_data = []
			for i in range(len(self.dataset[0])):
				row = []
				for j in range(variables):
					row.append(self.dataset[j][i])
				rowwise_data.append(row)
			tframe, tdata = treeview_table(dlg.content_frame, rowwise_data, self.data_labels[0:variables])
			tframe.grid(row=0, column=0, sticky=tk.NSEW)
			def save_data(*args):
				export_data_table(self.data_labels[0:variables], rowwise_data, sheetname="Data for univariate stats")
			dlg.dlg.bind("<Control-s>", save_data)
			dlg.show()

	def recalc(self):
		# Calculate and display statistics for each selected column.
		for ctl in self.data_page.winfo_children():
			ctl.destroy()
		for ctl in self.log_page.winfo_children():
			ctl.destroy()
		self.statdata = []
		self.logstatdata = []
		if self.data_labels is not None:
			for i in range(len(self.data_labels)):
				d = [float(dv) for dv in self.dataset[i] if dv is not None and not (type(dv) == type('') and dv.strip()=='')]
				logd = []
				if len(d) > 0:
					log_error = False
					for w in range(len(d)):
						try:
							logd.append(math.log10(d[w]))
						except:
							log_error = True
							break
					if log_error:
						logd = []
					if len(logd) > 0:
						meanlog = statistics.fmean(logd)
					#=== Descriptive statistics for untransformed data
					mean_d = statistics.fmean(d)
					dd = [self.data_labels[i], len(d), min(d), max(d), round_figs(mean_d, 4)]
					stdev_d = statistics.stdev(d)
					if len(d) > 1:
						dd.extend([round_figs(statistics.median(d), 4), statistics.mode(d)])
					else:
						dd.extend([None, None])
					# Geo. mean
					if len(logd) > 0:
						dd.append(round_figs(10**meanlog, 4))
					else:
						dd.append(None)
					# Std. dev.
					if len(d) > 1:
						dd.append(round_figs(stdev_d, 4))
					else:
						dd.append(None)
					# Coefficient of variation
					if len(d) > 1:
						dd.append(round_figs(stdev_d/mean_d, 4))
					else:
						dd.append(None)
					# Percentiles
					if len(d) > 4:
						dd.extend([round_figs(np.percentile(d, 5)), round_figs(np.percentile(d, 95))])
					else:
						dd.extend(["NC", "NC"])
					#=== Normality tests for untransformed data
					if len(d) > 4:
						da = np.array(d)
						normstats = []
						if stdev_d != 0.0:
							try:
								adval, adpval = normal_ad(da)
								normstats.append("%.2E" % adpval)
							except:
								normstats.append("NC")
							try:
								lfval, lfpval = kstest_normal(da)
								normstats.append("%.2E" % lfpval)
							except:
								normstats.append("NC")
						else:
							normstats = ["NC", "NC"]
						dd.extend(normstats)
					else:
						dd.extend([None, None])
					#=== Outlier evaluation for untransformed data
					if len(d) > 14:
						max_outliers = 5 if len(d) < 100 else 10
						try:
							# May be ZeroDivisionError if sd == 0
							dd.append(rosners_test(d, 0.05, max_outliers))
						except:
							dd.append("NC")
					else:
						dd.append(None)
					if len(d) > 5:
						dd.append(tukey_outliers(d))
					else:
						dd.append("NC")
					#
					#=== Descriptive statistics for log-transformed data
					logdd = [self.data_labels[i], len(logd)]
					mean_dd = statistics.fmean(logd)
					if len(logd) > 0:
						logdd.extend([round_figs(min(logd),4), round_figs(max(logd),4), round_figs(mean_dd, 4)])
					else:
						logdd.extend([None, None, None])
					if len(logd) > 1:
						logdd.extend([round_figs(statistics.median(logd), 4), statistics.mode(logd)])
					else:
						logdd.extend([None, None])
					if len(logd) > 1:
						stdevlog = statistics.stdev(logd)
						logdd.append(round_figs(stdevlog, 4))
					else:
						logdd.append(None)
					# Coefficient of variation
					if len(logd) > 1:
						logdd.append(round_figs(stdevlog/mean_dd, 4))
					else:
						logdd.append(None)
					# Percentiles
					if len(logd) > 4:
						logdd.extend([round_figs(np.percentile(logd, 5)), round_figs(np.percentile(logd, 95))])
					else:
						logdd.extend(["NC", "NC"])
					# Normality tests for log-transformed data
					if len(logd) > 4:
						logda = np.array(logd)
						lognormstats = []
						if stdevlog != 0.0:
							try:
								adval, adpval = normal_ad(logda)
								lognormstats.append("%.2E" % adpval)
							except:
								normstats.append("NC")
							try:
								lfval, lfpval = kstest_normal(logda)
								lognormstats.append("%.2E" % lfpval)
							except:
								lognormstats.append("NC")
						else:
							lognormstats = ["NC", "NC"]
						logdd.extend(lognormstats)
					else:
						logdd.extend([None, None])
					# Outlier evaluation for log-transformed data
					if len(logd) > 14:
						max_outliers = 5 if len(logd) < 100 else 10
						try:
							# May be ZeroDivisionError if sd == 0
							logdd.append(rosners_test(logd, 0.05, max_outliers))
						except:
							logdd.append("NC")
					else:
						logdd.append(None)
					if len(logd) > 5:
						logdd.append(tukey_outliers(logd))
					else:
						logdd.append("NC")
					self.statdata.append(dd)
					self.logstatdata.append(logdd)
		#if len(self.statdata) > 0:
		tframe, tdata = treeview_table(self.data_page, self.statdata, self.dnames)
		tframe.grid(row=0, column=0, sticky=tk.NSEW)
		ltframe, ltdata = treeview_table(self.log_page, self.logstatdata, self.logdnames)
		ltframe.grid(row=0, column=0, sticky=tk.NSEW)
		self.dlg.bind("<Control-s>", self.save_data)
		self.dlg.bind("<Control-z>", self.save_logdata)

	def clear_output(self):
		for ctl in self.data_page.winfo_children():
			ctl.destroy()
		for ctl in self.log_page.winfo_children():
			ctl.destroy()
		self.dlg.bind("<Control-s>")
		self.dlg.bind("<Control-z>")
		self.dataset = None
		tframe, tdata = treeview_table(self.log_page, [], self.dnames)
		tframe.grid(row=0, column=0, sticky=tk.NSEW)
		ltframe, ltdata = treeview_table(self.log_page, [], self.logdnames)
		ltframe.grid(row=0, column=0, sticky=tk.NSEW)
	def save_data(self, *args):
		export_data_table(self.dnames, self.statdata, sheetname="Stats for untransformed data")
	def save_logdata(self, *args):
		export_data_table(self.logdnames, self.logstatdata, sheetname="Stats for log-transformed data")
	def show(self):
		self.dlg.wait_window(self.dlg)
	def do_close(self, *args):
		self.parent.remove_univar(self)
		try:
			self.dlg.destroy()
		except:
			pass




class BivarStatsDialog(object):
	def __init__(self, parent, column_specs):
		self.parent = parent
		self.column_specs = column_specs
		self.alpha = 0.45
		self.theilsen = False
		self.dlg = tk.Toplevel()
		self.dlg.title("Bivariate Statistics")
		self.dlg.rowconfigure(0, weight=0)
		self.dlg.rowconfigure(1, weight=0)
		self.dlg.rowconfigure(2, weight=1)
		self.dlg.rowconfigure(3, weight=0)
		self.dlg.columnconfigure(0, weight=1)
		#self.dlg.bind("<FocusIn>", self.q_recalc)
		self.dlg.bind("<Alt-a>")
		self.dlg.bind("<Alt-s>")
		# Data
		self.dataset = None
		self.data_labels = None
		self.numeric_columns = [c[0] for c in self.column_specs if c[1] in ("int", "float")]
		self.numeric_columns.sort()
		self.quant_columns = [c[0] for c in self.column_specs if c[1] in ("int", "float", "date", "timestamp", "timestamptz")]
		self.quant_columns.sort()
		self.output_columns = ["Statistic", "Value"]
		self.statdata = []
		# Message
		prompt_frame = tk.Frame(self.dlg, borderwidth=5)
		prompt_frame.grid(row=0, column=0, sticky=tk.N+tk.EW, pady=(3,3))
		prompt_frame.columnconfigure(0, weight=1)
		msg_lbl = ttk.Label(prompt_frame, width=70, text="Select two variables.  If the X variable is a date/time, only the results of a Runs Test will be shown.")
		msg_lbl.grid(row=0, column=0, sticky=tk.W, padx=(6,6), pady=(3,3))
		def wrap_msg(event):
			msg_lbl.configure(wraplength=event.width - 5)
		msg_lbl.bind("<Configure>", wrap_msg)

		# Controls
		ctrl_frame = tk.Frame(self.dlg, borderwidth=5)
		ctrl_frame.grid(row=1, column=0, sticky=tk.N+tk.EW)

		self.x_var = tk.StringVar(ctrl_frame, "")
		x_lbl = ttk.Label(ctrl_frame, text="X column:")
		x_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.x_sel = ttk.Combobox(ctrl_frame, state="normal", textvariable=self.x_var, values=self.quant_columns, width=24)
		self.x_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.x_sel.bind("<<ComboboxSelected>>", self.q_recalc)

		self.y_var = tk.StringVar(ctrl_frame, "")
		y_lbl = ttk.Label(ctrl_frame, text="Y column:")
		y_lbl.grid(row=1, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.y_sel = ttk.Combobox(ctrl_frame, state="normal", textvariable=self.y_var, values=self.numeric_columns, width=24)
		self.y_sel.grid(row=1, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.y_sel.bind("<<ComboboxSelected>>", self.q_recalc)

		self.xlog_var = tk.StringVar(ctrl_frame, "0")
		self.xlog_ck = ttk.Checkbutton(ctrl_frame, text="Log X", state="normal", command=self.q_recalc, variable=self.xlog_var,
				onvalue="1", offvalue="0")
		self.xlog_ck.grid(row=0, column=2, sticky=tk.W, padx=(6,6), pady=(3,3))

		self.ylog_var = tk.StringVar(ctrl_frame, "0")
		self.ylog_ck = ttk.Checkbutton(ctrl_frame, text="Log Y", state="normal", command=self.q_recalc, variable=self.ylog_var,
				onvalue="1", offvalue="0")
		self.ylog_ck.grid(row=1, column=2, sticky=tk.W, padx=(6,6), pady=(3,3))

		self.sel_only_var = tk.StringVar(ctrl_frame, "0")
		self.sel_only_ck = ttk.Checkbutton(ctrl_frame, text="Selected data only", command=self.q_recalc, variable=self.sel_only_var,
				onvalue="1", offvalue="0")
		self.sel_only_ck.grid(row=2, column=0, columnspan=2, sticky=tk.W, padx=(6,3), pady=(3,3))

		# Frame for output summary table and plot
		self.content_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		self.content_frame.grid(row=2, column=0, sticky=tk.NSEW)
		self.dlg.rowconfigure(2, weight=1)
		self.dlg.columnconfigure(0, weight=1)
		self.content_frame.rowconfigure(0, weight=1)
		self.content_frame.columnconfigure(0, weight=1)

		# PanedWindow within content_frame: left is table, right is plot
		output_panes = ttk.PanedWindow(self.content_frame, orient=tk.HORIZONTAL)
		output_panes.grid(row=0, column=0, sticky=tk.NSEW)

		self.out_tbl_frm = tk.Frame(output_panes, borderwidth=3, relief=tk.RIDGE)
		self.out_tbl_frm.grid(row=0, column=0, sticky=tk.NSEW)
		self.out_tbl_frm.rowconfigure(0, weight=1)
		self.out_tbl_frm.columnconfigure(0, weight=1)
		output_panes.add(self.out_tbl_frm, weight=1)

		self.out_plt_frm = tk.Frame(output_panes, borderwidth=3, relief=tk.RIDGE)
		self.out_plt_frm.grid(row=0, column=1, sticky=tk.NSEW)
		self.out_plt_frm.rowconfigure(0, weight=1)
		self.out_plt_frm.columnconfigure(0, weight=1)
		output_panes.add(self.out_plt_frm, weight=1)
		self.plotfig = Figure(dpi=100)
		self.plotfig.set_figheight(2)
		self.plotfig.set_figwidth(2)
		self.plotfig_canvas = FigureCanvasTkAgg(self.plotfig, self.out_plt_frm)
		self.plot_nav = NavigationToolbar2Tk(self.plotfig_canvas, self.out_plt_frm)
		#self.plot_axes = self.plotfig.add_subplot(111)
		self.plotfig_canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
		self.plot_nav.update()

		# Initialize output frames
		self.clear_output()

		# Buttons
		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.columnconfigure(0, weight=1)
		btn_frame.grid(row=3, column=0, sticky=tk.EW, pady=(3,3))
		btn_frame.columnconfigure(0, weight=0)
		btn_frame.columnconfigure(1, weight=0)
		btn_frame.columnconfigure(2, weight=1)
		self.canceled = False
		self.help_btn = ttk.Button(btn_frame, text="Help", command=self.do_help, underline=0)
		self.help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		self.data_btn = ttk.Button(btn_frame, text="Source Data", state="disabled", command=self.show_data, underline=0)
		self.data_btn.grid(row=0, column=1, sticky=tk.W, padx=(3,3))
		close_btn = ttk.Button(btn_frame, text="Close", command=self.do_close, underline=0)
		close_btn.grid(row=0, column=2, sticky=tk.E, padx=(6,6))
		self.dlg.bind("<Alt-c>", self.do_close)
		self.dlg.bind("<Escape>", self.do_close)
		center_window(self.dlg)
		raise_window(self.dlg)

	def do_help(self, *args):
		webbrowser.open("https://mapdata.readthedocs.io/en/latest/dialogs.html#bivariate-statistics-dialog", new=2, autoraise=True)

	def get_data(self):
		self.data_btn["state"] = "disabled"
		self.dataset = None
		column_list = [self.x_var.get(), self.y_var.get()]
		# Get either only the selected data or all data.
		if self.sel_only_var.get() == "1":
			dataset = self.parent.get_sel_data(column_list)
		else:
			dataset = self.parent.get_all_data(column_list)
		if dataset is None or len(dataset[0]) == 0:
			self.dataset = None
			self.data_labels = None
			self.data_btn["state"] = "disabled"
		else:
			# Remove missing data
			column_indexes = range(len(dataset))
			clean_data = [[] for _ in dataset]
			for i in range(len(dataset[0])):
				ok = True
				for col in column_indexes:
					if dataset[col][i] is None or dataset[col][i] == '':
						ok = False
				if ok:
					for col in column_indexes:
						clean_data[col].append(dataset[col][i])
			dataset = None
			# Convert to floats for numeric data only
			for i in range(len(clean_data)):
				if column_list[i] in self.numeric_columns:
					clean_data[i] = [conv_float(v) for v in clean_data[i]]
			# Convert X to date or datetime values if necessary
			xvar_type = [self.column_specs[i][1] for i in range(len(self.column_specs)) if self.column_specs[i][0] == self.x_var.get()][0]
			if xvar_type == "date":
				clean_data[0] = [parse_date(x) for x in clean_data[0]]
			elif xvar_type == "timestamp":
				clean_data[0] = [parse_datetime(x) for x in clean_data[0]]
			elif xvar_type == "timestamptz":
				clean_data[0] = [parse_datetimetz(x) for x in clean_data[0]]
			# Log-transform data if specified.
			log_data = [[] for _ in clean_data]
			log_error = False
			if self.xlog_ck["state"] != "disabled" and self.xlog_var.get() == "1" and self.x_var.get() in self.numeric_columns:
				for i in range(len(clean_data[0])):
					try:
						log_data[0].append(math.log10(clean_data[0][i]))
					except:
						log_error = True
						self.xlog_var.set("0")
						self.xlog_ck["state"] = "disabled"
						break
				if not log_error:
					clean_data[0] = log_data[0]
					self.data_labels[0] = "Log10 of " + self.x_var.get()
			if self.ylog_ck["state"] != "disabled" and self.ylog_var.get() == "1" and len(clean_data) > 1 \
					and self.y_var.get() in self.numeric_columns:
				log_error = False
				for i in range(len(clean_data[1])):
					try:
						log_data[1].append(math.log10(clean_data[1][i]))
					except:
						log_error = True
						self.ylog_var.set("0")
						self.ylog_ck["state"] = "disabled"
						break
				if not log_error:
					clean_data[1] = log_data[1]
					self.data_labels[1] = "Log10 of " + self.y_var.get()
			log_data = None
			#
			self.dataset = sort_columns(clean_data)
			if self.xlog_ck["state"] != "disabled" and self.xlog_var.get() == "1":
				self.data_labels = ["Log10 of %s" % column_list[0]]
			else:
				self.data_labels = [column_list[0]]
			if self.ylog_ck["state"] != "disabled" and self.ylog_var.get() == "1":
				self.data_labels.append("Log10 of %s" % column_list[1])
			else:
				self.data_labels.append(column_list[1])
			self.data_btn["state"] = "normal"

	def clear_output(self):
		for ctl in self.out_tbl_frm.winfo_children():
			ctl.destroy()
		self.dlg.bind("<Alt-a>")
		self.dlg.bind("<Alt-s>")
		self.dlg.bind("<Control-s>")
		tframe, tdata = treeview_table(self.out_tbl_frm, [], ["Statistic", "Value"])
		tframe.grid(row=0, column=0, sticky=tk.NSEW)
		self.plotfig.clear()
		self.plot_axes = self.plotfig.add_subplot(111)
		self.plotfig_canvas.draw()

	def show_data(self, *args):
		if self.dataset is not None:
			dlg = MsgDialog2("Source Data", "Selected data:")
			variables = len(self.dataset)
			rowwise_data = []
			for i in range(len(self.dataset[0])):
				row = []
				for j in range(variables):
					row.append(self.dataset[j][i])
				rowwise_data.append(row)
			tframe, tdata = treeview_table(dlg.content_frame, rowwise_data, self.data_labels[0:variables])
			tframe.grid(row=0, column=0, sticky=tk.NSEW)
			def save_data(*args):
				export_data_table(self.data_labels[0:variables], rowwise_data, sheetname="Data for bivariate statistics")
			dlg.dlg.bind("<Control-s>", save_data)
			dlg.show()

	def q_recalc(self, get_data=True, *args):
		if self.x_var.get() != '' and self.y_var.get() != '':
			if get_data or self.dataset is None:
				self.get_data()
			if self.dataset is not None and len(self.dataset[0]) > 2:
				self.recalc()
			else:
				self.clear_output()

	def recalc(self):
		self.clear_output()
		for ctl in self.out_tbl_frm.winfo_children():
			ctl.destroy()
		self.statdata = []
		if self.dataset is not None:
			regr_run = False
			total_ss = 0.0
			x_is_numeric = self.x_var.get() in self.numeric_columns
			N = len(self.dataset[0])
			xpfx = "" if self.xlog_var.get() == '0' else "Log10 of "
			ypfx = "" if self.ylog_var.get() == '0' else "Log10 of "
			self.statdata.append(["X variable", xpfx + self.x_var.get()])
			self.statdata.append(["Y variable", ypfx + self.y_var.get()])
			self.statdata.append(["N", N])
			if N > 1:
				if x_is_numeric:
					xmean = statistics.fmean(self.dataset[0])
					xmid = xmean
				else:
					xmean = None
					xmid = (max(self.dataset[0]) - min(self.dataset[0]))/2
				ymean = statistics.fmean(self.dataset[1])
				xarray = np.array(self.dataset[0])
				yarray = np.array(self.dataset[1])
				if x_is_numeric:
					try:
						self.statdata.append(["Covariance", round_figs(statistics.covariance(self.dataset[0], self.dataset[1]), 3)])
					except:
						pass
					try:
						pearsonr = spstats.pearsonr(xarray, yarray)
						self.statdata.append(["Pearson's r", round_figs(pearsonr.statistic)])
						rp = pearsonr.pvalue
						if rp < 0.001:
							rp = "%.2E" % rp
						else:
							rp = round_figs(rp, 3)
						self.statdata.append(["p value for Pearson's r", rp])
					except:
						pass
					try:
						spearmanr = spstats.spearmanr(xarray, yarray)
						self.statdata.append(["Spearman's rho", round_figs(spearmanr.statistic)])
						sp = spearmanr.pvalue
						if sp < 0.001:
							sp = "%.2E" % sp
						else:
							sp = round_figs(sp, 3)
						self.statdata.append(["p value for Spearman's rho", sp])
					except TypeError:
						pass
					try:
						kendalltau = spstats.kendalltau(np.array(self.dataset[0]), np.array(self.dataset[1]))
						self.statdata.append(["Kendall's tau", round_figs(kendalltau.statistic)])
						ktau = kendalltau.pvalue
						if ktau < 0.001:
							ktau = "%.2E" % ktau
						else:
							ktau = round_figs(ktau, 3)
						self.statdata.append(["p value for Kendall's tau", ktau])
					except:
						pass
					# Linear regression
					ols_model = sm.OLS(np.array(self.dataset[1]), sm.add_constant(np.array(self.dataset[0])))
					ols_fit = ols_model.fit()
					slope = ols_fit.params[1]
					regr_run = True
					self.statdata.append(["OLS Regression slope", round_figs(slope)])
					self.statdata.append(["OLS Regression intercept", round_figs(ols_fit.params[0])])
					try:
						self.statdata.append(["Regression R squared", round_figs(ols_fit.rsquared)])
					except:
						pass
					try:
						self.statdata.append(["Regression adj. R squared", round_figs(ols_fit.rsquared_adj)])
					except:
						pass
					self.statdata.append(["Regression total SS", round_figs(ols_fit.centered_tss, 4)])
					total_ss = sum([(self.dataset[1][i] - ymean)**2 for i in range(N)])
					try:
						self.statdata.append(["Regression explained SS", round_figs(ols_fit.ess, 4)])
					except:
						pass
					try:
						self.statdata.append(["Regression residual SS", round_figs(ols_fit.ssr, 4)])
					except:
						pass
					self.statdata.append(["Regression p for slope=0", "%.2E" % ols_fit.pvalues[1]])
					self.statdata.append(["Regression p for intercept=0", "%.2E" % ols_fit.pvalues[0]])
					self.statdata.append(["Regression AIC", round_figs(ols_fit.aic, 3)])
					self.statdata.append(["Regression BIC", round_figs(ols_fit.bic, 3)])

				# Runs test
				#if not regr_run or total_ss == 0.0:
				#	runs_p = "NC"
				#else:
				rz, rp = runstest_1samp(self.dataset[1], cutoff="median")
				runs_p = "%.2E" % rp
				self.statdata.append(["Runs test for Y; p value", runs_p])

				if x_is_numeric:
					# Theil-Sen estimators
					ts_slope, ts_intercept, ts_high, ts_low = spstats.theilslopes(yarray, xarray)
					self.statdata.append(["Theil-Sen slope", round_figs(ts_slope)])
					self.statdata.append(["Theil-Sen 95% CI on slope", "(%s, %s)" % (round_figs(ts_high), round_figs(ts_low))])
					self.statdata.append(["Theil-Sen intercept", round_figs(ts_intercept)])

				# Update plot
				self.plot_axes.set_xlabel(self.data_labels[0])
				self.plot_axes.set_ylabel(self.data_labels[1])
				if regr_run:
					ols_ci = ols_fit.get_prediction().summary_frame()
					self.plot_axes.fill_between(self.dataset[0], list(ols_ci["mean_ci_lower"]), list(ols_ci["mean_ci_upper"]), color="antiquewhite", edgecolor="goldenrod", label="95% CI")
					self.plot_axes.axline((xmean, ymean), slope=slope, clip_on=True, label="Linear fit", color="darkorange", linestyle="dotted")
					self.plot_axes.legend()
				if self.theilsen and x_is_numeric:
					self.plot_axes.axline((statistics.median(self.dataset[0]), statistics.median(self.dataset[1])), slope=ts_slope, \
							clip_on=True, label="Theil-Sen line", color="darkgreen", linestyle="dotted", linewidth=2, alpha=0.65)
				self.plot_axes.scatter(self.dataset[0], self.dataset[1], alpha=self.alpha)
				self.plotfig_canvas.draw()
				self.plot_nav.update()
				self.dlg.bind("<Alt-a>", self.set_alpha)
				self.dlg.bind("<Alt-s>", self.set_theilsen)
		if len(self.statdata) > 0:
			tframe, tdata = treeview_table(self.out_tbl_frm, self.statdata, self.output_columns)
			tframe.grid(row=0, column=0, stick=tk.NSEW)
			self.dlg.bind("<Control-s>", self.save_table)
			self.dlg.minsize(width=900, height=500)

	def set_alpha(self, *args):
		dlg = OneFloatDialog(self.dlg, "Transparency", "Enter the opacity (alpha) value", min_value=0.0, max_value=1.0, initial=self.alpha)
		new_alpha = dlg.show()
		if new_alpha is not None:
			self.alpha = new_alpha
			self.q_recalc()
	def set_theilsen(self, *args):
		self.theilsen = not self.theilsen
		self.q_recalc()
	def save_table(self, *args):
		export_data_table(self.output_columns, self.statdata, sheetname="Selected map items")
	def show(self):
		#self.dlg.minsize(width=300, height=300)
		self.dlg.wait_window(self.dlg)
	def do_close(self, *args):
		self.parent.remove_bivar(self)
		try:
			self.dlg.destroy()
		except:
			pass



class CorrMatrixDialog(object):
	def __init__(self, parent, column_specs):
		self.parent = parent
		self.column_specs = column_specs
		self.dlg = tk.Toplevel()
		self.dlg.title("Correlation Matrix")
		self.dlg.rowconfigure(0, weight=0)
		self.dlg.rowconfigure(1, weight=0)
		self.dlg.rowconfigure(2, weight=1)
		self.dlg.rowconfigure(3, weight=0)
		self.dlg.columnconfigure(0, weight=1)
		#self.dlg.bind("<FocusIn>", self.q_redraw)
		# Data
		self.dataset = None
		self.data_labels = None
		self.numeric_columns = [c[0] for c in self.column_specs if c[1] in ("int", "float")]
		self.numeric_columns.sort()
		self.dlg.bind("<Control-s>")
		# Message
		prompt_frame = tk.Frame(self.dlg, borderwidth=5)
		prompt_frame.grid(row=0, column=0, sticky=tk.N+tk.EW, pady=(3,3))
		prompt_frame.columnconfigure(0, weight=1)
		msg_lbl = ttk.Label(prompt_frame, width=65, text="Select two or more variables from the left to see the correlation matrix.  Use Ctrl-click or Shift-click to select multiple rows.")
		msg_lbl.grid(row=0, column=0, sticky=tk.W, padx=(6,6), pady=(3,3))
		def wrap_msg(event):
			msg_lbl.configure(wraplength=event.width - 5)
		msg_lbl.bind("<Configure>", wrap_msg)

		# Controls
		# Top controls are only the 'Selected only' checkbox
		ctrl_frame = tk.Frame(self.dlg, borderwidth=5)
		ctrl_frame.rowconfigure(0, weight=0)
		ctrl_frame.columnconfigure(0, weight=1)
		ctrl_frame.grid(row=1, column=0, sticky=tk.N+tk.EW)
		self.sel_only_var = tk.StringVar(ctrl_frame, "0")
		self.sel_only_ck = ttk.Checkbutton(ctrl_frame, text="Selected data only", command=self.q_redraw, variable=self.sel_only_var,
				onvalue="1", offvalue="0")
		self.sel_only_ck.grid(row=0, column=0, sticky=tk.W, padx=(6,3), pady=(3,3))
		self.log10_var = tk.StringVar(ctrl_frame, "0")
		self.log10_ck = ttk.Checkbutton(ctrl_frame, text="Log10 transform data", command=self.q_redraw, variable=self.log10_var,
				onvalue="1", offvalue="0")
		self.log10_ck.grid(row=1, column=0, sticky=tk.W, padx=(6,3), pady=(3,3))

		# The data_frame encompasses the two panes of the variable frame and the content frame
		data_frame = tk.Frame(self.dlg, borderwidth=5)
		data_frame.rowconfigure(0, weight=1)
		data_frame.columnconfigure(0, weight=1)
		data_frame.grid(row=2, column=0, sticky=tk.NSEW)
		frame_panes = ttk.PanedWindow(data_frame, orient=tk.HORIZONTAL)
		frame_panes.grid(row=0, column=0, sticky=tk.NSEW)

		# Variable frame for list of quantitative columns/variables
		var_frame = tk.Frame(frame_panes, borderwidth=2, relief=tk.RIDGE)
		var_frame.grid(row=0, column=0, sticky=tk.NSEW)
		var_frame.rowconfigure(0, weight=1)
		var_frame.columnconfigure(0, weight=1)
		frame_panes.add(var_frame, weight=1)
		# Add multi-select list of variables to the leftmost pane
		self.column_frame, self.column_table = treeview_table(var_frame, rowset=[[v] for v in self.numeric_columns], \
				column_headers=['Variable'], select_mode=tk.EXTENDED, nrows=min(10, len(self.numeric_columns)))
		self.column_frame.grid(row=0, column=0, sticky=tk.NSEW)
		self.column_table.bind('<ButtonRelease-1>', self.q_redraw)

		# Content frame for correlation matrix figure
		self.content_frame = tk.Frame(frame_panes, borderwidth=3, relief=tk.RIDGE)
		self.content_frame.grid(row=0, column=1, sticky=tk.NSEW)
		self.content_frame.rowconfigure(0, weight=1)
		self.content_frame.columnconfigure(0, weight=1)
		frame_panes.add(self.content_frame, weight=12)
		self.plotfig = Figure(dpi=100)
		self.plotfig.set_figheight(2)
		self.plotfig.set_figwidth(2)
		self.plotfig_canvas = FigureCanvasTkAgg(self.plotfig, self.content_frame)
		self.plot_nav = NavigationToolbar2Tk(self.plotfig_canvas, self.content_frame)
		self.plot_axes = self.plotfig.add_subplot(111)
		self.plotfig_canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
		self.plot_nav.update()
		# initialize content frame with an empty plot
		self.clear_output()

		# Buttons
		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.columnconfigure(0, weight=1)
		btn_frame.grid(row=3, column=0, sticky=tk.EW, pady=(3,3))
		btn_frame.columnconfigure(0, weight=0)
		btn_frame.columnconfigure(1, weight=1)
		self.canceled = False
		self.help_btn = ttk.Button(btn_frame, text="Help", command=self.do_help, underline=0)
		self.help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.data_btn = ttk.Button(btn_frame, text="Source Data", state="disabled", command=self.show_data, underline=0)
		self.data_btn.grid(row=0, column=1, sticky=tk.W, padx=(3,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		close_btn = ttk.Button(btn_frame, text="Close", command=self.do_close, underline=0)
		close_btn.grid(row=0, column=2, sticky=tk.E, padx=(6,6))
		self.dlg.bind("<Alt-c>", self.do_close)
		self.dlg.bind("<Escape>", self.do_close)
		center_window(self.dlg)
		raise_window(self.dlg)

	def do_help(self, *args):
		webbrowser.open("https://mapdata.readthedocs.io/en/latest/dialogs.html#correlation-matrix-dialog", new=2, autoraise=True)

	def show_data(self, *args):
		if self.dataset is not None:
			dlg = MsgDialog2("Source Data", "Selected data:")
			variables = len(self.dataset)
			rowwise_data = []
			for i in range(len(self.dataset[0])):
				row = []
				for j in range(variables):
					row.append(self.dataset[j][i])
				rowwise_data.append(row)
			tframe, tdata = treeview_table(dlg.content_frame, rowwise_data, self.data_labels[0:variables])
			tframe.grid(row=0, column=0, sticky=tk.NSEW)
			def save_data(*args):
				export_data_table(self.data_labels[0:variables], rowwise_data, sheetname="Data for correlation matrix")
			dlg.dlg.bind("<Control-s>", save_data)
			dlg.show()

	def clear_output(self):
		self.plotfig.clear()
		self.plot_axes = self.plotfig.add_subplot(111)
		self.plotfig_canvas.draw()

	def q_redraw(self, get_data=True, *args):
		if self.dataset is None or get_data:
			self.clear_output()
			self.get_data()
		if self.dataset is not None and len(self.dataset) > 1 and len(self.dataset[0]) > 1:
			self.redraw()
		else:
			self.clear_output()
			self.data_btn["state"] = "disabled"

	def get_data(self):
		# Get the selected data into 'dataset'
		self.dataset = None
		column_list = []
		for sel_row in self.column_table.selection():
			datarow = self.column_table.item(sel_row)["values"]
			column_list.append(datarow[0])
		if len(column_list) > 0:
			# Get either only the selected data or all data.
			if self.sel_only_var.get() == "1":
				dataset = self.parent.get_sel_data(column_list)
			else:
				dataset = self.parent.get_all_data(column_list)
			if dataset is None or len(dataset[0]) == 0:
				self.dataset = None
				self.data_labels = None
			else:
				self.dataset = dataset
				self.data_labels = column_list
			# Remove missing data
			column_indexes = range(len(dataset))
			clean_data = [[] for _ in dataset]
			clean_labels = copy.copy(self.data_labels)
			for i in range(len(dataset[0])):
				ok = True
				for col in column_indexes:
					if dataset[col][i] is None or dataset[col][i] == '':
						ok = False
				if ok:
					for col in column_indexes:
						clean_data[col].append(dataset[col][i])
			dataset = None
			# Convert to floats
			for i in range(len(clean_data)):
				clean_data[i] = [conv_float(v) for v in clean_data[i]]
			# Log-transform data if specified.
			if self.log10_var.get() == '1':
				log_data = [[] for _ in clean_data]
				log_error = False
				nrows = len(clean_data[0])
				for v in range(len(clean_data)):
					for i in range(nrows):
						try:
							log_data[v].append(math.log10(clean_data[v][i]))
						except:
							log_error = True
							break
					if log_error:
						break
				if log_error:
					warning("Data can not all be log-transformed.", {})
					self.log10_var.set("0")
				else:
					clean_data = log_data
					clean_labels = ["Log10 of " + v for v in clean_labels]
			#
			self.dataset = clean_data
			self.data_labels = clean_labels
			self.data_btn["state"] = "normal"

	def redraw(self):
		# (Re)draw the correlation matrix
		if self.data_labels is not None and len(self.data_labels) > 1:
			nvar = len(self.data_labels)
			dmat = np.asarray(self.dataset)
			cormat = np.corrcoef(dmat)
			#caxes = self.plot_axes.matshow(cormat, cmap="PiYG", vmin=-1.0, vmax=1.0)
			caxes = self.plot_axes.matshow(cormat, cmap="BrBG", vmin=-1.0, vmax=1.0)
			self.plotfig.colorbar(caxes)
			self.plot_axes.set_xticks(range(nvar), labels=self.data_labels, rotation=25)
			self.plot_axes.set_yticks(range(nvar), labels=self.data_labels)
			for i in range(nvar):
				for j in range(nvar):
					v = cormat[i,j]
					c = "white" if abs(v) > 0.40 else "black"
					self.plot_axes.text(j, i, f"{cormat[i,j]:.2f}", ha="center", va="center", color=c)
			self.plotfig_canvas.draw()

	def show(self):
		self.dlg.wait_window(self.dlg)
	def do_close(self, *args):
		self.parent.remove_corrmat(self)
		try:
			self.dlg.destroy()
		except:
			pass




class CategCorrespDialog(object):
	def __init__(self, parent, column_specs):
		self.parent = parent
		self.column_specs = column_specs
		self.dlg = tk.Toplevel()
		self.dlg.title("Categorical Variable Correspondence")
		self.dlg.columnconfigure(0, weight=1)
		#self.dlg.bind("<FocusIn>", self.q_recalc)
		self.dlg.bind("<Control-s>")
		# Data
		self.dataset = None
		self.categ_columns2 = [c[0] for c in self.column_specs if c[1] in ("string", "boolean", "date")]
		self.categ_columns2.sort()
		self.column_headers = []
		self.statdata = []
		# Message
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.NSEW, pady=(3,3))
		prompt_frame.columnconfigure(0, weight=1)
		msg_lbl = ttk.Label(prompt_frame, width=70, text="Select two categorical variables to see the prevalence of all combinations.")
		msg_lbl.grid(row=0, column=0, sticky=tk.W, padx=(6,6), pady=(3,3))
		def wrap_msg(event):
			msg_lbl.configure(wraplength=event.width - 5)
		msg_lbl.bind("<Configure>", wrap_msg)

		# Controls
		ctrl_frame = tk.Frame(self.dlg)
		ctrl_frame.grid(row=1, column=0, sticky=tk.N+tk.EW)

		self.var1 = tk.StringVar(ctrl_frame, "")
		var1_lbl = ttk.Label(ctrl_frame, text="Variable 1:")
		var1_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.var1_sel = ttk.Combobox(ctrl_frame, state="normal", textvariable=self.var1, width=24)
		self.var1_sel["values"] = self.categ_columns2
		self.var1_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.var1_sel.bind("<<ComboboxSelected>>", self.check_var1)

		self.var2 = tk.StringVar(ctrl_frame, "")
		var2_lbl = ttk.Label(ctrl_frame, text="Variable 2:")
		var2_lbl.grid(row=0, column=2, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.var2_sel = ttk.Combobox(ctrl_frame, state="normal", textvariable=self.var2, width=24)
		self.var2_sel["values"] = self.categ_columns2
		self.var2_sel.grid(row=0, column=3, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.var2_sel.bind("<<ComboboxSelected>>", self.check_var2)

		self.sel_only_var = tk.StringVar(ctrl_frame, "0")
		self.sel_only_ck = ttk.Checkbutton(ctrl_frame, text="Selected data only", command=self.q_recalc, variable=self.sel_only_var,
				onvalue="1", offvalue="0")
		self.sel_only_ck.grid(row=1, column=0, columnspan=2, sticky=tk.W, padx=(6,3), pady=(3,3))

		# Frame for output table
		self.content_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		self.content_frame.grid(row=2, column=0, sticky=tk.NSEW)
		self.dlg.rowconfigure(2, weight=1)
		self.dlg.columnconfigure(0, weight=1)
		self.content_frame.rowconfigure(0, weight=1)
		self.content_frame.columnconfigure(0, weight=1)

		# Buttons
		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.columnconfigure(0, weight=1)
		btn_frame.grid(row=3, column=0, sticky=tk.EW, pady=(3,3))
		btn_frame.columnconfigure(0, weight=0)
		btn_frame.columnconfigure(1, weight=0)
		btn_frame.columnconfigure(2, weight=1)
		self.canceled = False
		self.help_btn = ttk.Button(btn_frame, text="Help", command=self.do_help, underline=0)
		self.help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		self.data_btn = ttk.Button(btn_frame, text="Source Data", state="disabled", command=self.show_data, underline=0)
		self.data_btn.grid(row=0, column=1, sticky=tk.W, padx=(3,3))
		close_btn = ttk.Button(btn_frame, text="Close", command=self.do_close, underline=0)
		close_btn.grid(row=0, column=2, sticky=tk.E, padx=(6,6))
		self.dlg.bind("<Alt-c>", self.do_close)
		self.dlg.bind("<Escape>", self.do_close)
		center_window(self.dlg)
		raise_window(self.dlg)

	def do_help(self, *args):
		webbrowser.open("https://mapdata.readthedocs.io/en/latest/dialogs.html#categorical-corrrespondence-dialog", new=2, autoraise=True)

	def clear_output(self):
		for ctl in self.content_frame.winfo_children():
			ctl.destroy()
		self.dlg.bind("<Control-s>")
		self.dataset = None

	def show_data(self, *args):
		# Show data that have been collected for co-occurrence display
		if self.dataset is not None:
			dlg = MsgDialog2("Source Data", "Original data:")
			variables = len(self.dataset)
			rowwise_data = []
			for i in range(len(self.dataset[0])):
				row = []
				for j in range(variables):
					row.append(self.dataset[j][i])
				rowwise_data.append(row)
			tframe, tdata = treeview_table(dlg.content_frame, rowwise_data, self.data_labels[0:variables])
			tframe.grid(row=0, column=0, sticky=tk.NSEW)
			def save_data(*args):
				export_data_table(self.data_labels[0:variables], rowwise_data, sheetname="Data for categorical correspondence")
			dlg.dlg.bind("<Control-s>", save_data)
			dlg.show()

	def check_var1(self, *args):
		self.clear_output()
		v1 = self.var1.get()
		v2 = self.var2.get()
		if v1 != '':
			if v2 != '':
				if v1 == v2:
					self.var2.set('')
				else:
					self.q_recalc()

	def check_var2(self, *args):
		self.clear_output()
		v1 = self.var1.get()
		v2 = self.var2.get()
		if v2 != '':
			if v1 != '':
				if v2 == v1:
					self.var1.set('')
				else:
					self.q_recalc()

	def q_recalc(self, get_data=True, *args):
		# Conditionally (re)calculate the co-occurrences
		can_recalc = self.var1.get() != '' and self.var2.get() != ''
		if can_recalc:
			if get_data or self.dataset is None:
				self.get_data()
			if self.dataset is not None:
				self.recalc()

	def get_data(self):
		# Get the selected data into 'dataset'
		self.data_btn["state"] = "disabled"
		self.dataset = None
		column_list = [self.var1.get(), self.var2.get()]
		# Get either only the selected data or all data.
		if self.sel_only_var.get() == "1":
			dataset = self.parent.get_sel_data(column_list)
		else:
			dataset = self.parent.get_all_data(column_list)
		if dataset is None or len(dataset[0]) == 0:
			self.dataset = None
			self.data_labels = None
			self.data_btn["state"] = "disabled"
		else:
			# DO NOT remove missing data for the categorical co-occurrence summary
			self.dataset = dataset
			self.data_labels = [self.var1.get(), self.var2.get()]
			self.data_btn["state"] = "normal"

	def recalc(self):
		# Put the data into a SQLite db for summarization
		self.db = sqlite3.connect(":memory:")
		cur = self.db.cursor()
		colnames = db_colnames(self.data_labels)
		cur.execute("create table cdata (%s);" % ",".join(colnames))
		tbldata = []
		for row_no in range(len(self.dataset[0])):
			row_vals = [self.dataset[0][row_no], self.dataset[1][row_no]]
			row_vals = [None if isinstance(x, str) and x.strip() == '' else x for x in row_vals]
			tbldata.append(row_vals)
		cur.executemany("insert into cdata values (?,?);", tbldata)
		# Create the summary
		sqlcmd = """select %s, count(*) as data_rows, round(100 * count(*)/total_rows, 3) as percent
from cdata cross join (select cast(count(*) as double) as total_rows from cdata)
group by 1,2 order by 1;""" % ",".join(colnames)
		result = cur.execute(sqlcmd)
		# Stuff the result into the content frame.
		self.output_columns = self.data_labels + ['Data rows', 'Percent']
		self.statdata = result.fetchall()
		tframe, tdata = treeview_table(self.content_frame, self.statdata, self.output_columns)
		cur.close()
		tframe.grid(row=0, column=0, sticky=tk.NSEW)
		self.dlg.bind("<Control-s>", self.save_table)

	def save_table(self, *args):
		export_data_table(self.output_columns, self.statdata, sheetname="Selected map items")
	def show(self):
		self.dlg.minsize(width=300, height=300)
		self.dlg.wait_window(self.dlg)
	def do_close(self, *args):
		self.parent.remove_categcorresp(self)
		try:
			self.dlg.destroy()
		except:
			pass



class MsgDialog(object):
	def __init__(self, title, message, parent=None, bgcolor=None, can_resize=True):
		if parent is not None:
			self.dlg = tk.Toplevel(parent)
		else:
			self.dlg = tk.Toplevel()
		if bgcolor is not None:
			self.dlg.configure(bg=bgcolor)
		self.dlg.title(title)
		self.dlg.rowconfigure(0, weight=1)
		self.dlg.columnconfigure(0, weight=2)
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.EW, padx=(6,6), pady=(6,6))
		prompt_frame.rowconfigure(0, weight=1)
		prompt_frame.columnconfigure(0, weight=2)
		msg_lbl = ttk.Label(prompt_frame, wraplength=100, text=message)
		msg_lbl.grid(row=0, column=0, sticky=tk.EW, padx=(6,6), pady=(3,3))
		def wrap_msg(event):
			msg_lbl.configure(wraplength=event.width - 5)
		msg_lbl.bind("<Configure>", wrap_msg)
		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.columnconfigure(0, weight=2)
		btn_frame.grid(row=1, column=0, sticky=tk.EW, pady=(0,0))
		btn_frame.columnconfigure(0, weight=1)
		# Buttons
		self.canceled = False
		ok_btn = ttk.Button(btn_frame, text="Close", command=self.do_select, underline=0)
		ok_btn.grid(row=0, column=0, sticky=tk.E, padx=(12,6))
		self.dlg.bind("<Alt-c>", self.do_select)
		self.dlg.bind("<Return>", self.do_select)
		self.dlg.bind("<Escape>", self.do_select)
		self.dlg.resizable(can_resize, can_resize)
		self.dlg.minsize(width=300, height=50)
		ok_btn.focus()
	def do_select(self, *args):
		self.dlg.destroy()
	def show(self, grab=False):
		if grab:
			self.dlg.grab_set()
		center_window(self.dlg)
		raise_window(self.dlg)
		self.dlg.attributes('-topmost', 'true')
		self.dlg.attributes('-topmost', 'false')
		self.dlg.wait_window(self.dlg)


class MsgDialog2(object):
	# With an extra content frame.
	def __init__(self, title, message, parent=None, bgcolor=None, can_resize=True):
		if parent is not None:
			self.dlg = tk.Toplevel(parent)
		else:
			self.dlg = tk.Toplevel()
		if bgcolor is not None:
			self.dlg.configure(bg=bgcolor)
		self.dlg.title(title)
		#self.dlg.rowconfigure(0, weight=1)
		self.dlg.columnconfigure(0, weight=1)
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.EW, padx=(6,6), pady=(6,6))
		prompt_frame.columnconfigure(0, weight=1)
		msg_lbl = ttk.Label(prompt_frame, wraplength=80, text=message)
		msg_lbl.grid(row=0, column=0, sticky=tk.EW, padx=(6,6), pady=(3,3))
		def wrap_msg(event):
			msg_lbl.configure(wraplength=event.width - 5)
		msg_lbl.bind("<Configure>", wrap_msg)
		self.content_frame = tk.Frame(self.dlg)
		self.content_frame.grid(row=1, column=0, sticky=tk.NSEW)
		self.dlg.rowconfigure(1, weight=1)
		self.content_frame.rowconfigure(0, weight=1)
		self.content_frame.columnconfigure(0, weight=1)
		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.columnconfigure(0, weight=1)
		btn_frame.grid(row=2, column=0, sticky=tk.EW, pady=(0,0))
		btn_frame.columnconfigure(0, weight=1)
		# Buttons
		self.canceled = False
		ok_btn = ttk.Button(btn_frame, text="Close", command=self.do_select, underline=0)
		ok_btn.grid(row=0, column=0, sticky=tk.E, padx=(12,6))
		self.dlg.bind("<Alt-c>", self.do_select)
		self.dlg.bind("<Return>", self.do_select)
		self.dlg.bind("<Escape>", self.do_select)
		self.dlg.resizable(can_resize, can_resize)
		self.dlg.minsize(width=300, height=50)
		#self.dlg.maxsize(width=1600, height=1000)
		ok_btn.focus()
	def do_select(self, *args):
		self.dlg.destroy()
	def show(self, grab=False):
		if grab:
			self.dlg.grab_set()
		center_window(self.dlg)
		raise_window(self.dlg)
		self.dlg.attributes('-topmost', 'true')
		self.dlg.attributes('-topmost', 'false')
		self.dlg.wait_window(self.dlg)


class OneEntryDialog(object):
	def __init__(self, parent, title, prompt):
		self.dlg = tk.Toplevel(parent)
		self.dlg.title(title)
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.NSEW, pady=(3,3))
		prompt_frame.columnconfigure(0, weight=1)
		msg_lbl = ttk.Label(prompt_frame, text=prompt)
		msg_lbl.grid(row=0, column=0, sticky=tk.W, padx=(6,6), pady=(6,6))
		self.entry_var = tk.StringVar(self.dlg, "")
		self.entry_var.trace('w', self.check_enable)
		self.val_entry = ttk.Entry(prompt_frame, width=50, textvariable=self.entry_var)
		self.val_entry.grid(row=1, column=0, sticky=tk.EW, padx=(6,6), pady=(3,3))
		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.columnconfigure(0, weight=1)
		btn_frame.grid(row=2, column=0, sticky=tk.EW, pady=(3,3))
		btn_frame.columnconfigure(0, weight=1)
		self.val_entry.focus()
		# Buttons
		self.canceled = False
		self.ok_btn = ttk.Button(btn_frame, text="OK", command=self.do_select, underline=0)
		self.ok_btn.grid(row=0, column=0, sticky=tk.E, padx=(6,3))
		self.ok_btn["state"] = tk.DISABLED
		self.dlg.bind('<Alt-o>', self.do_select)
		cancel_btn = ttk.Button(btn_frame, text="Cancel", command=self.do_cancel, underline=0)
		cancel_btn.grid(row=0, column=1, sticky=tk.E, padx=(3,6))
		self.dlg.bind("<Escape>", self.do_cancel)
		self.dlg.bind("<Alt-c>", self.do_cancel)
		self.dlg.bind("<Return>", self.do_select)
		self.dlg.bind("<Alt-o>", self.do_select)
	def check_enable(self, *args):
		if self.entry_var.get() != '':
			self.ok_btn["state"] = tk.NORMAL
		else:
			self.ok_btn["state"] = tk.DISABLED
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def do_select(self, *args):
		if self.entry_var.get() != '':
			self.canceled = False
			self.dlg.destroy()
	def show(self):
		self.dlg.grab_set()
		center_window(self.dlg)
		raise_window(self.dlg)
		self.dlg.resizable(True, False)
		self.dlg.attributes('-topmost', 'true')
		self.dlg.wait_window(self.dlg)
		if self.canceled:
			return None
		else:
			return self.entry_var.get()


class OneCheckboxDialog(object):
	def __init__(self, title, prompt, checkbox_value):
		self.dlg = tk.Toplevel()
		self.dlg.title(title)
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.NSEW, pady=(3,3))
		prompt_frame.columnconfigure(0, weight=1)
		self.check_var = tk.StringVar(self.dlg, "0")
		if checkbox_value:
			self.check_var.set("1")
		self.check_ck = ttk.Checkbutton(prompt_frame, text=prompt, state="normal", variable=self.check_var,
				onvalue="1", offvalue="0")
		self.check_ck.grid(row=0, column=0, sticky=tk.W, padx=(6,6), pady=(3,3))
		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.columnconfigure(0, weight=1)
		btn_frame.grid(row=2, column=0, sticky=tk.EW, pady=(3,3))
		btn_frame.columnconfigure(0, weight=1)
		self.check_ck.focus()
		# Buttons
		self.canceled = False
		self.ok_btn = ttk.Button(btn_frame, text="OK", command=self.do_select, underline=0)
		self.ok_btn.grid(row=0, column=0, sticky=tk.E, padx=(6,3))
		self.dlg.bind('<Alt-o>', self.do_select)
		cancel_btn = ttk.Button(btn_frame, text="Cancel", command=self.do_cancel, underline=0)
		cancel_btn.grid(row=0, column=1, sticky=tk.E, padx=(3,6))
		self.dlg.bind("<Escape>", self.do_cancel)
		self.dlg.bind("<Alt-c>", self.do_cancel)
		self.dlg.bind("<Return>", self.do_select)
		self.dlg.bind("<Alt-o>", self.do_select)
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def do_select(self, *args):
		if self.check_var.get() != '':
			self.canceled = False
			self.dlg.destroy()
	def show(self):
		self.dlg.grab_set()
		center_window(self.dlg)
		raise_window(self.dlg)
		self.dlg.resizable(True, False)
		self.dlg.attributes('-topmost', 'true')
		self.dlg.wait_window(self.dlg)
		if self.canceled:
			return None
		else:
			return self.check_var.get()


class OneIntDialog(object):
	def __init__(self, parent, title, prompt, min_value, max_value, initial):
		self.dlg = tk.Toplevel(parent)
		self.dlg.title(title)
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.NSEW, pady=(3,3))
		prompt_frame.columnconfigure(0, weight=1)
		msg_lbl = ttk.Label(prompt_frame, text=prompt)
		msg_lbl.grid(row=0, column=0, sticky=tk.W, padx=(6,6), pady=(6,6))
		self.entry_var = tk.IntVar(self.dlg, initial)
		self.val_entry = ttk.Spinbox(prompt_frame, textvariable=self.entry_var, from_=min_value, to=max_value)
		self.val_entry.grid(row=1, column=0, sticky=tk.EW, padx=(6,6), pady=(3,3))
		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.columnconfigure(0, weight=1)
		btn_frame.grid(row=2, column=0, sticky=tk.EW, pady=(3,3))
		btn_frame.columnconfigure(0, weight=1)
		self.val_entry.focus()
		# Buttons
		self.canceled = False
		self.ok_btn = ttk.Button(btn_frame, text="OK", command=self.do_select, underline=0)
		self.ok_btn.grid(row=0, column=0, sticky=tk.E, padx=(6,3))
		self.dlg.bind('<Alt-o>', self.do_select)
		cancel_btn = ttk.Button(btn_frame, text="Cancel", command=self.do_cancel, underline=0)
		cancel_btn.grid(row=0, column=1, sticky=tk.E, padx=(3,6))
		self.dlg.bind("<Escape>", self.do_cancel)
		self.dlg.bind("<Alt-c>", self.do_cancel)
		self.dlg.bind("<Return>", self.do_select)
		self.dlg.bind("<Alt-o>", self.do_select)
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def do_select(self, *args):
		self.canceled = False
		self.dlg.destroy()
	def show(self):
		self.dlg.grab_set()
		center_window(self.dlg)
		raise_window(self.dlg)
		self.dlg.resizable(True, False)
		self.dlg.attributes('-topmost', 'true')
		self.dlg.wait_window(self.dlg)
		if self.canceled:
			return None
		else:
			return self.entry_var.get()


class OneFloatDialog(object):
	def __init__(self, parent, title, prompt, min_value, max_value, initial):
		self.dlg = tk.Toplevel(parent)
		self.dlg.title(title)
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.NSEW, pady=(3,3))
		prompt_frame.columnconfigure(0, weight=1)
		msg_lbl = ttk.Label(prompt_frame, text=prompt)
		msg_lbl.grid(row=0, column=0, sticky=tk.W, padx=(6,6), pady=(6,6))
		self.entry_var = tk.DoubleVar(self.dlg, initial)
		self.val_entry = ttk.Spinbox(prompt_frame, textvariable=self.entry_var, from_=min_value, to=max_value)
		self.val_entry.grid(row=1, column=0, sticky=tk.EW, padx=(6,6), pady=(3,3))
		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.columnconfigure(0, weight=1)
		btn_frame.grid(row=2, column=0, sticky=tk.EW, pady=(3,3))
		btn_frame.columnconfigure(0, weight=1)
		self.val_entry.focus()
		# Buttons
		self.canceled = False
		self.ok_btn = ttk.Button(btn_frame, text="OK", command=self.do_select, underline=0)
		self.ok_btn.grid(row=0, column=0, sticky=tk.E, padx=(6,3))
		self.dlg.bind('<Alt-o>', self.do_select)
		cancel_btn = ttk.Button(btn_frame, text="Cancel", command=self.do_cancel, underline=0)
		cancel_btn.grid(row=0, column=1, sticky=tk.E, padx=(3,6))
		self.dlg.bind("<Escape>", self.do_cancel)
		self.dlg.bind("<Alt-c>", self.do_cancel)
		self.dlg.bind("<Return>", self.do_select)
		self.dlg.bind("<Alt-o>", self.do_select)
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def do_select(self, *args):
		self.canceled = False
		self.dlg.destroy()
	def show(self):
		self.dlg.grab_set()
		center_window(self.dlg)
		raise_window(self.dlg)
		self.dlg.resizable(True, False)
		self.dlg.attributes('-topmost', 'true')
		self.dlg.wait_window(self.dlg)
		if self.canceled:
			return None
		else:
			return self.entry_var.get()


class PlotConfigDialog(object):
	# Prompts for configuration settings for plot dialogs
	def __init__(self, show_regr_stats, wrap_width, wrap_underscores):
		self.show_regr_stats = show_regr_stats
		self.wrap_width = wrap_width
		self.wrap_underscores = wrap_underscores
		self.dlg = tk.Toplevel()
		self.dlg.title("Plot Configuration")
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.NSEW, pady=(3,3))
		prompt_frame.columnconfigure(0, weight=1)

		self.regr_check_var = tk.StringVar(self.dlg, "0" if not show_regr_stats else "1")
		self.regr_check_ck = ttk.Checkbutton(prompt_frame,
				text="Display regression statistics when drawing scatter plots", state="normal", variable=self.regr_check_var,
				onvalue="1", offvalue="0")
		self.regr_check_ck.grid(row=0, column=0, columnspan=2, sticky=tk.W, padx=(6,6), pady=(3,3))

		wrap_msg_lbl = ttk.Label(prompt_frame, text="Text wrapping width for plot tick labels:")
		wrap_msg_lbl.grid(row=1, column=0, sticky=tk.W, padx=(6,6), pady=(3,3))
		self.wrap_var = tk.IntVar(self.dlg, self.wrap_width)
		self.wrap_entry = ttk.Spinbox(prompt_frame, textvariable=self.wrap_var, width=5, from_=5, to=50)
		self.wrap_entry.grid(row=1, column=1, sticky=tk.W, padx=(6,6), pady=(3,3))

		self.wundr_check_var = tk.StringVar(self.dlg, "0" if not wrap_underscores else "1")
		self.wundr_check_ck = ttk.Checkbutton(prompt_frame,
				text="Wrap labels at underscores", state="normal", variable=self.wundr_check_var,
				onvalue="1", offvalue="0")
		self.wundr_check_ck.grid(row=2, column=0, columnspan=2, sticky=tk.W, padx=(6,6), pady=(3,3))

		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.columnconfigure(0, weight=1)
		btn_frame.grid(row=2, column=0, sticky=tk.EW, pady=(3,3))
		btn_frame.columnconfigure(0, weight=1)
		self.regr_check_ck.focus()
		# Buttons
		self.canceled = False
		self.ok_btn = ttk.Button(btn_frame, text="OK", command=self.do_select, underline=0)
		self.ok_btn.grid(row=0, column=0, sticky=tk.E, padx=(6,3))
		self.dlg.bind('<Alt-o>', self.do_select)
		cancel_btn = ttk.Button(btn_frame, text="Cancel", command=self.do_cancel, underline=0)
		cancel_btn.grid(row=0, column=1, sticky=tk.E, padx=(3,6))
		self.dlg.bind("<Escape>", self.do_cancel)
		self.dlg.bind("<Alt-c>", self.do_cancel)
		self.dlg.bind("<Return>", self.do_select)
		self.dlg.bind("<Alt-o>", self.do_select)
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def do_select(self, *args):
		self.canceled = False
		self.dlg.destroy()
	def show(self):
		self.dlg.grab_set()
		center_window(self.dlg)
		raise_window(self.dlg)
		self.dlg.resizable(True, False)
		self.dlg.attributes('-topmost', 'true')
		self.dlg.wait_window(self.dlg)
		if self.canceled:
			return {"show_regr_stats": self.show_regr_stats, "wrapwidth": self.wrap_width,
					"wrap_underscores": self.wrap_underscores}
		else:
			return {"show_regr_stats": self.regr_check_var.get() == '1', "wrapwidth": self.wrap_var.get(),
					"wrap_underscores": self.wundr_check_var.get()}



class GetEditorDialog(object):
	def __init__(self, parent, current_editor):
		self.dlg = tk.Toplevel(parent)
		self.dlg.title("Set Text Editor")
		self.dlg.columnconfigure(0, weight=1)
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.NSEW, pady=(3,3))
		prompt_frame.columnconfigure(0, weight=1)
		msg_lbl = ttk.Label(prompt_frame, wraplength=600, text="Choose the text editor to be used to edit SQL commands when pulling data from a database.  The editor is set from the EDITOR environment variable on startup, and may also be changed using a configuration file.")
		msg_lbl.grid(row=0, column=0, columnspan=2, sticky=tk.EW, padx=(6,6), pady=(6,3))
		def wrap_msg(event):
			msg_lbl.configure(wraplength=event.width - 5)
		msg_lbl.bind("<Configure>", wrap_msg)
		self.entry_var = tk.StringVar(self.dlg, current_editor or "")
		self.entry_var.trace('w', self.check_enable)
		self.val_entry = ttk.Entry(prompt_frame, width=60, textvariable=self.entry_var)
		self.val_entry.grid(row=1, column=0, sticky=tk.EW, padx=(6,6), pady=(3,6))
		fn_btn = ttk.Button(prompt_frame, text="Browse", command=self.get_fn, underline=0)
		fn_btn.grid(row=1, column=1, sticky=tk.W, padx=(3,3))
		self.dlg.bind("<Alt-b>", self.get_fn)
		# Buttons
		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.columnconfigure(0, weight=1)
		btn_frame.grid(row=2, column=0, sticky=tk.EW, pady=(3,3))
		btn_frame.columnconfigure(0, weight=1)
		self.val_entry.focus()
		self.canceled = False
		self.ok_btn = ttk.Button(btn_frame, text="OK", command=self.do_select, underline=0)
		self.ok_btn.grid(row=0, column=0, sticky=tk.E, padx=(6,3))
		self.ok_btn["state"] = tk.DISABLED
		cancel_btn = ttk.Button(btn_frame, text="Cancel", command=self.do_cancel, underline=0)
		cancel_btn.grid(row=0, column=1, sticky=tk.E, padx=(3,6))
		self.dlg.bind("<Escape>", self.do_cancel)
		self.dlg.bind("<Alt-c>", self.do_cancel)
		self.dlg.bind("<Return>", self.do_select)
		self.dlg.bind('<Alt-o>', self.do_select)
		self.check_enable()
	def get_fn(self, *args):
		fn = tkfiledialog.askopenfilename(parent=self.dlg)
		if fn is not None and fn != '' and fn != ():
			self.entry_var.set(fn)
	def check_enable(self, *args):
		if self.entry_var.get() != '':
			self.ok_btn["state"] = tk.NORMAL
		else:
			self.ok_btn["state"] = tk.DISABLED
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def do_select(self, *args):
		if self.entry_var.get() != '':
			self.canceled = False
			self.dlg.destroy()
	def show(self):
		self.dlg.grab_set()
		center_window(self.dlg)
		raise_window(self.dlg)
		self.dlg.resizable(True, False)
		self.dlg.attributes('-topmost', 'true')
		self.dlg.wait_window(self.dlg)
		if self.canceled:
			return None
		else:
			return self.entry_var.get()


class HelpHotkeysDialog(object):
	def __init__(self):
		self.dlg = tk.Toplevel()
		self.dlg.title("Hotkeys")
		self.dlg.rowconfigure(0, weight=1)
		self.dlg.columnconfigure(0, weight=2)
		keyframe = tk.Frame(self.dlg)
		keyframe.grid(row=0, column=0, sticky=tk.EW, padx=(6,6), pady=(6,6))
		keyframe.rowconfigure(0, weight=1)

		ttk.Label(keyframe, width=10, text="Ctrl-S").grid(row=0, column=0, sticky=tk.N+tk.W)
		ttk.Label(keyframe, width=50, justify=tk.LEFT, wraplength=350, text="Save the data tables shown by the plot dialog and the statistics dialogs.").grid(row=0, column=1)
		ttk.Label(keyframe, width=10, text="Ctrl-Z").grid(row=1, column=0, sticky=tk.N+tk.W)
		ttk.Label(keyframe, width=50, justify=tk.LEFT, wraplength=350, text="Save the table of univariate statistics for log-transformed data.").grid(row=1, column=1)
		ttk.Label(keyframe, width=10, text="Alt-A").grid(row=2, column=0, sticky=tk.N+tk.W)
		ttk.Label(keyframe, width=50, justify=tk.LEFT, wraplength=350, text="Change the opacity (alpha value) of symbols on scatter plots, line plots, stripcharts, and KD plots.").grid(row=2, column=1)
		ttk.Label(keyframe, width=10, text="Alt-B").grid(row=3, column=0, sticky=tk.N+tk.W)
		ttk.Label(keyframe, width=50, justify=tk.LEFT, wraplength=350, text="For histograms: change the number of bins used. For scatter plots: toggle on or off the display of vertical and horizontal lines delineating the Jenks Natural Breaks in X and Y variables, respectively. For line plots: toggle on or off the display of vertical lines delineating the Jenks Natural Breaks in the X variable.").grid(row=3, column=1)
		ttk.Label(keyframe, width=10, text="Alt-G").grid(row=4, column=0, sticky=tk.N+tk.W)
		ttk.Label(keyframe, width=50, justify=tk.LEFT, wraplength=350, text="Toggle the coloring of points on the Normal Q-Q plot to correspond to groups defined by the Jenks Natural Breaks method.").grid(row=4, column=1)
		ttk.Label(keyframe, width=10, text="Alt-L").grid(row=5, column=0, sticky=tk.N+tk.W)
		ttk.Label(keyframe, width=50, justify=tk.LEFT, wraplength=350, text="Toggle the display of a LOESS smoothing line on line plots and scatter plots.").grid(row=5, column=1)
		ttk.Label(keyframe, width=10, text="Alt-Q").grid(row=6, column=0, sticky=tk.N+tk.W)
		ttk.Label(keyframe, width=50, justify=tk.LEFT, wraplength=350, text="Opens the Query Data dialog from the map view.").grid(row=6, column=1)
		ttk.Label(keyframe, width=10, text="Alt-R").grid(row=7, column=0, sticky=tk.N+tk.W)
		ttk.Label(keyframe, width=50, justify=tk.LEFT, wraplength=350, text="For line plots and scatter plots: toggle the display of an ordinary least-squares linear regression line. For box plots, min-max plots, stripcharts and violin plots: rotate the X and Y axes.").grid(row=7, column=1)
		ttk.Label(keyframe, width=10, text="Alt-S").grid(row=8, column=0, sticky=tk.N+tk.W)
		ttk.Label(keyframe, width=50, justify=tk.LEFT, wraplength=350, text="Toggle the display of a Theil-Sen line on line and scatter plots, and the plot for the bivariate statistsics summary.").grid(row=8, column=1)
		ttk.Label(keyframe, width=10, text="Alt-T").grid(row=9, column=0, sticky=tk.N+tk.W)
		ttk.Label(keyframe, width=50, justify=tk.LEFT, wraplength=350, text="Add or change the plot title.").grid(row=9, column=1)
		ttk.Label(keyframe, width=10, text="Alt-X").grid(row=10, column=0, sticky=tk.N+tk.W)
		ttk.Label(keyframe, width=50, justify=tk.LEFT, wraplength=350, text="Change the label on the plot's X axis.").grid(row=10, column=1)
		ttk.Label(keyframe, width=10, text="Alt-Y").grid(row=11, column=0, sticky=tk.N+tk.W)
		ttk.Label(keyframe, width=50, justify=tk.LEFT, wraplength=350, text="Change the label on the plot's Y axis.").grid(row=11, column=1)

		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.columnconfigure(0, weight=2)
		btn_frame.grid(row=1, column=0, sticky=tk.EW, pady=(0,0))
		btn_frame.columnconfigure(0, weight=1)
		# Buttons
		self.canceled = False
		ok_btn = ttk.Button(btn_frame, text="Close", command=self.do_select, underline=0)
		ok_btn.grid(row=0, column=0, sticky=tk.E, padx=(12,6))
		self.dlg.bind("<Alt-c>", self.do_select)
		self.dlg.bind("<Return>", self.do_select)
		self.dlg.bind("<Escape>", self.do_select)
		self.dlg.resizable(False, False)
		#self.dlg.minsize(width=300, height=50)
		ok_btn.focus()
	def do_select(self, *args):
		self.dlg.destroy()
	def show(self, grab=False):
		if grab:
			self.dlg.grab_set()
		center_window(self.dlg)
		raise_window(self.dlg)
		self.dlg.attributes('-topmost', 'true')
		self.dlg.attributes('-topmost', 'false')
		self.dlg.wait_window(self.dlg)


class SelDataSrcDialog(object):
	def __init__(self, parent, mapui):
		self.parent = parent
		self.mapui = mapui
		self.canceled = False
		self.dlg = tk.Toplevel(parent)
		self.dlg.title("Select Mapping Data")
		self.dlg.protocol("WM_DELETE_WINDOW", self.do_cancel)
		self.dlg.columnconfigure(0, weight=1)
		self.rv = (None, None, None, None, None, None, None, None, None, None)
		# Prompt
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.NSEW, padx=(6,6), pady=(6,3))
		msg_lbl = ttk.Label(prompt_frame, width=30, wraplength=100, anchor=tk.W, justify=tk.LEFT, text="Select the type of data source to use.  You will then be prompted for details about the selected data source.")
		msg_lbl.grid(row=0, column=0, padx=(6,6), pady=(3,3))
		def wrap_msg(event):
			msg_lbl.configure(wraplength=event.width - 5)
		msg_lbl.bind("<Configure>", wrap_msg)
		# Select buttons
		sel_frame = tk.Frame(self.dlg)
		sel_frame.grid(row=1, column=0, sticky=tk.NSEW, pady=(6,9))
		csv_btn = ttk.Button(sel_frame, text=" CSV file  ", command=self.sel_csv)
		csv_btn.grid(row=0, column=0, sticky=tk.EW, padx=(3,3), pady=(3,3))
		ss_btn = ttk.Button(sel_frame,  text="Spreadsheet", command=self.sel_spreadsheet)
		ss_btn.grid(row=0, column=1, sticky=tk.EW, padx=(3,3), pady=(3,3))
		db_btn = ttk.Button(sel_frame,  text=" Database  ", command=self.sel_database)
		db_btn.grid(row=0, column=2, sticky=tk.EW, padx=(3,3), pady=(3,3))
		# Help and Cancel buttons
		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.columnconfigure(0, weight=1)
		btn_frame.grid(row=2, column=0, sticky=tk.S+tk.EW, padx=(3,3), pady=(3,3))
		btn_frame.columnconfigure(0, weight=1)
		self.canceled = False
		help_btn = ttk.Button(btn_frame, text="Help", command=self.do_help, underline=0)
		help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		cancel_btn = ttk.Button(btn_frame, text="Cancel", command=self.do_cancel)
		cancel_btn.grid(row=0, column=0, sticky=tk.E, padx=(3,6))
		self.dlg.bind("<Escape>", self.do_cancel)
	def do_help(self, *args):
		webbrowser.open("https://mapdata.readthedocs.io/en/latest/", new=2, autoraise=True)
	def sel_csv(self):
		dfd = DataFileDialog()
		self.rv = dfd.get_datafile()
		if self.rv[0] is not None:
			self.dlg.destroy()
	def sel_spreadsheet(self):
		dfd = ImportSpreadsheetDialog(self.dlg, self.mapui)
		self.rv = dfd.get_datafile()
		if self.rv[0] is not None:
			self.dlg.destroy()
	def sel_database(self):
		dbd = DbConnectDialog(self.dlg, self.mapui)
		self.rv = dbd.get_data()
		if self.rv[0] is not None:
			self.dlg.destroy()
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def select(self):
		self.dlg.grab_set()
		center_window(self.dlg)
		raise_window(self.dlg)
		self.dlg.resizable(False, False)
		self.dlg.focus()
		self.dlg.wait_window(self.dlg)
		return self.rv


class EncodedFile(object):
	# A class providing an open method for an encoded file, allowing reading
	# and writing using unicode, without explicit decoding or encoding.
	def __repr__(self):
		return u"EncodedFile(%r, %r)" % (self.filename, self.encoding)
	def __init__(self, filename, file_encoding):
		self.filename = filename
		self.encoding = file_encoding
		self.bom_length = 0
		def detect_by_bom(path, default_enc):
			with io.open(path, 'rb') as f:
				raw = f.read(4)
			for enc, boms, bom_len in (
							('utf-8-sig', (codecs.BOM_UTF8,), 3),
							('utf_16', (codecs.BOM_UTF16_LE, codecs.BOM_UTF16_BE), 2),
							('utf_32', (codecs.BOM_UTF32_LE, codecs.BOM_UTF32_BE), 4)):
				if any(raw.startswith(bom) for bom in boms):
					return enc, bom_len
			return default_enc, 0
		if os.path.exists(filename):
			self.encoding, self.bom_length = detect_by_bom(filename, file_encoding)
		self.fo = None
	def open(self, mode='r'):
		self.fo = io.open(file=self.filename, mode=mode, encoding="UTF8", newline=None)
		return self.fo
	def close(self):
		if self.fo is not None:
			self.fo.close()


class LineDelimiter(object):
	def __init__(self, delim, quote, escchar):
		self.delimiter = delim
		self.joinchar = delim if delim else u""
		self.quotechar = quote
		if quote:
			if escchar:
				self.quotedquote = escchar+quote
			else:
				self.quotedquote = quote+quote
		else:
			self.quotedquote = None
	def delimited(self, datarow, add_newline=True):
		global conf
		if self.quotechar:
			d_row = []
			for e in datarow:
				if isinstance(e, str):
					if (self.quotechar in e) or (self.delimiter is not None and self.delimiter in e) or (u'\n' in e) or (u'\r' in e):
						d_row.append(u"%s%s%s" % (self.quotechar, e.replace(self.quotechar, self.quotedquote), self.quotechar))
					else:
						d_row.append(e)
				else:
					if e is None:
						d_row.append('')
					else:
						d_row.append(e)
			text = self.joinchar.join([type(u"")(d) for d in d_row])
		else:
			d_row = []
			for e in datarow:
				if e is None:
					d_row.append('')
				else:
					d_row.append(e)
			text = self.joinchar.join([type(u"")(d) for d in d_row])
		if add_newline:
			text = text + u"\n"
		return text


def write_delimited_file(outfile, filefmt, column_headers, rowsource, file_encoding='utf8', append=False):
	delim = None
	quote = None
	escchar = None
	if filefmt.lower() == 'csv':
		delim = ","
		quote = '"'
		escchar = None
	elif filefmt.lower() in ('tab', 'tsv'):
		delim = "\t"
		quote = None
		escchar = None
	elif filefmt.lower() in ('tabq', 'tsvq'):
		delim = "\t"
		quote = '"'
		escchar = None
	elif filefmt.lower() in ('unitsep', 'us'):
		delim = chr(31)
		quote = None
		escchar = None
	elif filefmt.lower() == 'plain':
		delim = " "
		quote = ''
		escchar = None
	elif filefmt.lower() == 'tex':
		delim = "&"
		quote = ''
		escchar = None
	line_delimiter = LineDelimiter(delim, quote, escchar)
	fmode = "w" if not append else "a"
	ofile = EncodedFile(outfile, file_encoding).open(mode=fmode)
	fdesc = outfile
	if not (filefmt.lower() == 'plain' or append):
		datarow = line_delimiter.delimited(column_headers)
		ofile.write(datarow)
	for rec in rowsource:
		datarow = line_delimiter.delimited(rec)
		ofile.write(datarow)
	ofile.close()



class OdsFile(object):
	def __repr__(self):
		return u"OdsFile()"
	def __init__(self):
		self.filename = None
		self.wbk = None
		self.cell_style_names = []
	def open(self, filename):
		self.filename = filename
		if os.path.isfile(filename):
			self.wbk = odf.opendocument.load(filename)
			# Get a list of all cell style names used, so as not to re-define them.
			# Adapted from http://www.pbertrand.eu/reading-an-odf-document-with-odfpy/
			for sty in self.wbk.automaticstyles.childNodes:
				try:
					fam = sty.getAttribute("family")
					if fam == "table-cell":
						name = sty.getAttribute("name")
						if not name in self.cell_style_names:
							self.cell_style_names.append(name)
				except:
					pass
		else:
			self.wbk = odf.opendocument.OpenDocumentSpreadsheet()
	def define_body_style(self):
		st_name = "body"
		if not st_name in self.cell_style_names:
			body_style = odf.style.Style(name=st_name, family="table-cell")
			body_style.addElement(odf.style.TableCellProperties(attributes={"verticalalign":"top"}))
			self.wbk.styles.addElement(body_style)
			self.cell_style_names.append(st_name)
	def define_header_style(self):
		st_name = "header"
		if not st_name in self.cell_style_names:
			header_style = odf.style.Style(name=st_name, family="table-cell")
			header_style.addElement(odf.style.TableCellProperties(attributes={"borderbottom":"1pt solid #000000",
				"verticalalign":"bottom"}))
			self.wbk.styles.addElement(header_style)
			self.cell_style_names.append(st_name)
	def define_iso_datetime_style(self):
		st_name = "iso_datetime"
		if not st_name in self.cell_style_names:
			dt_style = odf.number.DateStyle(name="iso-datetime")
			dt_style.addElement(odf.number.Year(style="long"))
			dt_style.addElement(odf.number.Text(text=u"-"))
			dt_style.addElement(odf.number.Month(style="long"))
			dt_style.addElement(odf.number.Text(text=u"-"))
			dt_style.addElement(odf.number.Day(style="long"))
			# odfpy collapses text elements that have only spaces, so trying to insert just a space between the date
			# and time actually results in no space between them.  Other Unicode invisible characters
			# are also trimmed.  The delimiter "T" is used instead, and conforms to ISO-8601 specifications.
			dt_style.addElement(odf.number.Text(text=u"T"))
			dt_style.addElement(odf.number.Hours(style="long"))
			dt_style.addElement(odf.number.Text(text=u":"))
			dt_style.addElement(odf.number.Minutes(style="long"))
			dt_style.addElement(odf.number.Text(text=u":"))
			dt_style.addElement(odf.number.Seconds(style="long", decimalplaces="3"))
			self.wbk.styles.addElement(dt_style)
			self.define_body_style()
			dts = odf.style.Style(name=st_name, datastylename="iso-datetime", parentstylename="body", family="table-cell")
			self.wbk.automaticstyles.addElement(dts)
			self.cell_style_names.append(st_name)
	def define_iso_date_style(self):
		st_name = "iso_date"
		if st_name not in self.cell_style_names:
			dt_style = odf.number.DateStyle(name="iso-date")
			dt_style.addElement(odf.number.Year(style="long"))
			dt_style.addElement(odf.number.Text(text=u"-"))
			dt_style.addElement(odf.number.Month(style="long"))
			dt_style.addElement(odf.number.Text(text=u"-"))
			dt_style.addElement(odf.number.Day(style="long"))
			self.wbk.styles.addElement(dt_style)
			self.define_body_style()
			dts = odf.style.Style(name=st_name, datastylename="iso-date", parentstylename="body", family="table-cell")
			self.wbk.automaticstyles.addElement(dts)
			self.cell_style_names.append(st_name)
	def sheetnames(self):
		# Returns a list of the worksheet names in the specified ODS spreadsheet.
		return [sheet.getAttribute("name") for sheet in self.wbk.spreadsheet.getElementsByType(odf.table.Table)]
	def sheet_named(self, sheetname):
		# Return the sheet with the matching name.  If the name is actually an integer,
		# return that sheet number.
		if isinstance(sheetname, int):
			sheet_no = sheetname
		else:
			try:
				sheet_no = int(sheetname)
				if sheet_no < 1:
					sheet_no = None
			except:
				sheet_no = None
		if sheet_no is not None:
			for i, sheet in enumerate(self.wbk.spreadsheet.getElementsByType(odf.table.Table)):
				if i+1 == sheet_no:
					return sheet
			else:
				sheet_no = None
		if sheet_no is None:
			for sheet in self.wbk.spreadsheet.getElementsByType(odf.table.Table):
				if sheet.getAttribute("name").lower() == sheetname.lower():
					return sheet
		return None
	def sheet_data(self, sheetname, junk_header_rows=0):
		sheet = self.sheet_named(sheetname)
		if not sheet:
			warning("There is no sheet named %s" % sheetname, kwargs={})
			raise
		def row_data(sheetrow):
			# Adapted from http://www.marco83.com/work/wp-content/uploads/2011/11/odf-to-array.py
			cells = sheetrow.getElementsByType(odf.table.TableCell)
			rowdata = []
			for cell in cells:
				p_content = []
				repeat = cell.getAttribute("numbercolumnsrepeated")
				if not repeat:
					repeat = 1
					spanned = int(cell.getAttribute("numbercolumnsspanned") or 0)
					if spanned > 1:
						repeat = spanned
				ps = cell.getElementsByType(odf.text.P)
				if len(ps) == 0:
					for rr in range(int(repeat)):
						p_content.append(None)
				else:
					for p in ps:
						pval = type(u"")(p)
						if len(pval) == 0:
							for rr in range(int(repeat)):
								p_content.append(None)
						else:
							for rr in range(int(repeat)):
								p_content.append(pval)
				if len(p_content) == 0:
					for rr in range(int(repeat)):
						rowdata.append(None)
				elif p_content[0] != u'#':
					rowdata.extend(p_content)
			return rowdata
		rows = sheet.getElementsByType(odf.table.TableRow)
		if junk_header_rows > 0:
			rows = rows[junk_header_rows: ]
		return [row_data(r) for r in rows]
	def new_sheet(self, sheetname):
		# Returns a sheet (a named Table) that has not yet been added to the workbook
		return odf.table.Table(name=sheetname)
	def add_row_to_sheet(self, datarow, odf_table, header=False):
		if header:
			self.define_header_style()
			style_name = "header"
		else:
			self.define_body_style()
			style_name = "body"
		tr = odf.table.TableRow()
		odf_table.addElement(tr)
		for item in datarow:
			if isinstance(item, bool):
				# Booleans must be evaluated before numbers.
				# Neither of the first two commented-out lines actually work (a bug in odfpy?).
				# Booleans *can* be written as either integers or strings; integers are chosen below.
				#tc = odf.table.TableCell(booleanvalue='true' if item else 'false')
				#tc = odf.table.TableCell(valuetype="boolean", value='true' if item else 'false')
				tc = odf.table.TableCell(valuetype="boolean", value=1 if item else 0, stylename=style_name)
				#tc = odf.table.TableCell(valuetype="string", stringvalue='True' if item else 'False')
			elif isinstance(item, float) or isinstance(item, int):
				tc = odf.table.TableCell(valuetype="float", value=item, stylename=style_name)
			elif isinstance(item, datetime.datetime):
				self.define_iso_datetime_style()
				tc = odf.table.TableCell(valuetype="date", datevalue=item.strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3], stylename="iso_datetime")
			elif isinstance(item, datetime.date):
				self.define_iso_date_style()
				tc = odf.table.TableCell(valuetype="date", datevalue=item.strftime("%Y-%m-%d"), stylename="iso_date")
			elif isinstance(item, datetime.time):
				self.define_iso_datetime_style()
				timeval = datetime.datetime(1899, 12, 30, item.hour, item.minute, item.second, item.microsecond, item.tzinfo)
				tc = odf.table.TableCell(timevalue=timeval.strftime("PT%HH%MM%S.%fS"), stylename="iso_datetime")
				tc.addElement(odf.text.P(text=timeval.strftime("%H:%M:%S.%f")))
			elif isinstance(item, str):
				item = item.replace(u'\n', u' ').replace(u'\r', u' ')
				tc = odf.table.TableCell(valuetype="string", stringvalue=item, stylename=style_name)
			else:
				tc = odf.table.TableCell(value=item, stylename=style_name)
			if item is not None:
				tc.addElement(odf.text.P(text=item))
			tr.addElement(tc)
	def add_sheet(self, odf_table):
		self.wbk.spreadsheet.addElement(odf_table)
	def save_close(self):
		ofile = io.open(self.filename, "wb")
		self.wbk.write(ofile)
		ofile.close()
		self.filename = None
		self.wbk = None
	def close(self):
		self.filename = None
		self.wbk = None


def ods_data(filename, sheetname, junk_header_rows=0):
	# Returns the data from the specified worksheet as a list of headers and a list of lists of rows.
	wbk = OdsFile()
	try:
		wbk.open(filename)
	except:
		warning("%s is not a valid OpenDocument spreadsheet." % filename, kwargs={})
		raise
	try:
		alldata = wbk.sheet_data(sheetname, junk_header_rows)
	except:
		warning("%s is not a worksheet in %s." % (sheetname, filename), kwargs={})
		raise
	colhdrs = alldata[0]
	if any([x is None or len(x.strip())==0 for x in colhdrs]):
		if conf.del_empty_cols:
			blanks = [i for i in range(len(colhdrs)) if colhdrs[i] is None or len(colhdrs[i].strip())==0]
			while len(blanks) > 0:
				b = blanks.pop()
				for r in range(len(alldata)):
					del(alldata[r][b])
			colhdrs = alldata[0]
		else:
			if conf.create_col_hdrs:
				for i in range(len(colhdrs)):
					if colhdrs[i] is None or len(colhdrs[i]) == 0:
						colhdrs[i] = "Col%s" % str(i+1)
			else:
				warning("The input file %s, sheet %s has missing column headers." % (filename, sheetname), kwargs={})
				raise
	#if conf.clean_col_hdrs:
	#	colhdrs = clean_words(colhdrs)
	#if conf.trim_col_hdrs != 'none':
	#	colhdrs = trim_words(colhdrs, conf.trim_col_hdrs)
	#if conf.fold_col_hdrs != 'no':
	#	colhdrs = fold_words(colhdrs, conf.fold_col_hdrs)
	#if conf.dedup_col_hdrs:
	#	colhdrs = dedup_words(colhdrs)
	return colhdrs, alldata[1:]


def export_ods(outfile, hdrs, rows, append=False, querytext=None, sheetname=None, desc=None):
	# If not given, determine the worksheet name to use.  The pattern is "Sheetx", where x is
	# the first integer for which there is not already a sheet name.
	if append and os.path.isfile(outfile):
		wbk = OdsFile()
		wbk.open(outfile)
		sheet_names = wbk.sheetnames()
		name = sheetname or u"Sheet"
		sheet_name = name
		sheet_no = 1
		while True:
			if sheet_name not in sheet_names:
				break
			sheet_no += 1
			sheet_name = u"%s%d" % (name, sheet_no)
		wbk.close()
	else:
		sheet_name = sheetname or u"Sheet1"
		if os.path.isfile(outfile):
			os.unlink(outfile)
	wbk = OdsFile()
	wbk.open(outfile)
	# Add the data to a new sheet.
	tbl = wbk.new_sheet(sheet_name)
	wbk.add_row_to_sheet(hdrs, tbl, header=True)
	for row in rows:
		wbk.add_row_to_sheet(row, tbl)
	# Add sheet to workbook
	wbk.add_sheet(tbl)
	# Save and close the workbook.
	wbk.save_close()


class XlsFile(object):
	def __repr__(self):
		return u"XlsFile()"
	def __init__(self):
		self.filename = None
		self.encoding = None
		self.wbk = None
		self.datemode = 0
	def open(self, filename, encoding=None, read_only=False):
		self.filename = filename
		self.encoding = encoding
		self.read_only = read_only
		self.wbk = xlrd.open_workbook(filename, encoding_override=self.encoding)
		self.datemode = self.wbk.datemode
	def sheetnames(self):
		return self.wbk.sheet_names()
	def sheet_named(self, sheetname):
		# Return the sheet with the matching name.  If the name is actually an integer,
		# return that sheet number.
		if isinstance(sheetname, int):
			sheet_no = sheetname
		else:
			try:
				sheet_no = int(sheetname)
				if sheet_no < 1:
					sheet_no = None
			except:
				sheet_no = None
		if sheet_no is None:
			sheet = self.wbk.sheet_by_name(sheetname)
		else:
			# User-specified sheet numbers should be 1-based; xlrd sheet indexes are 0-based
			sheet = self.wbk.sheet_by_index(max(0, sheet_no-1))
		return sheet
	def sheet_data(self, sheetname, junk_header_rows=0):
		sheet = self.sheet_named(sheetname)
		# Don't rely on sheet.ncols and sheet.nrows, because Excel will count columns
		# and rows that have ever been filled, even if they are now empty.  Base the column count
		# on the number of contiguous non-empty cells in the first row, and process the data up to nrows until
		# a row is entirely empty.
		def row_data(sheetrow, columns=None):
			cells = sheet.row_slice(sheetrow)
			if columns:
				d = [cells[c] for c in range(columns)]
			else:
				d = [cell for cell in cells]
			datarow = []
			for c in d:
				if c.ctype == 0:
					# empty
					datarow.append(None)
				elif c.ctype == 1:
					datarow.append(c.value)
				elif c.ctype == 2:
					# float, but maybe should be int
					if c.value - int(c.value) == 0:
						datarow.append(int(c.value))
					else:
						datarow.append(c.value)
				elif c.ctype == 3:
					# date
					dt = xlrd.xldate_as_tuple(c.value, self.datemode)
					# Convert to time or datetime
					if not any(dt[:3]):
						# No date values
						datarow.append(datetime.time(*dt[3:]))
					else:
						datarow.append(datetime.datetime(*dt))
				elif c.ctype == 4:
					# Boolean
					datarow.append(bool(c.value))
				elif c.ctype == 5:
					# Error code
					datarow.append(xlrd.error_text_from_code(c.value))
				elif c.ctype == 6:
					# blank
					datarow.append(None)
				else:
					datarow.append(c.value)
			return datarow
		hdr_row = row_data(junk_header_rows)
		ncols = 0
		for c in range(len(hdr_row)):
			if not hdr_row[c]:
				break
			ncols += 1
		sheet_data = []
		for r in range(junk_header_rows, sheet.nrows - junk_header_rows):
			datarow = row_data(r, ncols)
			if datarow.count(None) == len(datarow):
				break
			sheet_data.append(datarow)
		return sheet_data


class XlsxFile(object):
	def __repr__(self):
		return u"XlsxFile()"
	def __init__(self):
		self.filename = None
		self.encoding = None
		self.wbk = None
		self.read_only = False
	def open(self, filename, encoding=None, read_only=False):
		self.filename = filename
		self.encoding = encoding
		self.read_only = read_only
		if os.path.isfile(filename):
			if read_only:
				self.wbk = openpyxl.load_workbook(filename, read_only=True)
			else:
				self.wbk = openpyxl.load_workbook(filename)
	def close(self):
		if self.wbk is not None:
			self.wbk.close()
			self.wbk = None
			self.filename = None
			self.encoding = None
	def sheetnames(self):
		return self.wbk.sheetnames
	def sheet_named(self, sheetname):
		# Return the sheet with the matching name.  If the name is actually an integer,
		# return that sheet number.
		if isinstance(sheetname, int):
			sheet_no = sheetname
		else:
			try:
				sheet_no = int(sheetname)
				if sheet_no < 1:
					sheet_no = None
			except:
				sheet_no = None
		if sheet_no is not None:
			# User-specified sheet numbers should be 1-based
			sheet = self.wbk[self.wbk.sheetnames[sheet_no - 1]]
		else:
			sheet = self.wbk[sheetname]
		return sheet
	def sheet_data(self, sheetname, junk_header_rows=0):
		sheet = self.sheet_named(sheetname)
		# Don't rely on sheet.max_column and sheet.max_row, because Excel will count columns
		# and rows that have ever been filled, even if they are now empty.  Base the column count
		# on the number of contiguous non-empty cells in the first row, and process the data up to nrows until
		# a row is entirely empty.
		# Get the header row, skipping junk rows
		rowsrc = sheet.iter_rows(max_row = junk_header_rows + 1, values_only = True)
		for hdr_row in rowsrc:
			pass
		# Get the number of columns
		ncols = 0
		for c in range(len(hdr_row)):
			if not hdr_row[c]:
				break
			ncols += 1
		# Get all the data rows
		sheet_data = []
		rowsrc = sheet.iter_rows(min_row = junk_header_rows + 1, values_only = True)
		for r in rowsrc:
			if not any(r):
				break
			sheet_data.append(list(r))
		for r in range(len(sheet_data)):
			rd = sheet_data[r]
			for c in range(len(rd)):
				if isinstance(rd[c], str):
					if rd[c] == '=FALSE()':
						rd[c] = False
					elif rd[c] == '=TRUE()':
						rd[c] = True
		return sheet_data


def xls_data(filename, sheetname, junk_header_rows, encoding=None):
	# Returns the data from the specified worksheet as a list of headers and a list of lists of rows.
	root, ext = os.path.splitext(filename)
	ext = ext.lower()
	if ext == ".xls":
		wbk = XlsFile()
	else:
		wbk = XlsxFile()
	try:
		wbk.open(filename, encoding, read_only=True)
	except:
		warning("%s is not a valid Excel spreadsheet." % filename, kwargs={})
		raise
	alldata = wbk.sheet_data(sheetname, junk_header_rows)
	if len(alldata) == 0:
		raise ErrInfo(type="cmd", other_msg="There are no data on worksheet %s of file %s." % (sheetname, filename))
	if ext == 'xlsx':
		wbk.close()
	if len(alldata) == 1:
		return alldata[0], []
	colhdrs = alldata[0]
	# Delete columns with missing headers
	if any([x is None or (isinstance(x, str) and len(x.strip())==0) for x in colhdrs]):
		blanks = [i for i in range(len(colhdrs)) if colhdrs[i] is None or len(colhdrs[i].strip())==0]
		while len(blanks) > 0:
			b = blanks.pop()
			for r in range(len(alldata)):
				del(alldata[r][b])
		colhdrs = alldata[0]
	#if conf.clean_col_hdrs:
	#	colhdrs = clean_words(colhdrs)
	#if conf.trim_col_hdrs != 'none':
	#	colhdrs = trim_words(colhdrs, conf.trim_col_hdrs)
	#if conf.fold_col_hdrs != 'no':
	#	colhdrs = fold_words(colhdrs, conf.fold_col_hdrs)
	#if conf.dedup_col_hdrs:
	#	colhdrs = dedup_words(colhdrs)
	return colhdrs, alldata[1:]



def file_data(filename, junk_headers=0):
	# Get headers and rows from the specified CSV file
	csvreader = CsvFile(filename, junk_header_lines=junk_headers)
	headers = csvreader.next()
	rows = []
	for line in csvreader:
		rows.append(line)
	return headers, rows


#***************************************************************************************************
#***************************  SQL Scripting Extensions  ********************************************
#***************************************************************************************************

# Support for SQL scripts used to obtain a data table from a database.
# These are a subset of features in execsql.py.

#===============================================================================================
#-----  GLOBAL VARIABLES FOR SQL INTERPRETYER

# Other variables are defined in the context of further class and function definitions.

# A list of errors found while processing the SQL script.  Each item in this list is
# a two-element list consiting of a) a description of the error, and b) the line number of the error.
script_errors = []

# The last command run.  This should be a ScriptCmd object.
last_command = None

# A compiled regex to match prefixed regular expressions, used to check
# for unsubstituted variables.  This is global rather than local to SqlStmt and
# MetacommandStmt objects because Python 2 can't deepcopy a compiled regex.
varlike = re.compile(r'!![$@&~#]?\w+!!', re.I)

# A ScriptExecSpec object for a script to be executed when the program halts due to an error.
# This is intially None, but may be set and re-set by metacommands.
err_halt_exec = None

# A ScriptExecSpec object for a script to be executed when the program halts due
# user cancellation.
# This is intially None, but may be set and re-set by metacommands.
cancel_halt_exec = None

# A stack of the CommandList objects currently in the queue to be executed.
# The list on the top of the stack is the currently executing script.
commandliststack = []

# A dictionary of CommandList objects (ordinarily created by
# BEGIN/END SCRIPT metacommands) that may be inserted into the
# commandliststack.
savedscripts = {}

# A stack of CommandList objects that are used when compiling the
# statements within a loop (between LOOP and END LOOP metacommands).
loopcommandstack = []
# A global flag to indicate that commands should be compiled into
# the topmost entry in the loopcommandstack rather than executed.
compiling_loop = False
# Compiled regex for END LOOP metacommand, which is immediate.
endloop_rx = re.compile(r'^\s*END\s+LOOP\s*$', re.I)
# Compiled regex for *start of* LOOP metacommand, for testing
# while compiling commands within a loop.
loop_rx = re.compile(r'\s*LOOP\s+', re.I)
# Nesting counter, to ensure loops are only ended when nesting
# level is zero.
loop_nest_level = 0

# A count of all of the commands run.
cmds_run = 0

# Pattern for deferred substitution, e.g.: "!{somevar}!"
defer_rx = re.compile(r'(!{([$@&~#]?[a-z0-9_]+)}!)', re.I)

#	End of global variables (1) for execsql interpreter
#===============================================================================================


#===============================================================================================
#-----  CONFIGURATION

class ConfigData(object):
	def __init__(self):
		self.db_encoding = None
		self.script_encoding = 'utf8'
		self.output_encoding = 'utf8'
		self.import_encoding = 'utf8'
		self.empty_rows = True
		self.del_empty_cols = False
		self.create_col_hdrs = False
		self.trim_col_hdrs = 'none'
		self.clean_col_hdrs = False
		self.fold_col_hdrs = 'no'
		self.dedup_col_hdrs = False
		self.trim_strings = False
		self.replace_newlines = False
		self.export_row_buffer = 1000

#	End of configuration for execsql interpreter
#===============================================================================================
	
#===============================================================================================
#-----  SUPPORT FUNCTIONS AND CLASSES (1)

def ins_rxs(rx_list, fragment1, fragment2):
	# Returns a tuple of all strings consisting of elements of the 'rx_list' tuple
	# inserted between 'fragment1' and 'fragment2'.  The fragments may themselves
	# be tuples.
	if type(fragment1) != tuple:
		fragment1 = (fragment1, )
	if fragment2 is None:
		fragment2 = ('', )
	if type(fragment2) != tuple:
		fragment2 = (fragment2, )
	rv = []
	for te in rx_list:
		for f1 in fragment1:
			for f2 in fragment2:
				rv.append(f1 + te + f2)
	return tuple(rv)

def ins_quoted_rx(fragment1, fragment2, rx):
	return ins_rxs((rx, r'"%s"' % rx), fragment1, fragment2)

def ins_schema_rxs(fragment1, fragment2, suffix=None):
	schema_exprs = (r'"(?P<schema>[A-Za-z0-9_\- ]+)"',
					r'(?P<schema>[A-Za-z0-9_\-]+)',
					r'\[(?P<schema>[A-Za-z0-9_\- ]+)\]'
					)
	if suffix:
		schema_exprs = tuple([s.replace("schema", "schema"+suffix) for s in schema_exprs])
	return ins_rxs(schema_exprs, fragment1, fragment2)

def ins_table_rxs(fragment1, fragment2, suffix=None):
	tbl_exprs = (r'(?:"(?P<schema>[A-Za-z0-9_\- ]+)"\.)?"(?P<table>[A-Za-z0-9_\-\# ]+)"',
					r'(?:(?P<schema>[A-Za-z0-9_\-]+)\.)?(?P<table>[A-Za-z0-9_\-\#]+)',
					r'(?:"(?P<schema>[A-Za-z0-9_\- ]+)"\.)?(?P<table>[A-Za-z0-9_\-\#]+)',
					r'(?:(?P<schema>[A-Za-z0-9_\-]+)\.)?"(?P<table>[A-Za-z0-9_\-\# ]+)"',
					r'(?:\[(?P<schema>[A-Za-z0-9_\- ]+)\]\.)?\[(?P<table>[A-Za-z0-9_\-\# ]+)\]',
					r'(?:(?P<schema>[A-Za-z0-9_\-]+)\.)?(?P<table>[A-Za-z0-9_\-\#]+)',
					r'(?:\[(?P<schema>[A-Za-z0-9_\- ]+)\]\.)?(?P<table>[A-Za-z0-9_\-\#]+)',
					r'(?:(?P<schema>[A-Za-z0-9_\-]+)\.)?\[(?P<table>[A-Za-z0-9_\-\# ]+)\]'
					)
	if suffix:
		tbl_exprs = tuple([s.replace("schema", "schema"+suffix).replace("table", "table"+suffix) for s in tbl_exprs])
	return ins_rxs(tbl_exprs, fragment1, fragment2)

def ins_table_list_rxs(fragment1, fragment2):
	tbl_exprs = (r'(?:(?P<tables>(?:"[A-Za-z0-9_\- ]+"\.)?"[A-Za-z0-9_\-\# ]+"(?:\s*,\s*(?:"[A-Za-z0-9_\- ]+"\.)?"[A-Za-z0-9_\-\# ]+")*))',
				r'(?:(?P<tables>(?:[A-Za-z0-9_\-]+\.)?[A-Za-z0-9_\-\#]+(?:\s*,\s*(?:[A-Za-z0-9_\-]+\.)?[A-Za-z0-9_\-\#]+)*))'
				)
	return ins_rxs(tbl_exprs, fragment1, fragment2)


def ins_fn_rxs(fragment1, fragment2, symbolicname="filename"):
	if os.name == 'posix':
		fns = (r'(?P<%s>[\w\.\-\\\/\'~`!@#$^&()+={}\[\]:;,]*[\w\.\-\\\/\'~`!@#$^&(+={}\[\]:;,])' % symbolicname, r'"(?P<%s>[\w\s\.\-\\\/\'~`!@#$^&()+={}\[\]:;,]+)"' % symbolicname)
	else:
		fns = (r'(?P<%s>([A-Z]\:)?[\w+\,()!@#$^&\+=;\'{}\[\]~`\.\-\\\/]*[\w+\,(!@#$^&\+=;\'{}\[\]~`\.\-\\\/])' % symbolicname, r'"(?P<%s>([A-Z]\:)?[\w+\,()!@#$^&\+=;\'{}\[\]~`\s\.\-\\\/]+)"' % symbolicname)
	return ins_rxs(fns, fragment1, fragment2)


dt_fmts = collections.deque((
			"%x %X",
			"%m/%d/%Y %H:%M",
			"%m/%d/%Y %H%M",
			"%m/%d/%Y %H:%M:%S",
			"%Y-%m-%d %H:%M:%S",
			"%Y-%m-%dT%H:%M:%S",
			"%Y-%m-%d %H%M",
			"%Y-%m-%d %H:%M",
			"%Y-%m-%d %I:%M%p",
			"%Y-%m-%d %I:%M %p",
			"%Y-%m-%d %I:%M:%S%p",
			"%Y-%m-%d %I:%M:%S %p",
			"%m/%d/%Y %I:%M%p",
			"%m/%d/%Y %I:%M %p",
			"%m/%d/%Y %I:%M:%S%p",
			"%m/%d/%Y %I:%M:%S %p",
			"%Y/%m/%d %H%M",
			"%Y/%m/%d %H:%M",
			"%Y/%m/%d %H:%M:%S",
			"%Y/%m/%d %I:%M%p",
			"%Y/%m/%d %I:%M %p",
			"%Y/%m/%d %I:%M:%S%p",
			"%Y/%m/%d %I:%M:%S %p",
			"%Y/%m/%d %X",
			"%c",
			"%b %d, %Y %X",
			"%b %d, %Y %I:%M %p",
			"%b %d %Y %X",
			"%b %d %Y %I:%M %p",
			"%d %b, %Y %X",
			"%d %b, %Y %I:%M %p",
			"%d %b %Y %X",
			"%d %b %Y %I:%M %p",
			"%b. %d, %Y %X",
			"%b. %d, %Y %I:%M %p",
			"%b. %d %Y %X",
			"%b. %d %Y %I:%M %p",
			"%d %b., %Y %X",
			"%d %b., %Y %I:%M %p",
			"%d %b. %Y %X",
			"%d %b. %Y %I:%M %p",
			"%B %d, %Y %X",
			"%B %d, %Y %I:%M %p",
			"%B %d %Y %X",
			"%B %d %Y %I:%M %p",
			"%d %B, %Y %X",
			"%d %B, %Y %I:%M %p",
			"%d %B %Y %X",
			"%d %B %Y %I:%M %p"
			))
def parse_datetime(datestr):
	if type(datestr) == datetime.datetime:
		return datestr
	if not isinstance(datestr, str):
		try:
			datestr = str(datestr)
		except:
			return None
	dt = None
	for i, f in enumerate(dt_fmts):
		try:
			dt = datetime.datetime.strptime(datestr, f)
		except:
			continue
		break
	if i:
		del dt_fmts[i]
		dt_fmts.appendleft(f)
	return dt

dtzrx = re.compile(r"(.+)\s*([+-])(\d{1,2}):?(\d{2})$")
timestamptz_fmts = collections.deque((
	"%Y-%m-%d %H%M%Z", "%Y-%m-%d %H%M %Z",
	"%m/%d/%Y%Z", "%m/%d/%Y %Z",
	"%m/%d/%y%Z", "%m/%d/%y %Z",
	"%m/%d/%Y %H%M%Z", "%m/%d/%Y %H%M %Z",
	"%m/%d/%Y %H:%M%Z", "%m/%d/%Y %H:%M %Z",
	"%Y-%m-%dT%H%M%Z", "%Y-%m-%dT%H%M %Z",
	"%Y-%m-%d %H:%M%Z", "%Y-%m-%d %H:%M %Z",
	"%Y-%m-%dT%H:%M%Z", "%Y-%m-%dT%H:%M %Z",
	"%Y-%m-%d %H:%M:%S%Z", "%Y-%m-%d %H:%M:%S %Z",
	"%Y-%m-%dT%H:%M:%S%Z", "%Y-%m-%dT%H:%M:%S %Z",
	"%Y-%m-%d %I:%M%p%Z", "%Y-%m-%d %I:%M%p %Z",
	"%Y-%m-%dT%I:%M%p%Z", "%Y-%m-%dT%I:%M%p %Z",
	"%Y-%m-%d %I:%M %p%Z", "%Y-%m-%d %I:%M %p %Z",
	"%Y-%m-%dT%I:%M %p%Z", "%Y-%m-%dT%I:%M %p %Z",
	"%Y-%m-%d %I:%M:%S%p%Z", "%Y-%m-%d %I:%M:%S%p %Z",
	"%Y-%m-%dT%I:%M:%S%p%Z", "%Y-%m-%dT%I:%M:%S%p %Z",
	"%Y-%m-%d %I:%M:%S %p%Z", "%Y-%m-%d %I:%M:%S %p %Z",
	"%Y-%m-%dT%I:%M:%S %p%Z", "%Y-%m-%dT%I:%M:%S %p %Z",
	"%c%Z", "%c %Z",
	"%x %X%Z", "%x %X %Z",
	"%m/%d/%Y %H:%M:%S%Z", "%m/%d/%Y %H:%M:%S %Z",
	"%m/%d/%Y %I:%M%p%Z", "%m/%d/%Y %I:%M%p %Z",
	"%m/%d/%Y %I:%M %p%Z", "%m/%d/%Y %I:%M %p %Z",
	"%m/%d/%Y %I:%M:%S%p%Z", "%m/%d/%Y %I:%M:%S%p %Z",
	"%m/%d/%Y %I:%M:%S %p%Z", "%m/%d/%Y %I:%M:%S %p %Z",
	"%Y/%m/%d %H%M%Z", "%Y/%m/%d %H%M %Z",
	"%Y/%m/%d %H:%M%Z", "%Y/%m/%d %H:%M %Z",
	"%Y/%m/%d %H:%M:%S%Z", "%Y/%m/%d %H:%M:%S %Z",
	"%Y/%m/%d %I:%M%p%Z", "%Y/%m/%d %I:%M%p %Z",
	"%Y/%m/%d %I:%M %p%Z", "%Y/%m/%d %I:%M %p %Z",
	"%Y/%m/%d %I:%M:%S%p%Z", "%Y/%m/%d %I:%M:%S%p %Z",
	"%Y/%m/%d %I:%M:%S %p%Z", "%Y/%m/%d %I:%M:%S %p %Z",
	"%Y/%m/%d %X%Z", "%Y/%m/%d %X %Z",
	"%b %d, %Y %X%Z", "%b %d, %Y %X %Z",
	"%b %d, %Y %I:%M %p%Z", "%b %d, %Y %I:%M %p %Z",
	"%b %d %Y %X%Z", "%b %d %Y %X %Z",
	"%b %d %Y %I:%M %p%Z", "%b %d %Y %I:%M %p %Z",
	"%d %b, %Y %X%Z", "%d %b, %Y %X %Z",
	"%d %b, %Y %I:%M %p%Z", "%d %b, %Y %I:%M %p %Z",
	"%d %b %Y %X%Z", "%d %b %Y %X %Z",
	"%d %b %Y %I:%M %p%Z", "%d %b %Y %I:%M %p %Z",
	"%b. %d, %Y %X%Z", "%b. %d, %Y %X %Z",
	"%b. %d, %Y %I:%M %%Z", "%b. %d, %Y %I:%M %p %Z",
	"%b. %d %Y %X%Z", "%b. %d %Y %X %Z",
	"%b. %d %Y %I:%M %p%Z", "%b. %d %Y %I:%M %p %Z",
	"%d %b., %Y %X%Z", "%d %b., %Y %X %Z",
	"%d %b., %Y %I:%M %p%Z", "%d %b., %Y %I:%M %p %Z",
	"%d %b. %Y %X%Z", "%d %b. %Y %X %Z",
	"%d %b. %Y %I:%M %p%Z", "%d %b. %Y %I:%M %p %Z",
	"%B %d, %Y %X%Z", "%B %d, %Y %X %Z",
	"%B %d, %Y %I:%M %p%Z", "%B %d, %Y %I:%M %p %Z",
	"%B %d %Y %X%Z", "%B %d %Y %X %Z",
	"%B %d %Y %I:%M %p%Z", "%B %d %Y %I:%M %p %Z",
	"%d %B, %Y %X%Z", "%d %B, %Y %X %Z",
	"%d %B, %Y %I:%M %p%Z", "%d %B, %Y %I:%M %p %Z",
	"%d %B %Y %X%Z", "%d %B %Y %X %Z",
	"%d %B %Y %I:%M %p%Z", "%d %B %Y %I:%M %p %Z"
	))
def parse_datetimetz(data):
	if type(data) == type(datetime.datetime.now()):
		if data.tzinfo is None or data.tzinfo.utcoffset(data) is None:
			return None
		return data
	if not isinstance(data, str):
		return None
	dt = None
	# Check for numeric timezone
	try:
		datestr, sign, hr, min = dtzrx.match(data).groups()
		dt = parse_datetime(datestr)
		if not dt:
			return None
		sign = -1 if sign=='-' else 1
		return datetime.datetime(dt.year, dt.month, dt.day, dt.hour, dt.minute, dt.second, tzinfo=Tz(sign, int(hr), int(min)))
	except:
		# Check for alphabetic timezone
		for i,f in enumerate(timestamptz_fmts):
			try:
				dt = datetime.datetime.strptime(data, f)
			except:
				continue
			break
		if i:
			del timestamptz_fmts[i]
			timestamptz_fmts.appendleft(f)
		return dt

date_fmts = collections.deque(("%x",
		"%Y-%m-%d",
		"%Y/%m/%d",
		"%m/%d/%Y",
		"%d/%m/%Y",
		"%b %d, %Y",
		"%b %d %Y",
		"%d %b, %Y",
		"%d %b %Y",
		"%b. %d, %Y",
		"%b. %d %Y",
		"%d %b., %Y",
		"%d %b. %Y",
		"%B %d, %Y",
		"%B %d %Y",
		"%d %B, %Y",
		"%d %B %Y"
		))
def parse_date(data):
	if data is None:
		return None
	if isinstance(data, datetime.date):
		return data
	if not isinstance(data, str):
		return None
	for i,f in enumerate(date_fmts):
		try:
			dt = datetime.datetime.strptime(data, f)
			dtt = datetime.date(dt.year, dt.month, dt.day)
		except:
			continue
		break
	else:
		return None
	if i:
		del date_fmts[i]
		date_fmts.appendleft(f)
	return dtt

def parse_boolean(data):
	if data is None:
		return None
	true_strings = ('yes', 'true', '1')
	bool_strings = ('yes', 'no', 'true', 'false', '1', '0')
	if type(data) == bool:
		return data
	elif isinstance(data, int) and data in (0, 1):
		return data == 1
	elif isinstance(data, str) and data.lower() in bool_strings:
		return data.lower() in true_strings
	else:
		return None



#	End of support functions (1)
#===============================================================================================


#===============================================================================================
#-----  STATUS RECORDING

class StatObj(object):
	# A generic object to maintain status indicators.  These status
	# indicators are primarily those used in the metacommand
	# environment rather than for the program as a whole.
	def __init__(self):
		self.halt_on_err = True
		self.sql_error = False
		self.halt_on_metacommand_err = True
		self.metacommand_error = False
		self.cancel_halt = True
		self.batch = BatchLevels()

# End of status recording class.
#===============================================================================================


#===============================================================================================
#-----  ERROR HANDLING

class ErrInfo(Exception):
	def __repr__(self):
		return u"ErrInfo(%r, %r, %r, %r)" % (self.type, self.command, self.exception, self.other)
	def __init__(self, type, command_text=None, exception_msg=None, other_msg=None):
		# Argument 'type' should be "db", "cmd", "log", "error", or "exception".
		# Arguments for each type are as follows:
		# 	"db"		: command_text, exception_msg
		# 	"cmd"	: command_text, <exception_msg | other_msg>
		# 	"log"	: other_msg [, exception_msg]
		# 	"error"	: other_msg [, exception_msg]
		#	"systemexit" : other_msg
		# 	"exception"	: exception_msg [, other_msg]
		self.type = type
		self.command = command_text
		self.exception = None if not exception_msg else exception_msg.replace(u'\n', u'\n     ')
		self.other = None if not other_msg else other_msg.replace(u'\n', u'\n     ')
		if last_command is not None:
			self.script_line_no = current_script_line()
			self.cmd = last_command.command.statement
			self.cmdtype = last_command.command_type
		else:
			self.script_file = None
			self.script_line_no = None
			self.cmd = None
			self.cmdtype = None
		self.error_message = None
		subvars.add_substitution("$ERROR_MESSAGE", self.errmsg())
	def script_info(self):
		if self.script_line_no:
			return u"Line %d of script" % self.script_line_no
		else:
			return None
	def cmd_info(self):
		if self.cmdtype:
			if self.cmdtype == "cmd":
				em = u"Metacommand: %s" % self.cmd
			else:
				em = u"SQL statement: \n         %s" % self.cmd.replace(u'\n', u'\n         ')
			return em
		else:
			return None
	def eval_err(self):
		if self.type == 'db':
			self.error_message = u"**** Error in SQL statement."
		elif self.type == 'cmd':
			self.error_message = u"**** Error in metacommand."
		elif self.type == 'log':
			self.error_message = u"**** Error in logging."
		elif self.type == 'error':
			self.error_message = u"**** General error."
		elif self.type == 'systemexit':
			self.error_message = u"**** Exit."
		elif self.type == 'exception':
			self.error_message = u"**** Exception."
		else:
			self.error_message = u"**** Error of unknown type: %s" % self.type
		sinfo = self.script_info()
		cinfo = self.cmd_info()
		if sinfo:
			self.error_message += u"\n     %s" % sinfo
		if self.exception:
			self.error_message += u"\n     %s" % self.exception
		if self.other:
			self.error_message += u"\n     %s" % self.other
		if self.command:
			self.error_message += u"\n     %s" % self.command
		if cinfo:
			self.error_message += u"\n     %s" % cinfo
		self.error_message += u"\n     Error occurred at %s UTC." % time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime())
		return self.error_message
	def write(self):
		errmsg = self.eval_err()
		output.write_err(errmsg)
		return errmsg
	def errmsg(self):
		return self.eval_err()


def exception_info():
	# Returns the exception type, value, source file name, source line number, and source line text.
	strace = traceback.extract_tb(sys.exc_info()[2])[-1:]
	traces = traceback.extract_tb(sys.exc_info()[2])
	xline = 0
	for trace in traces:
		if u"mapdata" in trace[0]:
			xline = trace[1]
	exc_message = u''
	exc_param = sys.exc_info()[1]
	if isinstance(exc_param, str):
		exc_message = exc_param
	else:
		if hasattr(exc_param, 'message') and isinstance(exc_param.message, str) and len(exc_param.message) > 0:
			exc_message = exc_param.message
		elif hasattr(exc_param, 'value') and isinstance(exc_param.value, str) and len(exc_param.value) > 0:
			exc_message = exc_param.value
		else:
			exc_message = type(u"")(exc_param)
	try:
		exc_message = type(u"")(exc_message)
	except:
		exc_message = repr(exc_message)
	xinfo = sys.exc_info()[0]
	xname = getattr(xinfo, "__name__", "")
	return xname, exc_message, strace[0][0], xline, strace[0][3]


def exception_desc():
	exc_type, exc_strval, exc_filename, exc_lineno, exc_linetext = exception_info()
	return u"%s: %s in %s on line %s of program." % (exc_type, exc_strval, exc_filename, exc_lineno)

#-----  End of ERROR HANDLING
#===============================================================================================



#===============================================================================================
#-----  DATABASE TYPES


class DbTypeError(Exception):
	def __init__(self, dbms_id, data_type, error_msg):
		self.dbms_id = dbms_id
		self.data_type = data_type
		self.error_msg = error_msg or "Unspecified error"
	def __repr__(self):
		return u"DbTypeError(%r, %r)" % (self.dbms_id, self.data_type, self.error_msg)
	def __str__(self):
		if self.data_type:
			return "%s DBMS type error with data type %s: %s" % (self.dbms_id, self.data_type.data_type_name, self.error_msg)
		else:
			return "%s DBMS type error: %s" % (self.dbms_id, self.error_msg)


class DbType(object):
	def __init__(self, DBMS_id, db_obj_quotes=u'""'):
		# The DBMS_id is the name by which this DBMS is identified.
		# db_obj_quotechars is a string of two characters that are the opening and closing quotes
		# for identifiers (schema, table, and column names) that need to be quoted.
		self.dbms_id = DBMS_id
		self.quotechars = db_obj_quotes
		# The dialect is a dictionary of DBMS-specific names for each column type.
		# Dialect keys are DataType classes.
		# Dialect objects are 4-tuples consisting of:
		#	0. a data type name (str)--non-null
		#	1. a Boolean indicating whether or not the length is part of the data type definition
		#		(e.g., for varchar)--non-null
		#	2. a name to use with the 'cast' operator as an alternative to the data type name--nullable.
		#	3. a function to perform a dbms-specific modification of the type conversion result produced
		#		by the 'from_data()' method of the data type.
		#	4. the precision for numeric data types.
		#	5. the scale for numeric data types.
		self.dialect = None
		# The dt_xlate dictionary translates one data type to another.
		# This is specifically needed for Access pre v. 4.0, which has no numeric type, and which
		# therefore requires the numeric data type to be treated as a float data type.
		self.dt_xlate = {}
	def __repr__(self):
		return u"DbType(%r, %r)" % (self.dbms_id, self.quotechars)
	def name_datatype(self, data_type, dbms_name, length_required=False, casting_name=None, conv_mod_fn=None, precision=None, scale=None):
		# data_type is a DataType class object.
		# dbms_name is the DBMS-specific name for this data type.
		# length_required indicates whether length information is required.
		# casting_name is an alternate to the data type name to use in SQL "cast(x as <casting_name>)" expressions.
		# conv_mod_fn is a function that modifies the result of data_type().from_data(x).
		if self.dialect is None:
			self.dialect = {}
		self.dialect[data_type] = (dbms_name, length_required, casting_name, conv_mod_fn, precision, scale)
	def datatype_name(self, data_type):
		# A convenience function to simplify access to data type namess.
		#if not isinstance(data_type, DataType):
		#	raise DbTypeError(self.dbms_id, None, "Unrecognized data type: %s" % data_type)
		try:
			return self.dialect[data_type][0]
		except:
			raise DbTypeError(self.dbms_id, data_type, "%s DBMS type has no specification for data type %s" % (self.dbms_id, data_type.data_type_name))
	def quoted(self, dbms_object):
		if re.search(r'\W', dbms_object):
			if self.quotechars[0] == self.quotechars[1] and self.quotechars[0] in dbms_object:
				dbms_object = dbms_object.replace(self.quotechars[0], self.quotechars[0]+self.quotechars[0])
			return self.quotechars[0] + dbms_object + self.quotechars[1]
		return dbms_object
	def spec_type(self, data_type):
		# Returns a translated data type or the original if there is no translation.
		if data_type in self.dt_xlate:
			return self.dt_xlate[data_type]
		return data_type
	def column_spec(self, column_name, data_type, max_len=None, is_nullable=False, precision=None, scale=None):
		# Returns a column specification as it would be used in a CREATE TABLE statement.
		# The arguments conform to those returned by Column().column_type
		#if not isinstance(data_type, DataType):
		#	raise DbTypeError(self.dbms_id, None, "Unrecognized data type: %s" % data_type)
		data_type = self.spec_type(data_type)
		try:
			dts = self.dialect[data_type]
		except:
			raise DbTypeError(self.dbms_id, data_type, "%s DBMS type has no specification for data type %s" % (self.dbms_id, data_type.data_type_name))
		if max_len and max_len > 0 and dts[1]:
			spec = "%s %s(%d)" % (self.quoted(column_name), dts[0], max_len)
		elif data_type.precspec and precision and scale:
			# numeric
			spec = "%s %s(%s,%s)" % (self.quoted(column_name), dts[0], precision, scale)
		else:
			spec = "%s %s" % (self.quoted(column_name), dts[0])
		if not is_nullable:
			spec += " NOT NULL"
		return spec

# Create a DbType object for each DBMS supported by execsql.

dbt_postgres = DbType("PostgreSQL")
dbt_sqlite = DbType("SQLite")
dbt_duckdb = DbType("DuckDB")
dbt_sqlserver = DbType("SQL Server")
dbt_mysql = DbType("MySQL")
dbt_firebird = DbType("Firebird")
dbt_oracle = DbType("Oracle")


#-----  End of DATABASE TYPES
#===============================================================================================



#===============================================================================================
#-----  DATABASE CONNECTIONS

class DatabaseNotImplementedError(Exception):
	def __init__(self, db_name, method):
		self.db_name = db_name
		self.method = method
	def __repr__(self):
		return u"DatabaseNotImplementedError(%r, %r)" % (self.db_name, self.method)
	def __str__(self):
		return "Method %s is not implemented for database %s" % (self.method, self.db_name)

class Database(object):
	def __init__(self, server_name, db_name, user_name=None, need_passwd=None, port=None, encoding=None):
		self.type = None
		self.server_name = server_name
		self.db_name = db_name
		self.user = user_name
		self.need_passwd = need_passwd
		self.password = None
		self.port = port
		self.encoding = encoding
		self.encode_commands = True
		self.paramstr = '?'
		self.conn = None
		self.autocommit = True
	def __repr__(self):
		return u"Database(%r, %r, %r, %r, %r, %r)" % (self.server_name, self.db_name, self.user,
				self.need_passwd, self.port, self.encoding)
	def name(self):
		if self.server_name:
			return "%s(server %s; database %s)" % (self.type.dbms_id, self.server_name, self.db_name)
		else:
			return "%s(file %s)" % (self.type.dbms_id, self.db_name)
	def open_db(self):
		raise DatabaseNotImplementedError(self.name(), 'open_db')
	def cursor(self):
		if self.conn is None:
			self.open_db()
		return self.conn.cursor()
	def close(self):
		if self.conn:
			self.conn.close()
			self.conn = None
	def paramsubs(self, paramcount):
		return ",".join((self.paramstr,) * paramcount)
	def execute(self, sql, paramlist=None):
		# A shortcut to self.cursor().execute() that handles encoding.
		# Whether or not encoding is needed depends on the DBMS.
		global subvars
		if type(sql) in (tuple, list):
			sql = u" ".join(sql)
		try:
			curs = self.cursor()
			if self.encoding and self.encode_commands and sys.version_info < (3,):
				curs.execute(sql.encode(self.encoding))
			else:
				if paramlist is None:
					curs.execute(sql)
				else:
					curs.execute(sql, paramlist)
			try:
				# DuckDB does not support the 'rowcount' attribute
				subvars.add_substitution("$LAST_ROWCOUNT", curs.rowcount)
			except:
				pass
		except Exception as e:
			try:
				self.rollback()
			except:
				pass
			raise e
	def autocommit_on(self):
		self.autocommit = True
	def autocommit_off(self):
		self.autocommit = False
	def commit(self):
		if self.conn and self.autocommit:
			self.conn.commit()
	def rollback(self):
		if self.conn:
			try:
				self.conn.rollback()
			except:
				pass
	def schema_qualified_table_name(self, schema_name, table_name):
		table_name = self.type.quoted(table_name)
		if schema_name:
			schema_name = self.type.quoted(schema_name)
			return u'%s.%s' % (schema_name, table_name)
		return table_name
	def select_data(self, sql):
		# Returns the results of the sql select statement.
		curs = self.cursor()
		try:
			curs.execute(sql)
		except:
			self.rollback()
			raise
		try:
			subvars.add_substitution("$LAST_ROWCOUNT", curs.rowcount)
		except:
			pass
		rows = curs.fetchall()
		return [d[0] for d in curs.description], rows
	def select_rowsource(self, sql):
		# Return 1) a list of column names, and 2) an iterable that yields rows.
		curs = self.cursor()
		try:
			# DuckDB cursors have no 'arraysize' attribute
			curs.arraysize = conf.export_row_buffer
		except:
			pass
		try:
			curs.execute(sql)
		except:
			self.rollback()
			raise
		try:
			subvars.add_substitution("$LAST_ROWCOUNT", curs.rowcount)
		except:
			pass
		def decode_row():
			while True:
				rows = curs.fetchmany()
				if not rows:
					break
				else:
					for row in rows:
						if self.encoding:
							if sys.version_info < (3,):
								yield [c.decode(self.encoding, "backslashreplace") if type(c) == type("") else c for c in row]
							else:
								yield [c.decode(self.encoding, "backslashreplace") if type(c) == type(b'') else c for c in row]
						else:
							yield row
		return [d[0] for d in curs.description], decode_row()
	def schema_exists(self, schema_name):
		curs = self.cursor()
		curs.execute(u"SELECT schema_name FROM information_schema.schemata WHERE schema_name = '%s';" % schema_name)
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def table_exists(self, table_name, schema_name=None):
		curs = self.cursor()
		sql = "select table_name from information_schema.tables where table_name = '%s'%s;" % (table_name, "" if not schema_name else " and table_schema='%s'" % schema_name)
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(),
								other_msg=u"Failed test for existence of table %s in %s" % (table_name, self.name()))
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def column_exists(self, table_name, column_name, schema_name=None):
		curs = self.cursor()
		sql = "select column_name from information_schema.columns where table_name='%s'%s and column_name='%s';" % (table_name, "" if not schema_name else " and table_schema='%s'" % schema_name, column_name)
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(),
							other_msg=u"Failed test for existence of column %s in table %s of %s" % (column_name, table_name, self.name()))
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def table_columns(self, table_name, schema_name=None):
		curs = self.cursor()
		sql = "select column_name from information_schema.columns where table_name='%s'%s order by ordinal_position;" % (table_name, "" if not schema_name else " and table_schema='%s'" % schema_name)
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(),
							other_msg=u"Failed to get column names for table %s of %s" % (table_name, self.name()))
		rows = curs.fetchall()
		curs.close()
		return [row[0] for row in rows]
	def view_exists(self, view_name, schema_name=None):
		curs = self.cursor()
		sql = "select table_name from information_schema.views where table_name = '%s'%s;" % (view_name, "" if not schema_name else " and table_schema='%s'" % schema_name)
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(),
							other_msg=u"Failed test for existence of view %s in %s" % (view_name, self.name()))
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def role_exists(self, rolename):
		raise DatabaseNotImplementedError(self.name(), 'role_exists')
	def drop_table(self, tablename):
		# The 'tablename' argument should be schema-qualified and quoted as necessary.
		self.execute(u"drop table if exists %s cascade;" % tablename)
		self.commit()


class SqlServerDatabase(Database):
	def __init__(self, server_name, db_name, user_name, need_passwd=False, port=1433, encoding='latin1', password=None):
		global pyodbc
		try:
			import pyodbc
		except:
			fatal_error(u"The pyodbc module is required.  See http://github.com/mkleehammer/pyodbc", kwargs={})
		self.type = dbt_sqlserver
		self.server_name = server_name
		self.db_name = db_name
		self.user = user_name
		self.need_passwd = need_passwd
		self.password = password
		self.port = port if port else 1433
		self.encoding = encoding or 'latin1'    # Default on installation of SQL Server
		self.encode_commands = True
		self.paramstr = '?'
		self.conn = None
		self.autocommit = True
		self.open_db()
	def __repr__(self):
		return u"SqlServerDatabase(%r, %r, %r, %r, %r, %r)" % (self.server_name, self.db_name, self.user,
				self.need_passwd, self.port, self.encoding)
	def open_db(self):
		if self.conn is None:
			if self.user and self.need_passwd and not self.password:
				raise ErrInfo("error", other_msg="Password required but not provided")
			# Use pyodbc to connect.  Try different driver versions from newest to oldest.
			ssdrivers = ('ODBC Driver 17 for SQL Server', 'ODBC Driver 13.1 for SQL Server',
					'ODBC Driver 13 for SQL Server', 'ODBC Driver 11 for SQL Server',
					'SQL Server Native Client 11.0', 'SQL Server Native Client 10.0',
					'SQL Native Client', 'SQL Server')
			for drv in ssdrivers:
				if self.user:
					if self.password:
						connstr = "DRIVER={%s};SERVER=%s;MARS_Connection=Yes; DATABASE=%s;Uid=%s;Pwd=%s" % (drv, self.server_name, self.db_name, self.user, self.password)
					else:
						connstr = "DRIVER={%s};SERVER=%s;MARS_Connection=Yes; DATABASE=%s;Uid=%s" % (drv, self.server_name, self.db_name, self.user)
				else:
					connstr = "DRIVER={%s};SERVER=%s;MARS_Connection=Yes; DATABASE=%s;Trusted_Connection=yes" % (drv, self.server_name, self.db_name)
				try:
					self.conn = pyodbc.connect(connstr)
				except:
					pass
				else:
					break
			if not self.conn:
				raise ErrInfo(type="error", other_msg=u"Can't open SQL Server database %s on %s" % (self.db_name, self.server_name))
			curs = self.conn.cursor()
			curs.execute("SET IMPLICIT_TRANSACTIONS OFF;")
			curs.execute("SET ANSI_NULLS ON;")
			curs.execute("SET ANSI_PADDING ON;")
			curs.execute("SET ANSI_WARNINGS ON;")
			curs.execute("SET QUOTED_IDENTIFIER ON;")
			self.conn.commit()
	def schema_exists(self, schema_name):
		curs = self.cursor()
		curs.execute(u"select * from sys.schemas where name = '%s';" % schema_name)
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def role_exists(self, rolename):
		curs = self.cursor()
		curs.execute(u"select name from sys.database_principals where type in ('R', 'S') and name = '%s';" % rolename)
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def drop_table(self, tablename):
		# SQL Server and Firebird will throw an error if there are foreign keys to the table.
		tablename = self.type.quoted(tablename)
		self.execute(u"drop table %s;" % tablename)


class PostgresDatabase(Database):
	def __init__(self, server_name, db_name, user_name, need_passwd=False, port=5432, new_db=False, encoding='UTF8', password=None):
		global psycopg2
		try:
			import psycopg2
		except:
			fatal_error(u"The psycopg2 module is required to connect to PostgreSQL.", kwargs={})
		self.type = dbt_postgres
		self.server_name = server_name
		self.db_name = db_name
		self.user = user_name
		self.need_passwd = need_passwd
		self.password = password
		self.port = port if port else 5432
		self.new_db = new_db
		self.encoding = encoding or 'UTF8'
		self.encode_commands = False
		self.paramstr = '%s'
		self.conn = None
		self.autocommit = True
		self.open_db()
	def __repr__(self):
		return u"PostgresDatabase(%r, %r, %r, %r, %r, %r, %r)" % (self.server_name, self.db_name, self.user,
				self.need_passwd, self.port, self.new_db, self.encoding)
	def open_db(self):
		def db_conn(db, db_name):
			if db.user and db.password:
				return psycopg2.connect(host=str(db.server_name), database=str(db_name), port=db.port, user=db.user, password=db.password)
			else:
				return psycopg2.connect(host=str(db.server_name), database=db_name, port=db.port)
		def create_db(db):
			conn = db_conn(db, 'postgres')
			conn.autocommit = True
			curs = conn.cursor()
			curs.execute("create database %s encoding '%s';" % (db.db_name, db.encoding))
			conn.close()
		if self.conn is None:
			try:
				if self.user and self.need_passwd and not self.password:
					raise ErrInfo("error", "Password required but not provided")
				if self.new_db:
					create_db(self)
				self.conn = db_conn(self, self.db_name)
			except ErrInfo:
				raise
			except:
				msg = u"Failed to open PostgreSQL database %s on %s" % (self.db_name, self.server_name)
				raise ErrInfo(type="exception", exception_msg=exception_desc(), other_msg=msg)
			# (Re)set the encoding to match the database.
			self.encoding = self.conn.encoding
	def role_exists(self, rolename):
		curs = self.cursor()
		curs.execute(u"select rolname from pg_roles where rolname = '%s';" % rolename)
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def table_exists(self, table_name, schema_name=None):
		curs = self.cursor()
		if schema_name is not None:
			sql = "select table_name from information_schema.tables where table_name = '%s'%s;" % (table_name, "" if not schema_name else " and table_schema='%s'" % schema_name)
		else:
			sql = """select table_name from information_schema.tables where table_name = '%s' and
			         table_schema in (select nspname from pg_namespace where oid = pg_my_temp_schema()
                     union
                     select trim(unnest(string_to_array(replace(setting, '"$user"', CURRENT_USER), ',')))
                     from pg_settings where name = 'search_path');""" % table_name
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(),
								other_msg=u"Failed test for existence of table %s in %s" % (table_name, self.name()))
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def view_exists(self, view_name, schema_name=None):
		curs = self.cursor()
		if schema_name is not None:
			sql = "select table_name from information_schema.views where table_name = '%s'%s;" % (view_name, "" if not schema_name else " and table_schema='%s'" % schema_name)
		else:
			sql = """select table_name from information_schema.views where table_name = '%s' and
			         table_schema in (select nspname from pg_namespace where oid = pg_my_temp_schema()
                     union
                     select trim(unnest(string_to_array(replace(setting, '"$user"', CURRENT_USER), ',')))
                     from pg_settings where name = 'search_path');""" % view_name
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(),
							other_msg=u"Failed test for existence of view %s in %s" % (view_name, self.name()))
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def vacuum(self, argstring):
		self.commit()
		self.conn.set_session(autocommit=True)
		self.conn.cursor().execute("VACUUM %s;" % argstring)
		self.conn.set_session(autocommit=False)


class OracleDatabase(Database):
	def __init__(self, server_name, db_name, user_name, need_passwd=False, port=5432, encoding='UTF8', password=None):
		global cx_Oracle
		try:
			import cx_Oracle
		except:
			fatal_error(u"The cx-Oracle module is required to connect to Oracle.   See https://pypi.org/project/cx-Oracle/", kwargs={})
		self.type = dbt_oracle
		self.server_name = server_name
		self.db_name = db_name
		self.user = user_name
		self.need_passwd = need_passwd
		self.password = password
		self.port = port if port else 1521
		self.encoding = encoding or 'UTF8'
		self.encode_commands = False
		self.paramstr = ':1'
		self.conn = None
		self.autocommit = True
		self.open_db()
	def __repr__(self):
		return u"OracleDatabase(%r, %r, %r, %r, %r, %r)" % (self.server_name, self.db_name, self.user,
				self.need_passwd, self.port, self.encoding)
	def open_db(self):
		def db_conn(db, db_name):
			dsn = cx_Oracle.makedsn(db.server_name, db.port, service_name=db_name)
			if db.user and db.password:
				return cx_Oracle.connect(user=db.user, password=db.password, dsn=dsn)
			else:
				return cx_Oracle.connect(dsn=dsn)
		if self.conn is None:
			try:
				if self.user and self.need_passwd and not self.password:
					raise ErrInfo("error", other_msg="Password required but not provided")
				self.conn = db_conn(self, self.db_name)
			except ErrInfo:
				raise
			except:
				msg = u"Failed to open Oracle database %s on %s" % (self.db_name, self.server_name)
				raise ErrInfo(type="exception", exception_msg=exception_desc(), other_msg=msg)
	def execute(self, sql, paramlist=None):
		# Strip any semicolon off the end and pass to the parent method.
		if sql[-1:] == ";":
			super(OracleDatabase, self).execute(sql[:-1], paramlist)
		else:
			super(OracleDatabase, self).execute(sql, paramlist)
	def select_data(self, sql):
		if sql[-1:] == ";":
			return super(OracleDatabase, self).select_data(sql[:-1])
		else:
			return super(OracleDatabase, self).select_data(sql)
	def select_rowsource(self, sql):
		if sql[-1:] == ";":
			return super(OracleDatabase, self).select_rowsource(sql[:-1])
		else:
			return super(OracleDatabase, self).select_rowsource(sql)
	def schema_exists(self, schema_name):
		raise DatabaseNotImplementedError(self.name(), 'schema_exists')
	def table_exists(self, table_name, schema_name=None):
		curs = self.cursor()
		sql = "select table_name from sys.all_tables where table_name = '%s'%s" % (table_name, "" if not schema_name else " and owner ='%s'" % schema_name)
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(),
								other_msg=u"Failed test for existence of table %s in %s" % (table_name, self.name()))
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def column_exists(self, table_name, column_name, schema_name=None):
		curs = self.cursor()
		sql = "select column_name from all_tab_columns where table_name='%s'%s and column_name='%s'" % (table_name, "" if not schema_name else " and owner ='%s'" % schema_name, column_name)
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(),
							other_msg=u"Failed test for existence of column %s in table %s of %s" % (column_name, table_name, self.name()))
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def table_columns(self, table_name, schema_name=None):
		curs = self.cursor()
		sql = "select column_name from all_tab_columns where table_name='%s'%s order by column_id" % (table_name, "" if not schema_name else " and owner='%s'" % schema_name)
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(),
							other_msg=u"Failed to get column names for table %s of %s" % (table_name, self.name()))
		rows = curs.fetchall()
		curs.close()
		return [row[0] for row in rows]
	def view_exists(self, view_name, schema_name=None):
		curs = self.cursor()
		sql = "select view_name from sys.all_views where view_name = '%s'%s" % (view_name, "" if not schema_name else " and owner ='%s'" % schema_name)
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(),
							other_msg=u"Failed test for existence of view %s in %s" % (view_name, self.name()))
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def role_exists(self, rolename):
		curs = self.cursor()
		curs.execute(u"select role from dba_roles where role = '%s' union " \
				" select username from all_users where username = '%s';" % (rolename, rolename))
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def drop_table(self, tablename):
		tablename = self.type.quoted(tablename)
		self.execute(u"drop table %s cascade constraints" % tablename)
	def paramsubs(self, paramcount):
		return ",".join(":"+str(d) for d in range(1, paramcount+1))


class SQLiteDatabase(Database):
	def __init__(self, SQLite_fn):
		global sqlite3
		self.type = dbt_sqlite
		self.server_name = None
		self.db_name = SQLite_fn
		self.user = None
		self.need_passwd = False
		self.encoding = 'UTF-8'
		self.encode_commands = False
		self.paramstr = '?'
		self.conn = None
		self.autocommit = True
		self.open_db()
	def __repr__(self):
		return u"SQLiteDabase(%r)" % self.db_name
	def open_db(self):
		if self.conn is None:
			try:
				self.conn = sqlite3.connect(self.db_name)
			except ErrInfo:
				raise
			except:
				raise ErrInfo(type="exception", exception_msg=exception_desc(), other_msg=u"Can't open SQLite database %s" % self.db_name)
		pragma_cols, pragma_data = self.select_data("pragma encoding;")
		self.encoding = pragma_data[0][0]
	def table_exists(self, table_name, schema_name=None):
		curs = self.cursor()
		sql = "select name from sqlite_master where type='table' and name='%s';" % table_name
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(), other_msg=u'Failed test for existence of SQLite table "%s";' % table_name)
		rows = curs.fetchall()
		return len(rows) > 0
	def column_exists(self, table_name, column_name, schema_name=None):
		curs = self.cursor()
		sql = "select %s from %s limit 1;" % (column_name, table_name)
		try:
			curs.execute(sql)
		except:
			return False
		return True
	def table_columns(self, table_name, schema_name=None):
		curs = self.cursor()
		sql = "select * from %s where 1=0;" % table_name
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(),
							other_msg=u"Failed to get column names for table %s of %s" % (table_name, self.name()))
		return [d[0] for d in curs.description]
	def view_exists(self, view_name):
		curs = self.cursor()
		sql = "select name from sqlite_master where type='view' and name='%s';" % view_name
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(), other_msg=u'Failed test for existence of SQLite view "%s";' % view_name)
		rows = curs.fetchall()
		return len(rows) > 0
	def schema_exists(self, schema_name):
		return False
	def drop_table(self, tablename):
		tablename = self.type.quoted(tablename)
		self.execute(u"drop table if exists %s;" % tablename)


class DuckDBDatabase(Database):
	def __init__(self, DuckDB_fn):
		global duckdb
		try:
			import duckdb
		except:
			fatal_error(u"The duckdb module is required.", kwargs={})
		self.type = dbt_duckdb
		self.server_name = None
		self.db_name = DuckDB_fn
		self.catalog_name = os.path.splitext(DuckDB_fn)[0]
		self.user = None
		self.need_passwd = False
		self.encoding = 'UTF-8'
		self.encode_commands = False
		self.paramstr = '?'
		self.conn = None
		self.autocommit = True
		self.open_db()
	def __repr__(self):
		return u"DuckDBDabase(%r)" % self.db_name
	def open_db(self):
		if self.conn is None:
			try:
				self.conn = duckdb.connect(self.db_name, read_only=False)
			except ErrInfo:
				raise
			except:
				raise ErrInfo(type="exception", exception_msg=exception_desc(), other_msg=u"Can't open DuckDB database %s" % self.db_name)
	def view_exists(self, view_name):
		# DuckDB information_schema has no 'views' table; views are listed in 'tables'
		return self.table_exists(view_name)
	def schema_exists(self, schema_name):
		# In DuckDB, the 'schemata' view is not limited to the current database.
		curs = self.cursor()
		curs.execute(u"SELECT schema_name FROM information_schema.schemata WHERE schema_name = '%s' and catalog_name = '%s';" % (schema_name, self.catalog_name))
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def execute(self, sql, paramlist=None):
		if type(sql) in (tuple, list):
			sql = u" ".join(sql)
		try:
			curs = self.cursor()
			if paramlist is None:
				curs.execute(sql)
			else:
				curs.execute(sql, paramlist)
			# DuckDB does not support the 'rowcount' attribute, so $LAST_ROWCOUNT is not set
		except Exception as e:
			try:
				self.rollback()
			except:
				raise ErrInfo(type="exception", exception_msg=exception_desc(), other_msg=u"Can't open DuckDB database %s" % self.db_name)


class MySQLDatabase(Database):
	def __init__(self, server_name, db_name, user_name, need_passwd=False, port=3306, encoding='latin1', password=None):
		global mysql_lib
		try:
			import pymysql as mysql_lib
		except:
			fatal_error(u"The pymysql module is required to connect to MySQL.   See https://pypi.python.org/pypi/PyMySQL", kwargs={})
		self.type = dbt_mysql
		self.server_name = str(server_name)
		self.db_name = str(db_name)
		self.user = str(user_name)
		self.need_passwd = need_passwd
		self.password = password
		self.port = 3306 if not port else port
		self.encoding = encoding or 'latin1'
		self.encode_commands = True
		self.paramstr = '%s'
		self.conn = None
		self.autocommit = True
		self.open_db()
	def __repr__(self):
		return u"MySQLDatabase(%r, %r, %r, %r, %r, %r)" % (self.server_name, self.db_name, self.user,
				self.need_passwd, self.port, self.encoding)
	def open_db(self):
		def db_conn():
			if self.user and self.password:
				return mysql_lib.connect(host=self.server_name, database=self.db_name, port=self.port, user=self.user, password=self.password, charset=self.encoding, local_infile=True)
			else:
				return mysql_lib.connect(host=self.server_name, database=self.db_name, port=self.port, charset=self.encoding, local_infile=True)
		if self.conn is None:
			try:
				if self.user and self.need_passwd and not self.password:
					raise ErrInfo("error", other_msg="Password required but not provided")
				self.conn = db_conn()
				self.execute("set session sql_mode='ANSI';")
			except ErrInfo:
				raise
			except:
				msg = u"Failed to open MySQL database %s on %s" % (self.db_name, self.server_name)
				raise ErrInfo(type="exception", exception_msg=exception_desc(), other_msg=msg)
	def schema_exists(self, schema_name):
		return False
	def role_exists(self, rolename):
		curs = self.cursor()
		curs.execute(u"select distinct user as role from mysql.user where user = '%s'" \
				" union select distinct role_name as role from information_schema.applicable_roles" \
				" where role_name = '%s'" % (rolename, rolename))
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0


class FirebirdDatabase(Database):
	def __init__(self, server_name, db_name, user_name, need_passwd=False, port=3050, encoding='latin1', password=None):
		global firebird_lib
		try:
			import fdb as firebird_lib
		except:
			fatal_error(u"The fdb module is required to connect to MySQL.   See https://pypi.python.org/pypi/fdb/", kwargs={})
		self.type = dbt_firebird
		self.server_name = str(server_name)
		self.db_name = str(db_name)
		self.user = str(user_name)
		self.need_passwd = need_passwd
		self.password = password
		self.port = 3050 if not port else port
		self.encoding = encoding or 'latin1'
		self.encode_commands = True
		self.paramstr = '?'
		self.conn = None
		self.autocommit = True
		self.open_db()
	def __repr__(self):
		return u"FirebirdDatabase(%r, %r, %r, %r, %r, %r)" % (self.server_name, self.db_name, self.user,
				self.need_passwd, self.port, self.encoding)
	def open_db(self):
		def db_conn():
			if self.user and self.password:
				return firebird_lib.connect(host=self.server_name, database=self.db_name, port=self.port, user=self.user, password=self.password, charset=self.encoding)
			else:
				return firebird_lib.connect(host=self.server_name, database=self.db_name, port=self.port, charset=self.encoding)
		if self.conn is None:
			try:
				if self.user and self.need_passwd and not self.password:
					raise ErrInfo("error", other_msg="Password required but not provided")
				self.conn = db_conn()
				#self.execute('set autoddl off;')
			except ErrInfo:
				raise
			except:
				msg = u"Failed to open Firebird database %s on %s" % (self.db_name, self.server_name)
				raise ErrInfo(type="exception", exception_msg=exception_desc(), other_msg=msg)
	def table_exists(self, table_name, schema_name=None):
		curs = self.cursor()
		sql = "SELECT RDB$RELATION_NAME FROM RDB$RELATIONS WHERE RDB$SYSTEM_FLAG=0 AND RDB$VIEW_BLR IS NULL AND RDB$RELATION_NAME='%s';" % table_name.upper()
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			e = ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(), other_msg=u"Failed test for existence of Firebird table %s" % table_name)
			try:
				self.rollback()
			except:
				pass
			raise e
		rows = curs.fetchall()
		self.conn.commit()
		curs.close()
		return len(rows) > 0
	def column_exists(self, table_name, column_name, schema_name=None):
		curs = self.cursor()
		sql = "select first 1 %s from %s;" % (column_name, table_name)
		try:
			curs.execute(sql)
		except:
			return False
		return True
	def table_columns(self, table_name, schema_name=None):
		curs = self.cursor()
		sql = "select first 1 * from %s;" % table_name
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(),
							other_msg=u"Failed to get column names for table %s of %s" % (table_name, self.name()))
		return [d[0] for d in curs.description]
	def view_exists(self, view_name, schema_name=None):
		curs = self.cursor()
		sql = "select distinct rdb$view_name from rdb$view_relations where rdb$view_name = '%s';" % view_name
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(), other_msg=u"Failed test for existence of Firebird view %s" % view_name)
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def schema_exists(self, schema_name):
		return False
	def role_exists(self, rolename):
		curs = self.cursor()
		curs.execute(u"SELECT DISTINCT USER FROM RDB$USER_PRIVILEGES WHERE USER = '%s' union " \
				" SELECT DISTINCT RDB$ROLE_NAME FROM RDB$ROLES WHERE RDB$ROLE_NAME = '%s';" % (rolename, rolename))
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def drop_table(self, tablename):
		# Firebird will thrown an error if there are foreign keys into the table.
		tablename = self.type.quoted(tablename)
		self.execute(u"DROP TABLE %s;" % tablename)
		self.conn.commit()


class DatabasePool(object):
	# Define an object that maintains a set of database connection objects, each with
	# a name (alias), and with the current and initial databases identified.
	def __init__(self):
		self.pool = {}
		self.initial_db = None
		self.current_db = None
		self.do_rollback = True
	def __repr__(self):
		return u"DatabasePool()"
	def add(self, db_alias, db_obj):
		db_alias = db_alias.lower()
		if db_alias == 'initial' and len(self.pool) > 0:
			raise ErrInfo(type="error", other_msg="You may not use the name 'INITIAL' as a database alias.")
		if len(self.pool) == 0:
			self.initial_db = db_alias
			self.current_db = db_alias
		if db_alias in self.pool:
			# Don't allow reassignment of a database that is used in any batch.
			if status.batch.uses_db(self.pool[db_alias]):
				raise ErrInfo(type="error", other_msg="You may not reassign the alias of a database that is currently used in a batch.")
			self.pool[db_alias].close()
		self.pool[db_alias] = db_obj
	def aliases(self):
		# Return a list of the currently defined aliases
		return list(self.pool)
	def current(self):
		# Return the current db object.
		return self.pool[self.current_db]
	def current_alias(self):
		# Return the alias of the current db object.
		return self.current_db
	def initial(self):
		return self.pool[self.initial_db]
	def aliased_as(self, db_alias):
		return self.pool[db_alias]
	def make_current(self, db_alias):
		# Change the current database in use.
		db_alias = db_alias.lower()
		if not db_alias in self.pool:
			raise ErrInfo(type="error", other_msg=u"Database alias '%s' is unrecognized; cannnot use it." % db_alias)
		self.current_db = db_alias
	def disconnect(self, alias):
		if alias == self.current_db or (alias == 'initial' and 'initial' in self.pool):
			raise ErrInfo(type="error", other_msg=u"Database alias %s can't be removed or redefined while it is in use." % alias)
		if alias in self.pool:
			self.pool[alias].close()
			del self.pool[alias]
	def closeall(self):
		for alias, db in self.pool.items():
			nm = db.name()
			try:
				if self.do_rollback:
					db.rollback()
				db.close()
			except:
				pass
		self.__init__()

# End of database connections
#===============================================================================================



#===============================================================================================
#-----  SCRIPTING

class BatchLevels(object):
	# A stack to keep a record of the databases used in nested batches.
	class Batch(object):
		def __init__(self):
			self.dbs_used = []
	def __init__(self):
		self.batchlevels = []
	def in_batch(self):
		return len(self.batchlevels) > 0
	def new_batch(self):
		self.batchlevels.append(self.Batch())
	def using_db(self, db):
		if len(self.batchlevels) > 0 and not db in self.batchlevels[-1].dbs_used:
			self.batchlevels[-1].dbs_used.append(db)
	def uses_db(self, db):
		if len(self.batchlevels) == 0:
			return False
		for batch in self.batchlevels:
			if db in batch.dbs_used:
				return True
	def rollback_batch(self):
		if len(self.batchlevels) > 0:
			b = self.batchlevels[-1]
			for db in b.dbs_used:
				db.rollback()
	def end_batch(self):
		b = self.batchlevels.pop()
		for db in b.dbs_used:
			db.commit()

class IfItem(object):
	# An object representing an 'if' level, with context data.
	def __init__(self, tf_value):
		self.tf_value = tf_value
		self.scriptline = current_script_line()
	def value(self):
		return self.tf_value
	def invert(self):
		self.tf_value = not self.tf_value
	def change_to(self, tf_value):
		self.tf_value = tf_value
	def script_line(self):
		return self.scriptline

class IfLevels(object):
	# A stack of True/False values corresponding to a nested set of conditionals,
	# with methods to manipulate and query the set of conditional states.
	# This stack is used by the IF metacommand and related metacommands.
	def __init__(self):
		self.if_levels = []
	def nest(self, tf_value):
		self.if_levels.append(IfItem(tf_value))
	def unnest(self):
		if len(self.if_levels) == 0:
			raise ErrInfo(type="error", other_msg="Can't exit an IF block; no IF block is active.")
		else:
			self.if_levels.pop()
	def invert(self):
		if len(self.if_levels) == 0:
			raise ErrInfo(type="error", other_msg="Can't change the IF state; no IF block is active.")
		else:
			self.if_levels[-1].invert()
	def replace(self, tf_value):
		if len(self.if_levels) == 0:
			raise ErrInfo(type="error", other_msg="Can't change the IF state; no IF block is active.")
		else:
			self.if_levels[-1].change_to(tf_value)
	def current(self):
		if len(self.if_levels) == 0:
			raise ErrInfo(type="error", other_msg="No IF block is active.")
		else:
			return self.if_levels[-1].value()
	def all_true(self):
		if self.if_levels == []:
			return True
		return all([tf.value() for tf in self.if_levels])
	def only_current_false(self):
		# Returns True if the current if level is false and all higher levels are True.
		# Metacommands such as ELSE and ENDIF are executed in this state.
		if len(self.if_levels) == 0:
			return False
		elif len(self.if_levels) == 1:
			return not self.if_levels[-1].value()
		else:
			return not self.if_levels[-1].value() and all([tf.value() for tf in self.if_levels[:-1]])
	def script_lines(self, top_n):
		# Returns a list of tuples containing the script name and line number
		# for the topmost 'top_n' if levels, in bottom-up order.
		if len(self.if_levels) < top_n:
			raise ErrInfo(type="error", other_msg="Invalid IF stack depth reference.")
		levels = self.if_levels[len(self.if_levels) - top_n:]
		return [lvl.script_line() for lvl in levels]

class CounterVars(object):
	# A dictionary of dynamically created named counter variables.  Counter
	# variables are created when first referenced, and automatically increment
	# the integer value returned with each reference.
	_COUNTER_RX = re.compile(r'!!\$(COUNTER_\d+)!!', re.I)
	def __init__(self):
		self.counters = {}
	def _ctrid(self, ctr_no):
		return u'counter_%d' % ctr_no
	def set_counter(self, ctr_no, ctr_val):
		self.counters[self._ctrid(ctr_no)] = ctr_val
	def remove_counter(self, ctr_no):
		ctr_id = self._ctrid(ctr_no)
		if ctr_id in self.counters:
			del self.counters[ctr_id]
	def remove_all_counters(self):
		self.counters = {}
	def substitute(self, command_str):
		# Substitutes any counter variable references with the counter
		# value and returns the modified command string and a flag
		# indicating whether any replacements were made.
		match_found = False
		m = self._COUNTER_RX.search(command_str, re.I)
		if m:
			ctr_id = m.group(1).lower()
			if not ctr_id in self.counters:
				self.counters[ctr_id] = 0
			new_count = self.counters[ctr_id] + 1
			self.counters[ctr_id] = new_count
			return command_str.replace(u'!!$'+m.group(1)+u'!!', str(new_count)), True
		return command_str, False
	def substitute_all(self, any_text):
		subbed = True
		any_subbed = False
		while subbed:
			any_text, subbed = self.substitute(any_text)
			if subbed:
				any_subbed = True
		return any_text, any_subbed

class SubVarSet(object):
	# A pool of substitution variables.  Each variable consists of a name and
	# a (string) value.  All variable names are stored as lowercase text.
	# This is implemented as a list of tuples rather than a dictionary to enforce
	# ordered substitution.
	def __init__(self):
		self.substitutions = []
		#List of acceptable single-character variable name prefixes
		self.prefix_list = ['$','&','@']
		# Regex for matching
		# Don't construct/compile on init because deepcopy() can't handle compiled regexes.
		# 'Regular' variables, dereferenced with "!!"
		self.var_rx = None
	def compile_var_rx(self):
		# Compile regex to validate variable name, using the prefix list
		# This is: any character from the prefix (optionally), followed by one or more word chars
		self.var_rx_str = r'^[' +  "".join(self.prefix_list) + r']?\w+$'
		self.var_rx = re.compile(self.var_rx_str, re.I)
	def var_name_ok(self, varname):
		if self.var_rx is None:
			self.compile_var_rx()
		return self.var_rx.match(varname) is not None
	def check_var_name(self, varname):
		if not self.var_name_ok(varname.lower()):
			raise ErrInfo("error", other_msg="Invalid variable name (%s) in this context." % varname)
	def remove_substitution(self, template_str):
		self.check_var_name(template_str)
		old_sub = template_str.lower()
		self.substitutions = [sub for sub in self.substitutions if sub[0] != old_sub]
	def add_substitution(self, varname, repl_str):
		self.check_var_name(varname)
		varname = varname.lower()
		self.remove_substitution(varname)
		self.substitutions.append((varname, repl_str))
	def append_substitution(self, varname, repl_str):
		self.check_var_name(varname)
		varname = varname.lower()
		oldsub = [x for x in self.substitutions if x[0] == varname]
		if len(oldsub) == 0:
			self.add_substitution(varname, repl_str)
		else:
			self.add_substitution(varname, "%s\n%s" % (oldsub[0][1], repl_str))
	def varvalue(self, varname):
		self.check_var_name(varname)
		vname = varname.lower()
		for vardef in self.substitutions:
			if vardef[0] == vname:
				return vardef[1]
		return None
	def increment_by(self, varname, numeric_increment):
		self.check_var_name(varname)
		varvalue = self.varvalue(varname)
		if varvalue is None:
			varvalue = "0"
			self.add_substitution(varname, varvalue)
		numvalue = as_numeric(varvalue)
		numinc = as_numeric(numeric_increment)
		if numvalue is None or numinc is None:
			newval = "%s+%s" % (varvalue, numeric_increment)
		else:
			newval = str(numvalue + numinc)
		self.add_substitution(varname, newval)
	def sub_exists(self, template_str):
		self.check_var_name(template_str)
		test_str = template_str.lower()
		return test_str in [s[0] for s in self.substitutions]
	def merge(self, other_subvars):
		# Return a new SubVarSet object with this object's variables
		# merged with the 'other_subvars' substitutions; the latter
		# takes precedence.
		# Also merges the prefix lists
		if other_subvars is not None:
			newsubs = SubVarSet()
			newsubs.substitutions = self.substitutions
			newsubs.prefix_list = list(set(self.prefix_list + other_subvars.prefix_list))
			newsubs.compile_var_rx()
			for vardef in other_subvars.substitutions:
				newsubs.add_substitution(vardef[0], vardef[1])
			return newsubs
		return self
	def substitute(self, command_str):
		# Replace any substitution variables in the command string.
		# This does only one round of replacements: if the first round of replacements
		# produces more substitution variables that could be replaced, those derived
		# matching strings are not replaced.  The second value returned by this
		# function indicates whether any substitutions were made, so that this
		# method can be called repeatedly.
		match_found = False
		if isinstance(command_str, str):
			for match, sub in self.substitutions:
				if sub is None:
					sub = ''
				sub = str(sub)
				if match[0] == "$":
					match = "\\"+match
				if os.name != 'posix':
					sub = sub.replace("\\", "\\\\")
				pat = "!!%s!!" % match
				patq = "!'!%s!'!" % match
				patdq = '!"!%s!"!' % match
				if re.search(pat, command_str, re.I):
					return re.sub(pat, sub, command_str, flags=re.I), True
				if re.search(patq, command_str, re.I):
					sub = sub.replace("'", "''")
					return re.sub(patq, sub, command_str, flags=re.I), True
				if re.search(patdq, command_str, re.I):
					sub = '"' + sub + '"'
					return re.sub(patdq, sub, command_str, flags=re.I), True
		return command_str, False
	def substitute_all(self, any_text):
		subbed = True
		any_subbed = False
		while subbed:
			any_text, subbed = self.substitute(any_text)
			if subbed:
				any_subbed = True
		return any_text, any_subbed


class LocalSubVarSet(SubVarSet):
	# A pool of local substitution variables.
	# Inherits everything from the base class except the allowed prefix list.
	# For local variables, only '~' is allowed as a prefix and MUST be present
	def __init__(self):
		SubVarSet.__init__(self)
		self.prefix_list = ['~']
	def compile_var_rx(self):
		# This is different from the base class because the prefix is required, not optional
		self.var_rx_str = r'^[' +  "".join(self.prefix_list) + r']\w+$'
		self.var_rx = re.compile(self.var_rx_str, re.I)


class ScriptArgSubVarSet(SubVarSet):
	# A pool of script argument names.
	# Inherits everything from the base class except the allowed prefix list.
	# For script arguments, only '#' is allowed as a prefix and MUST be present
	def __init__(self):
		SubVarSet.__init__(self)
		self.prefix_list = ['#']
	def compile_var_rx(self):
		# This is different from the base class because the prefix is required, not optional
		self.var_rx_str = r'^[' +  "".join(self.prefix_list) + r']\w+$'
		self.var_rx = re.compile(self.var_rx_str, re.I)


class MetaCommand(object):
	# A compiled metacommand that can be run if it matches a metacommand command string in the input.
	def __init__(self, rx, exec_func, description=None, run_in_batch=False, run_when_false=False, set_error_flag=True):
		# rx: a compiled regular expression
		# exec_func: a function object that carries out the work of the metacommand.
		#			This function must take keyword arguments corresponding to those named
		#			in the regex, and must return a value (which is used only for conditional
		#			metacommands) or None.
		# run_in_batch: determines whether a metacommand should be run inside a batch.  Only 'END BATCH'
		#			should be run inside a batch.
		# run_when_false: determines whether a metacommand should be run when the exec state is False.
		#			only 'ELSE', 'ELSEIF', 'ORIF', and 'ENDIF' should be run when False, and only when
		#			all higher levels are True.  This condition is evaluated by the script processor.
		# set_error_flag: When run, sets or clears status.metacommand_error.
		self.next_node = None
		self.rx = rx
		self.exec_fn = exec_func
		self.description = description
		self.run_in_batch = run_in_batch
		self.run_when_false = run_when_false
		self.set_error_flag = set_error_flag
		self.hitcount = 0
	def __repr__(self):
		return u"MetaCommand(%r, %r, %r, %r, %r)" % (self.rx.pattern, self.exec_fn, self.description,
				self.run_in_batch, self.run_when_false)
	def run(self, cmd_str):
		# Runs the metacommand if the command string matches the regex.
		# Returns a 2-tuple consisting of:
		#	0. True or False indicating whether the metacommand applies.  If False, the
		#		remaining return value is None and has no meaning.
		#	1. The return value of the metacommand function.
		#		Exceptions are caught and converted to ErrInfo exceptions.
		m = self.rx.match(cmd_str.strip())
		if m:
			cmdargs = m.groupdict()
			cmdargs['metacommandline'] = cmd_str
			er = None
			try:
				rv = self.exec_fn(**cmdargs)
			except ErrInfo as errinf:
				# This variable reassignment is required by Python 3;
				# if the line is "except ErrInfo as er:" then an
				# UnboundLocalError occurs at the "if er:" statement.
				er = errinf
			except:
				er = ErrInfo("cmd", command_text=cmd_str, exception_msg=exception_desc())
			if er:
				if status.halt_on_metacommand_err:
					raise er
				if self.set_error_flag:
					status.metacommand_error = True
					return True, None
			else:
				if self.set_error_flag:
					status.metacommand_error = False
				self.hitcount += 1
				return True, rv
		return False, None

class MetaCommandList(object):
	# The head node for a linked list of MetaCommand objects.
	def __init__(self):
		self.next_node = None
	def __iter__(self):
		n1 = self.next_node
		while n1 is not None:
			yield n1
			n1 = n1.next_node
	def insert_node(self, new_node):
		new_node.next_node = self.next_node
		self.next_node = new_node
	def add(self, matching_regexes, exec_func, description=None, run_in_batch=False, run_when_false=False, set_error_flag=True):
		# Creates a new Metacomman and adds it at the head of the linked list.
		if type(matching_regexes) in (tuple, list):
			self.regexes = [re.compile(rx, re.I) for rx in tuple(matching_regexes)]
		else:
			self.regexes = [re.compile(matching_regexes, re.I)]
		for rx in self.regexes:
			self.insert_node(MetaCommand(rx, exec_func, description, run_in_batch, run_when_false, set_error_flag))
	def eval(self, cmd_str):
		# Evaluates the given metacommand string (line from the SQL script).
		# Searches the linked list of MetaCommand objects.  If a match is found, the metacommand
		# is run, and that MetaCommand is moved to the head of the list.
		# Returns a 2-tuple consisting of:
		#	0. True or False indicating whether the metacommand applies.  If False, the
		#		remaining return value is None and has no meaning.
		#	1. The return value of the metacommand function.
		#		Exceptions are caught and converted to ErrInfo exceptions.
		n1 = self
		node_no = 0
		while n1 is not None:
			n2 = n1.next_node
			if n2 is not None:
				node_no += 1
				if if_stack.all_true() or n2.run_when_false:
					success, value = n2.run(cmd_str)
					if success:
						# Move n2 to the head of the list.
						n1.next_node = n2.next_node
						n2.next_node = self.next_node
						self.next_node = n2
						return True, value
			n1 = n2
		return False, None
	def get_match(self, cmd):
		# Tries to match the command 'cmd' to any MetaCommand.  If a match
		# is found, returns a tuple containing the MetaCommand object and
		# the match object; if not, returns None.
		n1 = self.next_node
		while n1 is not None:
			m = n1.rx.match(cmd.strip())
			if m is not None:
				return (n1, m)
			n1 = n1.next_node
		return None


# Global linked lists of MetaCommand objects (commands and conditional tests).
# These are filled in the 'MetaCommand Functions' and 'Conditional Tests for Metacommands' sections.
metacommandlist = MetaCommandList()
conditionallist = MetaCommandList()


class SqlStmt(object):
	# A SQL statement to be passed to a database to execute.
	# The purpose of storing a SQL statement as a SqlStmt object rather
	# than as a simple string is to allow the definition of a 'run()'
	# method that is different from the 'run()' method of a MetacommandStmt.
	# In effect, the SqlStmt and MetacommandStmt classes are both
	# subclasses of a Stmt class, but the Stmt class, and subclassing,
	# are not implemented because the Stmt class would be trivial: just
	# an assignment in the init method.
	def __init__(self, sql_statement):
		self.statement = re.sub(r'\s*;(\s*;\s*)+$', ';', sql_statement)
	def __repr__(self):
		return u"SqlStmt(%s)" % self.statement
	def run(self, localvars=None, commit=True):
		# Run the SQL statement on the current database.  The current database
		# is obtained from the global database pool "dbs".
		# 'localvars' must be a SubVarSet object.
		if if_stack.all_true():
			e = None
			status.sql_error = False
			cmd = substitute_vars(self.statement, localvars)
			if varlike.search(cmd):
				lno = current_script_line()
				script_errors.append(["There is a potential un-substituted variable in the command %s" % cmd, lno])
			try:
				db = dbs.current()
				db.execute(cmd)
				if commit:
					db.commit()
			except ErrInfo as errinfo:
				# This variable reassignment is required by Python 3;
				# if the line is "except ErrInfo as e:" then an
				# UnboundLocalError occurs at the "if e:" statement.
				e = errinfo
			except:
				e = ErrInfo(type="exception", exception_msg=exception_desc())
			if e:
				subvars.add_substitution("$LAST_ERROR", cmd)
				status.sql_error = True
				if status.halt_on_err:
					raise e
				return
			subvars.add_substitution("$LAST_SQL", cmd)
	def commandline(self):
		return self.statement


class MetacommandStmt(object):
	# A metacommand to be handled by execsql.
	def __init__(self, metacommand_statement):
		self.statement = metacommand_statement
	def __repr__(self):
		return u"MetacommandStmt(%s)" % self.statement
	def run(self, localvars=None, commit=False):
		# Tries all metacommands in "metacommandlist" until one runs.
		# Returns the result of the metacommand that was run, or None.
		# Arguments:
		#	localvars: a SubVarSet object.
		#	commit   : not used; included to allow an isomorphic interface with SqlStmt.run().
		errmsg = "Unknown metacommand"
		cmd = substitute_vars(self.statement, localvars)
		if if_stack.all_true() and varlike.search(cmd):
			lno = current_script_line()
			script_errors.append(["There is a potential un-substituted variable in the command %s" % cmd, lno])
		e = None
		try:
			applies, result = metacommandlist.eval(cmd)
			if applies:
				return result
		except ErrInfo as errinfo:
			# This variable reassignment is required by Python 3;
			# if the line is "except ErrInfo as e:" then an
			# UnboundLocalError occurs at the "if e:" statement.
			e = errinfo
		except:
			e = ErrInfo(type="exception", exception_msg=exception_desc())
		if e:
			status.metacommand_error = True
			subvars.add_substitution("$LAST_ERROR", cmd)
			if status.halt_on_metacommand_err:
				raise e
				#raise ErrInfo(type="cmd", command_text=cmd, other_msg=errmsg)
		if if_stack.all_true():
			# but nothing applies, because we got here.
			status.metacommand_error = True
			lno = current_script_line()
			script_errors.append(["%s:  %s" % (errmsg, cmd), lno])
			#raise ErrInfo(type="cmd", command_text=cmd, other_msg=errmsg)
		return None
	def commandline(self):
		# Returns the SQL or metacommand as in a script
		return  u"-- !x! " + self.statement


class ScriptCmd(object):
	# A SQL script object that is either a SQL statement (SqlStmt object)
	# or an execsql metacommand (MetacommandStmt object).
	# This is the basic uniform internal representation of a single
	# command or statement from an execsql script file.
	# The object attributes include source file information.
	# 'command_type' is "sql" or "cmd".
	# 'script_command' is either a SqlStmt or a MetacommandStmt object.
	def __init__(self, command_source_name, command_line_no, command_type, script_command):
		self.source = command_source_name
		self.line_no = command_line_no
		self.command_type = command_type
		self.command = script_command
	def __repr__(self):
		return u"ScriptCmd(%r, %r, %r, %r)" % (self.source, self.line_no, self.command_type, repr(self.command))
	def current_script_line(self):
		return self.line_no
	def commandline(self):
		# Returns the SQL or metacommand as in a script
		return self.command.statement if self.command_type == "sql" else u"-- !x! " + self.command.statement


class CommandList(object):
	# A list of ScriptCmd objects, including an index into the list, an
	# optional list of parameter names, and an optional set of parameter
	# values (SubvarSet).  This is the basic internal representation of
	# a list of interspersed SQL commands and metacommands.
	def __init__(self, cmdlist, listname, paramnames=None):
		# Arguments:
		#    cmdlist    : A Python list of ScriptCmd objects.  May be an empty list.
		#    listname   : A string to identify the list (e.g., a source file name or SCRIPT name).
		#    paramnames : A list of strings identifying parameters the script expects.
		# Parameter names will be used to check the names of actual arguments
		# if they are specified, but are optional: a sub-script may take
		# arguments even if parameter names have not been specified.
		if cmdlist is None:
			raise ErrInfo("error", other_msg="Initiating a command list without any commands.")
		self.listname = listname
		self.cmdlist = cmdlist
		self.cmdptr = 0
		self.paramnames = paramnames
		self.paramvals = None
		# Local variables must start with a tilde.  Other types are not allowed.
		self.localvars = LocalSubVarSet()
		self.init_if_level = None
	def add(self, script_command):
		# Adds the given ScriptCmd object to the end of the command list.
		self.cmdlist.append(script_command)
	def set_paramvals(self, paramvals):
		# Parameter values should ordinarily set immediately before the script
		# (command list) is run.
		# Arguments:
		#    paramvals : A SubVarSet object.
		self.paramvals = paramvals
		if self.paramnames is not None:
			# Check that all named parameters are provided.
			# Strip '#' off passed parameter names
			passed_paramnames = [p[0][1:] if p[0][0]=='#' else p[0][1:] for p in paramvals.substitutions]
			if not all([p in passed_paramnames for p in self.paramnames]):
				raise ErrInfo("error", other_msg="Formal and actual parameter name mismatch in call to %s." % self.listname)
	def current_command(self):
		if self.cmdptr > len(self.cmdlist) - 1:
			return None
		return self.cmdlist[self.cmdptr]
	def check_iflevels(self):
		if_excess = len(if_stack.if_levels) - self.init_if_level
		if if_excess > 0:
			sources = if_stack.script_lines(if_excess)
			src_msg = ", ".join(["input line %s" % src for src in sources])
			raise ErrInfo(type="error", other_msg="IF level mismatch at beginning and end of script; origin at or after: %s." % src_msg)
	def run_and_increment(self):
		global last_command
		global loop_nest_level
		cmditem = self.cmdlist[self.cmdptr]
		if compiling_loop:
			# Don't run this command, but save it or complete the loop and add the loop's set of commands to the stack.
			if cmditem.command_type == 'cmd' and loop_rx.match(cmditem.command.statement):
				loop_nest_level += 1
    			# Substitute any deferred substitution variables with regular substition var flags, e.g.: "!!somevar!!"
				m = defer_rx.findall(cmditem.command.statement)
				if m is not None:
					for dv in m:
						rep = "!!" +  dv[1] + "!!"
						cmditem.command.statement = cmditem.command.statement.replace(dv[0], rep)
				loopcommandstack[-1].add(cmditem)
			elif cmditem.command_type == 'cmd' and endloop_rx.match(cmditem.command.statement):
				if loop_nest_level == 0:
					endloop()
				else:
					loop_nest_level -= 1
					loopcommandstack[-1].add(cmditem)
			else:
				loopcommandstack[-1].add(cmditem)
		else:
			last_command = cmditem
			if cmditem.command_type == "sql" and status.batch.in_batch():
				status.batch.using_db(dbs.current())
			subvars.add_substitution("$CURRENT_SCRIPT", cmditem.source)
			subvars.add_substitution("$CURRENT_SCRIPT_PATH", os.path.dirname(os.path.abspath(cmditem.source)) + os.sep)
			subvars.add_substitution("$CURRENT_SCRIPT_NAME", os.path.basename(cmditem.source))
			subvars.add_substitution("$CURRENT_SCRIPT_LINE", str(cmditem.line_no))
			subvars.add_substitution("$SCRIPT_LINE", str(cmditem.line_no))
			cmditem.command.run(self.localvars.merge(self.paramvals), not status.batch.in_batch())
		self.cmdptr += 1
	def run_next(self):
		global last_command
		if self.cmdptr == 0:
			self.init_if_level = len(if_stack.if_levels)
		if self.cmdptr > len(self.cmdlist) - 1:
			self.check_iflevels()
			raise StopIteration
		self.run_and_increment()
	def __iter__(self):
		return self
	def __next__(self):
		if self.cmdptr > len(self.cmdlist) - 1:
			raise StopIteration
		scriptcmd = self.cmdlist[self.cmdptr]
		self.cmdptr += 1
		return scriptcmd


class CommandListWhileLoop(CommandList):
	# Subclass of CommandList() that will loop WHILE a condition is met.
	# Additional argument:
	#	loopcondition : A string containing the conditional for continuing the WHILE loop.
	def __init__(self, cmdlist, listname, paramnames, loopcondition):
		super(CommandListWhileLoop, self).__init__(cmdlist, listname, paramnames)
		self.loopcondition = loopcondition
	def run_next(self):
		global last_command
		if self.cmdptr == 0:
			self.init_if_level = len(if_stack.if_levels)
			if not CondParser(substitute_vars(self.loopcondition)).parse().eval():
				raise StopIteration
		if self.cmdptr > len(self.cmdlist) - 1:
			self.check_iflevels()
			self.cmdptr = 0
		else:
			self.run_and_increment()


class CommandListUntilLoop(CommandList):
	# Subclass of CommandList() that will loop UNTIL a condition is met.
	# Additional argument:
	#    loopcondition : A string containing the conditional for terminating the UNTIL loop.
	def __init__(self, cmdlist, listname, paramnames, loopcondition):
		super(CommandListUntilLoop, self).__init__(cmdlist, listname, paramnames)
		self.loopcondition = loopcondition
	def run_next(self):
		global last_command
		if self.cmdptr == 0:
			self.init_if_level = len(if_stack.if_levels)
		if self.cmdptr > len(self.cmdlist) - 1:
			self.check_iflevels()
			if CondParser(substitute_vars(self.loopcondition)).parse().eval():
				raise StopIteration
			self.cmdptr = 0
		else:
			self.run_and_increment()

class ScriptFile(EncodedFile):
	# A file reader that returns lines and records the line number.
	def __init__(self, scriptfname, file_encoding):
		super(ScriptFile, self).__init__(scriptfname, file_encoding)
		self.lno = 0
		self.f = self.open("r")
	def __repr__(self):
		return u"ScriptFile(%r, %r)" % (super(ScriptFile, self).filename, super(ScriptFile, self).encoding)
	def __iter__(self):
		return self
	def __next__(self):
		l = next(self.f)
		self.lno += 1
		return l


class ScriptExecSpec(object):
	# An object that stores the specifications for executing a SCRIPT,
	# for later use.  This is specifically intended to be used by
	# ON ERROR_HALT EXECUTE SCRIPT and ON CANCEL_HALT EXECUTE SCRIPT.
	args_rx = re.compile(r'(?P<param>#?\w+)\s*=\s*(?P<arg>(?:(?:[^"\'\[][^,\)]*)|(?:"[^"]*")|(?:\'[^\']*\')|(?:\[[^\]]*\])))', re.I)
	def __init__(self, **kwargs):
		self.script_id = kwargs["script_id"].lower()
		if self.script_id not in savedscripts.keys():
			raise ErrInfo("cmd", other_msg="There is no SCRIPT named %s." % self.script_id)
		self.arg_exp = kwargs["argexp"]
		self.looptype = kwargs["looptype"].upper() if "looptype" in kwargs and kwargs["looptype"] is not None else None
		self.loopcond = kwargs["loopcond"] if "loopcond" in kwargs else None
	def execute(self):
		# Copy the saved script because otherwise the memory-recovery nullification
		# of completed commands will erase the saved script commands.
		cl = copy.deepcopy(savedscripts[self.script_id])
		# If looping is specified, redirect to appropriate CommandList() subclass 
		if self.looptype is not None:
			cl = CommandListWhileLoop(cl.cmdlist, cl.listname, cl.paramnames, self.loopcond) if self.looptype == 'WHILE' else CommandListUntilLoop(cl.cmdlist, cl.listname, cl.paramnames, self.loopcond)
		# If there are any argument expressions, parse the arguments
		if self.arg_exp is not None:
			# Clean arg_exp
			all_args = re.findall(self.args_rx, self.arg_exp)
			all_cleaned_args = [(ae[0], wo_quotes(ae[1])) for ae in all_args]
			# Prepend '#' on each param name if the user did not include one
			all_prepared_args = [(ae[0] if ae[0][0]=='#' else '#' + ae[0], ae[1]) for ae in all_cleaned_args]
			scriptvarset = ScriptArgSubVarSet()
			for param, arg in all_prepared_args:
				scriptvarset.add_substitution(param, arg)
			cl.set_paramvals(scriptvarset)
		# If argument expressions were NOT found, confirm that the command list is not expecting named params
		else:
			# because if it IS, there's a problem.
			if cl.paramnames is not None:
				raise ErrInfo("error", other_msg="Missing expected parameters (%s) in call to %s." % (", ".join(cl.paramnames), cl.listname))
		commandliststack.append(cl)


# End of scripting classes.
#===============================================================================================



#===============================================================================================
#----- Parsers
#
# Parsers for conditional and numeric expressions.

#-------------------------------------------------------------------------------------
# Source string objects.  These are strings (metacommands arguments) with
# a pointer into the string.
#-------------------------------------------------------------------------------------
class SourceString(object):
	def __init__(self, source_string):
		self.str = source_string
		self.currpos = 0
	def eoi(self):
		# Returns True or False indicating whether or not there is any of
		# the source string left to be consumed.
		return self.currpos >= len(self.str)
	def eat_whitespace(self):
		while not self.eoi() and self.str[self.currpos] in [' ', '\t', '\n']:
			self.currpos += 1
	def match_str(self, str):
		# Tries to match the 'str' argument at the current position in the
		# source string.  Matching is case-insensitive.  If matching succeeds,
		# the matched string is returned and the internal pointer is incremented.
		# If matching fails, None is returned and the internal pointer is unchanged.
		self.eat_whitespace()
		if self.eoi():
			return None
		else:
			found = self.str.lower().startswith(str.lower(), self.currpos)
			if found:
				matched = self.str[self.currpos:self.currpos+len(str)]
				self.currpos += len(str)
				return matched
			else:
				return None
	def match_regex(self, regex):
		# Tries to match the 'regex' argument at the current position in the
		# source string.  If it succeeds, a dictionary of all of the named
		# groups is returned, and the internal pointer is incremented.
		self.eat_whitespace()
		if self.eoi():
			return None
		else:
			m = regex.match(self.str[self.currpos:])
			if m:
				self.currpos += m.end(0)
				return m.groupdict() or {}
			else:
				return None
	def match_metacommand(self, commandlist):
		# Tries to match text at the current position to any metacommand
		# in the specified commandlist. 
		# If it succeeds, the return value is a tuple of the MetaCommand object
		# and a dictionary of all of the named groups.  The internal pointer is
		# incremented past the match.
		self.eat_whitespace()
		if self.eoi():
			return None
		else:
			m = commandlist.get_match(self.str[self.currpos:])
			if m is not None:
				self.currpos += m[1].end(0)
				return (m[0], m[1].groupdict() or {})
			else:
				return None
	def remainder(self):
		return self.str[self.currpos:]
 
#-------------------------------------------------------------------------------------
#	Classes for AST operator types.
#-------------------------------------------------------------------------------------
class CondTokens(object):
	AND, OR, NOT, CONDITIONAL = range(4)

class NumTokens(object):
	MUL, DIV, ADD, SUB, NUMBER = range(5)

#-------------------------------------------------------------------------------------
#	AST for conditional expressions
#-------------------------------------------------------------------------------------
class CondAstNode(CondTokens, object):
	def __init__(self, type, cond1, cond2):
		# 'type' should be one of the constants AND, OR, NOT, CONDITIONAL.
		# For AND and OR types, 'cond1' and 'cond2' should be a subtree (a CondAstNode)
		# For NOT type, 'cond1' should be a CondAstNOde and 'cond2' should be None
		# For CONDITIONAL type, cond1' should be a tuple consisting of metacommand object and
		# its dictionary of named groups (mcmd, groupdict) and 'cond2' should be None.
		self.type = type
		self.left = cond1
		if type not in (self.CONDITIONAL, self.NOT):
			self.right = cond2
		else:
			self.right = None
	def eval(self):
		# Evaluates the subtrees and/or conditional value for this node,
		# returning True or False.
		if self.type == self.CONDITIONAL:
			exec_fn = self.left[0].exec_fn
			cmdargs = self.left[1]
			return exec_fn(**cmdargs)
		if self.type == self.NOT:
			return not self.left.eval()
		lcond = self.left.eval()
		if self.type == self.AND:
			if not lcond: return False
			return self.right.eval()
		if self.type == self.OR:
			if lcond: return True
			return self.right.eval()

#-------------------------------------------------------------------------------------
#	AST for numeric expressions
#-------------------------------------------------------------------------------------
class NumericAstNode(NumTokens, object):
	def __init__(self, type, value1, value2):
		# 'type' should be one of the constants MUL, DIV, ADD, SUB, OR NUMBER.
		# 'value1' and 'value2' should each be either a subtree (a
		# NumericAstNode) or (only 'value1' should be) a number.
		self.type = type
		self.left = value1
		if type != self.NUMBER:
			self.right = value2
		else:
			self.right = None
	def eval(self):
		# Evaluates the subtrees and/or numeric value for this node,
		# returning a numeric value.
		if self.type == self.NUMBER:
			return self.left
		else:
			lnum = self.left.eval()
			rnum = self.right.eval()
			if self.type == self.MUL:
				return lnum * rnum
			elif self.type == self.DIV:
				return lnum / rnum
			elif self.type == self.ADD:
				return lnum + rnum
			else:
				return lnum - rnum

#-------------------------------------------------------------------------------------
#	Conditional Parser
#-------------------------------------------------------------------------------------
class CondParserError(Exception):
	def __init__(self, msg):
		self.value = msg
	def __repr__(self):
		return "ConditionalParserError(%r)" % self.value


class CondParser(CondTokens, object):
	# Takes a conditional expression string.
	def __init__(self, condexpr):
		self.condexpr = condexpr
		self.cond_expr = SourceString(condexpr)
	def match_not(self):
		# Try to match 'NOT' operator. If not found, return None
		m1 = self.cond_expr.match_str('NOT')
		if m1 is not None:
			return self.NOT
		return None
	def match_andop(self):
		# Try to match 'AND' operator. If not found, return None
		m1 = self.cond_expr.match_str('AND')
		if m1 is not None:
			return self.AND
		return None
	def match_orop(self):
		# Try to match 'OR' operator. If not found, return None
		m1 = self.cond_expr.match_str('OR')
		if m1 is not None:
			return self.OR
		return None
	def factor(self): 
		m1 = self.match_not()
		if m1 is not None:
			m1 = self.factor()
			return CondAstNode(self.NOT, m1, None)
		# Find the matching metacommand -- get a tuple consisting of (metacommand, groupdict)
		m1 = self.cond_expr.match_metacommand(conditionallist)
		if m1 is not None:
			m1[1]["metacommandline"] = self.condexpr
			return CondAstNode(self.CONDITIONAL, m1, None)
		else:
			if self.cond_expr.match_str("(") is not None:
				m1 = self.expression()
				rp = self.cond_expr.match_str(")")
				if rp is None:
					raise CondParserError("Expected closing parenthesis at position %s of %s." % (self.cond_expr.currpos, self.cond_expr.str))
				return m1
			else:
				raise CondParserError("Can't parse a factor at position %s of %s." % (self.cond_expr.currpos, self.cond_expr.str))
	def term(self):
		m1 = self.factor()
		andop = self.match_andop()
		if andop is not None:
			m2 = self.term()
			return CondAstNode(andop, m1, m2)
		else:
			return m1
	def expression(self):
		e1 = self.term()
		orop = self.match_orop()
		if orop is not None:
			e2 = self.expression()
			return CondAstNode(orop, e1, e2)
		else:
			return e1
	def parse(self):
		exp = self.expression()
		if not self.cond_expr.eoi():
			raise CondParserError("Conditional expression parser did not consume entire string; remainder = %s." % self.cond_expr.remainder())
		return exp

#-------------------------------------------------------------------------------------
#		Numeric Parser
#-------------------------------------------------------------------------------------
class NumericParserError(Exception):
	def __init__(self, msg):
		self.value = msg
	def __repr__(self):
		return "NumericParserError(%r)" % self.value

class NumericParser(NumTokens, object):
	# Takes a numeric expression string
	def __init__(self, numexpr):
		self.num_expr = SourceString(numexpr)
		self.rxint = re.compile(r'(?P<int_num>[+-]?[0-9]+)')
		self.rxfloat = re.compile(r'(?P<float_num>[+-]?(?:(?:[0-9]*\.[0-9]+)|(?:[0-9]+\.[0-9]*)))')
	def match_number(self):
		# Try to match a number in the source string.
		# Return it if matched, return None if unmatched.
		m1 = self.num_expr.match_regex(self.rxfloat)
		if m1 is not None:
			return float(m1['float_num'])
		else:
			m2 = self.num_expr.match_regex(self.rxint)
			if m2 is not None:
				return int(m2['int_num'])
		return None
	def match_mulop(self):
		# Try to match a multiplication or division operator in the source string.
		# if found, return the matching operator type.  If not found, return None.
		m1 = self.num_expr.match_str("*")
		if m1 is not None:
			return self.MUL
		else:
			m2 = self.num_expr.match_str("/")
			if m2 is not None:
				return self.DIV
		return None
	def match_addop(self):
		# Try to match an addition or division operator in the source string.
		# if found, return the matching operator type.  If not found, return None.
		m1 = self.num_expr.match_str("+")
		if m1 is not None:
			return self.ADD
		else:
			m2 = self.num_expr.match_str("-")
			if m2 is not None:
				return self.SUB
		return None
	def factor(self):
		# Parses a factor out of the source string and returns the
		# AST node that is created.
		m1 = self.match_number()
		if m1 is not None:
			return NumericAstNode(self.NUMBER, m1, None)
		else:
			if self.num_expr.match_str("(") is not None:
				m1 = self.expression()
				rp = self.num_expr.match_str(")")
				if rp is None:
					raise NumericParserError("Expected closing parenthesis at position %s of %s." % (self.num_expr.currpos, self.num_expr.str))
				else:
					return m1
			else:
				raise NumericParserError("Can't parse a factor at position %s of %s." % (self.num_expr.currpos, self.num_expr.str))
	def term(self):
		# Parses a term out of the source string and returns the
		# AST node that is created.
		m1 = self.factor()
		mulop = self.match_mulop()
		if mulop is not None:
			m2 = self.term()
			return NumericAstNode(mulop, m1, m2)
		else:
			return m1
	def expression(self):
		# Parses an expression out of the source string and returns the
		# AST node that is created.
		e1 = self.term()
		if e1 is None:
			return
		addop = self.match_addop()
		if addop is not None:
			e2 = self.expression()
			return NumericAstNode(addop, e1, e2)
		else:
			return e1
	def parse(self):
		exp = self.expression()
		if not self.num_expr.eoi():
			raise NumericParserError("Numeric expression parser did not consume entire string; remainder = %s." % self.num_expr.remainder())
		return exp


# End of Parser classes
#===============================================================================================



#===============================================================================================
#-----  METACOMMAND FUNCTIONS


#****	DEBUG WRITE METACOMMANDLIST
# Undocumented; used to acquire data used to set the ordering of metacommands.
def x_debug_write_metacommands(**kwargs):
	output_dest = kwargs['filename']
	if output_dest is None or output_dest == 'stdout':
		ofile = output
	else:
		ofile = EncodedFile(output_dest, conf.output_encoding).open("w")
	for m in metacommandlist:
		ofile.write(u"(%s)  %s\n" % (m.hitcount, m.rx.pattern))

metacommandlist.add(ins_fn_rxs(r'^\s*DEBUG\s+WRITE\s+METACOMMANDLIST\s+TO\s+', r'\s*$'), x_debug_write_metacommands)


#****	BREAK
def x_break(**kwargs):
	global commandlistack
	global loopcommandstack
	if len(commandliststack) == 1:
		line_no = current_script_line()
		script_errors.append(["BREAK metacommand with no command nesting", line_no])
	else:
		if_stack.if_levels = if_stack.if_levels[:commandliststack[-1].init_if_level]
		commandliststack.pop()
	return None

metacommandlist.add(r'^\s*BREAK\s*$', x_break)


#****	CD
def x_cd(**kwargs):
	new_dir = unquoted(kwargs['dir'])
	if not os.path.isdir(new_dir):
		raise ErrInfo(type="cmd", command_text=kwargs['metacommandline'], other_msg='Directory does not exist')
	os.chdir(new_dir)
	lno = current_script_line()
	return None

metacommandlist.add(r'^\s*CD\s+(?P<dir>.+)\s*$', x_cd)


#****	SUB_LOCAL
# Define a local variable.  Local variables must start with a tilde.  As a convenience, one
# will be added if missing.
def x_sub_local(**kwargs):
	varname = kwargs['match']
	if varname[0] != '~':
		varname = '~' + varname
	global commandliststack
	commandliststack[-1].localvars.add_substitution(varname, kwargs['repl'])
	return None

metacommandlist.add(r'^\s*SUB_LOCAL\s+(?P<match>~?\w+)\s+(?P<repl>.+)$', x_sub_local, "SUB", "Define a local variable consisting of a string to match and a replacement for it.")


#****	WAIT_UNTIL
def x_wait_until(**kwargs):
	countdown = int(kwargs['seconds'])
	while countdown > 0:
		if xcmd_test(kwargs['condition']):
			return
		time.sleep(1)
		countdown -= 1
	if kwargs['end'].lower() == 'halt':
		return None
	return None

metacommandlist.add(r'^\s*WAIT_UNTIL\s+(?P<condition>.+)\s+(?P<end>HALT|CONTINUE)\s+AFTER\s+(?P<seconds>\d+)\s+SECONDS\s*$', x_wait_until)


#****	ON ERROR_HALT EXECUTE SCRIPT CLEAR
def x_error_halt_exec_clear(**kwargs):
	global err_halt_exec
	err_halt_exec = None

metacommandlist.add(r'^\s*ON\s+ERROR_HALT\s+EXEC\s+CLEAR\s*$', x_error_halt_exec_clear)


#****	RESET COUNTER
def x_reset_counter(**kwargs):
	ctr_no = int(kwargs["counter_no"])
	counters.remove_counter(ctr_no)

metacommandlist.add(r'^\s*RESET\s+COUNTER\s+(?P<counter_no>\d+)\s*$', x_reset_counter)


#****	RESET COUNTERS
def x_reset_counters(**kwargs):
	counters.remove_all_counters()

metacommandlist.add(r'^\s*RESET\s+COUNTERS\s*$', x_reset_counters)


#****	SET COUNTER
def x_set_counter(**kwargs):
	ctr_no = int(kwargs["counter_no"])
	ctr_expr = kwargs["value"]
	counters.set_counter(ctr_no, int(math.floor(NumericParser(ctr_expr).parse().eval())))

metacommandlist.add(r'^\s*SET\s+COUNTER\s+(?P<counter_no>\d+)\s+TO\s+(?P<value>[0-9+\-*/() ]+)\s*$', x_set_counter)


#****	TIMER
def x_timer(**kwargs):
	onoff = kwargs["onoff"].lower()
	if onoff == 'on':
		timer.start()
	else:
		timer.stop()

metacommandlist.add(r'^\s*TIMER\s+(?P<onoff>ON|OFF)\s*$', x_timer)


#****	BEGIN BATCH
def x_begin_batch(**kwargs):
	status.batch.new_batch()
	return None

metacommandlist.add(r'^\s*BEGIN\s+BATCH\s*$', x_begin_batch)


#****	END BATCH
def x_end_batch(**kwargs):
	status.batch.end_batch()
	return None

# Set a name so this can be found and evaluated during processing, when all other metacommands are ignored.
metacommandlist.add(r'^\s*END\s+BATCH\s*$', x_end_batch, "END BATCH", run_in_batch=True)


#****	ROLLBACK BATCH
def x_rollback(**kwargs):
	status.batch.rollback_batch()

metacommandlist.add(r'^\s*ROLLBACK(:?\s+BATCH)?\s*$', x_rollback, "ROLLBACK BATCH", run_in_batch=True)


#****	ERROR_HALT
def x_error_halt(**kwargs):
	flag = kwargs['onoff'].lower()
	if not flag in ('on', 'off', 'yes', 'no', 'true', 'false'):
		raise ErrInfo(type="cmd", command_text=kwargs["metacommandline"], other_msg=u"Unrecognized flag for error handling: %s" % flag)
	status.halt_on_err = flag in ('on', 'yes', 'true')
	return None

metacommandlist.add(r'\s*ERROR_HALT\s+(?P<onoff>ON|OFF|YES|NO|TRUE|FALSE)\s*$', x_error_halt)


#****	METACOMMAND_ERROR_HALT
def x_metacommand_error_halt(**kwargs):
	flag = kwargs['onoff'].lower()
	if not flag in ('on', 'off', 'yes', 'no', 'true', 'false'):
		raise ErrInfo(type="cmd", command_text=kwargs["metacommandline"], other_msg=u"Unrecognized flag for metacommand error handling: %s" % flag)
	status.halt_on_metacommand_err = flag in ('on', 'yes', 'true')
	return None

metacommandlist.add(r'\s*METACOMMAND_ERROR_HALT\s+(?P<onoff>ON|OFF|YES|NO|TRUE|FALSE)\s*$', x_metacommand_error_halt, set_error_flag=False)


#****	LOOP
def x_loop(**kwargs):
	global compiling_loop
	compiling_loop = True
	looptype = kwargs["looptype"].upper()
	loopcond = kwargs["loopcond"]
	listname = 'loop'+str(len(loopcommandstack)+1)
	if looptype == 'WHILE':
		loopcommandstack.append(CommandListWhileLoop([], listname, paramnames=None, loopcondition=loopcond))
	else:
		loopcommandstack.append(CommandListUntilLoop([], listname, paramnames=None, loopcondition=loopcond))

metacommandlist.add(r'^\s*LOOP\s+(?P<looptype>WHILE|UNTIL)\s*\(\s*(?P<loopcond>.+)\s*\)\s*$', x_loop)


#****	END LOOP
def endloop():
	if len(loopcommandstack) == 0:
		raise ErrInfo("error", other_msg="END LOOP metacommand without a matching preceding LOOP metacommand.")
	global compiling_loop
	compiling_loop = False
	commandliststack.append(loopcommandstack[-1])
	loopcommandstack.pop()


#****	SUB_EMPTY
def x_sub_empty(**kwargs):
	varname = kwargs['match']
	# Get subvarset assignment and cleansed variable name
	subvarset, varname = get_subvarset(varname, kwargs['metacommandline'])
	subvarset.add_substitution(varname, u'')
	return None

metacommandlist.add(r'^\s*SUB_EMPTY\s+(?P<match>[+~]?\w+)\s*$', x_sub_empty)


#****	SUB_ADD
def x_sub_add(**kwargs):
	varname = kwargs["match"]
	increment_expr = kwargs["increment"]
	# Get subvarset assignment and cleansed variable name
	subvarset, varname = get_subvarset(varname, kwargs['metacommandline'])
	subvarset.increment_by(varname, NumericParser(increment_expr).parse().eval())
	return None

metacommandlist.add(r'^\s*SUB_ADD\s+(?P<match>[+~]?\w+)\s+(?P<increment>[+\-0-9\.*/() ]+)\s*$', x_sub_add)


#****	SUB_APPEND
def x_sub_append(**kwargs):
	varname = kwargs["match"]
	# Get subvarset assignment and cleansed variable name
	subvarset, varname = get_subvarset(varname, kwargs['metacommandline'])
	subvarset.append_substitution(varname, kwargs['repl'])
	return None

metacommandlist.add(r'^\s*SUB_APPEND\s+(?P<match>[+~]?\w+)\s(?P<repl>(.|\n)*)$', x_sub_append)


#****	BLOCK ORIF
def x_if_orif(**kwargs):
	if if_stack.all_true():
		return None		# Short-circuit evaluation
	if if_stack.only_current_false():
		if_stack.replace(xcmd_test(kwargs['condtest']))
	return None

metacommandlist.add(r'^\s*ORIF\s*\(\s*(?P<condtest>.+)\s*\)\s*$', x_if_orif, run_when_false=True)


#****	EXTEND SCRIPT WITH SCRIPT
#****	APPEND SCRIPT
def x_extendscript(**kwargs):
	script1 = kwargs["script1"].lower()
	if script1 not in savedscripts:
		raise ErrInfo("cmd", other_msg="There is no SCRIPT named %s." % script1)
	script2 = kwargs["script2"].lower()
	if script1 not in savedscripts:
		raise ErrInfo("cmd", other_msg="There is no SCRIPT named %s." % script2)
	s1 = savedscripts[script1]
	s2 = savedscripts[script2]
	for cmd in s1.cmdlist:
		s2.add(cmd)
	if s1.paramnames is not None:
		if s2.paramnames is None:
			s2.paramnames = []
		for param in s1.paramnames:
			if param not in s2.paramnames:
				s2.paramnames.append(param)

metacommandlist.add(r'\s*EXTEND\s+SCRIPT\s+(?P<script2>\w+)\s+WITH\s+SCRIPT\s+(?P<script1>\w+)\s*$', x_extendscript)
metacommandlist.add(r'\s*APPEND\s+SCRIPT\s+(?P<script1>\w+)\s+TO\s+(?P<script2>\w+)\s*$', x_extendscript)


#****	EXTEND SCRIPT WITH METACOMMAND
def x_extendscript_metacommand(**kwargs):
	script = kwargs["script"].lower()
	if script not in savedscripts:
		raise ErrInfo("cmd", other_msg="There is no SCRIPT named %s." % script)
	script_line_no = current_script_line()
	savedscripts[script].add(ScriptCmd(script_file, script_line_no, 'cmd', MetacommandStmt(kwargs["cmd"])))

metacommandlist.add(r'\s*EXTEND\s+SCRIPT\s+(?P<script>\w+)\s+WITH\s+METACOMMAND\s+(?P<cmd>.+)\s*$', x_extendscript_metacommand)


#****	EXTEND SCRIPT WITH SQL
def x_extendscript_sql(**kwargs):
	script = kwargs["script"].lower()
	if script not in savedscripts:
		raise ErrInfo("cmd", other_msg="There is no SCRIPT named %s." % script)
	sql = kwargs["sql"]
	script_line_no = current_script_line()
	savedscripts[script].add(ScriptCmd(script_file, script_line_no , 'sql', SqlStmt(kwargs["sql"])))

metacommandlist.add(r'\s*EXTEND\s+SCRIPT\s+(?P<script>\w+)\s+WITH\s+SQL\s+(?P<sql>.+;)\s*$', x_extendscript_sql)


#****	ON ERROR_HALT EXECUTE SCRIPT
def x_error_halt_exec(**kwargs):
	global err_halt_exec
	err_halt_exec = ScriptExecSpec(**kwargs)

metacommandlist.add(r'^\s*ON\s+ERROR_HALT\s+EXEC(?:UTE)?\s+SCRIPT\s+(?P<script_id>\w+)(?:(?:\s+WITH)?(?:\s+ARG(?:UMENT)?S?)?\s*\(\s*(?P<argexp>#?\w+\s*=\s*(?:(?:[^"\'\[][^,\)]*)|(?:"[^"]*")|(?:\'[^\']*\')|(?:\[[^\]]*\]))(?:\s*,\s*#?\w+\s*=\s*(?:(?:[^"\'\[][^,\)]*)|(?:"[^"]*")|(?:\'[^\']*\')|(?:\[[^\]]*\])))*)\s*\))?\s*$', x_error_halt_exec)


#****	DEBUG_WRITE
def x_debug_write(**kwargs):
	msg = u'%s\n' % kwargs['text']
	print("%s\n" % msg)
	return None

metacommandlist.add(r'^\s*WRITE\s+\~(?P<text>([^\~]|\n)*)\~\s*$', x_debug_write)
metacommandlist.add(r'^\s*WRITE\s+\#(?P<text>([^\#]|\n)*)\#\s*$', x_debug_write)
metacommandlist.add(r'^\s*WRITE\s+\`(?P<text>([^\`]|\n)*)\`\s*$', x_debug_write)
metacommandlist.add(r'^\s*WRITE\s+\[(?P<text>([^\]]|\n)*)\]\s*$', x_debug_write)
metacommandlist.add(r'^\s*WRITE\s+\'(?P<text>([^\']|\n)*)\'\s*$', x_debug_write)
metacommandlist.add(r'^\s*WRITE\s+"(?P<text>([^"]|\n)*)"\s*$', x_debug_write)



#****	INCLUDE
def x_include(**kwargs):
	filename = kwargs['filename']
	if len(filename) > 1 and filename[0] == "~" and filename[1] == os.sep:
		filename = os.path.join(os.path.expanduser(r'~'), filename[2:])
	exists = kwargs['exists']
	if exists is not None:
		if os.path.isfile(filename):
			read_sqlfile(filename)
	else:
		if not os.path.isfile(filename):
			raise ErrInfo(type="error", other_msg="File %s does not exist." % filename)
		read_sqlfile(filename)
	return None

metacommandlist.add(ins_fn_rxs(r'^\s*INCLUDE(?P<exists>\s+IF\s+EXISTS?)?\s+', r'\s*$'), x_include)


#****	RM_SUB
def x_rm_sub(**kwargs):
	varname = kwargs["match"]
	subvarset = subvars if varname[0] != '~' else commandliststack[-1].localvars
	subvarset.remove_substitution(varname)
	return None

metacommandlist.add(r'^\s*RM_SUB\s+(?P<match>~?\w+)\s*$', x_rm_sub)


#****	BLOCK ELSEIF
def x_if_elseif(**kwargs):
	if if_stack.only_current_false():
		if_stack.replace(xcmd_test(kwargs['condtest']))
	else:
		if_stack.replace(False)
	return None

metacommandlist.add(r'^\s*ELSEIF\s*\(\s*(?P<condtest>.+)\s*\)\s*$', x_if_elseif, run_when_false=True)


#****	AUTOCOMMIT OFF
def x_autocommit_off(**kwargs):
	db = dbs.current()
	db.autocommit_off()

metacommandlist.add(r'^\s*AUTOCOMMIT\s+OFF\s*$', x_autocommit_off)


#****	AUTOCOMMIT ON
def x_autocommit_on(**kwargs):
	action = kwargs['action']
	if action is not None:
		action = action.lower()
	db = dbs.current()
	db.autocommit_on()
	if action is not None:
		if action == 'commit':
			db.commit()
		else:
			db.rollback()

metacommandlist.add(r'^\s*AUTOCOMMIT\s+ON(?:\s+WITH\s+(?P<action>COMMIT|ROLLBACK))?\s*$', x_autocommit_on)


#****	BLOCK ANDIF
def x_if_andif(**kwargs):
	if if_stack.all_true():
		if_stack.replace(if_stack.current() and xcmd_test(kwargs['condtest']))
	return None

metacommandlist.add(r'^\s*ANDIF\s*\(\s*(?P<condtest>.+)\s*\)\s*$', x_if_andif)


#****	SELECT_SUB
def x_selectsub(**kwargs):
	sql = u"select * from %s;" % kwargs["datasource"]
	db = dbs.current()
	line_no = current_script_line()
	nodatamsg = "There are no data in %s to use with the SELECT_SUB metacommand (line %d)." % (kwargs["datasource"], line_no)
	try:
		hdrs, rec = db.select_rowsource(sql)
	except ErrInfo:
		raise
	except:
		raise ErrInfo(type="exception", exception_msg=exception_desc(), other_msg="Can't get headers and rows from %s." % sql)
	# Remove any existing variables with these names
	for subvar in hdrs:
		subvar = u'@'+subvar
		if subvars.sub_exists(subvar):
			subvars.remove_substitution(subvar)
	try:
		row1 = next(rec)
	except StopIteration:
		row1 = None
	except:
		raise ErrInfo(type="exception", exception_msg=exception_desc(), other_msg=nodatamsg)
	if row1:
		for i, item in enumerate(row1):
			if item is None:
				item = u''
			if sys.version_info < (3,):
				item = unicode(item)
			else:
				item = str(item)
			match_str = u"@" + hdrs[i]
			subvars.add_substitution(match_str, item)
	return None

metacommandlist.add(r'^\s*SELECT_SUB\s+(?P<datasource>.+)\s*$', x_selectsub)


#****	SUBDATA
def x_subdata(**kwargs):
	varname = kwargs["match"]
	sql = u"select * from %s;" % kwargs["datasource"]
	db = dbs.current()
	line_no = current_script_line()
	errmsg = "There are no data in %s to use with the SUBDATA metacommand (line %d)." % (kwargs["datasource"], line_no)
	# Get subvarset assignment and cleansed variable name
	subvarset, varname = get_subvarset(varname, kwargs['metacommandline'])
	subvarset.remove_substitution(varname)
	# Exceptions should be trapped by the caller, so are re-raised here after settting status
	try:
		hdrs, rec = db.select_rowsource(sql)
	except ErrInfo:
		raise
	except:
		raise ErrInfo(type="exception", exception_msg=exception_desc(), other_msg="Can't get headers and rows from %s." % sql)
	try:
		row1 = next(rec)
	except:
		row1 = None
	if row1:
		dataval = row1[0]
		if dataval is None:
			dataval = u''
		if not isinstance(dataval, str):
			if sys.version_info < (3,):
				dataval = unicode(dataval)
			else:
				dataval = str(dataval)
		subvarset.add_substitution(varname, dataval)
	return None

metacommandlist.add(r'^\s*SUBDATA\s+(?P<match>[+~]?\w+)\s+(?P<datasource>.+)\s*$', x_subdata)


#****	IF
def x_if(**kwargs):
	tf_value = xcmd_test(kwargs['condtest'])
	if tf_value:
		line_no = current_script_line()
		metacmd = MetacommandStmt(kwargs['condcmd'])
		script_cmd = ScriptCmd(src, line_no, "cmd", metacmd)
		cmdlist = CommandList([script_cmd], "%s_%d" % (src, line_no))
		commandliststack.append(cmdlist)
	return None

metacommandlist.add(r'^\s*IF\s*\(\s*(?P<condtest>.+)\s*\)\s*{\s*(?P<condcmd>.+)\s*}\s*$', x_if)


#****	EXECUTE SCRIPT
def x_executescript(**kwargs):
	exists = kwargs["exists"]
	script_id = kwargs["script_id"].lower()
	if exists is None or (exists is not None and script_id in savedscripts):
		ScriptExecSpec(**kwargs).execute()

metacommandlist.add(r'^\s*EXEC(?:UTE)?\s+SCRIPT(?:\s+(?P<exists>IF\s+EXISTS))?\s+(?P<script_id>\w+)(?:(?:\s+WITH)?(?:\s+ARG(?:UMENT)?S?)?\s*\(\s*(?P<argexp>#?\w+\s*=\s*(?:(?:[^"\'\[][^,\)]*)|(?:"[^"]*")|(?:\'[^\']*\')|(?:\[[^\]]*\]))(?:\s*,\s*#?\w+\s*=\s*(?:(?:[^"\'\[][^,\)]*)|(?:"[^"]*")|(?:\'[^\']*\')|(?:\[[^\]]*\])))*)\s*\))?(?:\s+(?P<looptype>WHILE|UNTIL)\s*\(\s*(?P<loopcond>.+)\s*\))?\s*$', x_executescript)


#****	BLOCK ELSE
def x_if_else(**kwargs):
	if if_stack.all_true() or if_stack.only_current_false():
		if_stack.invert()
	return None

metacommandlist.add(r'^\s*ELSE\s*$', x_if_else, run_when_false=True)


#****	SUB
def x_sub(**kwargs):
	varname = kwargs['match']
	# Get subvarset assignment and cleansed variable name
	subvarset, varname = get_subvarset(varname, kwargs['metacommandline'])
	subvarset.add_substitution(varname, kwargs['repl'])
	return None

metacommandlist.add(r'^\s*SUB\s+(?P<match>[+~]?\w+)\s+(?P<repl>.+)$', x_sub, "SUB", "Define a string to match and a replacement for it.")


#****	BLOCK IF
def x_if_block(**kwargs):
	if if_stack.all_true():
		test_result = xcmd_test(kwargs['condtest'])
		if_stack.nest(test_result)
	else:
		if_stack.nest(False)
	return None

metacommandlist.add(r'^\s*IF\s*\(\s*(?P<condtest>.+)\s*\)\s*$', x_if_block, run_when_false=True)


#****	BLOCK ENDIF
def x_if_end(**kwargs):
	if_stack.unnest()
	return None

metacommandlist.add(r'^\s*ENDIF\s*$', x_if_end, run_when_false=True)


#****	DEBUG WRITE METACOMMANDLIST
# Undocumented; used to acquire data used to set the ordering of metacommands.
def x_debug_write_metacommands(**kwargs):
	output_dest = kwargs['filename']
	if output_dest is None or output_dest == 'stdout':
		ofile = output
	else:
		ofile = EncodedFile(output_dest, conf.output_encoding).open("w")
	for m in metacommandlist:
		ofile.write(u"(%s)  %s\n" % (m.hitcount, m.rx.pattern))

metacommandlist.add(ins_fn_rxs(r'^\s*DEBUG\s+WRITE\s+METACOMMANDLIST\s+TO\s+', r'\s*$'), x_debug_write_metacommands)


#	End of metacommand definitions.
#===============================================================================================


#===============================================================================================
#-----  CONDITIONAL TESTS FOR METACOMMANDS


def xf_contains(**kwargs):
	s1 = kwargs["string1"]
	s2 = kwargs["string2"]
	if kwargs["ignorecase"] and kwargs["ignorecase"].lower() == "i":
		s1 = s1.lower()
		s2 = s2.lower()
	return s2 in s1

conditionallist.add(r'^\s*CONTAINS\s*\(\s*(?P<string1>[^ )]+)\s*,\s*(?P<string2>[^ )]+)(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_contains)
conditionallist.add(r'^\s*CONTAINS\s*\(\s*"(?P<string1>[^"]+)"\s*,\s*(?P<string2>[^ )]+)(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_contains)
conditionallist.add(r'^\s*CONTAINS\s*\(\s*(?P<string1>[^ )]+)\s*,\s*"(?P<string2>[^"]+)"(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_contains)
conditionallist.add(r"^\s*CONTAINS\s*\(\s*(?P<string1>[^ )]+)\s*,\s*'(?P<string2>[^']+)'\s*(?:\s*,\s*(?P<ignorecase>I))?\)", xf_contains)
conditionallist.add(r"^\s*CONTAINS\s*\(\s*(?P<string1>[^ )]+)\s*,\s*`(?P<string2>[^`]+)`(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_contains)
conditionallist.add(r"^\s*CONTAINS\s*\(\s*`(?P<string1>[^`]+)`\s*,\s*(?P<string2>[^ )]+)(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_contains)
conditionallist.add(r"^\s*CONTAINS\s*\(\s*'(?P<string1>[^']+)'\s*,\s*(?P<string2>[^ )]+)\s*(?:\s*,\s*(?P<ignorecase>I))?\)", xf_contains)
conditionallist.add(r'^\s*CONTAINS\s*\(\s*"(?P<string1>[^"]+)"\s*,\s*"(?P<string2>[^"]+)"(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_contains)
conditionallist.add(r"^\s*CONTAINS\s*\(\s*'(?P<string1>[^']+)'\s*,\s*'(?P<string2>[^']+)'(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_contains)
conditionallist.add(r"^\s*CONTAINS\s*\(\s*'(?P<string1>[^']+)'\s*,\s*\"(?P<string2>[^\"]+)\"(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_contains)
conditionallist.add(r"^\s*CONTAINS\s*\(\s*\"(?P<string1>[^\"]+)\"\s*,\s*'(?P<string2>[^']+)'(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_contains)
conditionallist.add(r"^\s*CONTAINS\s*\(\s*`(?P<string1>[^`]+)`\s*,\s*`(?P<string2>[^`]+)`(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_contains)
conditionallist.add(r'^\s*CONTAINS\s*\(\s*`(?P<string1>[^`]+)`\s*,\s*"(?P<string2>[^"]+)"(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_contains)
conditionallist.add(r'^\s*CONTAINS\s*\(\s*"(?P<string1>[^"]+)"\s*,\s*`(?P<string2>[^`]+)`(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_contains)
conditionallist.add(r"^\s*CONTAINS\s*\(\s*`(?P<string1>[^`]+)`\s*,\s*'(?P<string2>[^']+)'(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_contains)
conditionallist.add(r"^\s*CONTAINS\s*\(\s*'(?P<string1>[^']+)'\s*,\s*`(?P<string2>[^`]+)`(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_contains)


def xf_startswith(**kwargs):
	s1 = kwargs["string1"]
	s2 = kwargs["string2"]
	if kwargs["ignorecase"] and kwargs["ignorecase"].lower() == "i":
		s1 = s1.lower()
		s2 = s2.lower()
	return s1[:len(s2)] == s2

conditionallist.add(r'^\s*STARTS_WITH\s*\(\s*(?P<string1>[^ )]+)\s*,\s*(?P<string2>[^ )]+)(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_startswith)
conditionallist.add(r'^\s*STARTS_WITH\s*\(\s*"(?P<string1>[^"]+)"\s*,\s*(?P<string2>[^ )]+)(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_startswith)
conditionallist.add(r'^\s*STARTS_WITH\s*\(\s*(?P<string1>[^ )]+)\s*,\s*"(?P<string2>[^"]+)"(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_startswith)
conditionallist.add(r"^\s*STARTS_WITH\s*\(\s*(?P<string1>[^ )]+)\s*,\s*'(?P<string2>[^']+)'\s*(?:\s*,\s*(?P<ignorecase>I))?\)", xf_startswith)
conditionallist.add(r"^\s*STARTS_WITH\s*\(\s*(?P<string1>[^ )]+)\s*,\s*`(?P<string2>[^`]+)`(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_startswith)
conditionallist.add(r"^\s*STARTS_WITH\s*\(\s*`(?P<string1>[^`]+)`\s*,\s*(?P<string2>[^ )]+)(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_startswith)
conditionallist.add(r"^\s*STARTS_WITH\s*\(\s*'(?P<string1>[^']+)'\s*,\s*(?P<string2>[^ )]+)\s*(?:\s*,\s*(?P<ignorecase>I))?\)", xf_startswith)
conditionallist.add(r'^\s*STARTS_WITH\s*\(\s*"(?P<string1>[^"]+)"\s*,\s*"(?P<string2>[^"]+)"(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_startswith)
conditionallist.add(r"^\s*STARTS_WITH\s*\(\s*'(?P<string1>[^']+)'\s*,\s*'(?P<string2>[^']+)'(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_startswith)
conditionallist.add(r"^\s*STARTS_WITH\s*\(\s*'(?P<string1>[^']+)'\s*,\s*\"(?P<string2>[^\"]+)\"(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_startswith)
conditionallist.add(r"^\s*STARTS_WITH\s*\(\s*\"(?P<string1>[^\"]+)\"\s*,\s*'(?P<string2>[^']+)'(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_startswith)
conditionallist.add(r"^\s*STARTS_WITH\s*\(\s*`(?P<string1>[^`]+)`\s*,\s*`(?P<string2>[^`]+)`(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_startswith)
conditionallist.add(r'^\s*STARTS_WITH\s*\(\s*`(?P<string1>[^`]+)`\s*,\s*"(?P<string2>[^"]+)"(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_startswith)
conditionallist.add(r'^\s*STARTS_WITH\s*\(\s*"(?P<string1>[^"]+)"\s*,\s*`(?P<string2>[^`]+)`(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_startswith)
conditionallist.add(r"^\s*STARTS_WITH\s*\(\s*`(?P<string1>[^`]+)`\s*,\s*'(?P<string2>[^']+)'(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_startswith)
conditionallist.add(r"^\s*STARTS_WITH\s*\(\s*'(?P<string1>[^']+)'\s*,\s*`(?P<string2>[^`]+)`(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_startswith)



def xf_endswith(**kwargs):
	s1 = kwargs["string1"]
	s2 = kwargs["string2"]
	if kwargs["ignorecase"] and kwargs["ignorecase"].lower() == "i":
		s1 = s1.lower()
		s2 = s2.lower()
	return s1[-len(s2):] == s2

conditionallist.add(r'^\s*ENDS_WITH\s*\(\s*(?P<string1>[^ )]+)\s*,\s*(?P<string2>[^ )]+)(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_endswith)
conditionallist.add(r'^\s*ENDS_WITH\s*\(\s*"(?P<string1>[^"]+)"\s*,\s*(?P<string2>[^ )]+)(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_endswith)
conditionallist.add(r'^\s*ENDS_WITH\s*\(\s*(?P<string1>[^ )]+)\s*,\s*"(?P<string2>[^"]+)"(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_endswith)
conditionallist.add(r"^\s*ENDS_WITH\s*\(\s*(?P<string1>[^ )]+)\s*,\s*'(?P<string2>[^']+)'\s*(?:\s*,\s*(?P<ignorecase>I))?\)", xf_endswith)
conditionallist.add(r"^\s*ENDS_WITH\s*\(\s*(?P<string1>[^ )]+)\s*,\s*`(?P<string2>[^`]+)`(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_endswith)
conditionallist.add(r"^\s*ENDS_WITH\s*\(\s*`(?P<string1>[^`]+)`\s*,\s*(?P<string2>[^ )]+)(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_endswith)
conditionallist.add(r"^\s*ENDS_WITH\s*\(\s*'(?P<string1>[^']+)'\s*,\s*(?P<string2>[^ )]+)\s*(?:\s*,\s*(?P<ignorecase>I))?\)", xf_endswith)
conditionallist.add(r'^\s*ENDS_WITH\s*\(\s*"(?P<string1>[^"]+)"\s*,\s*"(?P<string2>[^"]+)"(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_endswith)
conditionallist.add(r"^\s*ENDS_WITH\s*\(\s*'(?P<string1>[^']+)'\s*,\s*'(?P<string2>[^']+)'(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_endswith)
conditionallist.add(r"^\s*ENDS_WITH\s*\(\s*'(?P<string1>[^']+)'\s*,\s*\"(?P<string2>[^\"]+)\"(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_endswith)
conditionallist.add(r"^\s*ENDS_WITH\s*\(\s*\"(?P<string1>[^\"]+)\"\s*,\s*'(?P<string2>[^']+)'(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_endswith)
conditionallist.add(r"^\s*ENDS_WITH\s*\(\s*`(?P<string1>[^`]+)`\s*,\s*`(?P<string2>[^`]+)`(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_endswith)
conditionallist.add(r'^\s*ENDS_WITH\s*\(\s*`(?P<string1>[^`]+)`\s*,\s*"(?P<string2>[^"]+)"(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_endswith)
conditionallist.add(r'^\s*ENDS_WITH\s*\(\s*"(?P<string1>[^"]+)"\s*,\s*`(?P<string2>[^`]+)`(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_endswith)
conditionallist.add(r"^\s*ENDS_WITH\s*\(\s*`(?P<string1>[^`]+)`\s*,\s*'(?P<string2>[^']+)'(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_endswith)
conditionallist.add(r"^\s*ENDS_WITH\s*\(\s*'(?P<string1>[^']+)'\s*,\s*`(?P<string2>[^`]+)`(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_endswith)


def xf_hasrows(**kwargs):
	queryname = kwargs["queryname"]
	sql = u"select count(*) from %s;" % queryname
	# Exceptions should be trapped by the caller, so are re-raised here after settting status
	try:
		hdrs, rec = dbs.current().select_data(sql)
	except ErrInfo:
		raise
	except:
		raise ErrInfo("db", sql, exception_msg=exception_desc())
	nrows = rec[0][0]
	return nrows > 0

conditionallist.add(r'^\s*HASROWS\((?P<queryname>[^)]+)\)', xf_hasrows)
conditionallist.add(r'^\s*HAS_ROWS\((?P<queryname>[^)]+)\)', xf_hasrows)

def xf_sqlerror(**kwargs):
	return status.sql_error

conditionallist.add(r'^\s*sql_error\(\s*\)', xf_sqlerror)

def xf_fileexists(**kwargs):
	filename = kwargs["filename"]
	return os.path.isfile(filename.strip())

conditionallist.add(ins_fn_rxs(r'^FILE_EXISTS\(\s*', r'\)'), xf_fileexists)

def xf_direxists(**kwargs):
	dirname = kwargs["dirname"]
	return os.path.isdir(dirname.strip())

conditionallist.add(r'^DIRECTORY_EXISTS\(\s*("?)(?P<dirname>[^")]+)\1\)', xf_direxists)

def xf_schemaexists(**kwargs):
	schemaname = kwargs["schema"]
	return dbs.current().schema_exists(schemaname)

conditionallist.add((
	r'^SCHEMA_EXISTS\(\s*(?P<schema>[A-Za-z0-9_\-\: ]+)\s*\)',
	r'^SCHEMA_EXISTS\(\s*"(?P<schema>[A-Za-z0-9_\-\: ]+)"\s*\)'
	), xf_schemaexists)


def xf_tableexists(**kwargs):
	schemaname = kwargs["schema"]
	tablename = kwargs["tablename"]
	return  dbs.current().table_exists(tablename.strip(), schemaname)

conditionallist.add((
	r'^TABLE_EXISTS\(\s*(?:(?P<schema>[A-Za-z0-9_\-\/\: ]+)\.)?(?P<tablename>[A-Za-z0-9_\-\/\: ]+)\)',
	r'^TABLE_EXISTS\(\s*(?:\[(?P<schema>[A-Za-z0-9_\-\/\: ]+)\]\.)?\[(?P<tablename>[A-Za-z0-9_\-\/\: ]+)\]\)',
	r'^TABLE_EXISTS\(\s*(?:"(?P<schema>[A-Za-z0-9_\-\/\: ]+)"\.)?"(?P<tablename>[A-Za-z0-9_\-\/\: ]+)"\)',
	r'^TABLE_EXISTS\(\s*(?:(?P<schema>[A-Za-z0-9_\-\/]+)\.)?(?P<tablename>[A-Za-z0-9_\-\/]+)\)'
	), xf_tableexists)

def xf_roleexists(**kwargs):
	rolename = kwargs["role"]
	return dbs.current().role_exists(rolename)

conditionallist.add((
	r'^ROLE_EXISTS\(\s*(?P<role>[A-Za-z0-9_\-\:\$ ]+)\s*\)',
	r'^ROLE_EXISTS\(\s*"(?P<role>[A-Za-z0-9_\-\:\$ ]+)"\s*\)'
	), xf_roleexists)


def xf_sub_defined(**kwargs):
	varname = kwargs["match_str"]
	subvarset = subvars if varname[0] not in ('~','#') else commandliststack[-1].localvars if varname[0] == '~' else commandliststack[-1].paramvals
	return subvarset.sub_exists(varname) if subvarset else False

conditionallist.add(r'^SUB_DEFINED\s*\(\s*(?P<match_str>[\$&@~#]?\w+)\s*\)', xf_sub_defined)


def xf_sub_empty(**kwargs):
	varname = kwargs["match_str"]
	subvarset = subvars if varname[0] not in ('~','#') else commandliststack[-1].localvars if varname[0] == '~' else commandliststack[-1].paramvals
	if not subvarset.sub_exists(varname):
		raise ErrInfo(type="cmd", command_text=kwargs["metacommandline"], other_msg=u"Unrecognized substitution variable name: %s" % varname)
	return subvarset.varvalue(varname) == u''

conditionallist.add(r'^SUB_EMPTY\s*\(\s*(?P<match_str>[\$&@~#]?\w+)\s*\)', xf_sub_empty)

def xf_script_exists(**kwargs):
	script_id = kwargs["script_id"].lower()
	return script_id in savedscripts

conditionallist.add(r'^\s*SCRIPT_EXISTS\s*\(\s*(?P<script_id>\w+)\s*\)', xf_script_exists)


def xf_equals(**kwargs):
	import unicodedata
	s1 = unicodedata.normalize('NFC', kwargs["string1"]).lower().strip('"')
	s2 = unicodedata.normalize('NFC', kwargs["string2"]).lower().strip('"')
	converters = (int, float, parse_datetime, parse_datetimetz, parse_date, parse_boolean)
	converted = False
	for convf in converters:
		try:
			v1 = convf(s1)
			v2 = convf(s2)
			if not(v1 is None and v2 is None):
				converted = True
				break
		except:
			continue
	if converted:
		return v1 == v2
	else:
		return s1 == s2

conditionallist.add(r'^\s*EQUAL(S)?\s*\(\s*(?P<string1>[^ )]+)\s*,\s*(?P<string2>[^ )]+)\s*\)', xf_equals)
conditionallist.add(r'^\s*EQUAL(S)?\s*\(\s*"(?P<string1>[^"]+)"\s*,\s*(?P<string2>[^ )]+)\s*\)', xf_equals)
conditionallist.add(r'^\s*EQUAL(S)?\s*\(\s*(?P<string1>[^ )]+)\s*,\s*"(?P<string2>[^"]+)"\s*\)', xf_equals)
conditionallist.add(r'^\s*EQUAL(S)?\s*\(\s*"(?P<string1>[^"]+)"\s*,\s*"(?P<string2>[^"]+)"\s*\)', xf_equals)

def xf_identical(**kwargs):
	s1 = kwargs["string1"].strip('"')
	s2 = kwargs["string2"].strip('"')
	return s1 == s2

conditionallist.add(r'^\s*IDENTICAL\s*\(\s*(?P<string1>[^ ,)]+)\s*,\s*(?P<string2>[^ )]+)\s*\)', xf_identical)
conditionallist.add(r'^\s*IDENTICAL\s*\(\s*"(?P<string1>[^"]+)"\s*,\s*(?P<string2>[^ )]+)\s*\)', xf_identical)
conditionallist.add(r'^\s*IDENTICAL\s*\(\s*(?P<string1>[^ ,]+)\s*,\s*"(?P<string2>[^"]+)"\s*\)', xf_identical)
conditionallist.add(r'^\s*IDENTICAL\s*\(\s*"(?P<string1>[^"]+)"\s*,\s*"(?P<string2>[^"]+)"\s*\)', xf_identical)

def xf_isnull(**kwargs):
	item = kwargs["item"].strip().strip(u'"')
	return item == u""

conditionallist.add(r'^\s*IS_NULL\(\s*(?P<item>"[^"]*")\s*\)', xf_isnull)

def xf_iszero(**kwargs):
	val = kwargs["value"].strip()
	try:
		v = float(val)
	except:
		raise ErrInfo(type="cmd", command_text=kwargs["metacommandline"], other_msg="The value {%s} is not numeric." % val)
	return v == 0

conditionallist.add(r'^\s*IS_ZERO\(\s*(?P<value>[^)]*)\s*\)', xf_iszero)

def xf_isgt(**kwargs):
	val1 = kwargs["value1"].strip()
	val2 = kwargs["value2"].strip()
	try:
		v1 = float(val1)
		v2 = float(val2)
	except:
		raise ErrInfo(type="cmd", command_text=kwargs["metacommandline"], other_msg="Values {%s} and {%s} are not both numeric." % (val1, val2))
	return v1 > v2

conditionallist.add(r'^\s*IS_GT\(\s*(?P<value1>[^)]*)\s*,\s*(?P<value2>[^)]*)\s*\)', xf_isgt)


def xf_isgte(**kwargs):
	val1 = kwargs["value1"].strip()
	val2 = kwargs["value2"].strip()
	try:
		v1 = float(val1)
		v2 = float(val2)
	except:
		raise ErrInfo(type="cmd", command_text=kwargs["metacommandline"], other_msg="Values {%s} and {%s} are not both numeric." % (val1, val2))
	return v1 >= v2

conditionallist.add(r'^\s*IS_GTE\(\s*(?P<value1>[^)]*)\s*,\s*(?P<value2>[^)]*)\s*\)', xf_isgte)


def xf_boolliteral(**kwargs):
	return unquoted(kwargs["value"].strip()).lower() in ('true', 'yes', '1')

conditionallist.add((
	r'^\s*(?P<value>1)\s*',
	r'^\s*(?P<value>"1")\s*',
	r'^\s*(?P<value>0)\s*',
	r'^\s*(?P<value>"0")\s*',
	r'^\s*(?P<value>Yes)\s*',
	r'^\s*(?P<value>"Yes")\s*',
	r'^\s*(?P<value>No)\s*',
	r'^\s*(?P<value>"No")\s*',
	r'^\s*(?P<value>"False")\s*',
	r'^\s*(?P<value>False)\s*',
	r'^\s*(?P<value>"True")\s*',
	r'^\s*(?P<value>True)\s*'
	), xf_boolliteral)


def xf_istrue(**kwargs):
	return unquoted(kwargs["value"].strip()).lower() in ('yes', 'y', 'true', 't', '1')

conditionallist.add(r'^\s*IS_TRUE\(\s*(?P<value>[^)]*)\s*\)', xf_istrue)

def xf_dbms(**kwargs):
	dbms = kwargs["dbms"]
	return dbs.current().type.dbms_id.lower() == dbms.strip().lower()

conditionallist.add((
	r'^\s*DBMS\(\s*(?P<dbms>[A-Z0-9_\-\(\/\\\. ]+)\s*\)',
	r'^\s*DBMS\(\s*"(?P<dbms>[A-Z0-9_\-\(\)\/\\\. ]+)"\s*\)'
	), xf_dbms)


def xf_dbname(**kwargs):
	dbname = kwargs["dbname"]
	return dbs.current().name().lower() == dbname.strip().lower()

                           
conditionallist.add((
	r'^\s*DATABASE_NAME\(\s*(?P<dbname>[A-Z0-9_;\-\(\/\\\. ]+)\s*\)', 
	r'^\s*DATABASE_NAME\(\s*"(?P<dbname>[A-Z0-9_;\-\(\)\/\\\. ]+)"\s*\)'), xf_dbname)

def xf_viewexists(**kwargs):
	viewname = kwargs["viewname"]
	return dbs.current().view_exists(viewname.strip())

conditionallist.add(r'^\s*VIEW_EXISTS\(\s*("?)(?P<viewname>[^")]+)\1\)', xf_viewexists)


def xf_columnexists(**kwargs):
	tablename = kwargs["tablename"]
	schemaname = kwargs["schema"]
	columnname = kwargs["columnname"]
	return dbs.current().column_exists(tablename.strip(), columnname.strip(), schemaname)

conditionallist.add((
	r'^COLUMN_EXISTS\(\s*(?P<columnname>[A-Za-z0-9_\-\:]+)\s+IN\s+(?:(?P<schema>[A-Za-z0-9_\-\: ]+)\.)?(?P<tablename>[A-Za-z0-9_\-\: ]+)\)',
	r'^COLUMN_EXISTS\(\s*(?P<columnname>[A-Za-z0-9_\-\:]+)\s+IN\s+(?:\[(?P<schema>[A-Za-z0-9_\-\: ]+)\]\.)?\[(?P<tablename>[A-Za-z0-9_\-\: ]+)\]\)', 
	r'^COLUMN_EXISTS\(\s*(?P<columnname>[A-Za-z0-9_\-\:]+)\s+IN\s+(?:"(?P<schema>[A-Za-z0-9_\-\: ]+)"\.)?"(?P<tablename>[A-Za-z0-9_\-\: ]+)"\)',
	r'^COLUMN_EXISTS\(\s*"(?P<columnname>[A-Za-z0-9_\-\: ]+)"\s+IN\s+(?:(?P<schema>[A-Za-z0-9_\-\: ]+)\.)?(?P<tablename>[A-Za-z0-9_\-\: ]+)\)',
	r'^COLUMN_EXISTS\(\s*"(?P<columnname>[A-Za-z0-9_\-\: ]+)"\s+IN\s+(?:\[(?P<schema>[A-Za-z0-9_\-\: ]+)\]\.)?\[(?P<tablename>[A-Za-z0-9_\-\: ]+)\]\)',
	r'^COLUMN_EXISTS\(\s*"(?P<columnname>[A-Za-z0-9_\-\: ]+)"\s+IN\s+(?:"(?P<schema>[A-Za-z0-9_\-\: ]+)"\.)?"(?P<tablename>[A-Za-z0-9_\-\: ]+)"\)'
	), xf_columnexists)

def xf_aliasdefined(**kwargs):
	alias = kwargs["alias"]
	return alias in dbs.aliases()

conditionallist.add(r'^\s*ALIAS_DEFINED\s*\(\s*(?P<alias>\w+)\s*\)', xf_aliasdefined)


def xf_metacommanderror(**kwargs):
	return status.metacommand_error

conditionallist.add(r'^\s*metacommand_error\(\s*\)', xf_metacommanderror)


def xcmd_test(teststr):
	result = CondParser(teststr).parse().eval()
	if result is not None:
		return result
	else:
		raise ErrInfo(type="cmd", command_text=teststr, other_msg="Unrecognized conditional")


#	End of conditional tests for metacommands.
#===============================================================================================




#===============================================================================================
#-----  SUPPORT FUNCTIONS (2)

def set_system_vars():
	# (Re)define the system substitution variables that are not script-specific.
	global subvars
	subvars.add_substitution("$ERROR_HALT_STATE", "ON" if status.halt_on_err else "OFF")
	subvars.add_substitution("$METACOMMAND_ERROR_HALT_STATE", "ON" if status.halt_on_metacommand_err else "OFF")
	subvars.add_substitution("$CURRENT_TIME", datetime.datetime.now().strftime("%Y-%m-%d %H:%M"))
	subvars.add_substitution("$CURRENT_DIR", os.path.abspath(os.path.curdir))
	subvars.add_substitution("$CURRENT_PATH", os.path.abspath(os.path.curdir) + os.sep)
	subvars.add_substitution("$AUTOCOMMIT_STATE", "ON" if dbs.current().autocommit else "OFF")
	subvars.add_substitution("$RANDOM", str(random.random()))
	subvars.add_substitution("$UUID", str(uuid.uuid4()))


def substitute_vars(command_str, localvars = None):
	# Substitutes global variables, global counters, and local variables
	# into the command string until no more substitutions can be made.
	# Returns the modified command_str.
	global subvars
	global counters
	if localvars is not None:
		subs = subvars.merge(localvars)
	else:
		subs = subvars
	cmdstr = copy.copy(command_str)
	# Substitute variables and counters until no more substitutions are made.
	subs_made = True
	while subs_made:
		subs_made = False
		cmdstr, subs_made = subs.substitute_all(cmdstr)
		cmdstr, any_subbed = counters.substitute_all(cmdstr)
		subs_made = subs_made or any_subbed
	m = defer_rx.findall(cmdstr)
    # Substitute any deferred substitution variables with regular substition var flags, e.g.: "!!somevar!!"
	if m is not None:
		for dv in m:
			rep = "!!" +  dv[1] + "!!"
			cmdstr = cmdstr.replace(dv[0], rep)
	return cmdstr


def current_script_line():
	if len(commandliststack) > 0:
		current_cmds = commandliststack[-1]
		if current_cmds.current_command() is not None:
			return current_cmds.current_command().current_script_line()
		else:
			return len(current_cmds.cmdlist)
	else:
		return 0


def wo_quotes(argstr):
	# Strip first and last quotes off an argument.
	argstr = argstr.strip()
	if argstr[0]=='"' and argstr[-1]=='"' or argstr[0]=="'" and argstr[-1]=="'" or argstr[0]=="[" and argstr[-1]=="]":
		return argstr[1:-1]
	return argstr


def get_subvarset(varname, metacommandline):
	# Supports the exec functions for the substitution metacommands that allow
	# substitution variables with a "+" prefix, to reference outer scope local
	# variables
	subvarset = None
	# Outer scope variable
	if varname[0] == '+':
		varname = re.sub('^[+]', '~', varname)
		for cl in reversed(commandliststack[0:-1]):
			if cl.localvars.sub_exists(varname):
				subvarset = cl.localvars
				break
		# Raise error if local variable not found anywhere down in commandliststack
		if not subvarset:
			raise ErrInfo(type="cmd", command_text=metacommandline, other_msg="Outer-scope referent variable (%s) has no matching local variable (%s)." % (re.sub('^[~]', '+', varname), varname)) # Global or local variable else:
	# Global or local variable
	else:
		subvarset = subvars if varname[0] != '~' else commandliststack[-1].localvars
	return subvarset, varname

# End of support functions (2).
#===============================================================================================



#===============================================================================================
#-----  GLOBAL OBJECT INITIALIZATION FOR EXECSQL INTERPRETER

# Status object with status-related attributes.
status = StatObj()

# Stack of conditional levels to support IF metacommands.
if_stack = IfLevels()

# Global counter variables.
counters = CounterVars()

# Global substitution variables.  (There may also be SCRIPT-specific
# substitution variables used as parameters.)
subvars = SubVarSet()
for k in os.environ.keys():
	try:
		subvars.add_substitution(u"&"+k, os.environ[k])
	except:
		# Ignore "ProgramFiles(x86)" on Windows and any others with invalid characters.
		pass
subvars.add_substitution("$LAST_ROWCOUNT", None)

dt_now = datetime.datetime.now()
subvars.add_substitution("$SCRIPT_START_TIME", dt_now.strftime("%Y-%m-%d %H:%M"))
subvars.add_substitution("$DATE_TAG", dt_now.strftime("%Y%m%d"))
subvars.add_substitution("$DATETIME_TAG", dt_now.strftime("%Y%m%d_%H%M"))
subvars.add_substitution("$LAST_SQL", "")
subvars.add_substitution("$LAST_ERROR", "")
subvars.add_substitution("$ERROR_MESSAGE", "")
subvars.add_substitution("$PATHSEP", os.sep)
osys = sys.platform
if osys.startswith('linux'):
	osys = 'linux'
elif osys.startswith('win'):
	osys = 'windows'
subvars.add_substitution("$OS", osys)

conf = ConfigData()

# Storage for all the (named) databases that are opened.  Databases are added in 'main()'
# and by the CONNECT metacommand.
dbs = DatabasePool()

#	End of global object initialization for execsql interpreter.
#===============================================================================================



def process_sql(sql_commands):
	# Read lines from the list of SQL commands, create a list of ScriptCmd objects,
	# and append the list to the top of the stack of script commands.
	# The filename (fn) and line number are stored with each command.
	# Arguments:
	#    sql_file_name:  The name of the execql script to read and store.
	# Return value:
	#    No return value.
	# Side effects:
	#    1. The script that is read is appended to the global 'commandliststack'.
	#    2. Items may be added to the global 'savedscripts' if there are any
	#       BEGIN/END SCRIPT commands in the file.
	#
	# Lines containing execsql command statements must begin with "-- !x!"
	# Currently this routine knows only three things about SQL:
	#	1. Lines that start with "--" are comments.
	#	2. Lines that end with ";" terminate a SQL statement.'
	#	3. Lines that start with "/*" begin a block comment, and lines that
	#		end with "*/" end a block comment.
	# The following metacommands are executed IMMEDIATELY during this process:
	#	* BEGIN SCRIPT <scriptname>
	#	* END SCRIPT
	#	* BEGIN SQL
	#	* END SQL
	#
	# May update the script_errors global list.
	#
	# Returns True if there are no errors or only warnings.  Returns False
	# if there are any fatal errors.
	beginscript = re.compile(r'^\s*--\s*!x!\s*(?:BEGIN|CREATE)\s+SCRIPT\s+(?P<scriptname>\w+)(?:(?P<paramexpr>\s*\S+.*))?$', re.I)
	endscript = re.compile(r'^\s*--\s*!x!\s*END\s+SCRIPT(?:\s+(?P<scriptname>\w+))?\s*$', re.I)
	beginsql = re.compile(r'^\s*--\s*!x!\s*BEGIN\s+SQL\s*$', re.I)
	endsql = re.compile(r'^\s*--\s*!x!\s*END\s+SQL\s*$', re.I)
	execline = re.compile(r'^\s*--\s*!x!\s*(?P<cmd>.+)$', re.I)
	cmtline = re.compile(r'^\s*--')
	in_block_cmt = False
	in_block_sql = False
	sqllist = []
	sqlline = 0
	subscript_stack = []
	sql_file_name = ""
	scriptfilename = ""
	file_lineno = 0
	currcmd = ''
	for line in sql_commands:
		file_lineno += 1
		# Remove trailing whitespace but not leading whitespace; this may be a plpythonu command in Postgres, where leading whitespace is significant.
		line = line.rstrip()
		is_comment_line = False
		comment_match = cmtline.match(line)
		metacommand_match = execline.match(line)
		if len(line) > 0:
			if in_block_cmt:
				is_comment_line = True
				if len(line) > 1 and line[-2:] == u"*/":
					in_block_cmt = False
			else:
				# Not in block comment
				if len(line.strip()) > 1 and line.strip()[0:2] == u"/*":
					in_block_cmt = True
					is_comment_line = True
					if line.strip()[-2:] == u"*/":
						in_block_cmt = False
				else:
					if comment_match:
						is_comment_line = not metacommand_match
			if not is_comment_line:
				if metacommand_match:
					if beginsql.match(line):
						in_block_sql = True
					if in_block_sql:
						if endsql.match(line):
							in_block_sql = False
							if len(currcmd) > 0:
								cmd = ScriptCmd(sql_file_name, sqlline, 'sql', SqlStmt(currcmd))
								if len(subscript_stack) == 0:
									sqllist.append(cmd)
								else:
									subscript_stack[-1].add(cmd)
								currcmd = ''
					else:
						if len(currcmd) > 0:
							script_errors.append(["Incomplete SQL statement", sqlline])
						begs = beginscript.match(line)
						if not begs:
							ends = endscript.match(line)
						if begs:
							# This is a BEGIN SCRIPT metacommand.
							scriptname = begs.group('scriptname').lower()
							paramnames = None
							paramexpr = begs.group('paramexpr')
							if paramexpr:
								withparams = re.compile(r'(?:\s+WITH)?(?:\s+PARAM(?:ETER)?S)?\s*\(\s*(?P<params>\w+(?:\s*,\s*\w+)*)\s*\)\s*$', re.I)
								wp = withparams.match(paramexpr)
								if not wp:
									raise ErrInfo(type="cmd", command_text=line, other_msg="Invalid BEGIN SCRIPT metacommand on line %s of file %s." % (file_lineno, sql_file_name))
								else:
									param_rx = re.compile(r'\w+', re.I)
									paramnames = re.findall(param_rx, wp.group('params'))
							# If there are no parameter names to pass, paramnames will be None
							subscript_stack.append(CommandList([], scriptname, paramnames))
						elif ends:
							# This is an END SCRIPT metacommand.
							endscriptname = ends.group('scriptname')
							if endscriptname is not None:
								endscriptname = endscriptname.lower()
							if len(subscript_stack) == 0:
								raise ErrInfo(type="cmd", command_text=line, other_msg="Unmatched END SCRIPT metacommand on line %s of file %s." % (file_lineno, sql_file_name))
							if len(currcmd) > 0:
								raise ErrInfo(type="cmd", command_text=line, other_msg="Incomplete SQL statement\n  (%s)\nat END SCRIPT metacommand on line %s of file %s." % (currcmd, file_lineno, sql_file_name))
							if endscriptname is not None and endscriptname != scriptname:
								raise ErrInfo(type="cmd", command_text=line, other_msg="Mismatched script name in the END SCRIPT metacommand on line %s of file %s." % (file_lineno, sql_file_name))
							sub_script = subscript_stack.pop()
							savedscripts[sub_script.listname] = sub_script
						else:
							# This is a non-IMMEDIATE metacommand.
							cmd = ScriptCmd(sql_file_name, file_lineno, 'cmd', MetacommandStmt(metacommand_match.group('cmd').strip()))
							if len(subscript_stack) == 0:
								sqllist.append(cmd)
							else:
								subscript_stack[-1].add(cmd)
				else:
					# This line is not a comment and not a metacommand, therefore should be
					# part of a SQL statement.
					cmd_end = True if line[-1] == ';' else False
					if line[-1] == '\\':
						line = line[:-1].strip()
					if currcmd == '':
						sqlline = file_lineno
						currcmd = line
					else:
						currcmd = u"%s \n%s" % (currcmd, line)
					if cmd_end and not in_block_sql:
						cmd = ScriptCmd(sql_file_name, sqlline, 'sql', SqlStmt(currcmd.strip()))
						if len(subscript_stack) == 0:
							sqllist.append(cmd)
						else:
							subscript_stack[-1].add(cmd)
						currcmd = ''
	if len(subscript_stack) > 0:
		raise ErrInfo(type="error", other_msg="Unmatched BEGIN SCRIPT metacommand at end of file %s." % sql_file_name)
	if len(currcmd) > 0:
		raise ErrInfo(type="error", other_msg="Incomplete SQL statement starting on line %s at end of file %s." % (sqlline, sql_file_name))
	if len(sqllist) > 0:
		# The file might be all comments or just a BEGIN/END SCRIPT metacommand.
		commandliststack.append(CommandList(sqllist, scriptfilename))

def read_sqlfile(sql_file_name):
	with io.open(sql_file_name, "r") as f:
		script_cmds = f.read()
	process_sql(script_cmds.splitlines())


def runscripts(db, sql_text):
	# Repeatedly run the next statement from the script at the top of the
	# command list stack until there are no more statements.
	# Metacommands may modify the stack or the commands in a stack entry.
	# This is the central script processing routine for execsql script extensions.
	global commandliststack
	global cmds_run
	global script_errors
	dbs.add('mapdata_connection', db)
	subvars.add_substitution("$CURRENT_DBMS", db.type.dbms_id)
	subvars.add_substitution("$CURRENT_DATABASE", db.name())
	process_sql(sql_text.splitlines())
	if len(script_errors) > 0:
		dlg = MsgDialog2("Warnings", "The following unexpected conditions were encountered while parsing the SQL script.", can_resize=True)
		hdrs = ["Warning message", "Line no"]
		tframe, tdata = treeview_table(dlg.content_frame, script_errors, hdrs)
		tframe.grid(row=0, column=0, sticky=tk.NSEW)
		dlg.show(grab=True)
	script_errors = []
	while len(commandliststack) > 0:
		current_cmds = commandliststack[-1]
		set_system_vars()
		try:
			current_cmds.run_next()
		except StopIteration:
			commandliststack.pop()
		except ErrInfo:
			commandliststack.pop()
			raise
		except:
			commandliststack.pop()
			raise ErrInfo(type="exception", exception_msg=exception_desc())
		cmds_run += 1
	if len(script_errors) > 0:
		dlg = MsgDialog2("Warnings", "The following unexpected conditions were encountered while running the SQL script.", can_resize=True)
		hdrs = ["Warning message", "Line no"]
		tframe, tdata = treeview_table(dlg.content_frame, script_errors, hdrs)
		tframe.grid(row=0, column=0, sticky=tk.NSEW)
		dlg.show(grab=True)


#***************************************************************************************************
#***************************  End of SQL Scripting Extensions  *************************************
#***************************************************************************************************




class DbConnectDialog(object):
	FILE, SERVER, FILE_PW = range(3)
	def __init__(self, parent, mapui):
		self.parent = parent
		self.mapui = mapui
		self.exit_status = 0	# Canceled
		self.exit_svr = None	# For caller
		self.exit_db = None	# For caller
		self.xpos = None
		self.ypos = None
		self.scriptfilepath = None
		# Values of db_params indicate whether server information is needed.
		self.db_params = {u"PostgreSQL": self.SERVER, u"SQLite": self.FILE, u"DuckDB": self.FILE,
							u"SQL Server": self.SERVER, u"MySQL": self.SERVER, u"Firebird": self.SERVER,
							u"MariaDB": self.SERVER, u"Oracle": self.SERVER}
		self.dlg = tk.Toplevel(parent)
		self.title = "Database Table for Map Display"
		self.dlg.title(self.title)
		self.dlg.protocol("WM_DELETE_WINDOW", self.do_cancel)
		self.dlg.geometry("650x260")
		self.headers = None
		self.header_list = None
		self.datarows = None

		# Main frames
		msgframe = tk.Frame(self.dlg)
		msgframe.grid(column=0, row=0, padx=6, pady=2, sticky=tk.EW)
		# Database selection is in one wizard pane, table and script selection are in a second, and column selection is in a third wizard pane
		wiz1_frame = tk.Frame(self.dlg)
		wiz1_frame.grid(column=0, row=1, sticky=tk.NSEW)
		wiz2_frame = tk.Frame(self.dlg)
		wiz2_frame.grid(column=0, row=1, sticky=tk.NSEW)
		wiz3_frame = tk.Frame(self.dlg)
		wiz3_frame.grid(column=0, row=1, sticky=tk.NSEW)
		self.dlg.rowconfigure(0, weight=0)
		self.dlg.rowconfigure(1, weight=1)
		self.dlg.columnconfigure(0, weight=1)
		wiz1_frame.rowconfigure(0, weight=1)
		wiz1_frame.columnconfigure(0, weight=1)
		wiz2_frame.rowconfigure(0, weight=1)
		wiz2_frame.columnconfigure(0, weight=1)
		wiz3_frame.rowconfigure(0, weight=1)
		wiz3_frame.columnconfigure(0, weight=1)

		# Populate message frame
		msg_label = ttk.Label(msgframe, text="The database, table, and columns to be used for mapping must be specified.", anchor=tk.W, justify=tk.LEFT, wraplength=500)
		msg_label.grid(column=0, row=0, sticky=tk.EW)


		# Wizard page 1
		# Database selector
		# On the left side will be a combobox to choose the database type.
		# On the right side will be a prompt for the server, db, user name, and pw,
		# or for the filename (and possibly user name and pw).  Each of these alternative
		# types of prompts will be in its own frame, which will be in the same place.
		# Only one will be shown, controlled by the item in the self.db_params dictionary.
		# A separate frame for the table name is below the database parameters frame.
		dbframe = tk.Frame(wiz1_frame)
		dbtypeframe = tk.Frame(dbframe)
		rightframe = tk.Frame(dbframe)
		paramframe = tk.Frame(rightframe)
		self.serverparamframe = tk.Frame(paramframe)
		self.fileparamframe = tk.Frame(paramframe)
		self.filepwparamframe = tk.Frame(paramframe)
		w1btnframe = tk.Frame(wiz1_frame, borderwidth=3, relief=tk.RIDGE)

		# Grid wiz1 frame widgets
		def param_choices(*args, **kwargs):
			svr_params = self.db_params[self.db_type_var.get()]
			if svr_params == self.SERVER:
				self.fileparamframe.grid_remove()
				self.filepwparamframe.grid_remove()
				self.serverparamframe.grid()
			elif svr_params == self.FILE_PW:
				self.serverparamframe.grid_remove()
				self.fileparamframe.grid_remove()
				self.filepwparamframe.grid()
			else:
				self.serverparamframe.grid_remove()
				self.filepwparamframe.grid_remove()
				self.fileparamframe.grid()
			check_w1enable()

		def check_w1enable(*args):
			dbms = self.db_type_var.get()
			if dbms != '':
				dbtype = self.db_params[dbms]
				if dbtype == self.SERVER:
					if self.server.get() != '' and self.db.get != '':
						w1next_btn["state"] = tk.NORMAL
					else:
						w1next_btn["state"] = tk.DISABLED
				elif dbtype == self.FILE_PW:
					if self.db_file.get() != '' and self.user.get() != '' and self.pw.get() != '':
						w1next_btn["state"] = tk.NORMAL
					else:
						w1next_btn["state"] = tk.DISABLED
				else:	# self.FILE
					if self.db_file.get() != '':
						w1next_btn["state"] = tk.NORMAL
					else:
						w1next_btn["state"] = tk.DISABLED

		dbframe.grid(column=0, row=0, sticky=tk.NSEW)
		dbtypeframe.grid(column=0, row=0, padx=5, sticky=tk.NW)
		rightframe.grid(column=1, row=0, padx=5, sticky=tk.N + tk.EW)
		paramframe.grid(column=0, row=0, padx=5, sticky=tk.N + tk.EW)
		# Put serverparamframe, fileparamframe, and filepwparamframe in the same place in paramframe.
		# Leave only serverparamframe visible.
		self.fileparamframe.grid(row=0, column=0, sticky=tk.N + tk.EW)
		self.fileparamframe.grid_remove()
		self.filepwparamframe.grid(row=0, column=0, sticky=tk.N + tk.EW)
		self.filepwparamframe.grid_remove()
		self.serverparamframe.grid(row=0, column=0, sticky=tk.N + tk.EW)
		w1btnframe.grid(column=0, row=2, sticky=tk.S+tk.EW)
		w1btnframe.columnconfigure(0, weight=1)

		# Populate dbframe
		self.db_type_var = tk.StringVar()
		self.encoding = tk.StringVar()
		self.table_var = tk.StringVar()
		self.table_var.trace('w', check_w1enable)
		# Database type selection
		ttk.Label(dbtypeframe, text="DBMS:").grid(column=0, row=0, padx=3, pady=3, sticky=tk.NE)
		dbmss = [k for k in self.db_params.keys()]
		dbmss.sort()
		self.db_choices = ttk.Combobox(dbtypeframe, textvariable=self.db_type_var, width=12,
						values=dbmss)
		self.db_choices.bind("<<ComboboxSelected>>", param_choices)
		self.db_choices.config(state='readonly')
		self.db_choices.grid(column=1, row=0, padx=3, pady=3, sticky=tk.NW)
		self.db_choices.focus()
		ttk.Label(dbtypeframe, text="Encoding:").grid(column=0, row=1, padx=3, pady=3, sticky=tk.NE)
		self.db_choices.set('PostgreSQL')
		enc_choices = ttk.Combobox(dbtypeframe, textvariable=self.encoding, width=12,
						values=('UTF8', 'Latin1', 'Win1252'))
		enc_choices.set('UTF8')
		enc_choices.grid(column=1, row=1, padx=3, pady=3, sticky=tk.NW)
		# Database parameter entry frames
		self.server = tk.StringVar()
		self.server.trace('w', check_w1enable)
		self.port = tk.StringVar()
		self.db = tk.StringVar()
		self.db.trace('w', check_w1enable)
		self.user = tk.StringVar()
		self.user.trace('w', check_w1enable)
		self.pw = tk.StringVar()
		self.pw.trace('w', check_w1enable)
		self.db_file = tk.StringVar()
		self.db_file.trace('w', check_w1enable)

		# Server databases
		ttk.Label(self.serverparamframe, text="Server:").grid(column=0, row=0, padx=3, pady=3, sticky=tk.E)
		ttk.Entry(self.serverparamframe, width=30, textvariable=self.server).grid(column=1, row=0, padx=3, pady=3, sticky=tk.W)
		ttk.Label(self.serverparamframe, text="Database:").grid(column=0, row=1, padx=3, pady=3, sticky=tk.E)
		ttk.Entry(self.serverparamframe, width=30, textvariable=self.db).grid(column=1, row=1, padx=3, pady=3, sticky=tk.W)
		ttk.Label(self.serverparamframe, text="User:").grid(column=0, row=2, padx=3, pady=3, sticky=tk.E)
		ttk.Entry(self.serverparamframe, width=30, textvariable=self.user).grid(column=1, row=2, padx=3, pady=3, sticky=tk.W)
		ttk.Label(self.serverparamframe, text="Password:").grid(column=0, row=3, padx=3, pady=3, sticky=tk.E)
		ttk.Entry(self.serverparamframe, width=30, textvariable=self.pw, show="*").grid(column=1, row=3, padx=3, pady=3, sticky=tk.W)
		ttk.Label(self.serverparamframe, text="Port:").grid(column=0, row=4, padx=3, pady=3, sticky=tk.E)
		ttk.Entry(self.serverparamframe, width=4, textvariable=self.port).grid(column=1, row=4, padx=3, pady=3, sticky=tk.W)

		# File databases
		ttk.Label(self.fileparamframe, text="Database file:").grid(column=0, row=0, padx=3, pady=3, sticky=tk.NW)
		ttk.Entry(self.fileparamframe, width=40, textvariable=self.db_file).grid(column=0, row=1, padx=3, pady=3, sticky=tk.NW)
		ttk.Button(self.fileparamframe, text="Browse...", command=self.set_sel_fn).grid(column=1, row=1)

		# File databases with user name and password
		ttk.Label(self.filepwparamframe, text="Database file:").grid(column=0, row=0, columnspan=2, padx=3, pady=3, sticky=tk.NW)
		ttk.Entry(self.filepwparamframe, width=40, textvariable=self.db_file).grid(column=0, row=1, columnspan=2, padx=3, pady=3, sticky=tk.NW)
		ttk.Button(self.filepwparamframe, text="Browse...", command=self.set_sel_fn).grid(column=2, row=1)
		ttk.Label(self.filepwparamframe, text="User:").grid(column=0, row=2, padx=3, pady=3, sticky=tk.E)
		ttk.Entry(self.filepwparamframe, width=30, textvariable=self.user).grid(column=1, row=2, padx=3, pady=3, sticky=tk.W)
		ttk.Label(self.filepwparamframe, text="Password:").grid(column=0, row=3, padx=3, pady=3, sticky=tk.E)
		ttk.Entry(self.filepwparamframe, width=30, textvariable=self.pw, show="*").grid(column=1, row=3, padx=3, pady=3, sticky=tk.W)

		# Put serverparamframe, fileparamframe, and filepwparamframe in the same place in paramframe
		self.fileparamframe.grid(row=0, column=0, sticky=tk.NW)
		self.fileparamframe.grid_remove()
		self.filepwparamframe.grid(row=0, column=0, sticky=tk.NW)
		self.filepwparamframe.grid_remove()
		self.serverparamframe.grid(row=0, column=0, sticky=tk.NW)
		self.db_type_var.set(u"PostgreSQL")

		def w1_next(*args):
			self.dlg.bind("<Alt-p>", load_script)
			self.dlg.bind("<Alt-s>", save_script)
			self.dlg.bind("<Alt-e>", edit_sql)
			wiz2_frame.lift()
			# The following conditional fails
			#if w1next_btn["state"] == tk.NORMAL:
			#	wiz2_frame.lift()
			self.dlg.bind("<Alt-n>")
			self.dlg.bind("<Alt-n>", w2_next)
			self.dlg.bind("<Alt-b>")
			self.dlg.bind("<Alt-b>", w2_back)
			self.table_entry.focus()

		# Populate w1btnframe
		w1help_btn = ttk.Button(w1btnframe, text="Help", command=self.do_help, underline=0)
		w1help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		w1next_btn = ttk.Button(w1btnframe, text="Next", command=w1_next, underline=0)
		w1next_btn.grid(row=0, column=1, sticky=tk.E, padx=3)
		w1cancel_btn = ttk.Button(w1btnframe, text="Cancel", command=self.do_cancel, underline=0)
		w1cancel_btn.grid(row=0, column=2, sticky=tk.E, padx=(3,6))
		w1next_btn["state"] = tk.DISABLED
		self.dlg.bind("<Alt-n>", w1_next)
		self.dlg.bind("<Alt-c>", self.do_cancel)
		self.dlg.bind("<Escape>", self.do_cancel)
		

		# Wizard page 2
		# Database table and optional query
		def check_w2enable(*args):
			if self.table_var.get() != '':
				w2next_btn["state"] = tk.NORMAL
			else:
				w2next_btn["state"] = tk.DISABLED
		def w2_back(*args):
			self.dlg.unbind("<Alt-p>")
			self.dlg.unbind("<Alt-s>")
			self.dlg.unbind("<Alt-e>")
			wiz1_frame.lift()
			self.dlg.bind("<Alt-n>", w1_next)
			self.dlg.bind("<Alt-b>")
			self.db_choices.focus()
		def w2_next(*args):
			if self.table_var.get() != '':
				self.mapui.loading_dlg.display("Querying database")
				sql = "select * from %s;" % self.table_var.get()
				#conn = None
				db = None
				# Open database, get table data and column headers, populate wiz2 comboboxes
				dbms = self.db_type_var.get()
				if dbms == 'PostgreSQL':
					need_pw = self.pw.get() != ''
					user = self.user.get() if self.user.get() != '' else None
					pw = self.pw.get() if self.pw.get() != '' else None
					port = 5432 if self.port.get() == '' else int(self.port.get())
					try:
						db = PostgresDatabase(self.server.get(), self.db.get(), user, need_pw, port, password=pw)
					except ErrInfo as e:
						warning(e.eval_err(), kwargs={'parent': self.dlg})
						db = None
					except:
						warning("Cannot open the Postgres database.", kwargs={'parent': self.dlg})
						db = None

				elif dbms == 'SQLite':
					try:
						db = SQLiteDatabase(self.db_file.get())
					except ErrInfo as e:
						warning(e.eval_err(), kwargs={'parent': self.dlg})
						db = None
					except:
						warning("Cannot open the file %s as a SQLite database." % self.db_file.get(), kwargs={'parent': self.dlg})
						db = None

				elif dbms == 'DuckDB':
					try:
						db = DuckDBDatabase(self.db_file.get())
					except ErrInfo as e:
						warning(e.eval_err(), kwargs={'parent': self.dlg})
						db = None
					except:
						warning("Cannot open the file %s as a DuckDB database." % self.db_file.get(), kwargs={'parent': self.dlg})
						db = None

				elif dbms == 'MariaDB' or dbms == 'MySQL':
					need_pw = self.pw.get() != ''
					user = self.user.get() if self.user.get() != '' else None
					pw = self.pw.get() if self.pw.get() != '' else None
					port = 3306 if self.port.get() == '' else int(self.port.get())
					try:
						db = MySQLDatabase(self.server.get(), self.db.get(), user, need_pw, port, password=pw)
					except ErrInfo as e:
						warning(e.eval_err(), kwargs={'parent': self.dlg})
						db = None
					except:
						warning("Cannot open the MariaDB/MySQL database.", kwargs={'parent': self.dlg})
						db = None

				elif dbms == 'SQL Server':
					need_pw = self.pw.get() != ''
					user = self.user.get() if self.user.get() != '' else None
					pw = self.pw.get() if self.pw.get() != '' else None
					port = 1433 if self.port.get() == '' else int(self.port.get())
					try:
						db = SqlServerDatabase(self.server.get(), self.db.get(), user, need_pw, port, password=pw)
					except ErrInfo as e:
						warning(e.eval_err(), kwargs={'parent': self.dlg})
						db = None
					except:
						warning("Cannot open the SQL Server database.", kwargs={'parent': self.dlg})
						db = None

				elif dbms == 'Oracle':
					need_pw = self.pw.get() != ''
					user = self.user.get() if self.user.get() != '' else None
					pw = self.pw.get() if self.pw.get() != '' else None
					port = 1521 if self.port.get() == '' else int(self.port.get())
					try:
						db = SqlServerDatabase(self.server.get(), self.db.get(), user, need_pw, port, password=pw)
					except ErrInfo as e:
						warning(e.eval_err(), kwargs={'parent': self.dlg})
						db = None
					except:
						warning("Cannot open the Oracle database.", kwargs={'parent': self.dlg})
						db = None

				elif dbms == 'Firebird':
					need_pw = self.pw.get() != ''
					user = self.user.get() if self.user.get() != '' else None
					pw = self.pw.get() if self.pw.get() != '' else None
					port = 3050 if self.port.get == '' else int(self.port.get())
					try:
						db = FirebirdDatabase(self.server.get(), self.db.get(), user, need_pw, port, password=pw)
					except ErrInfo as e:
						warning(e.eval_err(), kwargs={'parent': self.dlg})
						db = None
					except:
						warning("Cannot open the Firebird database.", kwargs={'parent': self.dlg})
						db = None

				else:
					warning("Unrecognized DBMS type", kwargs={'parent': self.dlg})

				if db is not None:
					script_error = False
					script_text = self.script_text.get("1.0", "end")
					if len(script_text) > 0:
						try:
							runscripts(db, script_text)
						except ErrInfo as e:
							warning(e.eval_err(), kwargs={'parent':self.dlg})
							script_error = True
						except:
							script_error = True
							raise
					if not script_error:
						try:
							self.headers, self.rows = db.select_data(sql)
						except ErrInfo as e:
							warning(e.eval_err(), kwargs={'parent': self.dlg})
						except:
							warning("Cannot select data from table %s." % self.table_var.get(), kwargs={'parent':self.dlg})
						else:
							self.header_list = list(self.headers)
							# Set list box values
							self.id_sel["values"] = self.header_list
							self.lat_sel["values"] = self.header_list
							self.lon_sel["values"] = self.header_list
							self.sym_sel["values"] = self.header_list
							self.col_sel["values"] = self.header_list
							self.dlg.unbind("<Alt-p>")
							self.dlg.unbind("<Alt-s>")
							self.dlg.unbind("Alt-e>")
							wiz3_frame.lift()
							self.dlg.bind("<Alt-n>")
							self.dlg.bind("<Alt-b>")
							self.dlg.bind("<Alt-b>", w3_back)
							self.dlg.bind('<Alt-o>', self.do_select)
					db.close()
				self.mapui.loading_dlg.hide()

		def load_script(*args):
			fn = tkfiledialog.askopenfilename(parent=self.dlg, title="SQL script file to open", filetypes=([('SQL script files', '.sql')]))
			if not (fn is None or fn == '' or fn == ()):
				path, filename = os.path.split(os.path.abspath(fn))
				self.scriptfilepath = path
				with open(fn, "r") as f:
					sql = f.read()
				self.script_text.insert("end", sql)
		def save_script(*args):
			outfile = tkfiledialog.asksaveasfilename(initialdir=self.scriptfilepath, parent=self.dlg, title="SQL script file to save", filetypes=[('SQL script files', '.sql')])
			if not (outfile is None or outfile == ''):
				sql = self.script_text.get("1.0", "end")
				with open(outfile, "w") as f:
					f.write(sql)
		def edit_sql(*args):
			td = tempfile.TemporaryDirectory()
			edit_fn = os.path.join(td.name, "mapfile_temp.sql")
			with open(edit_fn, "w") as f:
				f.write(self.script_text.get("1.0", "end"))
			returncode = subprocess.call([editor, edit_fn])
			if returncode == 0:
				with open(edit_fn, "r") as f:
					sql = f.read()
				self.script_text.delete("1.0", "end")
				self.script_text.insert("end", sql)
			else:
				warning("Failure attempting to edit the SQL with %s" % editor, kwargs={'parent':self.dlg})

		w2req_frame = ttk.LabelFrame(wiz2_frame, text="Required")
		w2req_frame.grid(row=0, column=0, sticky=tk.EW+tk.N, padx=(6,6), pady=(3,3))
		w2req_frame.columnconfigure(1, weight=1)
		#
		tbl_label = ttk.Label(w2req_frame, text="Table:")
		tbl_label.grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.table_var = tk.StringVar(w2req_frame, '')
		self.table_var.trace('w', check_w2enable)
		self.table_entry = ttk.Entry(w2req_frame, width=30, textvariable=self.table_var)
		self.table_entry.grid(row=0, column=1, padx=(3,6), pady=3, sticky=tk.W)
		#
		w2opt_frame = ttk.LabelFrame(wiz2_frame, text="Optional")
		w2opt_frame.grid(row=1, column=0, sticky=tk.EW+tk.N, padx=(6,6), pady=(3,3))
		w2opt_frame.columnconfigure(1, weight=1)
		#
		self.script_text = tk.Text(w2opt_frame, width=40, height=4)
		self.script_text.grid(row=0, column=1, columnspan=2, rowspan=4, sticky=tk.NSEW, padx=(3,0), pady=(3,3))
		scr_label = ttk.Label(w2opt_frame, text="Script:")
		scr_label.grid(row=0, column=0, sticky=tk.NE, padx=(6,3), pady=(2,2))
		load_btn = ttk.Button(w2opt_frame, text="Open", command=load_script, underline=1)
		load_btn.grid(row=1, column=0, sticky=tk.E, padx=(6,3))
		save_btn = ttk.Button(w2opt_frame, text="Save", command=save_script, underline=0)
		save_btn.grid(row=2, column=0, sticky=tk.E, padx=(3,3))
		edit_btn = ttk.Button(w2opt_frame, text="Edit", command=edit_sql, underline=0)
		edit_btn.grid(row=3, column=0, sticky=tk.E, padx=(3,3), pady=(0,2))
		if editor is None:
			edit_btn["state"] = tk.DISABLED
		else:
			edit_btn["state"] = tk.NORMAL
		sbar = tk.Scrollbar(w2opt_frame)
		sbar.grid(row=0, column=2, rowspan=4, sticky=tk.NS, padx=(0,3), pady=(3,3))
		sbar.config(command=self.script_text.yview)
		self.script_text.config(yscrollcommand = sbar.set)
		#
		w2btn_frame = tk.Frame(wiz2_frame, borderwidth=3, relief=tk.RIDGE)
		w2btn_frame.grid(row=2, column=0, sticky=tk.S+tk.EW, padx=(3,3), pady=(2,2))
		w2btn_frame.columnconfigure(0, weight=1)
		#
		w2help_btn = ttk.Button(w2btn_frame, text="Help", command=self.do_help, underline=0)
		w2help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		w2prev_btn = ttk.Button(w2btn_frame, text="Back", command=w2_back, underline=0)
		w2prev_btn.grid(row=0, column=1, sticky=tk.E, padx=3)
		w2next_btn = ttk.Button(w2btn_frame, text="Next", command=w2_next, underline=0)
		w2next_btn.grid(row=0, column=2, sticky=tk.E, padx=3)
		w2cancel_btn = ttk.Button(w2btn_frame, text="Cancel", command=self.do_cancel, underline=0)
		w2cancel_btn.grid(row=0, column=3, sticky=tk.E, padx=(3,6))
		w2next_btn["state"] = tk.DISABLED


		# Wizard page 3
		# Column selectors
		def check_w3enable(*args):
			if self.lat_var.get() != '' and self.lon_var.get() != '':
				w3ok_btn["state"] = tk.NORMAL
			else:
				w3ok_btn["state"] = tk.DISABLED
		w3req_frame = ttk.LabelFrame(wiz3_frame, text="Required")
		w3req_frame.grid(row=0, column=0, sticky=tk.EW, padx=(6,6), pady=(3,3))
		w3req_frame.columnconfigure(0, weight=1)
		#
		lat_label = ttk.Label(w3req_frame, text="Latitude column:")
		lat_label.grid(row=0, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.lat_var = tk.StringVar(w3req_frame, '')
		self.lat_var.trace('w', check_w3enable)
		self.lat_sel = ttk.Combobox(w3req_frame, state="readonly", textvariable=self.lat_var, values=self.header_list, width=15)
		self.lat_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,30), pady=(3,3))
		#
		lon_label = ttk.Label(w3req_frame, text="Longitude column:")
		lon_label.grid(row=0, column=2, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.lon_var = tk.StringVar(w3req_frame, '')
		self.lon_var.trace('w', check_w3enable)
		self.lon_sel = ttk.Combobox(w3req_frame, state="readonly", textvariable=self.lon_var, values=self.header_list, width=15)
		self.lon_sel.grid(row=0, column=3, sticky=tk.W, padx=(3,6), pady=(3,3))

		w3opt_frame = ttk.LabelFrame(wiz3_frame, text="Optional")
		w3opt_frame.grid(row=1, column=0, sticky=tk.EW, padx=(6,6), pady=(6,3))
		w3opt_frame.columnconfigure(0, weight=1)
		#
		id_label = ttk.Label(w3opt_frame, text="Label column:")
		id_label.grid(row=0, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.id_var = tk.StringVar(w3opt_frame, '')
		self.id_sel = ttk.Combobox(w3opt_frame, state="readonly", textvariable=self.id_var, values=self.header_list, width=12)
		self.id_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,30), pady=(3,3))
		#
		crs_label = ttk.Label(w3opt_frame, text="CRS:")
		crs_label.grid(row=0, column=2, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.crs_var = tk.IntVar(w3opt_frame, 4326)
		self.crs_var.trace('w', check_w3enable)
		self.crs_sel = ttk.Entry(w3opt_frame, width=8, textvariable=self.crs_var)
		self.crs_sel.grid(row=0, column=3, sticky=tk.W, padx=(3,6), pady=(3,3))
		#
		sym_label = ttk.Label(w3opt_frame, text="Symbol column:")
		sym_label.grid(row=1, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.sym_var = tk.StringVar(w3opt_frame, '')
		self.sym_sel = ttk.Combobox(w3opt_frame, state="readonly", textvariable=self.sym_var, values=self.header_list, width=12)
		self.sym_sel.grid(row=1, column=1, sticky=tk.W, padx=(3,30), pady=(3,3))
		#
		col_label = ttk.Label(w3opt_frame, text="Color column:")
		col_label.grid(row=1, column=2, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.col_var = tk.StringVar(w3opt_frame, '')
		self.col_sel = ttk.Combobox(w3opt_frame, state="readonly", textvariable=self.col_var, values=self.header_list, width=12)
		self.col_sel.grid(row=1, column=3, sticky=tk.W, padx=(3,6), pady=(3,3))
		#
		desc_label = ttk.Label(w3opt_frame, text="Description:")
		desc_label.grid(row=2, column=0, sticky=tk.E, padx=(3,3), pady=(3,6))
		self.desc_var = tk.StringVar(w3opt_frame, '')
		desc_entry = ttk.Entry(w3opt_frame, width=60, textvariable=self.desc_var)
		desc_entry.grid(row=2, column=1, columnspan=3, sticky=tk.W, padx=(3,3), pady=(3,6))

		def w3_back(*args):
			self.dlg.bind("<Alt-p>", load_script)
			self.dlg.bind("<Alt-s>", save_script)
			self.dlg.bind("<Alt-e>", edit_sql)
			self.dlg.bind("<Alt-o>")
			wiz2_frame.lift()
			self.dlg.bind("<Alt-n>", w2_next)
			self.dlg.bind("<Alt-b>")
			self.dlg.bind("<Alt-b>", w2_back)
			self.table_entry.focus()

		w3btn_frame = tk.Frame(wiz3_frame, borderwidth=3, relief=tk.RIDGE)
		w3btn_frame.grid(row=2, column=0, sticky=tk.S+tk.EW, padx=(3,3), pady=(3,3))
		w3btn_frame.columnconfigure(0, weight=1)
		#
		w3help_btn = ttk.Button(w3btn_frame, text="Help", command=self.do_help, underline=0)
		w3help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		w3prev_btn = ttk.Button(w3btn_frame, text="Back", command=w3_back, underline=0)
		w3prev_btn.grid(row=0, column=1, sticky=tk.E, padx=3)
		w3ok_btn = ttk.Button(w3btn_frame, text="OK", command=self.do_select, underline=0)
		w3ok_btn.grid(row=0, column=2, sticky=tk.E, padx=3)
		w3cancel_btn = ttk.Button(w3btn_frame, text="Cancel", command=self.do_cancel, underline=0)
		w3cancel_btn.grid(row=0, column=3, sticky=tk.E, padx=(3,6))
		w3ok_btn["state"] = tk.DISABLED

		wiz1_frame.lift()

		self.canceled = True
		# Limit resizing
		self.dlg.resizable(False, False)
		center_window(self.dlg)
	def set_sel_fn(self):
		fn = tkfiledialog.askopenfilename(parent=self.fileparamframe, title=self.title)
		if fn is not None and fn != '' and fn != ():
			self.db_file.set(fn)
			#self.clearstatus()
	def do_select(self, *args):
		if self.table_var.get() != '' and self.lat_var.get() != '' and self.lon_var.get() != '':
			self.canceled = False
			self.dlg.destroy()
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def do_help(self, *args):
		webbrowser.open("https://mapdata.readthedocs.io/en/latest/dialogs.html#open-database-data-source", new=2, autoraise=True)
	def get_data(self):
		self.dlg.grab_set()
		self.dlg.focus_force()
		self.db_choices.focus()
		self.dlg.wait_window(self.dlg)
		if self.canceled:
			return (None, None, None, None, None, None, None, None, None, None)
		else:
			return self.table_var.get(), self.id_var.get(), self.lat_var.get(), self.lon_var.get(), \
					self.crs_var.get(), self.sym_var.get(), self.col_var.get(), self.desc_var.get(), \
					self.headers, self.rows





def read_all_config(datafile=None):
	global config_files
	config_files = []
	if os.name == 'posix':
		sys_config_file = os.path.join("/etc", config_file_name)
	else:
		sys_config_file = os.path.join(os.path.expandvars(r'%APPDIR%'), config_file_name)
	if os.path.isfile(sys_config_file):
		config_files.append(sys_config_file)
	program_dir_config = os.path.join(os.path.abspath(sys.argv[0]), config_file_name)
	if os.path.isfile(program_dir_config) and not program_dir_config in config_files:
		config_files.append(program_dir_config)
	user_config_file = os.path.join(os.path.expanduser(r'~/.config'), config_file_name)
	if os.path.isfile(user_config_file) and not user_config_file in config_files:
		config_files.append(user_config_file)
	if datafile is not None:
		data_config_file = os.path.join(os.path.abspath(datafile), config_file_name)
		if os.path.isfile(data_config_file) and not data_config_file in config_files:
			config_files.append(data_config_file)
	startdir_config_file = os.path.join(os.path.abspath(os.path.curdir), config_file_name)
	if os.path.isfile(startdir_config_file) and not startdir_config_file in config_files:
		config_files.append(startdir_config_file)
	files_read = []
	for config_file in config_files:
		files_read.append(config_file)
		read_config(config_file)


def read_config(configfile):
	_BASEMAP_SECTION = "basemap_tile_servers"
	_APIKEYS_SECTION = "api_keys"
	_SYMBOL_SECTION = "symbols"
	_DEFAULTS_SECTION = "defaults"
	_MISC_SECTION = "misc"
	cp = ConfigParser()
	cp.read(configfile)
	# Tile servers
	if cp.has_section(_BASEMAP_SECTION):
		basemap_sources = cp.items(_BASEMAP_SECTION)
		for name, url in basemap_sources:
			if url is None:
				if name in bm_servers and len(bm_servers) > 1:
					del(bm_servers[name])
			else:
				bm_servers[name.capitalize()] = url
	# API keys
	if cp.has_section(_APIKEYS_SECTION):
		apikeys = cp.items(_APIKEYS_SECTION)
		for name, apikey in apikeys:
			if apikey is None:
				if name in api_keys and len(api_keys) > 1:
					del(api_keys[name])
			else:
				api_keys[name.capitalize()] = apikey
	# Symbols
	if cp.has_section(_SYMBOL_SECTION):
		symbols = cp.items(_SYMBOL_SECTION)
		for name, filename in symbols:
			import_symbol(name, filename)
	# Defaults
	if cp.has_option(_DEFAULTS_SECTION, "multiselect"):
		global multiselect
		err = False
		try:
			multi = cp.getboolean(_DEFAULTS_SECTION, "multiselect")
		except:
			err = True
			warning("Invalid argument to the 'multiselect' configuration option", kwargs={})
		if not err:
			multiselect = "1" if multi else "0"
	if cp.has_option(_DEFAULTS_SECTION, "basemap"):
		global initial_basemap
		bm = cp.get(_DEFAULTS_SECTION, "basemap")
		if bm is None or bm not in bm_servers:
			warning("Invalid argument to the 'basemap' configuration option", kwargs={})
		else:
			initial_basemap = bm
	if cp.has_option(_DEFAULTS_SECTION, "location_marker"):
		global location_marker
		loc_mkr = cp.get(_DEFAULTS_SECTION, "location_marker")
		if loc_mkr is not None:
			location_marker = loc_mkr
	if cp.has_option(_DEFAULTS_SECTION, "location_color"):
		global location_color
		loc_color = cp.get(_DEFAULTS_SECTION, "location_color")
		if loc_color is not None:
			if loc_color not in color_names:
				warning("Invalid argument to the 'location_color' configuration option", kwargs={})
			else:
				location_color = loc_color
	if cp.has_option(_DEFAULTS_SECTION, "use_data_marker"):
		global use_data_marker
		loc_mkr = cp.getboolean(_DEFAULTS_SECTION, "use_data_marker")
		if loc_mkr is not None:
			use_data_marker = loc_mkr
	if cp.has_option(_DEFAULTS_SECTION, "use_data_color"):
		global use_data_color
		loc_clr = cp.getboolean(_DEFAULTS_SECTION, "use_data_color")
		if loc_clr is not None:
			use_data_color = loc_clr
	if cp.has_option(_DEFAULTS_SECTION, "select_symbol"):
		global select_symbol
		default_symbol = cp.get(_DEFAULTS_SECTION, "select_symbol")
		if default_symbol is not None:
			if default_symbol not in icon_xbm:
				warning("Unrecognized symbol name for the 'select_symbol' configuration option", kwargs={})
			else:
				select_symbol = default_symbol
	if cp.has_option(_DEFAULTS_SECTION, "select_color"):
		global select_color
		sel_color = cp.get(_DEFAULTS_SECTION, "select_color")
		if sel_color is not None:
			if sel_color not in color_names:
				warning("Invalid argument to the 'multiselect' configuration option", kwargs={})
			else:
				select_color = sel_color
	if cp.has_option(_DEFAULTS_SECTION, "label_color"):
		global label_color
		lbl_color = cp.get(_DEFAULTS_SECTION, "label_color")
		if lbl_color is not None:
			if lbl_color not in color_names:
				warning("Invalid argument to the 'label_color' configuration option", kwargs={})
			else:
				label_color = lbl_color
	if cp.has_option(_DEFAULTS_SECTION, "label_font"):
		global label_font
		lbl_font = cp.get(_DEFAULTS_SECTION, "label_font")
		if lbl_font is not None:
			if lbl_font not in list(tk.font.families()):
				warning("Invalid argument to the 'label_font' configuration option", kwargs={})
			else:
				label_font = lbl_font
	if cp.has_option(_DEFAULTS_SECTION, "label_size"):
		global label_size
		err = False
		try:
			lbl_size = cp.getint(_DEFAULTS_SECTION, "label_size")
		except:
			err = True
			warning("Invalid argument to the 'label_size' configuration option", kwargs={})
		if not err:
			if lbl_size is not None and lbl_size > 6:
				label_size = lbl_size
	if cp.has_option(_DEFAULTS_SECTION, "label_bold"):
		global label_bold
		err = False
		try:
			lbl_bold = cp.getboolean(_DEFAULTS_SECTION, "label_bold")
		except:
			err = True
			warning("Invalid argument to the 'label_bold' configuration option", kwargs={})
		if not err:
			if lbl_bold is not None:
				label_bold = lbl_bold
	if cp.has_option(_DEFAULTS_SECTION, "label_position"):
		global label_position
		lbl_position = cp.get(_DEFAULTS_SECTION, "label_position")
		if lbl_position is not None:
			lbl_position = lbl_position.lower()
			if lbl_position not in ("above", "below"):
				warning("Invalid argument to the 'label_position' configuration option", kwargs={})
			else:
				label_position = lbl_position
	if cp.has_option(_DEFAULTS_SECTION, "show_regression_stats"):
		global show_regression_stats
		err = False
		try:
			srs = cp.getboolean(_DEFAULTS_SECTION, "show_regression_stats")
		except:
			err = True
			warning("Invalid argument to the 'show_regression_stats' configuration option", kwargs={})
		if not err:
			show_regression_stats = srs
	if cp.has_option(_DEFAULTS_SECTION, "wrapwidth"):
		global wrapwidth
		err = False
		try:
			wr = cp.getint(_DEFAULTS_SECTION, "wrapwidth")
		except:
			err = True
			warning("Invalid argument to the 'wrapwidth' configuration option", kwargs={})
		if not err:
			wrapwidth = wr
	if cp.has_option(_MISC_SECTION, "temp_dbfile"):
		global temp_dbfile
		err = False
		try:
			dbfile = cp.getboolean(_MISC_SECTION, "temp_dbfile")
		except:
			err = True
			warning("Invalid argument to the 'temp_dbfile' configuration option", kwargs={})
		if not err:
			temp_dbfile = dbfile
	if cp.has_option(_MISC_SECTION, "editor"):
		global editor
		err = False
		try:
			ed = cp.get(_MISC_SECTION, "editor")
		except:
			err = True
			warning("Invalid argument to the 'editor' configuration option", kwargs={})
		if not err:
			editor = ed



def import_symbol(symbol_name, filename):
	with open(filename, mode='r') as f:
		symbol_def = f.read()
	icon_xbm[symbol_name] = symbol_def




def clparser():
	desc_msg = "Display an interactive map with points read from a data file or database. Version %s, %s" % (version, vdate)
	parser = argparse.ArgumentParser(description=desc_msg)
	parser.add_argument('-f', '--file', default=None,
			help="The name of a CSV or spreadsheet file containing latitude and longitude coordinates")
	parser.add_argument('-m', '--message',
			dest='message', default='Map display.',
			help='A message to display above the map')
	parser.add_argument('-t', '--sheet', default=None,
			help="The name of a worksheet when the data source is a spreadsheet")
	parser.add_argument('-i', '--identifier', default='location_id', dest='id',
			help="The name of the column in the data file containing location identifiers or labels (default: location_id)")
	parser.add_argument('-x', '--lon', default='x_coord', dest='lon',
			help="The name of the column in the data file containing longitude values (default: x_coord)")
	parser.add_argument('-y', '--lat', default='y_coord', dest='lat',
			help="The name of the column in the data file containg latitude values (default: y_coord)")
	parser.add_argument('-s', '--symbol', default=None, dest='symbol',
			help="The name of the column in the data file containing symbol names")
	parser.add_argument('-c', '--color', default=None, dest='color',
			help="The name of the column in the data file containing color names")
	parser.add_argument('-p', '--projection', default=4326,
			help="The coordinate reference system (CRS) if the data are projected (default: 4326, i.e., no projection)")
	parser.add_argument('-g', '--image', dest='imagefile', default=None,
			help="The name of an image file to which the map will be exported--no UI will be created.")
	parser.add_argument('-w', '--imagewait', default=12,
			help="The time in seconds to wait before exporting the map to an image file.")
	return parser



def main():
	args = clparser().parse_args()
	if args.file is None or args.lat is None or args.lon is None:
		fn = lat_col = lon_col = id_col = sym_col = col_col = crs = sheet = msg = headers = rows = imagefile = None
		imagewait = 12
	else:
		fn = args.file
		if not os.path.exists(fn):
			win = tk.Tk()
			win.withdraw()
			fatal_error("File %s does not exist" % fn, kwargs={})
		sheet = args.sheet
		lat_col = args.lat
		lon_col = args.lon
		id_col = args.id
		sym_col = args.symbol
		col_col = args.color
		crs = args.projection
		msg = args.message
		imagefile = args.imagefile
		imagewait = args.imagewait
	read_all_config(fn)
	app = MapUI(fn, msg, lat_col, lon_col, crs, sheet, id_col, sym_col, col_col, map_export_file=imagefile,
			export_time_sec=imagewait)
	app.win.mainloop()


main()


